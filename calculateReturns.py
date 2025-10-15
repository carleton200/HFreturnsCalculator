import sys
import os
import json
import subprocess
import traceback
import sqlite3
import requests
import calendar
import warnings
import pandas as pd
import time
import copy
import re
import pyxirr
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, wait
import queue
import threading
import logging, functools
from dateutil.relativedelta import relativedelta
from multiprocessing import Pool, freeze_support, Manager
from PyQt5.QtWidgets import (
    QApplication, QWidget, QStackedWidget, QVBoxLayout,
    QLabel, QLineEdit, QPushButton, QFormLayout,
    QRadioButton, QButtonGroup, QComboBox, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QProgressBar, QTableView, QCheckBox, QMessageBox,
    QScrollArea, QFileDialog, QListWidgetItem, QListWidget, QDialog, QSizePolicy, QGridLayout,
    QFrame, QTextEdit, QHeaderView, QDateEdit, QSplitter
)
from PyQt5.QtGui import QBrush, QColor, QDesktopServices
from PyQt5.QtCore import Qt, QTimer, QAbstractTableModel, QModelIndex, pyqtSignal, QPoint, QUrl, QDate

currentVersion = "1.1.6"
demoMode = True
ownershipCorrect = True
importInterval = relativedelta(hours=2)

executor = ThreadPoolExecutor()
APIexecutor = ThreadPoolExecutor(max_workers=5) #limits overcalling
gui_queue = queue.Queue()

def poll_queue():
    try:
        while True:
            callback = gui_queue.get_nowait()
            if callback:
                try:
                    callback()  # Run the GUI update in the main thread
                except Exception as e:
                    trace = traceback.format_exc()
                    print(f"Error occured while attempting to run background gui update: {e}. \n traceback: \n {trace}")
    except queue.Empty:
        pass

# Determine assets path, works in PyInstaller bundle or script
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
    ASSETS_DIR = os.path.join(BASE_DIR, '_internal','assets')
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    ASSETS_DIR = os.path.join(BASE_DIR, 'assets')
if not os.path.exists(BASE_DIR):
    os.makedirs(BASE_DIR)

DATABASE_PATH = os.path.join(ASSETS_DIR, 'Acc_Tran.db')
TRAN_DATABASE_PATH = os.path.join(ASSETS_DIR, 'tranCalc.db')
HELP_PATH = os.path.join(ASSETS_DIR,"helpInfo.txt")



logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    filename=ASSETS_DIR + "/systemLogs.log",
    filemode="a"
)
def log_exceptions(method):
    @functools.wraps(method)
    def wrapper(*args, **kwargs):
        try:
            return method(*args, **kwargs)
        except Exception as e:
            logging.exception(f"Error in {method.__qualname__}: {e}")
            raise  # Re-raise the exception after logging
    return wrapper
def attach_logging_to_class(cls):
    for attr_name, attr_value in cls.__dict__.items():
        if callable(attr_value):  # Only wrap methods
            setattr(cls, attr_name, log_exceptions(attr_value))
    return cls
if getattr(sys, 'frozen', False): #Executables will log print statements
    class PrintToLogger:
        def write(self, msg):
            msg = msg.strip()
            if msg:
                logging.info(msg)

        def flush(self):
            pass

    sys.stdout = PrintToLogger()

dynamoAPIenvName = "Dynamo_API"
mainURL = "https://api.dynamosoftware.com/api/v2.2"

nameHier = {
                "Family Branch" : {"api" : "Parent investor", "dynHigh" : "Parentinvestor", "local" : "Family Branch"},
                "Unfunded" : {"api" : "Remaining commitment change", "dynLow" : "RemainingCommitmentChange", "local" : "Unfunded", "value" : "CashFlowSys"},
                "Commitment" : {"api" : "Amount" , "dynLow" : "ValueInSystemCurrency", "local" : "Commitment"},
                "Transaction Time" : {"dynLow" : "TransactionTiming"},
                "sleeve" : {"sleeve" : "sleeve", "fund" : "Name", "local" : "subAssetSleeve"},
                "CashFlow" : {"dynLow" : "CashFlowSys", "dynHigh" : "CashFlowSys"}, 
                "Value" : {"local" : "NAV", "api" : "Value in system currency", "dynLow" : "ValueInSystemCurrency", "dynHigh" : "ValueInSystemCurrency"},
                "Classification" : {"local" : "Classification" , "dynLow" : "Target nameExposureHFClassificationLevel2"},
                "FundClass" : {"dynLow" : "Fundclass" , "dynHigh" : "Fundclass"},
            }

#TODO: make this database stored variable later
assetClass1Order = ["Illiquid", "Liquid","Cash"]
assetClass2Order = ["Direct Private Equity", "Private Equity", "Direct Real Assets", "Real Assets", "Public Equity", "Long/Short", "Absolute Return", "Fixed Income", "Cash"] 
commitmentChangeTransactionTypes = ["Commitment", "Transfer of commitment", "Transfer of commitment (out)", "Secondary - Original commitment (by secondary seller)"]
ignoreInvTranTypes = [""]
headerOptions = ["Return","NAV", "Monthly Gain", "Ownership" , "MDdenominator", "Commitment", "Unfunded", "IRR ITD", "%"]
dataOptions = ["Investor","Family Branch","Classification", "dateTime"]
tranAppHeaderOptions = ["Transaction Sum"]
tranAppDataOptions = ["Investor","Family Branch", "dateTime"]
assetLevelLinks = {1: {"Display" : "Asset Level 1", "Link" : "assetClass"}, 2: {"Display" : "Asset Level 2", "Link" : "subAssetClass"}, 3: {"Display" : "Asset Level 3", "Link" : "subAssetSleeve"}}
displayLinks = {"assetClass" : "Asset Level 1", "subAssetClass" : "Asset Level 2" , "subAssetSleeve" : "Asset Level 3"}
for link in displayLinks.copy(): #builds out in reverse so it can work both ways
    displayLinks[displayLinks.get(link)] = link
balanceTypePriority = ["Actual", "Adjusted", "Manager Estimate"]
yearOptions = (1,2,3,5,7,10,12,15,20)

timeOptions = ["MTD","QTD","YTD", "ITD", "IRR ITD"] + [f"{y}YR" for y in yearOptions]
percent_headers = {option for option in timeOptions}
for header in ("Return","Ownership"):
    percent_headers.add(header)

calculationPingTime = 2
#-------------------------------- Common functions --------------------------------
warnings.simplefilter("error",RuntimeWarning)
def calculate_xirr(cash_flows, dates, guess : float = None):
    try:
        if cash_flows[-1] == 0:
            #indicates closed fund. Remove the NAV as the cashflows should show the fund emptying
            if len(cash_flows) > 2 and cash_flows[-2] != 0:
                cash_flows = cash_flows[:-1]
                dates = dates[:-1]
            else:
                return None #if only two cashflows, it is just a singular investment
        if not( any(cf > 0 for cf in cash_flows) and any(cf < 0 for cf in cash_flows)):
            return None #indicates no returns yet or no investments
        result = pyxirr.xirr(dates, cash_flows)
        if result:
            return result * 100
        else:
            return None
    except pyxirr.InvalidPaymentsError as e:
        print(f"Skipping XIRR calculation due to InvalidPaymentsError: {e} \n Cash flows: {cash_flows} \n Dates: {dates}")
        return None
    except RuntimeWarning as e:
        #print(f"Skipping XIRR calculation due to RuntimeWarning: {e}")
        return None
    except Exception as e:
        print(f"Skipping XIRR calculation due to Exception: {e} \n Cash flows: {cash_flows} \n Dates: {dates}")
        return None
def descendingNavSort(input : dict):
    return sorted(input.keys(), key=lambda x: float(input.get(x,0.0)) * -1)

#-----------------------------Database Manager ---------------------------#
class DatabaseManager:
    """Thread-safe SQLite database manager.

    Uses a single connection with check_same_thread=False and an RLock to
    serialize access. Suitable for simple concurrent usage via a thread pool.
    """

    def __init__(self, db_path: str) -> None:
        self.db_path = db_path
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        self._lock = threading.RLock()
        self._conn = sqlite3.connect(self.db_path, check_same_thread=False)
        self._conn.execute("PRAGMA journal_mode=WAL;")
        self._conn.execute("PRAGMA foreign_keys=ON;")
        self._conn.commit()
        self.instantiateTables()


    def instantiateTables(self) -> None:
        """Instantiate the tables in the database."""
        with self._lock:
            self._conn.execute("""
                CREATE TABLE IF NOT EXISTS calculations (
                    dateTime TEXT,
                    Investor TEXT,
                    Pool TEXT,
                    Fund TEXT,
                    assetClass TEXT,
                    subAssetClass TEXT,
                    NAV TEXT,
                    MonthlyGain TEXT,
                    Return TEXT,
                    MDdenominator TEXT,
                    Ownership TEXT,
                    Commitment TEXT,
                    Unfunded TEXT,
                    IRRITD TEXT,
                    Classification TEXT,
                    Return REAL,
                    NAV REAL,
                    MonthlyGain REAL,
                    Ownership REAL,
                    MDdenominator REAL,
                    Commitment REAL,
                    Unfunded REAL,
                    IRRITD REAL,
                    PRIMARY KEY (dateTime, Investor, FamilyBranch, Classification)
                )
            """)
            self._conn.execute("""
                CREATE TABLE IF NOT EXISTS options (
                    grouping TEXT,
                    id TEXT,
                    value TEXT,
                    PRIMARY KEY (grouping, id)
                )
                """)
            self._conn.execute("""
                CREATE TABLE IF NOT EXISTS benchmarkLinks (
                    benchmark TEXT,
                    asset TEXT,
                    assetLevel INTEGER
                )
                """)
            self._conn.commit()
            
    def fetchOptions(self, grouping : str, update: bool = False):
        if not hasattr(self, "options"):
            self.options = {}
        if not hasattr(self.options, grouping) or update:
            with self._lock:
                cursor = self._conn.cursor()
                cursor.execute("SELECT * FROM options WHERE grouping = ?", (grouping,))
                headers = [d[0] for d in cursor.description]
                options = [dict(zip(headers, row)) for row in cursor.fetchall()]
                self.options[grouping] = {row["id"] : row["value"] for row in options}
                cursor.close()
        return self.options[grouping]
    def saveAsset3Visibility(self, visibility : list):
        with self._lock:
            cursor = self._conn.cursor()
            cursor.execute("DELETE FROM options WHERE grouping = ?", ("asset3Visibility",))
            for vis in visibility:
                cursor.execute("INSERT INTO options (grouping, id, value) VALUES (?, ?, ?)", ("asset3Visibility", vis, "hide"))
            self._conn.commit()
            cursor.close()
        self.options["asset3Visibility"] = {vis : "hide" for vis in visibility}
        logging.info(f"Saved asset3Visibility: {visibility}")
        print(f"Saved asset3Visibility: {visibility}")
    def fetchBenchmarkLinks(self, update: bool = False):
        if not hasattr(self, "benchmarkLinks") or update:
            with self._lock:
                cursor = self._conn.cursor()
                cursor.execute("SELECT * FROM benchmarkLinks")
                headers = [d[0] for d in cursor.description]
                self.benchmarkLinks = [dict(zip(headers, row)) for row in cursor.fetchall()]
                cursor.close()
        return self.benchmarkLinks
    def fetchBenchmarks(self, update: bool = False):
        if not hasattr(self, "benchmarks") or update:
            with self._lock:
                cursor = self._conn.cursor()
                cursor.execute("SELECT DISTINCT [Index] FROM benchmarks")
                self.benchmarks = [row[0] for row in cursor.fetchall()]
                cursor.close()
        return self.benchmarks
    def close(self) -> None:
        try:
            with self._lock:
                self._conn.close()
        except Exception:
            pass

@attach_logging_to_class
class returnsApp(QWidget):
    def __init__(self, start_index=0):
        super().__init__()
        self.setWindowTitle('Returns Calculator')
        self.setGeometry(100, 100, 1000, 600)

        os.makedirs(ASSETS_DIR, exist_ok=True)
        self.db = DatabaseManager(DATABASE_PATH)
        self.start_index = start_index
        self.api_key = None
        self.filterCallLock = False
        self.cancel = False
        self.lock = threading.Lock()
        self.db._lock = self.lock #attach the lock to the database manager
        self.tableWindows = {}
        self.dataTimeStart = datetime(2000,1,1)
        self.earliestChangeDate = datetime(datetime.now().year,datetime.now().month + 1,datetime.now().day)
        self.poolChangeDates = {"active" : False}
        self.currentTableData = None
        self.fullLevelOptions = {}
        self.buildTableCancel = None
        self.buildTableFuture = None
        self.cFundsCalculated = False
        self.previousGrouping = set()

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_from_queue)
        self.queue = []

        # main stack
        self.main_layout = QVBoxLayout()
        self.appStyle = ("""
                        QWidget#borderFrame {
                            border: 2px solid #3E85E9;
                            border-radius: 4px;
                            padding: 4px;
                        }
                        QWidget#titleBox {
                            border: 4px solid #0665EA;
                            border-radius: 5px;
                            padding: 4px;
                        }
                        QWidget#mainPage, QMessageBox, QDialog {
                            background-color: #383838
                        }
                        QLineEdit[status="disabled"]{
                            background-color: #383838;
                            color : #383838;
                        }
                        QPushButton {
                            background-color: #3E85E9;
                            border: 2px solid transparent;
                            border-radius: 12px;
                            padding: 4px
                        }
                        QPushButton:hover {
                                background-color: #1771EE;
                        }
                        QPushButton#exportBtn {
                            background-color: #51AE2B;
                        }
                        QPushButton#exportBtn:hover {
                            background-color: #429321;
                        }
                        QPushButton#cancelBtn {
                            background-color: #D63131;
                        }
                        QPushButton#helpBtn {
                            background-color: #FFDE59;
                            color: black;
                        }
                        QLabel, QRadioButton, QCheckBox, QProgressBar {
                            color: white
                        }
                        QTableWidget, QWidget#subPanel, QHeaderView::corner, QTableCornerButton::section {
                        background-color : #514F4F
                        }
                        QHeaderView::section {
                            background-color: #A8A2A2;
                        }
                        QListWidget {
                            background-color : #514F4F;
                            color: white
                        }
                        QLineEdit{
                            border: 2px solid transparent;
                            border-radius: 12px;
                            background-color: #514F4F;
                            color : white;
                        }
                        QComboBox {
                            background-color: #514F4F;
                            color : white;
                        }
                    """)
        self.setStyleSheet(self.appStyle)
        self.setObjectName("mainPage")
        self.checkVersion()
        self.stack = QStackedWidget()
        self.init_global_widgets()

        self.init_api_key_page() #0
        self.init_returns_page() #1
        self.init_calculation_page() #2

        self.stack.setCurrentIndex(start_index)
        self.main_layout.addWidget(self.stack)
        self.setLayout(self.main_layout)
    def init_global_widgets(self):
        headerBox = QWidget()
        headerLayout = QHBoxLayout()
        self.lastImportLabel = QLabel("Last Data Import: ")
        headerLayout.addWidget(self.lastImportLabel)
        if ownershipCorrect:
            headerLayout.addStretch()
            headerLayout.addWidget(QLabel("                              Notice: \n Investor ownership has automatic corrections due to Dynamo issues"))
        headerLayout.addStretch()
        headerLayout.addWidget(QLabel(f"Version: {currentVersion}"))
        self.helpBtn = QPushButton("Help")
        self.helpBtn.clicked.connect(self.helpClicked)
        self.helpBtn.setObjectName("helpBtn")
        headerLayout.addWidget(self.helpBtn)
        headerBox.setLayout(headerLayout)
        self.main_layout.addWidget(headerBox)
        self.apiLoadingBarBox = QWidget()
        t2 = QVBoxLayout()
        t2.addWidget(QLabel("Pulling transaction and account data from server..."))
        self.apiLoadingBar = QProgressBar()
        self.apiLoadingBar.setRange(0,100)
        t2.addWidget(self.apiLoadingBar)
        self.apiLoadingBarBox.setLayout(t2)
        self.apiLoadingBarBox.setVisible(False)
        self.main_layout.addWidget(self.apiLoadingBarBox)
        loadLay = QGridLayout()
        self.calculationLoadingBar = QProgressBar()
        self.calculationLoadingBar.setRange(0,100)
        self.calculationLabel = QLabel()
        self.cancelCalcBtn = QPushButton("Cancel Calculations")
        self.cancelCalcBtn.setObjectName("cancelBtn")
        self.cancelCalcBtn.setEnabled(False)
        self.cancelCalcBtn.clicked.connect(self.cancelCalc)
        loadLay.addWidget(self.calculationLabel,0,0,1,5)
        loadLay.addWidget(self.calculationLoadingBar, 1,0, 1,5)
        loadLay.addWidget(self.cancelCalcBtn, 2, 2)
        self.calculationLoadingBox = QWidget()
        self.calculationLoadingBox.setLayout(loadLay)
        self.calculationLoadingBox.setVisible(False)
        self.main_layout.addWidget(self.calculationLoadingBox)

    def init_api_key_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        self.api_label = QLabel('Enter Dynamo API Key:')
        self.api_input = QLineEdit()
        btn = QPushButton('Submit')
        btn.clicked.connect(self.check_api_key)
        layout.addWidget(self.api_label)
        layout.addWidget(self.api_input)
        layout.addWidget(btn)
        page.setLayout(layout)
        self.stack.addWidget(page)

    def init_calculation_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        self.info_label = QLabel('Results')
        layout.addWidget(self.info_label)
        btn_to_form = QPushButton('Go to Results')
        exportBtn = QPushButton('Export Data')
        exportBtn.setObjectName("exportBtn")
        exportBtn.clicked.connect(self.exportCalculations)
        layout.addWidget(exportBtn)
        btn_to_form.clicked.connect(lambda: self.stack.setCurrentIndex(1))
        layout.addWidget(btn_to_form)
        

        hl = QHBoxLayout()
        self.calculationTable = QTableView(); self.calculationTable.setSortingEnabled(True)
        hl.addWidget(self.calculationTable)
        layout.addLayout(hl)

        

        page.setLayout(layout)
        self.stack.addWidget(page)

    def init_returns_page(self):
        
        page = QWidget()
        layout = QVBoxLayout()
        controlsBox = QWidget()
        controlsLayout = QHBoxLayout()
        controlsLayout.addStretch(1)
        self.importButton = QPushButton('Reimport Data')
        self.importButton.clicked.connect(self.beginImport)
        clearButton = QPushButton('Full Recalculation')
        clearButton.clicked.connect(self.resetData)
        if not demoMode:
            controlsLayout.addWidget(clearButton, stretch=0)
        controlsLayout.addWidget(self.importButton, stretch=0)
        btn_to_results = QPushButton('See Calculation Database')
        btn_to_results.clicked.connect(self.show_results)
        controlsLayout.addWidget(btn_to_results, stretch=0)
        tranAppBtn = QPushButton('Transaction App')
        tranAppBtn.clicked.connect(self.openTranApp)
        controlsLayout.addWidget(tranAppBtn, stretch=0)
        self.exportBtn = QPushButton("Export Current Table to Excel")
        self.exportBtn.clicked.connect(self.exportCurrentTable)
        self.exportBtn.setObjectName("exportBtn")
        controlsLayout.addWidget(self.exportBtn, stretch=0)
        controlsLayout.addStretch(1)
        controlsBox.setLayout(controlsLayout)
        layout.addWidget(controlsBox)

        optionsBox = QWidget()
        optionsBox.setObjectName("borderFrame")
        optionsGrid = QGridLayout()
        optionsTitle = QLabel("Options")
        optionsTitle.setObjectName("titleBox")
        optionsGrid.addWidget(optionsTitle,0,0,4,1)
        self.tableBtnGroup = QButtonGroup()
        self.complexTableBtn = QRadioButton("Complex Table")
        self.monthlyTableBtn = QRadioButton("Monthly Table")
        buttonBox = QWidget()
        buttonLayout = QVBoxLayout()
        for idx, rb in enumerate((self.complexTableBtn,self.monthlyTableBtn)):
            self.tableBtnGroup.addButton(rb)
            #rb.toggled.connect(self.updateTableType)
            buttonLayout.addWidget(rb)
        self.returnOutputType = QComboBox()
        self.returnOutputType.addItems(headerOptions)
        self.returnOutputType.currentTextChanged.connect(self.buildReturnTable)
        self.dataTypeBox = QWidget()
        dataTypeLayout = QHBoxLayout()
        dataTypeLayout.addWidget(QLabel("Data type:"))
        dataTypeLayout.addWidget(self.returnOutputType)
        self.dataTypeBox.setLayout(dataTypeLayout)
        buttonLayout.addWidget(self.dataTypeBox)
        buttonBox.setLayout(buttonLayout)
        optionsGrid.addWidget(buttonBox, 0,1,2,1)
        self.tableBtnGroup.buttonClicked.connect(self.buildReturnTable)
        self.complexTableBtn.setChecked(True)
        
        self.dataStartSelect = simpleMonthSelector()
        self.dataEndSelect = simpleMonthSelector()
        for idx, [text, CB] in enumerate((["Start: ", self.dataStartSelect], ["End: ", self.dataEndSelect])):
            optionsGrid.addWidget(QLabel(text),idx,2)
            optionsGrid.addWidget(CB,idx,3)
        optionsGrid.addWidget(QLabel("Benchmarks:"),0,4)
        self.benchmarkSelection = MultiSelectBox()
        self.benchmarkSelection.popup.closed.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.benchmarkSelection,1,4)
        optionsGrid.addWidget(QLabel("Group by: "),0,5)
        self.sortHierarchy = MultiSelectBox()
        self.sortHierarchy.hierarchyMode()
        self.sortHierarchy.setCheckedItems(["assetClass","subAssetClass"])
        self.sortHierarchy.popup.closed.connect(self.groupingChange)
        optionsGrid.addWidget(self.sortHierarchy,1,5)
        self.consolidateFundsBtn = QCheckBox("Consolidate Funds")
        self.consolidateFundsBtn.setChecked(True)
        self.consolidateFundsBtn.clicked.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.consolidateFundsBtn,0,6)
        self.exitedFundsBtn = QCheckBox("Show Exited Funds (Cannot turn off)")
        self.exitedFundsBtn.setChecked(False)
        self.exitedFundsBtn.setEnabled(False) #remove later
        self.exitedFundsBtn.setChecked(True)  #remove later
        optionsGrid.addWidget(self.exitedFundsBtn,1,6)
        self.headerSort = SortButtonWidget()
        self.headerSort.popup.popup_closed.connect(self.headerSortClosed)
        optionsGrid.addWidget(self.headerSort,0,7)
        self.sortStyle = QPushButton("Sort Style: NAV")
        self.sortStyle.clicked.connect(self.sortStyleClicked)
        optionsGrid.addWidget(self.sortStyle,1,7)
        # Add a horizontal line across the optionsGrid after the top row of options
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        optionsGrid.addWidget(line, 2, 1, 1, optionsGrid.columnCount() - 1)
        self.assetClass3Visibility = MultiSelectBox()
        self.assetClass3Visibility.popup.closed.connect(self.assetClass3VisibilityChanged)
        optionsGrid.addWidget(QLabel("Hidden Asset Level 3s:"),3,1)
        optionsGrid.addWidget(self.assetClass3Visibility,4,1)
        self.showBenchmarkLinksBtn = QCheckBox("Show Benchmark Links")
        self.showBenchmarkLinksBtn.setChecked(True)
        self.showBenchmarkLinksBtn.clicked.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.showBenchmarkLinksBtn,3,2)
        self.linkBenchmarksBtn = QPushButton("Link Benchmarks")
        self.linkBenchmarksBtn.clicked.connect(self.linkBenchmarks)
        optionsGrid.addWidget(self.linkBenchmarksBtn,4,2)
        optionsBox.setLayout(optionsGrid)
        layout.addWidget(optionsBox)

        mainFilterBox = QWidget()
        mainFilterBox.setObjectName("borderFrame")
        mainFilterLayout = QGridLayout()
        filterTitle = QLabel("Filters")
        filterTitle.setObjectName("titleBox")
        mainFilterLayout.addWidget(filterTitle,0,0,2,1)

        self.filterOptions = [
                            {"key": "Classification", "name": "HF Classification", "dataType" : None, "dynNameLow" : "Target nameExposureHFClassificationLevel2"},
                            {"key" : nameHier["Family Branch"]["local"], "name" : nameHier["Family Branch"]["local"], "dataType" : None, "dynNameLow" : None, "dynNameHigh" : nameHier["Family Branch"]["dynHigh"]},
                            {"key": "Investor",       "name": "Investor", "dataType" : "Investor", "dynNameLow" : None, "dynNameHigh" : "Source name"},
                            {"key": "assetClass",     "name": "Asset Level 1", "dataType" : "Total Asset", "dynNameLow" : "ExposureAssetClass", "dynNameHigh" : "ExposureAssetClass"},
                            {"key": "subAssetClass",  "name": "Asset Level 2", "dataType" : "Total subAsset", "dynNameLow" : "ExposureAssetClassSub-assetClass(E)", "dynNameHigh" : "ExposureAssetClassSub-assetClass(E)"},
                            {"key" : nameHier["sleeve"]["local"], "name" : "Asset Level 3", "dataType" : "Total sleeve", "dynNameLow" : nameHier["sleeve"]["local"]},
                            {"key": "Pool",           "name": "Pool", "dataType" : "Total Pool" , "dynNameLow" : "Source name", "dynNameHigh" : "Target name"},
                            {"key": "Fund",           "name": "Fund/Investment", "dataType" : "Total Fund" , "dynNameLow" : "Target name"},
                            
                        ]
        self.filterBtnExclusions = ["Investor","Classification", nameHier["Family Branch"]["local"]]
        self.highOnlyFilters = ["Investor", nameHier["Family Branch"]["local"]]
        self.filterDict = {}
        self.filterRadioBtnDict = {}
        self.filterBtnGroup = QButtonGroup()
        self.filterBtnGroup.setExclusive(False)
        for col, filter in enumerate(self.filterOptions, start=1):
            if filter["key"] not in self.filterBtnExclusions:
                #investor level is not filterable. It is total portfolio or shows the investors data
                self.filterRadioBtnDict[filter["key"]] = QCheckBox(f"{filter["name"]}:")
                self.filterRadioBtnDict[filter["key"]].setChecked(True)
                self.filterBtnGroup.addButton(self.filterRadioBtnDict[filter["key"]])
                mainFilterLayout.addWidget(self.filterRadioBtnDict[filter["key"]],0, col)
            else:
                mainFilterLayout.addWidget(QLabel(f"{filter["name"]}:"), 0, col)
            if filter["key"] != "Fund":
                self.sortHierarchy.addItem(filter["key"])
            self.filterDict[filter["key"]] = MultiSelectBox()
            self.filterDict[filter["key"]].popup.closed.connect(lambda: self.filterUpdate())
            mainFilterLayout.addWidget(self.filterDict[filter["key"]],1,col)
        self.sortHierarchy.setCheckedItems(["assetClass","subAssetClass"])
        self.filterBtnGroup.buttonToggled.connect(self.filterBtnUpdate)
        mainFilterBox.setLayout(mainFilterLayout)
        layout.addWidget(mainFilterBox)
        t1 = QVBoxLayout() #build table loading bar
        self.buildTableLoadingBox = QWidget()
        self.tableLoadingLabel = QLabel("Building returns table...")
        t1.addWidget(self.tableLoadingLabel)
        self.buildTableLoadingBar = QProgressBar()
        self.buildTableLoadingBar.setRange(0,8)
        t1.addWidget(self.buildTableLoadingBar)
        self.buildTableLoadingBox.setLayout(t1)
        self.buildTableLoadingBox.setVisible(False)
        layout.addWidget(self.buildTableLoadingBox)
        self.returnsTable = SmartStretchTable() #table
        self.returnsTable.setSelectionMode(QTableWidget.ContiguousSelection)  # Required
        self.returnsTable.setSelectionBehavior(QTableWidget.SelectItems)
        layout.addWidget(self.returnsTable)
        unDataBox = QWidget()
        unDataLayout = QHBoxLayout()
        unDataLayout.addStretch(1)
        self.viewUnderlyingDataBtn = QPushButton("View Underlying Data")
        self.viewUnderlyingDataBtn.clicked.connect(self.viewUnderlyingData)
        unDataLayout.addWidget(self.viewUnderlyingDataBtn,stretch=0)
        unDataLayout.addStretch(1)
        unDataBox.setLayout(unDataLayout)
        layout.addWidget(unDataBox)
        


        page.setLayout(layout)
        self.stack.addWidget(page)

        self.pullLevelNames()
        self.updateMonthOptions()
        if self.start_index != 0:
            self.filterUpdate()
        self.dataEndSelect.currentTextChanged.connect(self.buildReturnTable)
        self.dataStartSelect.currentTextChanged.connect(self.buildReturnTable)
    def init_data_processing(self):
        self.calcSubmitted = False
        self.lastImportDB = load_from_db("history")
        if len(self.lastImportDB) != 1:
            self.lastImportDB = None
        if self.lastImportDB is None:
            print("No previous import found")
            #pull data is there is no data pulled yet
            executor.submit(lambda: self.pullData())
        else:
            lastImportString = self.lastImportDB[0]["lastImport"]
            lastImport = datetime.strptime(lastImportString, "%B %d, %Y @ %I:%M %p")  
            self.lastImportLabel.setText(f"Last Data Import: {lastImportString}")
            now = datetime.now()
            if lastImport.month != now.month or now > (lastImport + importInterval):
                print(f"Reimporting due to time elapsing. \n     Last import: {lastImport}\n    Current time: {now}")
                #pull data if in a new month or 1 days have elapsed
                executor.submit(self.pullData)
            elif self.lastImportDB[0]["lastImport"] != self.lastImportDB[0].get("lastCalculation", "None"):
                self.earliestChangeDate = datetime.strptime(self.lastImportDB[0].get("changeDate"), "%B %d, %Y @ %I:%M %p")
                self.processFunds()
                self.calculateReturn()
            else:
                calculations = load_from_db("calculations")
                self.processFunds()
                if calculations != []:
                    self.populate(self.calculationTable,calculations)
                    self.buildReturnTable()
                else:
                    self.calculateReturn()
    def watchForUpdateTime(self):
        try:
            print("Checking for update required.")
            lastImportString = self.lastImportDB[0]["lastImport"]
            lastImport = datetime.strptime(lastImportString, "%B %d, %Y @ %I:%M %p")  
            now = datetime.now()
            if lastImport.month != now.month or now > (lastImport + importInterval):
                print(f"Reimporting due to time elapsing. \n     Last import: {lastImport}\n    Current time: {now}")
                #pull data if in a new month or 1 days have elapsed
                executor.submit(self.pullData)
        except:
            print("Background watch failed")

    def helpClicked(self,*_):
        #try:
            with open(HELP_PATH, 'r', encoding='utf-8') as f:
                text = f.read()
            helpMessage = displayWindow(parentSource=self, text=text, title="Help Page")
            helpMessage.show()
            self.helpPage = helpMessage
        # except:
        #     QMessageBox.warning(self,"Error","Error opening help page.")
    def exportCalculations(self,*_):
        window = exportWindow(parentSource=self)
        window.show()
        self.exportWindow = window
    def openTranApp(self,*_):
        tranApp = transactionApp(apiKey=self.api_key)
        tranApp.stack.setCurrentIndex(1)
        tranApp.init_data_processing()
        tranApp.show()
        self.tranApp = tranApp
    def cancelCalc(self, *_):
        _ = updateStatus("DummyFail",99,lock=self.lock, status="Failed")
        self.cancel = True
    def viewUnderlyingData(self,*_):
        row = self.returnsTable.currentRow()
        col = self.returnsTable.currentColumn()
        key = list(self.filteredReturnsTableData.keys())[row]
        vh_item = self.returnsTable.verticalHeaderItem(row)
        row = vh_item.text() if vh_item else f"Row {row}"

        # Get the horizontal (column) header text
        hh_item = self.returnsTable.horizontalHeaderItem(col)
        col = hh_item.text() if hh_item else f"Column {col}"
        self.selectedCell = {"entity": row, "month" : col, "rowKey" : key, "dataType" : self.filteredReturnsTableData[key]["dataType"] }
        try:
            window = underlyingDataWindow(parentSource=self)
            self.udWindow = window
            if window.success:
                window.show()
        except Exception as e:
            print(f"Error in data viewing window: {e} {traceback.format_exc()}")
    def exportCurrentTable(self,*_):
        # helper to darken a 6-digit hex color by a given factor
        def darken_color(hex_color, factor=0.01):
            h = hex_color.strip("#")
            r = int(h[0:2], 16)
            g = int(h[2:4], 16)
            b = int(h[4:6], 16)
            dr = max(0, int(r * factor))
            dg = max(0, int(g * factor))
            db = max(0, int(b * factor))
            return f"{dr:02X}{dg:02X}{db:02X}"
        # 1) prompt user
        path, _ = QFileDialog.getSaveFileName(
            self, "Save as…", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        def processExport():
            try:
                data = self.filteredReturnsTableData

                # 2) determine hierarchy levels present
                all_types = {row.get("dataType") for row in data.values()}
                if self.sortHierarchy.checkedItems() != []:
                    full_hierarchy = ["Total"] + ["Total " + level for level in self.sortHierarchy.checkedItems()] + ["Total Fund"]
                else:
                    full_hierarchy = ["Total", "Total assetClass", "Total Fund"]
                hierarchy_levels = [lvl for lvl in full_hierarchy if lvl in all_types]
                num_hier = 1

                # 3) dynamic data columns minus "dataType"
                all_cols = list(self.filteredHeaders)

                sorted_cols = self.orderColumns(all_cols)                

                # 4) create workbook or add sheet if already exists
                if os.path.exists(path):
                    wb = load_workbook(path)
                    # Create a unique sheet name for export
                    base_name = "Export"
                    i = 1
                    while True:
                        sheet_name = f"{base_name}{i}"
                        if sheet_name not in wb.sheetnames:
                            break
                        i += 1
                    ws = wb.create_sheet(sheet_name)
                else:
                    wb = Workbook()
                    ws = wb.active

                rowStart = 3
                # 5) header row
                for idx, colname in enumerate(sorted_cols, start=num_hier+1):
                    ws.cell(row=rowStart, column=idx, value=colname)

                split_cell = f"{get_column_letter(num_hier+1)}4"
                ws.freeze_panes = split_cell

                # 7) populate rows
                for r, (row_name, row_dict) in enumerate(data.items(), start=rowStart + 1):
                    row_name, code = self.separateRowCode(row_name)
                    dtype = row_dict.get("dataType")
                    if dtype != "benchmark": #keeps benchmark as the previous hierarchy level
                        level = hierarchy_levels.index(dtype) if dtype in hierarchy_levels else 0
                        # fills
                        data_color = "FFFFFF"
                        if dtype != "Total Fund":
                            depth      = code.count("::") if dtype != "Total" else code.count("::") - 1
                            maxDepth   = max(len(self.sortHierarchy.checkedItems()),1) + 1
                            data_color = darken_color(data_color,depth/maxDepth/3 + 2/3)

                        if r % 2 == 1:
                            data_color = darken_color(data_color,0.93)
                        header_color = darken_color(data_color, 0.9)
                        data_fill   = PatternFill("solid", data_color, data_color)
                        header_fill = PatternFill("solid", header_color, header_color)

                    data_start = 2
                    # spread header fill across hierarchy cols
                    cell = ws.cell(row=r, column=1, value=row_name)
                    cell.fill = header_fill
                    if dtype == "benchmark":
                        cell.font = Font(color="0000FF")
                    cell.alignment = Alignment(indent=level)

                    # data cells with proper formatting
                    for c, colname in enumerate(sorted_cols, start=data_start):
                        val = row_dict.get(colname, None)
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.fill = data_fill
                        if dtype == "benchmark":
                            cell.font = Font(color="0000FF")
                        if isinstance(val, (int, float)):
                            if colname not in percent_headers:
                                # show with commas, two decimals
                                cell.number_format = "#,##0.00"
                            else:
                                # interpret val as percentage (e.g. 10.5 → 10.5%)
                                cell.value = val / 100.0
                                cell.number_format = "0.00%"

                # 8) autofit column widths
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is not None:
                            text = str(cell.value)
                            max_len = max(max_len, len(text))
                    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

                filterSelections = {}
                for filter in self.filterOptions:
                    selections = self.filterDict.get(filter.get("key")).checkedItems()
                    if selections != []:
                        filterSelections[filter.get("key")] = ", ".join(selections)
                if filterSelections:
                    cell = ws.cell(row=1,column=1,value="Filters:")
                    cell.font = Font(bold=True)
                    cell = ws.cell(row=2,column=1,value="Selections:")
                    cell.font = Font(bold=True)
                    for idx, filter in enumerate(filterSelections, start=2):
                        ws.cell(row=1,column=idx, value=filter)
                        cell = ws.cell(row=2,column=idx, value=filterSelections.get(filter))
                        cell.alignment = Alignment(wrap_text=True)

            
                wb.save(path)
            except Exception as e:
                gui_queue.put(lambda error=e, trace = traceback.format_exc(): QMessageBox.critical(self, "Save error", trace))
            else:
                gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
                gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
        executor.submit(processExport)
    def processFunds(self):
        self.cFundsCalculated = True
        self.sleeveFundLinks = {}
        self.cFundToFundLinks = {}
        self.pools = []
        sleeves = set()
        funds = load_from_db("funds")
        if funds != []:
            consolidatorFunds = {}
            for row in funds: #find sleeve values and consolidated funds
                assetClass = row["assetClass"]
                subAssetClass = row["subAssetClass"]
                sleeve = row["sleeve"]
                sleeves.add(sleeve)
                if row.get("Fundpipelinestatus") is not None and "Z - Placeholder" in row.get("Fundpipelinestatus"):
                    consolidatorFunds[row["Name"]] = {"cFund" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass, "sleeve" : sleeve}
                    self.cFundToFundLinks[row["Name"]] = []
                if row["sleeve"] not in self.sleeveFundLinks:
                    self.sleeveFundLinks[row["sleeve"]] = [row["Name"]]
                else:
                    self.sleeveFundLinks[row["sleeve"]].append(row["Name"])
                if row["Fundpipelinestatus"] == "I - Internal":
                    self.pools.append({"poolName" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass})
            self.consolidatedFunds = {}
            for row in funds: #assign funds to their consolidators
                if row.get("Parentfund") in consolidatorFunds:
                    self.consolidatedFunds[row["Name"]] = consolidatorFunds.get(row.get("Parentfund"))
                    self.cFundToFundLinks[row.get("Parentfund")].append(row["Name"])
            self.fullLevelOptions["subAssetSleeve"] = list(sleeves)
        else:
            self.consolidatedFunds = {}
    def filterBtnUpdate(self, button, checked):
        if not self.filterCallLock:
            self.buildTableLoadingBox.setVisible(True)
            self.buildTableLoadingBar.setValue(1)
            self.filterCallLock = True
            reloadRequired = False
            for filter in self.filterOptions:
                if filter["key"] not in self.filterBtnExclusions:
                    if not self.filterRadioBtnDict[filter["key"]].isChecked():
                        if self.filterDict[filter["key"]].checkedItems() != []:
                            reloadRequired = True #rebuild the table only if filter selections are being removed
                        self.filterDict[filter["key"]].clearSelection()
                        self.filterDict[filter["key"]].setEnabled(False)
                        self.filterDict[filter["key"]].line_edit.setProperty("status","disabled")
                        self.filterDict[filter["key"]].line_edit.style().unpolish(self.filterDict[filter["key"]].line_edit)
                        self.filterDict[filter["key"]].line_edit.style().polish(self.filterDict[filter["key"]].line_edit)
                    else:
                        self.filterDict[filter["key"]].setEnabled(True)
                        self.filterDict[filter["key"]].line_edit.setProperty("status","enabled")
                        self.filterDict[filter["key"]].line_edit.style().unpolish(self.filterDict[filter["key"]].line_edit)
                        self.filterDict[filter["key"]].line_edit.style().polish(self.filterDict[filter["key"]].line_edit)
            self.filterCallLock = False
            if reloadRequired or self.currentTableData is None:
                self.buildReturnTable()
            else:
                self.populateReturnsTable(self.currentTableData)
    def resetData(self,*_):
        if not self.testAPIconnection():
            QMessageBox.warning(self,"API Failure", "API connection has failed. Server is down or API key is bad. \n Previous calculations are left in place for viewing.")
            return
        for table in ("calculations","positions_low","positions_high","transactions_low","transactions_high"):
            save_to_db(table,None,action="clear") #reset all tables so everything will be fresh data
        self.poolChangeDates = {"active" : False}
        executor.submit(self.pullData)
    def beginImport(self, *_):
        executor.submit(self.pullData)
    def updateMonthOptions(self):
        start = self.dataTimeStart
        end = datetime.now() - relativedelta(months=1) + relativedelta(hours=8)
        #ends on the previous month. Adds a few hours so index will still be before it and count as a month on the 1st
        index = start
        monthList = []
        while index < end:
            monthList.append(datetime.strftime(index,"%B %Y"))
            index += relativedelta(months=1)
        self.dataEndSelect.addItems(monthList)
        self.dataEndSelect.setCurrentText(monthList[-1])
        self.dataStartSelect.addItems(monthList)
        self.dataStartSelect.setCurrentText(monthList[0])
    def linkBenchmarks(self,*_):
        #open the link benchmarks window as its own window that is non blocking
        self.linkBenchmarksWindow = linkBenchmarksWindow(parentSource=self)
        self.linkBenchmarksWindow.show()
    def buildReturnTable(self, *_):
        self.buildTableLoadingBox.setVisible(True)
        self.buildTableLoadingBar.setValue(2)
        if not self.cFundsCalculated:
            self.processFunds()
        def buildTable(cancelEvent):
            try:
                print("Building return table...")
                self.currentTableData = None #resets so a failed build won't be used
                
                if self.tableBtnGroup.checkedButton().text() == "Complex Table":
                    gui_queue.put(lambda: self.returnOutputType.setCurrentText("Return"))
                    gui_queue.put(lambda: self.dataTypeBox.setVisible(False))
                else:
                    gui_queue.put(lambda: self.dataTypeBox.setVisible(True))
                if self.filterDict["Investor"].checkedItems() == [] and self.filterDict[nameHier["Family Branch"]["local"]].checkedItems() == []:
                    #if no investor level selections,show full portfolio information
                    parameters = ["Total Fund"]
                    condStatement = " WHERE [Investor] = ? "
                else:
                    #show investor level fund data
                    condStatement = " WHERE"
                    parameters = []
                    if self.filterDict["Investor"].checkedItems() != []:
                        paramTemp = self.filterDict["Investor"].checkedItems()
                        placeholders = ','.join('?' for _ in paramTemp) 
                        condStatement += f" [Investor] IN ({placeholders}) "
                        for param in paramTemp:
                            parameters.append(param)
                    if self.filterDict[nameHier["Family Branch"]["local"]].checkedItems() != []:
                        paramTemp = self.filterDict[nameHier["Family Branch"]["local"]].checkedItems()
                        placeholders = ','.join('?' for _ in paramTemp)
                        if condStatement == " WHERE":
                            condStatement += f" [{nameHier["Family Branch"]["local"]}] IN ({placeholders}) "
                        else:
                            condStatement += f" AND [{nameHier["Family Branch"]["local"]}] IN ({placeholders}) "
                        for param in paramTemp:
                            parameters.append(param)
                for filter in self.filterOptions:
                    if filter["key"] != "Investor" and filter["key"] != nameHier["Family Branch"]["local"]:
                        if self.filterDict[filter["key"]].checkedItems() != []:
                            paramTemp = self.filterDict[filter["key"]].checkedItems()
                            for param in paramTemp:
                                parameters.append(param)
                            placeholders = ','.join('?' for _ in paramTemp)
                            condStatement += f" AND [{filter["key"]}] IN ({placeholders})"
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(3))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                data = load_from_db("calculations",condStatement, tuple(parameters), lock=self.lock)
                output = {"Total##()##" : {}}
                if self.benchmarkSelection.checkedItems() != [] or self.showBenchmarkLinksBtn.isChecked():
                    output = self.applyBenchmarks(output)
                output , data = self.calculateUpperLevels(output,data)
                for benchmark in self.pendingBenchmarks: #remove the benchmarks used only in benchmark links
                    if benchmark not in self.benchmarkChoices and benchmark + self.buildCode([]) in output.keys():
                        output.pop(benchmark + self.buildCode([]))
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(4))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                complexOutput = copy.deepcopy(output)
                multiPoolFunds = {}
                dataOutputType = self.returnOutputType.currentText()
                for entry in data:
                    if (datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") >  datetime.strptime(self.dataEndSelect.currentText(),"%B %Y") or 
                        datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") <  datetime.strptime(self.dataStartSelect.currentText(),"%B %Y")):
                        #don't build in data outside the selection
                        continue
                    date = datetime.strftime(datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S"), "%B %Y")
                    Dtype = entry["Calculation Type"]
                    level = entry["rowKey"]
                    if dataOutputType == "IRR ITD" and ((self.consolidateFundsBtn.isChecked() and entry.get("Fund") in self.consolidatedFunds) or entry.get("Calculation Type") != "Total Fund"):
                        #this only skips for monthly table
                        continue #cannot use IRR ITD for consolidated funds
                    
                    if level not in output.keys():
                        output[level] = {}
                    if entry.get(dataOutputType) not in (None,"None",""):
                        if date not in output[level].keys():
                            #creates value if not exists. If it is not return percent, sums the values
                            output[level][date] = float(entry.get(dataOutputType))
                        elif dataOutputType not in ("Return", "Ownership"):
                            output[level][date] += float(entry.get(dataOutputType))
                        else: #should only reach here if two calculations exist of the same exact row which needs special handling of the return
                            if level not in multiPoolFunds:
                                multiPoolFunds[level] = [entry,]
                            else:
                                multiPoolFunds[level].append(entry)
                    if "dataType" not in output[level].keys():
                        output[level]["dataType"] = Dtype
                    if self.tableBtnGroup.checkedButton().text() == "Complex Table" and date == self.dataEndSelect.currentText():
                        if level not in complexOutput.keys():
                            complexOutput[level] = {}
                        if "dataType" not in complexOutput[level].keys():
                            complexOutput[level]["dataType"] = Dtype
                        if headerOptions[0] not in complexOutput[level].keys() and headerOptions:
                            for option in headerOptions:
                                if option == "IRR ITD" and ((self.consolidateFundsBtn.isChecked() and entry.get("Fund") in self.consolidatedFunds) or entry.get("Calculation Type") != "Total Fund"):
                                    continue #cannot use IRR ITD for consolidated funds
                                complexOutput[level][option] = float(entry[option] if entry.get(option) not in (None,"None","") else 0)
                        else:
                            for option in headerOptions:
                                if option not in ("Ownership", "IRR ITD"):
                                    #do not sum IRR ITD or ownership in any scenario. Should only occur for consolidation anyways
                                    complexOutput[level][option] += float(entry[option] if entry.get(option) not in (None,"None","") else 0)
                        if entry.get("Ownership") not in (None,"None") and (self.filterDict["Investor"].checkedItems() != [] or self.filterDict["Family Branch"].checkedItems() != []):
                            if "Ownership" not in complexOutput[level].keys():
                                complexOutput[level]["Ownership"] = float(entry["Ownership"])
                            else:
                                complexOutput[level]["Ownership"] += float(entry["Ownership"])
                            # else:
                            #     complexOutput[level]["Ownership"] += float(entry["Ownership"])
                if multiPoolFunds and dataOutputType == "Return": #must iterate through data again to correct for returns of multi pool funds
                    multiData = {}
                    for rowKey in multiPoolFunds: #instantiate multiData with the row
                        multiData[rowKey] = {}
                        # date = multiPoolFunds.get(rowKey).get("dateTime")
                        # multiData[rowKey][date] = {"MDdenominator" : 0, "Monthly Gain" : 0}
                    for entry in data:
                        if (datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") >  datetime.strptime(self.dataEndSelect.currentText(),"%B %Y") or 
                            datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") <  datetime.strptime(self.dataStartSelect.currentText(),"%B %Y")):
                            #don't build in data outside the selection
                            continue
                        if entry.get("rowKey") in multiData: #only occurs for the multifunds
                            #sums all gains and MDden for a row for a month
                            dateTime = entry.get("dateTime")
                            if dateTime not in multiData[entry.get("rowKey")]:
                                multiData[entry.get("rowKey")][entry.get("dateTime")] = {"MDdenominator" : float(entry.get("MDdenominator")), "Monthly Gain" : float(entry.get("Monthly Gain"))}
                            else:
                                multiData[entry.get("rowKey")][entry.get("dateTime")]["MDdenominator"] += float(entry.get("MDdenominator"))
                                multiData[entry.get("rowKey")][entry.get("dateTime")]["Monthly Gain"] += float(entry.get("Monthly Gain"))
                    for rowKey in multiData: #set proper return values
                        for date in multiData.get(rowKey):
                            strDate = datetime.strftime(datetime.strptime(date, "%Y-%m-%d %H:%M:%S"), "%B %Y")
                            MDden = multiData.get(rowKey).get(date).get("MDdenominator")
                            returnVal = multiData.get(rowKey).get(date).get("Monthly Gain") / MDden * 100 if MDden != 0 else 0
                            output[rowKey][strDate] = returnVal
                            if self.tableBtnGroup.checkedButton().text() == "Complex Table" and strDate == self.dataEndSelect.currentText():
                                complexOutput[rowKey]["Return"] = returnVal
                if self.tableBtnGroup.checkedButton().text() == "Complex Table":
                    for rowKey in complexOutput:
                        complexOutput[rowKey]["%"] = (complexOutput[rowKey].get("NAV",0.0) / complexOutput["Total##()##"].get("NAV",0.0) * 100 if complexOutput["Total##()##"]["NAV"] != 0 else 0 )if complexOutput[rowKey].get("NAV",0.0) != 0 else None
                else:
                    for rowKey in output:
                        output[rowKey]["%"] = (output[rowKey]["NAV"] / output["Total##()##"].get("NAV",0.0) * 100 if output["Total##()##"].get("NAV",0.0) != 0 else 0 ) if output[rowKey].get("NAV",0.0) != 0 else None
                
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(5))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                if self.tableBtnGroup.checkedButton().text() == "Complex Table":
                    output = self.calculateComplexTable(output,complexOutput)
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(6))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                outputKeys = output.keys()
                deleteKeys = []
                for key in outputKeys:
                    if len(output[key].keys()) == 0:
                        deleteKeys.append(key)
                for key in deleteKeys:
                    output.pop(key)
                gui_queue.put(lambda: self.populateReturnsTable(output))
                self.currentTableData = output
            except Exception as e:
                tracebackMsg = traceback.format_exc()
                gui_queue.put(lambda error = e: QMessageBox.warning(self, "Error building returns table", f"Error: {error}. {error.args}. Data entry: \n  \n Traceback:  \n {tracebackMsg}"))
                gui_queue.put(lambda: self.buildTableLoadingBox.setVisible(False))
        if self.buildTableCancel:
            self.buildTableCancel.set()
        if self.buildTableFuture and not self.buildTableFuture.done():
            self.buildTableFuture.cancel()

        cancelEvent = threading.Event()
        self.buildTableCancel = cancelEvent
        self.stack.setCurrentIndex(1)
        future = executor.submit(buildTable, cancelEvent)
        self.buildTableFuture = future
    def calculateComplexTable(self,monthOutput,complexOutput):
        endTime = datetime.strptime(self.dataEndSelect.currentText(),"%B %Y")
        MTDtime = datetime.strftime(endTime,"%B %Y")
        QTDtimes = [datetime.strftime(endTime - relativedelta(months=i),"%B %Y") for i in range(int((endTime.month)) % 3 if (int(endTime.month)) % 3 != 0 else 3)]
        YTDtimes = [datetime.strftime(endTime - relativedelta(months=i),"%B %Y") for i in range(int((endTime.month)) % 12 if (int(endTime.month)) % 12 != 0 else 12)]
        YR_times = {}
        for yr in yearOptions:
            YR_times[yr] = [datetime.strftime(endTime - relativedelta(months=i),"%B %Y") for i in range(12 * yr)]
        for level in monthOutput.keys():
            if MTDtime in monthOutput[level].keys():
                complexOutput[level]["MTD"] = monthOutput[level][MTDtime]
            if all(month in monthOutput[level].keys() for month in QTDtimes):
                complexOutput[level]["QTD"] = 1
                for month in QTDtimes:
                    complexOutput[level]["QTD"] *= (1 + float(monthOutput[level][month]) / 100)
                complexOutput[level]["QTD"] = (complexOutput[level]["QTD"] -1) * 100
            if all(month in monthOutput[level].keys() for month in YTDtimes):
                complexOutput[level]["YTD"] = 1
                for month in YTDtimes:
                    complexOutput[level]["YTD"] *= (1 + float(monthOutput[level][month]) / 100)
                complexOutput[level]["YTD"] = (complexOutput[level]["YTD"] -1) * 100
            for yearKey in YR_times.keys():
                if all(month in monthOutput[level].keys() for month in YR_times[yearKey]):
                    headerKey = f"{yearKey}YR"
                    complexOutput[level][headerKey] = 1
                    for month in YR_times[yearKey]:
                        complexOutput[level][headerKey] *= (1 + float(monthOutput[level][month]) / 100 )
                    complexOutput[level][headerKey] = ((complexOutput[level][headerKey] ** (1/int(yearKey)) ) - 1 ) * 100 if complexOutput[level][headerKey] > 0 else -1 * ((abs(complexOutput[level][headerKey]) ** (1/int(yearKey)) ) - 1)* 100
            try:
                if monthOutput[level].get("dataType","") != "benchmark":
                    monthCount = 0
                    if MTDtime in monthOutput[level].keys():
                        #only runs ITD if it is a current fund (MTD month exists)
                        ITDmonths = list(monthOutput[level].keys())
                        ITDmonths = [m for m in ITDmonths if m != "dataType"]
                        ITDmonths = sorted([datetime.strptime(date,"%B %Y") for date in ITDmonths])
                        
                        ITDmonths = [datetime.strftime(date,"%B %Y") for date in ITDmonths]
                        if len(ITDmonths) >= 2:
                            #only calculates if more than previous month
                            #ITDmonths = ITDmonths[1:] #remove first month?? 
                            complexOutput[level]["ITD"] = 1
                            for month in ITDmonths:
                                if month != "dataType" and datetime.strptime(month,"%B %Y") <= datetime.strptime(self.dataEndSelect.currentText(),"%B %Y"):
                                    monthCount += 1
                                    complexOutput[level]["ITD"] *= (1 + float(monthOutput[level][month]) / 100 )
                            complexOutput[level]["ITD"] = ((complexOutput[level]["ITD"] ** (12/int(monthCount)) ) - 1 ) * 100 if complexOutput[level]["ITD"] > 0 else -1 * ((abs(complexOutput[level]["ITD"]) ** (1/int(monthCount)) ) - 1)* 100
                        else:
                            #ITD is just the previous month if no more months are found
                            complexOutput[level]["ITD"] = monthOutput[level][MTDtime]
                else:
                    complexOutput[level]["ITD"] = monthOutput[level]["ITD"]
            except Exception as e:
                pass

        return complexOutput
    def applyBenchmarks(self, output):
        if self.showBenchmarkLinksBtn.isChecked(): #activate the benchmark links so they are all used if relevant
            benchmarkLinks = self.db.fetchBenchmarkLinks()
            self.pendingBenchmarks = set(link.get("benchmark") for link in benchmarkLinks)
        self.benchmarkChoices = self.benchmarkSelection.checkedItems()
        allBenchmarkChoices = set(set(self.benchmarkChoices) | set(self.pendingBenchmarks))
        code = self.buildCode([])
        placeholders = ','.join('?' for _ in allBenchmarkChoices)
        benchmarks = load_from_db("benchmarks",f"WHERE [Index] IN ({placeholders})",tuple(allBenchmarkChoices), lock=self.lock)
        for bench in benchmarks:
            name = bench["Index"] + code
            if (datetime.strptime(bench["Asofdate"], "%Y-%m-%dT%H:%M:%S") < datetime.strptime(self.dataStartSelect.currentText(), "%B %Y") or
                datetime.strptime(bench["Asofdate"], "%Y-%m-%dT%H:%M:%S") > datetime.strptime(self.dataEndSelect.currentText(), "%B %Y") + relativedelta(months=1) ) :
                continue #skip if outside selected range
            date = datetime.strftime(datetime.strptime(bench["Asofdate"], "%Y-%m-%dT%H:%M:%S"), "%B %Y")
            if output.get(name) is None:
                output[name] =  {}
            if self.tableBtnGroup.checkedButton().text() != "Complex Table" and self.returnOutputType.currentText() == "Return": #show monthly return benchmarks
                output[name][date] = float(bench.get("MTDnet",0) if bench.get("MTDnet",0) !=  "None" else 0) * 100
                if output[name].get("dataType") is None:
                    output[name]["dataType"] = "benchmark"
            elif self.tableBtnGroup.checkedButton().text() == "Complex Table" and date == self.dataEndSelect.currentText():
                #populate the complex fields
                if output[name].get("dataType") is None:
                    output[name]["dataType"] = "benchmark"
                for time in ("MTD","QTD","YTD"):
                    if bench.get(f"{time}net", "None") != "None":
                        output[name][time] = float(bench.get(f"{time}net")) * 100
                if bench.get("ITDTWR","None") != "None":
                    output[name]["ITD"] = float(bench.get("ITDTWR")) * 100
                for year in yearOptions:
                    if bench.get(f"Last{year}yrnet","None") not in ("None",None):
                        output[name][f"{year}YR"] = float(bench.get(f"Last{year}yrnet")) * 100
        return output
    def buildCode(self, path):
            code = f"##({"::".join(path)})##"
            return code
    def calculateUpperLevels(self, tableStructure,data):
        def applyLinkedBenchmarks(struc,code, levelName, option):
            for entry in benchmarkLinks:
                if levelName == assetLevelLinks[entry.get("assetLevel")].get("Link") and option == entry.get("asset"):
                    benchmark = entry.get("benchmark")
                    if benchmark + self.buildCode([]) in struc.keys(): #pull the benchmark data in if it exists
                        temp = struc[benchmark + self.buildCode([])]
                        struc[benchmark + code] = temp.copy()
                    else:
                        struc[benchmark + code] = {} #place table space for that level selection. Will not populate if previous failed
            return struc
        def buildLevel(levelName,levelIdx, struc,data,path : list, insertedOption = None):
            levelIdx += 1
            entryTemplate = {"dateTime" : None, "Calculation Type" : "Total " + levelName, "Pool" : None, "Fund" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Investor" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None, "IRR ITD" : None}
            for header in headerOptions:
                if header not in ("Ownership", "IRR ITD"):
                    entryTemplate[header] = 0

            #check for filtering. If none, use all options
            options = []
            for entry in data: #all available data
                if entry[levelName] not in options: #
                    options.append(entry[levelName])
            if levelName == "subAssetClass":
                # Sort options so those in assetClass2Order come first (in the given order),
                # then anything else not in assetClass2Order appears after (sorted alphabetically)
                options = [v for v in assetClass2Order if v in options] + sorted([v for v in options if v not in assetClass2Order])
            elif levelName == "assetClass":
                options = [v for v in assetClass1Order if v in options] + sorted([v for v in options if v not in assetClass1Order])
            else:
                options.sort()
            newTotalEntries = []
            if len(sortHierarchy) > levelIdx: #more hierarchy levels to parse
                highTotals = [] #all total values made on the level
                for option in options:
                    tempPath = path.copy()
                    tempPath.append(option)
                    
                    highEntries = {}
                    name = option if levelName != "assetClass" or option != "Cash" else "Cash "
                    code = self.buildCode(tempPath)
                    struc[name + code] = {} #place table space for that level selection
                    if self.showBenchmarkLinksBtn.isChecked():
                        struc = applyLinkedBenchmarks(struc,code, levelName, option)
                                
                                
                    levelData = []
                    for entry in data: #separates out only relevant data
                        if entry[levelName] == option:
                            levelData.append(entry)
                    if len(sortHierarchy) > levelIdx + 1 and levelName == "subAssetClass" and sortHierarchy[levelIdx] == "subAssetSleeve" and option in self.db.fetchOptions("asset3Visibility").keys():
                        #will skip the subAssetSleeve for hidden ones and send the entire section of data to the next level
                        print(f"skipping for {option}")
                        tempPath.append("hiddenLayer")
                        struc, lowTotals, fullEntries = buildLevel(sortHierarchy[levelIdx + 1],levelIdx + 1,struc,levelData,tempPath)
                    else:
                        struc, lowTotals, fullEntries = buildLevel(sortHierarchy[levelIdx],levelIdx,struc,levelData,tempPath, insertedOption = option)
                    newTotalEntries.extend(fullEntries)
                    for total in lowTotals:
                        if total["dateTime"] not in highEntries.keys():
                            highEntries[total["dateTime"]] = copy.deepcopy(entryTemplate)
                            highEntries[total["dateTime"]]["rowKey"] = name + code
                            for label in dataOptions:
                                highEntries[total["dateTime"]][label] = total[label]
                            if levelName not in ("Investor","Family Branch"):
                                highEntries[total["dateTime"]][levelName] = total[levelName] if total[levelName] != "Cash" or levelName != "assetClass" else "Cash "
                                if levelName == "subAssetClass":
                                    highEntries[total["dateTime"]]["assetClass"] = total["assetClass"] if total["assetClass"] != "Cash" else "Cash "
                        for header in headerOptions:
                            if header not in ("Ownership", "IRR ITD"):
                                highEntries[total["dateTime"]][header] += float(total[header])
                            elif header == "Ownership" and levelName in ("Pool", "Investor", "Family Branch") and total.get(header) not in (None,"None","",0) and "Pool" in sortHierarchy[:levelIdx]:
                                if highEntries[total["dateTime"]].get(header) is None:
                                    highEntries[total["dateTime"]][header] = float(total[header]) #initialize
                                else:
                                    highEntries[total["dateTime"]][header] += float(total[header]) #aggregate pool ownerships
                    for month in highEntries.keys():
                        highEntries[month]["Return"] = highEntries[month]["Monthly Gain"] / highEntries[month]["MDdenominator"] * 100 if highEntries[month]["MDdenominator"] != 0 else 0
                        highTotals.append(highEntries[month])
                newTotalEntries.extend(highTotals)       
                #high totals: all totals for the exact level
                #newTotalEntries: all totals for every level being tracked
                return struc, highTotals, newTotalEntries
            else: #occurs at level of fund parent
                newEntriesLow = []
                totalDataLow = []
                if levelName == "subAssetSleeve" and sortHierarchy[levelIdx - 2] == "subAssetClass" and insertedOption in self.db.fetchOptions("asset3Visibility").keys():
                    print(f"skipping for {insertedOption}")
                    options = ["hiddenLayer"]
                for option in options:
                    tempPath = path.copy()
                    tempPath.append(option)
                    totalEntriesLow = {}
                    name = option if levelName != "assetClass" or option != "Cash" else "Cash "
                    code = self.buildCode(tempPath)
                    if option != "hiddenLayer":
                        struc[name + code] = {} #place table space for that level selection
                        if self.showBenchmarkLinksBtn.isChecked():
                            struc = applyLinkedBenchmarks(struc,code, levelName, option)
                        levelData = []
                        for entry in data: #separates out only relevant data
                            if entry[levelName] == option:
                                levelData.append(entry)
                    else:
                        levelData = data #dont filter the data for hidden layer
                    #gui_queue.put(lambda rows = levelData, name = option: self.openTableWindow(rows,f"data for: {name}"))
                    nameList = {}
                    investorsAccessed = {}
                    for entry in levelData:
                        fundName = entry["Fund"] if not self.consolidateFundsBtn.isChecked() or entry["Fund"] not in self.consolidatedFunds or entry["Fund"] in self.filterDict["Fund"].checkedItems() else self.consolidatedFunds.get(entry["Fund"]).get("cFund")
                        nameList.setdefault(fundName + code,0.0)
                        if "NAV" in self.sortStyle.text() and datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") == datetime.strptime(self.dataEndSelect.currentText(),"%B %Y"):
                            nameList[fundName + code] += float(entry.get("NAV",0.0)) if entry.get("NAV",0.0) not in (None,"None","" ,0) else 0.0
                        temp = entry.copy()
                        temp["rowKey"] = fundName + code
                        totalDataLow.append(temp)
                        if entry["dateTime"] not in totalEntriesLow:
                            totalEntriesLow[entry["dateTime"]] = copy.deepcopy(entryTemplate)
                            totalEntriesLow[entry["dateTime"]]["rowKey"] =name + code
                            for label in dataOptions:
                                totalEntriesLow[entry["dateTime"]][label] = entry[label]
                            if levelName not in ("Investor","Family Branch"):
                                totalEntriesLow[entry["dateTime"]][levelName] = entry[levelName] if entry[levelName] != "Cash" or levelName != "assetClass" else "Cash "
                                if levelName == "subAssetClass":
                                    totalEntriesLow[entry["dateTime"]]["assetClass"] = entry["assetClass"] if entry["assetClass"] != "Cash" else "Cash "
                        for header in headerOptions:
                            if header not in ("Ownership", "IRR ITD") and entry.get(header) not in (None,"None",""):
                                totalEntriesLow[entry["dateTime"]][header] += float(entry[header])
                            elif header == "Ownership" and levelName in ("Investor", "Family Branch") and "Pool" in sortHierarchy and entry.get(header) not in (None,"None","") and float(entry.get(header)) != 0:
                                investor = entry.get("Investor")
                                if totalEntriesLow[entry["dateTime"]].get(header) is None:
                                    totalEntriesLow[entry["dateTime"]][header] = float(entry[header]) #assign investor to ownership based on fund
                                    investorsAccessed[entry["dateTime"]] = [investor,]
                                elif investor not in investorsAccessed.get(entry["dateTime"], []): #accounts for family branch level to add the investor level ownerships
                                    totalEntriesLow[entry["dateTime"]][header] += float(entry[header])
                                    investorsAccessed[entry["dateTime"]].append(investor)
                    if "alphabetical" in self.sortStyle.text().lower():
                        for name in sorted(nameList.keys()): #sort by alphabetical order
                            struc[name] = {}
                    else:
                        for name in descendingNavSort(nameList): #sort by descending NAV
                            struc[name] = {}
                    for month in totalEntriesLow.keys():
                        totalEntriesLow[month]["Return"] = totalEntriesLow[month]["Monthly Gain"] / totalEntriesLow[month]["MDdenominator"] * 100 if totalEntriesLow[month]["MDdenominator"] != 0 else 0
                        newEntriesLow.append(totalEntriesLow[month])
                totalDataLow.extend(newEntriesLow)
                return struc, newEntriesLow, totalDataLow

        if self.showBenchmarkLinksBtn.isChecked():
            benchmarkLinks = self.db.fetchBenchmarkLinks()
        sortHierarchy = self.sortHierarchy.checkedItems()
        levelIdx = 0
        tableStructure, highestEntries, newEntries = buildLevel(sortHierarchy[levelIdx],levelIdx,tableStructure,data, [])
        trueTotalEntries = {}
        for total in highestEntries:
            if total["dateTime"] not in trueTotalEntries.keys():
                trueTotalEntries[total["dateTime"]] = {"dateTime" : None, "Calculation Type" : "Total", "Pool" : None, "Fund" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Investor" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None, "IRR ITD" : None}
                trueTotalEntries[total["dateTime"]]["rowKey"] = "Total" + self.buildCode([])
                for header in headerOptions:
                    if header != "Ownership":
                        trueTotalEntries[total["dateTime"]][header] = 0
                for label in dataOptions:
                    trueTotalEntries[total["dateTime"]][label] = total[label]
            for header in headerOptions:
                if header not in ("Ownership", "IRR ITD"):
                    trueTotalEntries[total["dateTime"]][header] += float(total[header])
        for month in trueTotalEntries.keys():
            trueTotalEntries[month]["Return"] = trueTotalEntries[month]["Monthly Gain"] / trueTotalEntries[month]["MDdenominator"] * 100 if trueTotalEntries[month]["MDdenominator"] != 0 else 0
            newEntries.append(trueTotalEntries[month])
        #data.extend(newEntries)
        return tableStructure,newEntries
                    
    def filterUpdate(self):
        from functools import partial

        self.buildTableLoadingBar.setValue(0)
        self.tableLoadingLabel.setText("Waiting on database connection...")
        self.buildTableLoadingBox.setVisible(True)

        if self.filterCallLock:
            return

        def resetOptions(key, new_options):
            multiBox = self.filterDict[key]
            old_options = set(multiBox._checkboxes.keys())
            new_options_set = set(new_options)

            if old_options != new_options_set:
                currentSelections = multiBox.checkedItems()
                multiBox.clearItems()
                multiBox.addItems(sorted(new_options))
                for text in currentSelections:
                    if text in new_options_set:
                        multiBox.setCheckedItem(text)

        def build_condition(exclude_key):
            conditions = []
            params = []
            for f in self.filterOptions:
                key = f["key"]
                if key == exclude_key or key in self.highOnlyFilters:
                    continue
                selected = self.filterDict[key].checkedItems()
                if selected:
                    placeholders = ','.join('?' for _ in selected)
                    conditions.append(f"[{f['dynNameLow']}] IN ({placeholders})")
                    params.extend(selected)
            clause = "WHERE " + " AND ".join(conditions) if conditions else ""
            return clause, tuple(params)

        def exitFunc():
            self.filterCallLock = False
            gui_queue.put(lambda: self.tableLoadingLabel.setText("Building returns table..."))
            gui_queue.put(lambda: self.buildReturnTable())

        def processFilter():
            try:
                self.filterCallLock = True

                currentChoices = {
                    key: self.filterDict[key].checkedItems()
                    for key in self.filterDict
                    if key not in self.highOnlyFilters
                }

                if all(not choices for choices in currentChoices.values()):
                    for key in currentChoices:
                        gui_queue.put(partial(resetOptions, key, self.fullLevelOptions[key]))
                    exitFunc()
                    return

                for targetFilter in self.filterOptions:
                    key = targetFilter["key"]
                    if key in self.highOnlyFilters:
                        continue

                    condSQL, condParams = build_condition(exclude_key=key)
                    lowAccounts = load_from_db("positions_low", condSQL, condParams, lock=self.lock)

                    options = {
                        f["key"]: set()
                        for f in self.filterOptions
                        if f["key"] not in self.highOnlyFilters
                    }
                    for account in lowAccounts:
                        for f in self.filterOptions:
                            k = f["key"]
                            if k in options:
                                value = account.get(f["dynNameLow"])
                                if value is not None:
                                    options[k].add(value)

                    gui_queue.put(partial(resetOptions, key, sorted(options[key])))

            except Exception as e:
                gui_queue.put(lambda: QMessageBox.warning(self, "Filter Error", f"Error occurred updating filters:\n{e}"))

            exitFunc()

        executor.submit(processFilter)
    def assetClass3VisibilityChanged(self):
        hiddenItems = self.assetClass3Visibility.checkedItems()
        self.db.saveAsset3Visibility(hiddenItems)
        self.buildReturnTable()
    def updateMonths(self):
        start = self.dataTimeStart
        end = datetime.now()
        index = start
        monthList = []
        while index < end:
            monthList.append(index)
            index += relativedelta(months=1)
        dbDates = []
        for monthDT in monthList:
            month = int(monthDT.month)
            year = int(monthDT.year)
            lastDayCurrent = calendar.monthrange(int(year),month)[1]
            lastDayCurrent   = str(lastDayCurrent).zfill(2)
            if month - 1 > 0:
                prevMonth =  month - 1
                prevMyear = year
            else:
                prevMonth = 12
                prevMyear = str(int(year) - 1)
            lastDayPrev = calendar.monthrange(int(prevMyear),prevMonth)[1]
            lastDayPrev   = str(lastDayPrev).zfill(2)
            prevMonth = str(prevMonth).zfill(2)
            month = str(month).zfill(2)
            
            tranStart = f"{year}-{month}-01T00:00:00.000Z"
            bothEnd = f"{year}-{month}-{lastDayCurrent}T00:00:00.000Z"
            accountStart = f"{prevMyear}-{prevMonth}-{lastDayPrev}T00:00:00.000Z"

            
            dateString = monthDT.strftime("%B %Y")

            monthEntry = {"dateTime" : monthDT, "Month" : dateString, "tranStart" : tranStart.removesuffix(".000Z"), "endDay" : bothEnd.removesuffix(".000Z"), "accountStart" : accountStart.removesuffix(".000Z")}
            dbDates.append(monthEntry)
        save_to_db("Months",dbDates)

    def pullInvestorNames(self):
        accountsHigh = load_from_db('positions_high', lock=self.lock)
        if accountsHigh is not None:
            investors = []
            familyBranches = []
            for account in accountsHigh:
                if account["Source name"] not in investors:
                    investors.append(account["Source name"])
                if account[nameHier["Family Branch"]["dynHigh"]] not in familyBranches:
                    familyBranches.append(account[nameHier["Family Branch"]["dynHigh"]])
            investors.sort()
            familyBranches.sort()
            self.allInvestors = investors
            self.filterDict["Investor"].addItems(investors)
            self.allFamilyBranches = familyBranches
            self.filterDict[nameHier["Family Branch"]["local"]].addItems(familyBranches)
            self.fullLevelOptions["Investor"] = self.allInvestors
            self.fullLevelOptions["Family Branch"] = self.allFamilyBranches
        else:
            self.allInvestors = []
            self.allFamilyBranches = []
    def pullLevelNames(self):
        allOptions = {}
        fundPoolLink = {}
        for filter in self.filterOptions:
            if filter["key"] not in self.highOnlyFilters:
                allOptions[filter["key"]] = []
        accountsHigh = load_from_db("positions_high")
        if accountsHigh is not None:
            for account in accountsHigh:
                for filter in self.filterOptions:
                    if (filter["key"] in allOptions and "dynNameHigh" in filter.keys() and
                        account[filter["dynNameHigh"]] is not None and
                        account[filter["dynNameHigh"]] not in allOptions[filter["key"]]):
                        allOptions[filter["key"]].append(account[filter["dynNameHigh"]])
        else:
            print("no investor to pool accounts found")
        accountsLow = load_from_db("positions_low")
        if accountsLow is not None:
            for lowAccount in accountsLow:
                for filter in self.filterOptions:
                    if (filter["key"] in allOptions and "dynNameLow" in filter.keys() and
                        lowAccount[filter["dynNameLow"]] is not None and
                        lowAccount[filter["dynNameLow"]] not in allOptions[filter["key"]]):
                        allOptions[filter["key"]].append(lowAccount[filter["dynNameLow"]])
                fundPoolLink[lowAccount["Target name"]] = lowAccount["Source name"]
        else:
            print("no pool to fund accounts found")
        self.fullLevelOptions = {}
        for filter in self.filterOptions:
            if filter["key"] in allOptions:
                allOptions[filter["key"]].sort()
                self.filterDict[filter["key"]].addItems(allOptions[filter["key"]])
                self.fullLevelOptions[filter["key"]] = allOptions[filter["key"]]
        self.filterDict["Classification"].setCheckedItem("HFC")
        self.assetClass3Visibility.addItems(self.fullLevelOptions["subAssetClass"])
        hiddenItems = self.db.fetchOptions("asset3Visibility")
        for item in hiddenItems.keys():
            self.assetClass3Visibility.setCheckedItem(item)
        self.fundPoolLinks = fundPoolLink
        self.pullInvestorNames()
        self.pullBenchmarks()

    def pullBenchmarks(self):
        benchmarks = load_from_db("benchmarks")
        benchNames = []
        for bench in benchmarks:
            if bench["Index"] not in benchNames:
                benchNames.append(bench["Index"])
        self.benchmarkSelection.addItems(benchNames)
    def groupingChange(self):
        groupOpts = self.sortHierarchy.checkedItems()
        if groupOpts == []:
            self.sortHierarchy.setCheckedItems(["assetClass","subAssetClass"])
        self.filterCallLock = True
        for filt in ("Investor", "Family Branch"):
            if filt in groupOpts and self.filterDict[filt].checkedItems() == []:
                self.filterDict[filt].selectAll()
                self.previousGrouping.add(filt)
            elif filt in self.previousGrouping: #removes the selections if they stop grouping by investor/family
                self.filterDict[filt].clearSelection()
                self.previousGrouping.remove(filt)
            
            
        self.filterCallLock = False
        self.buildReturnTable()
    def testAPIconnection(self, key=None):
        apiKey = self.api_key if key is None else key
        headers = {
            "Authorization": f"Bearer {apiKey}",
            "Content-Type":  "application/json"
        }
        payload = {
            "advf": [{ "_name": "Fund" }],
            "mode": "compact",
            "page": {"size": 0}
        }
        resp = requests.get(f"{mainURL}/Entity", headers=headers, json=payload)
        if resp.status_code == 200:
            return True
        else:
            return False
    def check_api_key(self, *_):
        key = self.api_input.text().strip()
        if key:
            if self.testAPIconnection(key=key):
                self.api_label.setText('API key valid. Saving to system...')
                subprocess.run(['setx',dynamoAPIenvName,key], check=True)
                os.environ[dynamoAPIenvName] = key
                self.api_key = key
                self.stack.setCurrentIndex(1)
                self.init_data_processing()
            else:
                self.api_label.setText('Invalid API key or Dynamo is not responding')
        else:
            self.api_label.setText('API key cannot be empty')

    def show_results(self,*_):
        self.stack.setCurrentIndex(2)

    def pullData(self):
        if not self.testAPIconnection():
            gui_queue.put(lambda: QMessageBox.warning(self,"API Failure", "API connection has failed. Server is down or API key is bad. \n Previous calculations are left in place for viewing."))
            return
        def checkNewestData(table, rows):
            def buildKey(record):
                value = record[nameHier["Value"]["dynHigh"] if "position" in table else nameHier["CashFlow"]["dynLow"]]
                value = 0 if value is None or value == "None" else value
                key = (
                        record['Source name'] if record['Source name'] is not None else "None",
                        record['Target name'] if record['Target name'] is not None else "None",
                        round(float(value)) if table != "positions_high" else 0,               # normalize to float
                        record['Date'].replace(' ', 'T')      # normalize format if needed
                    )
                return key
            try:
                diffCount = 0
                differences = []
                newRows = []
                previous = load_from_db(table) or []

                # Build a set of tuple‐keys for the old data
                oldRecords = set()
                for rec in previous:
                    oldRecords.add(buildKey(rec))

                newRecords = set()
                earliest = None
                for rec in rows:
                    value = rec[nameHier["Value"]["dynHigh"] if "position" in table else nameHier["CashFlow"]["dynLow"]]
                    value = 0 if value is None or value == "None" else value
                    key = buildKey(rec)
                    newRecords.add(key)
                    if key in oldRecords:
                        continue
                    diffCount += 1
                    newRows.append(rec)
                    differences.append(rec)
                    differences.append({"Source name" : key[0],"Target name" : key[1],nameHier["Value"]["dynLow"] : key[2],"Date" : key[3]})
                    # parse the date for comparison
                    dt = datetime.strptime(rec['Date'], "%Y-%m-%dT%H:%M:%S")
                    if earliest is None or dt < earliest:
                        earliest = dt
                    poolTag = "Target name" if "high" in table else "Source name"
                    if dt < self.poolChangeDates.get(rec.get(poolTag),datetime.now()): 
                        self.poolChangeDates[rec.get(poolTag)] = dt # sets each pool value to earliest and instantiates if not existing
                for oldRec in oldRecords:
                    if oldRec not in newRecords: #find if a new record no longer exists in the old. Means old data is removed and must be redone
                        self.foundRetroChange = True
                        self.poolChangeDates["active"] = False
                        print(f"Retroactive changes found in {table}. Resetting whole table.")
                        break
                
                if earliest and not self.foundRetroChange:
                    if earliest < self.earliestChangeDate:
                        self.earliestChangeDate = earliest
                if self.foundRetroChange: #push full api data and reset calc date to redo all data
                    self.earliestChangeDate =  self.dataTimeStart
                    return rows, False
                print(f"Differences in {table} : {diffCount} of {len(rows)}")
                if diffCount > 0 and not demoMode:
                    def openWindow():
                        window = tableWindow(parentSource=self,all_rows=differences,table=table)
                        self.tableWindows[table] = window
                        window.show()
                    gui_queue.put(lambda: openWindow())
                return newRows, True
            except Exception as e:
                print(f"Error searching old data: {e}")
        try:
            self.earliestChangeDate = datetime(datetime.now().year,datetime.now().month + 1,datetime.now().day)
            gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(True))
            gui_queue.put(lambda: self.importButton.setEnabled(False))
            self.updateMonths()
            completeLock = threading.Lock()
            self.apiFutures = set()
            self.complete = float(0)
            totalCalls = float(6)
            self.pullInvestorNames()
            apiData = {
                "tranCols": "Investment in, Investing Entity, Transaction Type, Effective date, Asset Class (E), Sub-asset class (E), HF Classification, Remaining commitment change, Transaction timing, Amount in system currency, Cash flow change (USD), Parent investor",
                "tranName": "InvestmentTransaction",
                "tranSort": "Effective date:desc",
                "accountCols": "As of Date, Balance Type, Asset Class, Sub-asset class, Investing entity, Investment in, HF Classification, Parent investor, Value in system currency, Fund class",
                "accountName": "InvestmentPosition",
                "accountSort": "As of Date:desc",
                "fundCols" : "Fund Name, Asset class category, Parent fund, Fund Pipeline Status",
                "benchCols" : (f"Index, As of date, MTD %, QTD %, YTD %, ITD cumulative %, ITD TWRR %, "
                               f"{', '.join(f'Last {y} yr %' for y in yearOptions)}"), 
            }
            calculationsTest = load_from_db("calculations")
            if calculationsTest != []:
                skipCalculations = True
                self.poolChangeDates["active"] = True
                self.foundRetroChange = False
            else:
                skipCalculations = False
            for i in range(2):
                cols_key = 'accountCols' if i == 1 else 'tranCols'
                name_key = 'accountName' if i == 1 else 'tranName'
                sort_key = 'accountSort' if i == 1 else 'tranSort'
                headers = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": apiData[cols_key],
                    "x-sort": apiData[sort_key]
                }
                for j in range(2): #0: fund level, 1: pool to high investor level
                    investmentLevel = "Investing entity" if j == 0 else "Investment in"
                    if i == 0: #transaction
                        if j == 0: #fund level
                            payload = {
                                        "advf": {
                                            "e": [
                                                {
                                                    "_name": "InvestmentTransaction",
                                                    "e": [
                                                        {
                                                            "_name": "InvestorAccount",
                                                            "_not": True
                                                        },
                                                        {
                                                            "_name": "Fund",
                                                            "rule": [
                                                                {
                                                                    "_op": "is",
                                                                    "_prop": "Fund Pipeline Status",
                                                                    "values": [
                                                                        {
                                                                            "id": "d33af081-c4c8-431b-a98b-de9eaf576324",
                                                                            "es": "L_FundPipelineStatus",
                                                                            "name": "I - Internal"
                                                                        }
                                                                    ]
                                                                }
                                                            ]
                                                        }
                                                    ],
                                                    "rule": [
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Cash flow change (USD)"
                                                        },
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Investing entity"
                                                        }
                                                    ]
                                                },
                                                {
                                                    "_name": "InvestmentTransaction",
                                                    "e": [
                                                        {
                                                            "_name": "InvestorAccount",
                                                            "_not": True
                                                        },
                                                        {
                                                            "_name": "Fund",
                                                            "rule": [
                                                                {
                                                                    "_op": "is",
                                                                    "_prop": "Fund Pipeline Status",
                                                                    "values": [
                                                                        {
                                                                            "id": "d33af081-c4c8-431b-a98b-de9eaf576324",
                                                                            "es": "L_FundPipelineStatus",
                                                                            "name": "I - Internal"
                                                                        }
                                                                    ]
                                                                }
                                                            ]
                                                        }
                                                    ],
                                                    "rule": [
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Investing entity"
                                                        },
                                                        {
                                                            "_op": "any_item",
                                                            "_prop": "Transaction type",
                                                            "values": [
                                                                [
                                                                    {
                                                                        "id": "5327639c-8160-4d85-9b23-8c6bf60c5406",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Commitment"
                                                                    },
                                                                    {
                                                                        "id": "37339e7c-1c24-4d13-9d17-86d0efe079b3",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Transfer of commitment"
                                                                    },
                                                                    {
                                                                        "id": "0f8f8671-8579-49d7-b604-05300b6a3990",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Transfer of commitment (out)"
                                                                    },
                                                                    {
                                                                        "id": "5e098d83-70b0-4135-a629-aff19048fb1c",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Secondary - Original commitment (by secondary seller)"
                                                                    }
                                                                ]
                                                            ]
                                                        }
                                                    ]
                                                }
                                            ]
                                        },
                                        "mode": "compact"
                                    }
                        else: #investor level
                            payload = {
                                        "advf": {
                                            "e": [
                                                {
                                                    "_name": "InvestmentTransaction",
                                                    "e": [
                                                        {
                                                            "_name": "InvestorAccount"
                                                        }
                                                    ],
                                                    "rule": [
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Cash flow change (USD)"
                                                        },
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Investing entity"
                                                        }
                                                    ]
                                                }
                                            ]
                                        },
                                        "mode": "compact"
                                    }
                        
                    else: #account (position)
                        if j == 0: #fund level
                            payload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "InvestmentPosition",
                                                "e": [
                                                    {
                                                        "_name": "Fund",
                                                        "rule": [
                                                            {
                                                                "_op": "is",
                                                                "_prop": "Fund Pipeline Status",
                                                                "values": [
                                                                    {
                                                                        "id": "d33af081-c4c8-431b-a98b-de9eaf576324",
                                                                        "es": "L_FundPipelineStatus",
                                                                        "name": "I - Internal"
                                                                    }
                                                                ]
                                                            }
                                                        ]
                                                    },
                                                    {
                                                        "_name": "InvestorAccount",
                                                        "_not": True
                                                    }
                                                ],
                                                "rule": [
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investing entity"
                                                    }
                                                ]
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
                        else: #investor level
                            payload = {
                                            "advf": {
                                                "e": [
                                                    {
                                                        "_name": "InvestmentPosition",
                                                        "e": [
                                                            {
                                                                "_name": "InvestorAccount"
                                                            }
                                                        ],
                                                        "rule": [
                                                            {
                                                                "_op": "not_null",
                                                                "_prop": "Investing entity"
                                                            }
                                                        ]
                                                    }
                                                ]
                                            },
                                            "mode": "compact"
                                        }
                    def bgPullData(payload=payload, headers=headers, i=i, j=j):
                        rows = []
                        idx = 0
                        while rows in ([],None) and idx < 3: #if call fails, tries again
                            idx += 1
                            response = requests.post(f"{mainURL}/Search", headers=headers, data=json.dumps(payload))
                            if response.status_code == 200:
                                try:
                                    data = response.json()
                                except ValueError:
                                    continue
                                if isinstance(data, dict):
                                    rows = data.get('data', data.get('rows', []))
                                elif isinstance(data, list):
                                    rows = data
                                else:
                                    rows = []

                                keys_to_remove = {'_id', '_es'}
                                rows = [
                                    {k: v for k, v in row.items() if k not in keys_to_remove}
                                    for row in rows
                                ]
                            else:
                                print(f"Error in API call. Code: {response.status_code}. {response}")
                                try:
                                    print(f"Error: {response.json()}")
                                except:
                                    pass
                        if i == 1:
                            if j == 0:
                                if skipCalculations: #separate out only new rows to alter db
                                    rows, good = checkNewestData('positions_low',rows)
                                for row in rows:
                                    row[nameHier["Unfunded"]["local"]] = 0
                                    row[nameHier["Commitment"]["local"]] = 0
                                    row[nameHier["sleeve"]["local"]] = None
                                if skipCalculations and good:
                                    save_to_db('positions_low', rows, action="add", lock=self.lock)
                                else:
                                    save_to_db('positions_low', rows, lock=self.lock)
                            else:
                                if skipCalculations: #separate out only new rows to alter db
                                    rows, good = checkNewestData('positions_high',rows)
                                    if good:
                                        save_to_db('positions_high', rows, action="add", lock=self.lock)
                                    else:
                                        save_to_db('positions_high', rows, lock=self.lock)
                                else:
                                    save_to_db('positions_high', rows, lock=self.lock)
                        else:
                            if j == 0:
                                if skipCalculations: #separate out only new rows to alter db
                                    rows, good = checkNewestData('transactions_low',rows)
                                    if good:
                                        save_to_db('transactions_low', rows, action="add", lock=self.lock)
                                    else:
                                        save_to_db('transactions_low', rows, lock=self.lock)
                                else:
                                    save_to_db('transactions_low', rows, lock=self.lock)
                            else:
                                if skipCalculations: #separate out only new rows to alter db
                                    rows, good = checkNewestData('transactions_high',rows)
                                    if good:
                                        save_to_db('transactions_high', rows, action="add", lock=self.lock)
                                    else:
                                        save_to_db('transactions_high', rows, lock=self.lock)
                                else:
                                    save_to_db('transactions_high', rows, lock=self.lock)
                        with completeLock:
                            self.complete += 1
                        frac = self.complete/totalCalls
                        gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
                    try:
                        submitAPIcall(self,bgPullData)
                    except Exception as e:
                        print(f"Failure to run background thread API call: {e} \n {e.args}")
            fundPayload = {
                            "advf": {
                                "e": [
                                    {
                                        "_name": "Fund"
                                    }
                                ]
                            },
                            "mode": "compact"
                        }
            fundHeaders = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": apiData["fundCols"],
                }
            def bgFundPull():
                response = requests.post(f"{mainURL}/Search", headers=fundHeaders, data=json.dumps(fundPayload))
                if response.status_code == 200:
                    try:
                        data = response.json()
                        if isinstance(data, dict):
                            rows = data.get('data', data.get('rows', []))
                        elif isinstance(data, list):
                            rows = data
                        else:
                            rows = []
                        keys_to_remove = {'_id', '_es'}
                        rows = [{k: v for k, v in row.items() if k not in keys_to_remove} for row in rows]
                        consolidatorFunds = {}
                        for idx, row in enumerate(rows): #find sleeve values and consolidated funds
                            assetCat = row["ExposureAssetClassCategory"]
                            if assetCat is not None and assetCat.count(" > ") == 3:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = assetCat.split(" > ")[2]
                                sleeve = assetCat.split(" > ")[3]
                            elif assetCat is not None and assetCat.count(" > ") == 2:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = assetCat.split(" > ")[2]
                                sleeve = None
                            elif assetCat is not None and assetCat.count(" > ") == 1:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = None
                                sleeve = None
                            else:
                                assetClass = None
                                subAssetClass = None
                                sleeve = None
                            if row.get("Fundpipelinestatus") is not None and "Z - Placeholder" in row.get("Fundpipelinestatus"):
                                consolidatorFunds[row["Name"]] = {"cFund" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass, "sleeve" : sleeve}
                            rows[idx][nameHier["sleeve"]["sleeve"]] =  sleeve
                            rows[idx]["assetClass"] = assetClass
                            rows[idx]["subAssetClass"] = subAssetClass
                        self.consolidatedFunds = {}
                        for row in rows: #assign funds to their consolidators
                            if row.get("Parentfund") in consolidatorFunds:
                                self.consolidatedFunds[row["Name"]] = consolidatorFunds.get(row.get("Parentfund"))
                        if rows != []:
                            save_to_db("funds",rows,lock=self.lock)
                    except Exception as e:
                        print(f"Error proccessing fund API data : {e} {e.args}.  {traceback.format_exc()}")
                    
                else:
                    print(f"Error in API call for fund. Code: {response.status_code}. {response}. {traceback.format_exc()}")
                with completeLock:
                    self.complete += 1
                frac = self.complete/totalCalls
                gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
            submitAPIcall(self,bgFundPull)
            benchmarkPayload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "IndexPerformance"
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
            benchmarkHeaders = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": apiData["benchCols"],
                }
            def bgBenchPull():
                response = requests.post(f"{mainURL}/Search", headers=benchmarkHeaders, data=json.dumps(benchmarkPayload))
                if response.status_code == 200:
                    try:
                        data = response.json()
                        if isinstance(data, dict):
                            rows = data.get('data', data.get('rows', []))
                        elif isinstance(data, list):
                            rows = data
                        else:
                            rows = []
                        keys_to_remove = {'_id', '_es'}
                        rows = [{k: v for k, v in row.items() if k not in keys_to_remove} for row in rows]
                        save_to_db("benchmarks",rows, lock=self.lock)
                    except Exception as e:
                        print(f"Error proccessing benchmark API data : {e} {e.args}.  {traceback.format_exc()}")
                    
                else:
                    print(f"Error in API call for benchmarks. Code: {response.status_code}. {response}. {traceback.format_exc()}")
                with completeLock:
                    self.complete += 1
                frac = self.complete/totalCalls
                gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
            submitAPIcall(self,bgBenchPull)

            wait(self.apiFutures)
            if skipCalculations:
                print("Earliest change: ", self.earliestChangeDate)
                if self.poolChangeDates.get("active", False):
                    print(f"Changes dates by pools:")
                    for pool in self.poolChangeDates:
                        print(f"        {pool} : {self.poolChangeDates.get(pool)}")
            gui_queue.put(lambda: self.apiLoadingBar.setValue(100))
            
            while not gui_queue.empty(): #wait to assure database has been updated in main thread before continuing
                time.sleep(0.2)
            


            currentTime = datetime.now().strftime("%B %d, %Y @ %I:%M %p")
            changeDate = datetime.strftime(self.earliestChangeDate, "%B %d, %Y @ %I:%M %p")
            save_to_db(None,None,query="UPDATE history SET [lastImport] = ?, [changeDate] = ?", inputs=(currentTime,changeDate), action="replace", lock=self.lock)
            self.lastImportDB[0]["lastImport"] = currentTime
            self.lastImportDB[0]["changeDate"] = changeDate
            self.lastImportLabel.setText(f"Last Data Import: {currentTime}")
            gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(False))
            gui_queue.put(lambda: self.calculateReturn())
        except Exception as e:
            QMessageBox.warning(self,"Error Importing Data", f"Error pulling data from dynamo: {e} , {e.args}")
        gui_queue.put(lambda: self.importButton.setEnabled(True))
        gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(False))
    def openTableWindow(self, rows, name = "Table"):
        window = tableWindow(parentSource=self,all_rows=rows,table=name)
        self.tableWindows[name] = window
        window.show()
    def calculateReturn(self):
        def initalizeCalc():
            try:
                calculationStart = datetime.now()
                gui_queue.put(lambda: self.importButton.setEnabled(False))
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(True))
                self.updateMonths()
                gui_queue.put(lambda: self.pullLevelNames())
                print("Calculating return....")
                fundListDB = load_from_db("funds")
                fundList = {}
                for fund in fundListDB:
                    fundList[fund["Name"]] = fund[nameHier["sleeve"]["sleeve"]]
                months = load_from_db("Months", f"ORDER BY [dateTime] ASC",lock=self.lock)
                calculations = []
                monthIdx = 0
                if load_from_db("calculations",lock=self.lock) == []:
                    noCalculations = True
                else:
                    noCalculations = False

                if self.earliestChangeDate > datetime.now() and not noCalculations:
                    #if no new data exists, use old calculations
                    calculations = load_from_db("calculations")
                    keys = []
                    for row in calculations:
                        for key in row.keys():
                            if key not in keys:
                                keys.append(key)
                    gui_queue.put( lambda: self.populate(self.calculationTable,calculations,keys = keys))
                    gui_queue.put( lambda: self.buildReturnTable())
                    gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                    gui_queue.put(lambda: self.importButton.setEnabled(True))
                    apiPullTime = self.lastImportDB[0].get("lastImport")
                    save_to_db(None,None,query="UPDATE history SET [lastCalculation] = ?", inputs=(apiPullTime,), action="replace", lock=self.lock)
                    self.lastImportDB[0]["lastCalculation"] = apiPullTime
                    print("Calculations skipped.")
                    return
                
                # proces pool section----------------------------------------------------------------
                save_to_db("progress",None,action="reset", lock=self.lock)
                self.initializeProgressDB()

                # ------------------- build data cache ----------------------
                tables = ["positions_low", "transactions_low", "positions_high", "transactions_high", "calculations"]
                table_rows = {t: load_from_db(t,lock=self.lock) for t in tables}
                cache = {}
                for table, rows in table_rows.items():
                    for row in rows:
                        if table in ("positions_low", "transactions_low"):
                            poolKey = row.get("Source name")
                        elif table in ("positions_high", "transactions_high"):
                            poolKey = row.get("Target name")
                        else:
                            poolKey = row.get("Pool")
                        if poolKey is None:
                            continue
                        for m in months:
                            if table == "calculations":
                                if row.get("dateTime") != m["dateTime"]:
                                    continue
                            else:
                                start = m["accountStart"] if table in ("positions_low", "positions_high") else m["tranStart"]
                                date = row.get("Date")
                                if not (start <= date <= m["endDay"]):
                                    continue
                            cache.setdefault(poolKey, {}).setdefault(table, {}).setdefault(m["dateTime"], []).append(row)
                
                self.cachedPoolCalculations = []
                if self.poolChangeDates.get("active",False): #iterate through pools that have custom calculation dates
                    runPools = []
                    for idx, pool in enumerate(self.pools):
                        if pool.get("poolName") in self.poolChangeDates or idx == 0: #if there is a date to calculate from. Needs at least one pool to run (idx 0)
                            runPools.append(pool)
                        else: #otherwise, get the calculations and avoid building a worker thread for nothing
                            for month in cache.get(pool.get("poolName")).get("calculations", {}):
                                self.cachedPoolCalculations.extend(cache.get(pool.get("poolName")).get("calculations", {}).get(month)) #add all calculations for the pool
                    self.pools = runPools #only run calculatable pools
                for idx, pool in enumerate(self.pools):
                    self.pools[idx]["cache"] = cache.get(pool.get("poolName"))
                    if self.poolChangeDates.get("active",False): #if the pool changes have been calculated, use it or set to current date if no changes occured
                        self.pools[idx]["earliestChangeDate"] = self.poolChangeDates.get(pool.get("poolName"),datetime.now())
                    else: #if pool changes have not been calculated but calculation requirements were imported, set to earliest global date
                        self.pools[idx]["earliestChangeDate"] =  self.earliestChangeDate 
                    newMonths = []
                    if not noCalculations: #if there are calculations, find all months before the data pull, and then pull those calculations
                        for month in months:
                            #if the calculations for the month have already been complete, pull the old data
                            if self.pools[idx]["earliestChangeDate"] > datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S"):
                                pass
                            else:
                                newMonths.append(month)
                    else:
                        newMonths = months
                    _ = updateStatus(pool.get("poolName"),len(newMonths),self.lock,status="Initialization")
                def initializeWorkerPool():
                    self.manager = Manager()
                    self.lock = self.manager.Lock()
                    self.workerStatusQueue = self.manager.Queue()
                    self.workerDBqueue = self.manager.Queue()
                    self.calcFailedFlag = self.manager.Value('b', False)
                    self.cancelCalcBtn.setEnabled(True) #only allows cancelling once the lock for the db exists

                    self.pool = Pool()
                    self.futures = []
                    executor.submit(self.watch_db)

                    commonData = {"noCalculations" : noCalculations,
                                    "months" : months, "fundList" : fundList
                                    }
                    
                    self.calcStartTime = datetime.now()
                    for pool in self.pools:
                        res = self.pool.apply_async(processPool, args=(pool, commonData,self.workerStatusQueue, self.workerDBqueue, self.calcFailedFlag))
                        self.futures.append(res)
                    self.pool.close()

                    self.timer.start(int(calculationPingTime * 0.25) * 1000) #check at 0.75 the ping time to prevent queue buildup
                gui_queue.put(lambda: initializeWorkerPool()) #puts on main thread
            except Exception as e:
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                gui_queue.put(lambda: self.importButton.setEnabled(True))
                print(f"Error occured running calculations: {e}")
                print("e.args:", e.args)
                # maybe also:
                print(traceback.format_exc())
        executor.submit(initalizeCalc)
    def initializeProgressDB(self):
        with self.lock:
            conn = sqlite3.connect(DATABASE_PATH)
            c = conn.cursor()
            c.execute("""
                CREATE TABLE IF NOT EXISTS progress (
                    pool STRING PRIMARY KEY,
                    completed INTEGER NOT NULL,
                    total INTEGER NOT NULL,
                    status STRING NOT NULL
                )
            """)
            conn.commit()
            conn.close()
    def watch_db(self):
        conn = sqlite3.connect(DATABASE_PATH)
        c = conn.cursor()
        while True:
            count = 0
            while not self.workerStatusQueue.empty() and count < 300:
                count += 1 #count to allow the loading bar to take the lock and update
                vars = self.workerStatusQueue.get()
                try:
                    failed = updateStatus(vars[0],vars[1],self.lock,status=vars[2],connection=conn)
                    if failed:
                        self.calcFailedFlag = failed
                except Exception as e:
                    trace = traceback.format_exc()
                    print(f"Error occured while attempting to run background worker status update: {e}. \n traceback: \n {trace}")
            try:
                with self.lock:
                    c.execute("SELECT * FROM progress")
                    cols = [d[0] for d in c.description]
                    statusLines = [dict(zip(cols, row)) for row in c.fetchall()]
                failed = []
                completed = []
                complete = 0
                total = 0
                for line in statusLines:
                    complete += line.get("completed",0)
                    total += line.get("total",0)
                    if line["status"] == "Failed":
                        failed.append(line)
                    elif line["status"] == "Completed":
                        completed.append(line)
                if len(failed) > 0:
                    print(f"Halting progress watch due to worker '{failed[0].get("pool","Bad Pull")}' failure.")
                    self.queue.append(-86) #will halt the queue
                    break
                elif len(completed) == len(self.pools):
                    print("All workers have declared complete.")
                    self.queue.append(100) #backup in case the numbers below fail
                    break
                if total != 0:
                    percent = int((complete / total) * 100)
                    self.queue.append(percent)
                    if complete >= total:
                        break
            except Exception as e:
                print(f"Error watching database: {e}")
                print(traceback.format_exc())
                pass
            time.sleep(calculationPingTime * 0.01)
        conn.close()
    def updateWorkerDB(self):
        try:
            time.sleep(8) #give time for calculations to be saved first in other thread
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
        except:
            print("connection failed")
        dbFailure = False
        maxFails = 4
        print("Initiating background database updates...")
        while True:
            try:
                results = self.workerDBqueue.get_nowait()  # non-blocking, safe for fixed queues
                data = results.get("data")
                failCount = 0
                while True:
                    try:
                        if results.get("type") == "insert":
                            save_to_db(data[0], data[1], action=data[2], connection=conn, lock=self.lock)
                            break
                        elif results.get("type") == "update":
                            with self.lock:
                                cursor.executemany(data[0], data[1])
                                conn.commit()
                                break
                        else:
                            print(f"\n\n Database data was not handled correctly: {results} \n\n")
                            break
                    except:
                        failCount += 1
                        print(f"Error updating database. Attempt {failCount} of {maxFails}")
                        if failCount > maxFails:
                            print("Error occured in delayed database updates. Calculation date will be reset")
                            dbFailure = True
                            break
            except queue.Empty:
                break  # all done; queue drained
            except Exception as e:
                print(f"Error occurred updating database from worker threads: {e}, {e.args}")
        print("Background database updates complete")
        if dbFailure: #will force a recalculation on the next opening since the database won't be accurate
            save_to_db(None,None,query="UPDATE history SET [lastCalculation] = ?", inputs=("Database Failure",), action="replace", lock=self.lock)
            self.lastImportDB[0]["lastCalculation"] = "Database Failure"
        conn.close()
    def update_from_queue(self):
        if self.queue:
            while self.queue: #cycle through the queue options to get most up to date value. Breaks out if complete or halted
                val = self.queue.pop(0)
                if val in (-86,100):
                    break
            self.calculationLoadingBar.setValue(val)
            timeElapsed = datetime.now() - self.calcStartTime
            secsElapsed = timeElapsed.total_seconds()
            loadingFraction = float(val) / 100 #decimal format percentage
            if loadingFraction > 0:
                est_total_secs = secsElapsed / loadingFraction
                secs_remaining = est_total_secs - secsElapsed
            else:
                secs_remaining = 0
            mins, secs = divmod(int(secs_remaining), 60)
            time_str = f"{mins}m {secs}s" # format as “Xm Ys” or “MM:SS”
            self.calculationLabel.setText(f"Estimated time remaining: {time_str}")
            if val >= 100:
                self.timer.stop()
                executor.submit(self.calcCompletion)
            elif val == -86:
                self.timer.stop()
                if self.cancel:
                    QMessageBox.warning(self,"Calculation Halted", "Calculations are being halted.")
                    self.cancel = False
                else:
                    QMessageBox.warning(self,"Calculation Failure", "A worker thread has failed. Calculations will not be properly completed.")
                self.pool.terminate()
                self.pool.join()
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                gui_queue.put(lambda: self.importButton.setEnabled(True))
                
    def calcCompletion(self):
        try:
            print("Checking worker completion...")
            executor.submit(self.updateWorkerDB)
            self.pool.join()
            print("All workers finished")
            
            calculations = []
            for fut in self.futures:
                try:
                    calculations.extend(fut.get())
                except Exception as e:
                    print(f"Error appending calculations: {e}")
            calculations.extend(self.cachedPoolCalculations)
            keys = []
            for row in calculations:
                for key in row.keys():
                    if key not in keys:
                        keys.append(key)
            save_to_db("calculations",calculations, keys=keys, lock=self.lock)
            try:
                apiPullTime = self.lastImportDB[0].get("lastImport")
                save_to_db(None,None,query="UPDATE history SET [lastCalculation] = ?", inputs=(apiPullTime,), action="replace", lock=self.lock)
            except:
                print("failed to update last calculation time")
            gui_queue.put( lambda: self.populate(self.calculationTable,calculations,keys = keys))
            gui_queue.put( lambda: self.buildReturnTable())
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
            gui_queue.put(lambda: self.importButton.setEnabled(True))
            print("Calculations complete.")
            save_to_db("progress",None,action="reset", lock=self.lock)
        except:
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
            gui_queue.put(lambda: self.importButton.setEnabled(True))
            print(f"Error occured processing calculation results. Resetting... ")
            print(traceback.format_exc())
    def checkVersion(self):
        self.currentVersionAccess = False
        self.globalVersion = None
        try:
            row = load_from_db("history")[0]
            self.globalVersion = row["currentVersion"]
            if row["currentVersion"] == currentVersion:
                self.currentVersionAccess = True
            else:
                QMessageBox.warning(self,"Outdated Version", f"Your current version ({currentVersion}) is outdated. The current version is {row["currentVersion"]}. \n" + 
                                    " Please request the newer version. \n Your version will run, but cannot access the shared data. This will be signifigantly slower and may have bugs/errors.")
        except:
            QMessageBox.warning(self,"Error checking version", f"An error occured checking that your version is up to date. This session has been granted limited access. \n "+
                                "Your session will run, but cannot access the shared data. This will be signifigantly slower and may have bugs/errors. \n \n Try restarting the app." +
                                " If this error persists, contact an admin.")
    def separateRowCode(self, label):
        header = re.sub(r'##\(.*\)##', '', label, flags=re.DOTALL)
        code = re.findall(r'##\(.*\)##', label, flags=re.DOTALL)[0]
        return header, code
    def sortStyleClicked(self, *args):
        if "NAV" in self.sortStyle.text():
            self.sortStyle.setText("Sort Style: Alphabetical")
        else:
            self.sortStyle.setText("Sort Style: NAV")
        self.buildReturnTable()
    def headerSortClosed(self):
        self.populateReturnsTable(self.currentTableData)
    def orderColumns(self,keys, exceptions = []):
        mode = self.tableBtnGroup.checkedButton().text()
        if mode == "Monthly Table":
            dates = [datetime.strptime(k, "%B %Y") for k in keys]
            dates = sorted(dates, reverse=True)
            keys = [d.strftime("%B %Y") for d in dates]
        elif mode == "Complex Table":
            newOrder = ["%", "NAV", "Commitment", "Unfunded","MTD","QTD","YTD"] + [f"{y}YR" for y in yearOptions] + ["ITD"]
            ordered = [h for h in newOrder if h in keys]
            ordered += [h for h in keys if h not in newOrder and h not in exceptions]
            keys = ordered
        return keys
    def populateReturnsTable(self, origRows: dict):
        self.buildTableLoadingBar.setValue(7)
        mode = self.tableBtnGroup.checkedButton().text()
        if not origRows:
            # nothing to show
            self.returnsTable.clear()
            self.returnsTable.setRowCount(0)
            self.returnsTable.setColumnCount(0)
            self.buildTableLoadingBox.setVisible(False)
            return

        rows = copy.deepcopy(origRows) #prevents alteration of self.returnsTableData
        for f in self.filterOptions:
            if f["key"] not in self.filterBtnExclusions and not self.filterRadioBtnDict[f["key"]].isChecked():
                to_delete = [k for k,v in rows.items() if v["dataType"] == "Total " + f["key"]]
                for k in to_delete:
                    rows.pop(k)
        to_delete = []
        for row in rows.keys():
            row_label, _ = self.separateRowCode(row)
            if row_label == "hiddenLayer":
                to_delete.append(row)
        for row in to_delete:
            rows.pop(row)
        
        self.filteredReturnsTableData = copy.deepcopy(rows) #prevents removal of dataType key for data lookup

        # 1) Build a flat list of row-entries:
        #    each entry = (fund_label, unique_code, row_dict)
        row_entries = []
        for fund_label, row_dict in rows.items():
            row_label, code = self.separateRowCode(fund_label)
            row_entries.append((row_label, code, row_dict))

        # 2) Determine columns exactly as before, using cleanedRows for header order
        cleaned = {fund: d.copy() for fund, _, d in row_entries}
        for d in cleaned.values():
            d.pop("dataType", None)

        if not self.headerSort.active or mode == "Monthly Table":
            col_keys = set()
            for d in cleaned.values():
                col_keys |= set(d.keys())
            col_keys = list(col_keys)

            exceptions = ["Return", "Ownership", "MDdenominator", "Monthly Gain"]
            col_keys = self.orderColumns(col_keys, exceptions=exceptions)
            if mode == "Complex Table":
                allKeys = col_keys.copy()
                allKeys.extend(exceptions) #all key options for the header selections
                self.headerSort.set_items(allKeys,[item for item in allKeys if item not in exceptions])
                self.headerSort.setEnabled(True)
            else:
                self.headerSort.setEnabled(False)
        else:
            col_keys = self.headerSort.popup.get_checked_sorted_items()
            self.headerSort.setEnabled(True)
        self.filteredHeaders = col_keys
        # 3) Resize & set horizontal headers (we no longer call setVerticalHeaderLabels)
        self.returnsTable.setRowCount(len(row_entries))
        self.returnsTable.setColumnCount(len(col_keys))
        self.returnsTable.setHorizontalHeaderLabels(col_keys)


        # 4) Populate each row
        for r, (fund_label, code, row_dict) in enumerate(row_entries):
            # pull & remove dataType for coloring
            dataType = row_dict.pop("dataType", "")

            if dataType != "benchmark": #benchmark will use previous rounds color
                startColor = (160, 160, 160)
                if dataType == "Total":
                    bg =  QColor(*tuple(
                        int(startColor[i] * 0.8)
                        for i in range(3)
                    ))
                elif dataType == "benchmark":
                    bg = QColor(*(182, 205, 245))
                else:
                    depth      = code.count("::") if dataType != "Total Fund" else code.count("::") + 1
                    # if len(re.findall(r'##\((.*?)\)##', code, flags=re.DOTALL)[0]) > 0:
                    #     depth -= 1
                    maxDepth   = max(len(self.sortHierarchy.checkedItems()),1)
                    cRange     = 255 - startColor[0]
                    ratio      = (depth / maxDepth) if maxDepth != 0 else 1
                    color = tuple(
                        int(startColor[i] + cRange * ratio)
                        for i in range(3)
                    )
                    bg = QColor(*color)
            
                

            # — vertical header: only show the fund, stash the code —
            hdr = QTableWidgetItem(fund_label)
            hdr.setData(Qt.UserRole, code)
            if bg:
                hdr.setBackground(QBrush(bg))
                if dataType == "benchmark":
                    hdr.setBackground(QBrush(QColor("0000FF")))
            self.returnsTable.setVerticalHeaderItem(r, hdr)

            # — fill cells —
            for c, col in enumerate(col_keys):
                raw = row_dict.get(col, "")
                if raw not in (None, "", "None"):
                    try:
                        v = round(float(raw), 2)
                        if c in percent_headers or (mode == "Monthly Table" and self.returnOutputType.currentText() in percent_headers):
                            text = f"{v:.2f}%"
                        else:
                            text = f"{v:,.2f}"
                    except:
                        text = str(raw)
                else:
                    text = ""

                item = QTableWidgetItem(text)
                if text:
                    # store raw number for sorting or later retrieval
                    item.setData(Qt.UserRole, v)
                if bg:
                    item.setBackground(QBrush(bg))
                if dataType == "benchmark":
                    item.setForeground(QColor(0,0,255))
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.returnsTable.setItem(r, c, item)
        self.buildTableLoadingBox.setVisible(False)
    def populate(self, table, rows, keys = None):
        if not rows:
            return
        if keys is None:
            headers = list(rows[0].keys())
        else:
            headers = list(keys)

        self.calcTableModel = DictListModel(rows,headers, self)
        table.setModel(self.calcTableModel)

def submitAPIcall(self, fn, *args, **kwargs):
    fut = APIexecutor.submit(fn, *args, **kwargs)
    self.apiFutures.add(fut)
    fut.add_done_callback(self.apiFutures.discard)  # remove when done
    return

@attach_logging_to_class
class transactionApp(QWidget):
    def __init__(self, start_index=0, apiKey = None):
        super().__init__()
        self.setWindowTitle('Transaction Compare App')
        self.setGeometry(100, 100, 1000, 600)

        os.makedirs(ASSETS_DIR, exist_ok=True)
        self.start_index = start_index
        self.api_key = apiKey
        self.filterCallLock = False
        self.cancel = False
        self.lock = None
        self.tableWindows = {}
        self.dataTimeStart = datetime(2000,1,1)
        self.earliestChangeDate = datetime(datetime.now().year,datetime.now().month + 1,datetime.now().day)
        self.poolChangeDates = {"active" : False}
        self.currentTableData = None
        self.fullLevelOptions = {}
        self.buildTableCancel = None
        self.buildTableFuture = None
        self.cFundsCalculated = False
        self.previousGrouping = []

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_from_queue)
        self.queue = []

        # main stack
        self.main_layout = QVBoxLayout()
        appStyle = ("""
                        QWidget#borderFrame {
                            border: 2px solid #3E85E9;
                            border-radius: 4px;
                            padding: 4px;
                        }
                        QWidget#titleBox {
                            border: 4px solid #0665EA;
                            border-radius: 5px;
                            padding: 4px;
                        }
                        QWidget#mainPage, QMessageBox, QDialog {
                            background-color: #383838
                        }
                        QPushButton {
                            background-color: #3E85E9;
                            border: 2px solid transparent;
                            border-radius: 12px;
                            padding: 4px
                        }
                        QPushButton:hover {
                                background-color: #1771EE;
                        }
                        QPushButton#exportBtn {
                            background-color: #51AE2B;
                        }
                        QPushButton#exportBtn:hover {
                            background-color: #429321;
                        }
                        QPushButton#cancelBtn {
                            background-color: #D63131;
                        }
                        QLabel, QRadioButton, QCheckBox, QProgressBar {
                            color: white
                        }
                        QTableWidget, QWidget#subPanel, QHeaderView::corner, QTableCornerButton::section {
                        background-color : #514F4F
                        }
                        QHeaderView::section {
                            background-color: #A8A2A2;
                        }
                        QListWidget {
                            background-color : #514F4F;
                            color: white
                        }
                        QLineEdit{
                            border: 2px solid transparent;
                            border-radius: 12px;
                            background-color: #514F4F;
                            color : white;
                        }
                        QComboBox {
                            background-color: #514F4F;
                            color : white;
                        }
                    """)
        self.setStyleSheet(appStyle)
        self.setObjectName("mainPage")
        self.stack = QStackedWidget()
        self.init_global_widgets()

        self.init_api_key_page() #0
        self.init_returns_page() #1
        self.init_calculation_page() #2

        self.stack.setCurrentIndex(start_index)
        self.main_layout.addWidget(self.stack)
        self.setLayout(self.main_layout)
    def init_global_widgets(self):
        headerBox = QWidget()
        headerLayout = QHBoxLayout()
        self.lastImportLabel = QLabel("Last Data Import: ")
        headerLayout.addWidget(self.lastImportLabel)
        headerLayout.addStretch()
        headerLayout.addWidget(QLabel(f"Version: {currentVersion}"))
        headerBox.setLayout(headerLayout)
        self.main_layout.addWidget(headerBox)
        self.apiLoadingBarBox = QWidget()
        t2 = QVBoxLayout()
        t2.addWidget(QLabel("Pulling transaction and account data from server..."))
        self.apiLoadingBar = QProgressBar()
        self.apiLoadingBar.setRange(0,100)
        t2.addWidget(self.apiLoadingBar)
        self.apiLoadingBarBox.setLayout(t2)
        self.apiLoadingBarBox.setVisible(False)
        self.main_layout.addWidget(self.apiLoadingBarBox)
        loadLay = QGridLayout()
        self.calculationLoadingBar = QProgressBar()
        self.calculationLoadingBar.setRange(0,100)
        self.calculationLabel = QLabel()
        self.cancelCalcBtn = QPushButton("Cancel Calculations")
        self.cancelCalcBtn.setObjectName("cancelBtn")
        self.cancelCalcBtn.setEnabled(False)
        self.cancelCalcBtn.clicked.connect(self.cancelCalc)
        loadLay.addWidget(self.calculationLabel,0,0,1,5)
        loadLay.addWidget(self.calculationLoadingBar, 1,0, 1,5)
        loadLay.addWidget(self.cancelCalcBtn, 2, 2)
        self.calculationLoadingBox = QWidget()
        self.calculationLoadingBox.setLayout(loadLay)
        self.calculationLoadingBox.setVisible(False)
        self.main_layout.addWidget(self.calculationLoadingBox)

    def init_api_key_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        self.api_label = QLabel('Enter Dynamo API Key:')
        self.api_input = QLineEdit()
        btn = QPushButton('Submit')
        btn.clicked.connect(self.check_api_key)
        layout.addWidget(self.api_label)
        layout.addWidget(self.api_input)
        layout.addWidget(btn)
        page.setLayout(layout)
        self.stack.addWidget(page)

    def init_calculation_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        self.info_label = QLabel('Results')
        layout.addWidget(self.info_label)
        btn_to_form = QPushButton('Go to Results')
        btn_to_form.clicked.connect(lambda: self.stack.setCurrentIndex(1))
        layout.addWidget(btn_to_form)
        

        hl = QHBoxLayout()
        self.calculationTable = QTableView(); self.calculationTable.setSortingEnabled(True)
        hl.addWidget(self.calculationTable)
        layout.addLayout(hl)

        

        page.setLayout(layout)
        self.stack.addWidget(page)

    def init_returns_page(self):
        
        page = QWidget()
        layout = QVBoxLayout()
        controlsBox = QWidget()
        controlsLayout = QHBoxLayout()
        controlsLayout.addStretch(1)
        self.importButton = QPushButton('Reimport Data')
        self.importButton.clicked.connect(self.beginImport)
        clearButton = QPushButton('Clear All Cached Data')
        clearButton.clicked.connect(self.resetData)
        if not demoMode:
            controlsLayout.addWidget(clearButton, stretch=0)
        controlsLayout.addWidget(self.importButton, stretch=0)
        btn_to_results = QPushButton('See Calculation Database')
        btn_to_results.clicked.connect(self.show_results)
        controlsLayout.addWidget(btn_to_results, stretch=0)
        self.exportBtn = QPushButton("Export Current Table to Excel")
        self.exportBtn.clicked.connect(self.exportCurrentTable)
        self.exportBtn.setObjectName("exportBtn")
        controlsLayout.addWidget(self.exportBtn, stretch=0)
        controlsLayout.addStretch(1)
        controlsBox.setLayout(controlsLayout)
        layout.addWidget(controlsBox)

        optionsBox = QWidget()
        optionsBox.setObjectName("borderFrame")
        optionsGrid = QGridLayout()
        optionsTitle = QLabel("Options")
        optionsTitle.setObjectName("titleBox")
        optionsGrid.addWidget(optionsTitle,0,0,2,1)
        self.tableBtnGroup = QButtonGroup()
        self.complexTableBtn = QRadioButton("Complex Table")
        self.complexTableBtn.setVisible(False)
        
        
        self.monthlyTableBtn = QRadioButton("Monthly Table")
        self.monthlyTableBtn.setVisible(False)
        buttonBox = QWidget()
        buttonLayout = QVBoxLayout()
        for idx, rb in enumerate((self.monthlyTableBtn,self.complexTableBtn)):
            self.tableBtnGroup.addButton(rb)
            #rb.toggled.connect(self.updateTableType)
            buttonLayout.addWidget(rb)
        self.monthlyTableBtn.setChecked(True)
        self.complexTableBtn.setEnabled(False)
        self.returnOutputType = QComboBox()
        self.returnOutputType.addItems(tranAppHeaderOptions)
        self.returnOutputType.currentTextChanged.connect(self.buildReturnTable)
        self.returnOutputType.setVisible(False)
        self.dataTypeBox = QWidget()
        dataTypeLayout = QHBoxLayout()
        dataTypeLayout.addWidget(self.returnOutputType)
        self.dataTypeBox.setLayout(dataTypeLayout)
        buttonLayout.addWidget(self.dataTypeBox)
        buttonBox.setLayout(buttonLayout)
        optionsGrid.addWidget(buttonBox, 0,1,2,1)
        self.tableBtnGroup.buttonClicked.connect(self.buildReturnTable)
        
        self.dataStartSelect = simpleMonthSelector()
        self.dataEndSelect = simpleMonthSelector()
        for idx, [text, CB] in enumerate((["Start: ", self.dataStartSelect], ["End: ", self.dataEndSelect])):
            optionsGrid.addWidget(QLabel(text),idx,2)
            optionsGrid.addWidget(CB,idx,3)
        self.sortHierarchy = MultiSelectBox()
        self.sortHierarchy.hierarchyMode()
        self.sortHierarchy.addItem("Pool")
        self.sortHierarchy.setCheckedItem("Pool")
        self.sortHierarchy.setEnabled(False)
        self.sortHierarchy.setVisible(False)
        self.sortHierarchy.popup.closed.connect(self.groupingChange)
        optionsGrid.addWidget(self.sortHierarchy,1,5)
        self.consolidateFundsBtn = QRadioButton("Consolidate Funds")
        self.consolidateFundsBtn.setChecked(True)
        self.consolidateFundsBtn.setVisible(False)
        self.consolidateFundsBtn.clicked.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.consolidateFundsBtn,0,6)
        self.exitedFundsBtn = QRadioButton("Show Exited Funds (Cannot turn off)")
        self.exitedFundsBtn.setChecked(False)
        self.exitedFundsBtn.setEnabled(False) #remove later
        self.exitedFundsBtn.setChecked(True)  #remove later
        self.exitedFundsBtn.setVisible(False)
        optionsGrid.addWidget(self.exitedFundsBtn,1,6)
        self.headerSort = SortButtonWidget()
        self.headerSort.popup.popup_closed.connect(self.headerSortClosed)
        self.headerSort.setVisible(False)
        optionsGrid.addWidget(self.headerSort,0,7,2,1)
        optionsBox.setLayout(optionsGrid)
        layout.addWidget(optionsBox)

        mainFilterBox = QWidget()
        mainFilterBox.setObjectName("borderFrame")
        mainFilterLayout = QGridLayout()
        filterTitle = QLabel("Filters")
        filterTitle.setObjectName("titleBox")
        mainFilterLayout.addWidget(filterTitle,0,0,2,1)

        self.filterOptions = [
                            {"key": "Pool",           "name": "Pool", "dataType" : "Total Pool" , "dynNameLow" : "Source name", "dynNameHigh" : "Target name"},                            
                        ]
        self.filterBtnExclusions = ["Investor","Classification", nameHier["Family Branch"]["local"]]
        self.highOnlyFilters = ["Investor", nameHier["Family Branch"]["local"]]
        self.filterDict = {}
        self.filterRadioBtnDict = {}
        self.filterBtnGroup = QButtonGroup()
        self.filterBtnGroup.setExclusive(False)
        for col, filter in enumerate(self.filterOptions, start=1):
            if filter["key"] not in self.filterBtnExclusions:
                #investor level is not filterable. It is total portfolio or shows the investors data
                self.filterRadioBtnDict[filter["key"]] = QCheckBox(f"{filter["name"]}:")
                self.filterRadioBtnDict[filter["key"]].setChecked(True)
                self.filterBtnGroup.addButton(self.filterRadioBtnDict[filter["key"]])
                mainFilterLayout.addWidget(self.filterRadioBtnDict[filter["key"]],0, col)
            else:
                mainFilterLayout.addWidget(QLabel(f"{filter["name"]}:"), 0, col)
            if filter["key"] != "Fund":
                self.sortHierarchy.addItem(filter["key"])
            self.filterDict[filter["key"]] = MultiSelectBox()
            self.filterDict[filter["key"]].popup.closed.connect(lambda: self.filterUpdate())
            
            mainFilterLayout.addWidget(self.filterDict[filter["key"]],1,col)
        self.filterBtnGroup.buttonToggled.connect(self.filterBtnUpdate)
        mainFilterBox.setLayout(mainFilterLayout)
        layout.addWidget(mainFilterBox)
        t1 = QVBoxLayout() #build table loading bar
        self.buildTableLoadingBox = QWidget()
        t1.addWidget(QLabel("Building transactions table..."))
        self.buildTableLoadingBar = QProgressBar()
        self.buildTableLoadingBar.setRange(0,8)
        t1.addWidget(self.buildTableLoadingBar)
        self.buildTableLoadingBox.setLayout(t1)
        self.buildTableLoadingBox.setVisible(False)
        layout.addWidget(self.buildTableLoadingBox)
        self.returnsTable = SmartStretchTable() #table
        self.returnsTable.setSelectionMode(QTableWidget.ContiguousSelection)  # Required
        self.returnsTable.setSelectionBehavior(QTableWidget.SelectItems)
        layout.addWidget(self.returnsTable)
        unDataBox = QWidget()
        unDataLayout = QHBoxLayout()
        unDataLayout.addStretch(1)
        self.viewUnderlyingDataBtn = QPushButton("View Underlying Data")
        self.viewUnderlyingDataBtn.clicked.connect(self.viewUnderlyingData)
        unDataLayout.addWidget(self.viewUnderlyingDataBtn,stretch=0)
        unDataLayout.addStretch(1)
        unDataBox.setLayout(unDataLayout)
        layout.addWidget(unDataBox)
        


        page.setLayout(layout)
        self.stack.addWidget(page)

        self.pullLevelNames()
        self.updateMonthOptions()
        if self.start_index != 0:
            self.filterUpdate()
        self.dataEndSelect.currentTextChanged.connect(self.buildReturnTable)
        self.dataStartSelect.currentTextChanged.connect(self.buildReturnTable)
    def init_data_processing(self):
        self.calcSubmitted = False
        lastImportDB = load_from_db("history", db=TRAN_DATABASE_PATH) if len(load_from_db("history", db=TRAN_DATABASE_PATH)) == 1 else None
        if lastImportDB is None:
            print("No previous import found")
            #pull data is there is no data pulled yet
            executor.submit(lambda: self.pullData())
        else:
            lastImportString = lastImportDB[0]["lastImport"]
            lastImport = datetime.strptime(lastImportString, "%B %d, %Y @ %I:%M %p")  
            self.lastImportLabel.setText(f"Last Data Import: {lastImportString}")
            now = datetime.now()
            if lastImport.month != now.month or now > (lastImport + relativedelta(hours=2)):
                print(f"Reimporting due to two hour data gap. \n     Last import: {lastImport}\n    Current time: {now}")
                #pull data if in a new month or 1 days have elapsed
                executor.submit(self.pullData)
            elif lastImportDB[0]["lastImport"] != lastImportDB[0].get("lastCalculation", "None"):
                self.earliestChangeDate = datetime.strptime(lastImportDB[0].get("changeDate"), "%B %d, %Y @ %I:%M %p")
                self.processFunds()
                self.calculateReturn()
            else:
                calculations = load_from_db("calculations", db=TRAN_DATABASE_PATH)
                self.processFunds()
                if calculations != []:
                    self.populate(self.calculationTable,calculations)
                    self.buildReturnTable()
                else:
                    self.calculateReturn()
    def cancelCalc(self, *_):
        _ = updateStatus("DummyFail",99,lock=self.lock, status="Failed", db=TRAN_DATABASE_PATH)
        self.cancel = True
    def viewUnderlyingData(self,*_):
        row = self.returnsTable.currentRow()
        col = self.returnsTable.currentColumn()
        key = list(self.filteredReturnsTableData.keys())[row]
        vh_item = self.returnsTable.verticalHeaderItem(row)
        row = vh_item.text() if vh_item else f"Row {row}"

        # Get the horizontal (column) header text
        hh_item = self.returnsTable.horizontalHeaderItem(col)
        col = hh_item.text() if hh_item else f"Column {col}"
        self.selectedCell = {"entity": row, "month" : col, "rowKey" : key, "dataType" : self.filteredReturnsTableData[key]["dataType"] }
        try:
            window = underlyingDataWindow(parentSource=self, db=TRAN_DATABASE_PATH)
            self.udWindow = window
            if window.success:
                window.show()
        except Exception as e:
            print(f"Error in data viewing window: {e} {traceback.format_exc()}")
    def exportCurrentTable(self,*_):
        # helper to darken a 6-digit hex color by a given factor
        def darken_color(hex_color, factor=0.01):
            h = hex_color.strip("#")
            r = int(h[0:2], 16)
            g = int(h[2:4], 16)
            b = int(h[4:6], 16)
            dr = max(0, int(r * factor))
            dg = max(0, int(g * factor))
            db = max(0, int(b * factor))
            return f"{dr:02X}{dg:02X}{db:02X}"
        # 1) prompt user
        path, _ = QFileDialog.getSaveFileName(
            self, "Save as…", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        def processExport():
            try:
                data = self.currentTableData  # dict of dicts

                # 2) determine hierarchy levels present
                all_types = {row.get("dataType") for row in data.values()}
                if self.sortHierarchy.checkedItems() != []:
                    full_hierarchy = ["Total"] + ["Total " + level for level in self.sortHierarchy.checkedItems()] + ["Total Pool"]
                else:
                    full_hierarchy = ["Total", "Total assetClass", "Total Pool"]
                hierarchy_levels = [lvl for lvl in full_hierarchy if lvl in all_types]
                num_hier = len(hierarchy_levels)

                # 3) dynamic data columns minus "dataType"
                all_cols = {
                    k for row in data.values() for k in row.keys()
                    if k != "dataType"
                }

                sorted_cols = self.orderColumns(all_cols)

                # 4) create workbook or add sheet if already exists
                if os.path.exists(path):
                    wb = load_workbook(path)
                    # Create a unique sheet name for export
                    import datetime
                    base_name = "Export"
                    i = 1
                    while True:
                        sheet_name = f"{base_name}{i}"
                        if sheet_name not in wb.sheetnames:
                            break
                        i += 1
                    ws = wb.create_sheet(sheet_name)
                else:
                    wb = Workbook()
                    ws = wb.active

                rowStart = 4
                # 5) header row
                for idx, lvl in enumerate(hierarchy_levels, start=1):
                    ws.cell(row=rowStart, column=idx, value=lvl)
                for idx, colname in enumerate(sorted_cols, start=num_hier+1):
                    ws.cell(row=rowStart, column=idx, value=colname)

                split_cell = f"{get_column_letter(num_hier+1)}4"
                ws.freeze_panes = split_cell

                # 7) populate rows
                for r, (row_name, row_dict) in enumerate(data.items(), start=rowStart + 1):
                    row_name, code = self.separateRowCode(row_name)
                    dtype = row_dict.get("dataType")
                    level = hierarchy_levels.index(dtype) if dtype in hierarchy_levels else 0

                    # fills
                    # Default: light gray
                    data_color = "A0A0A0"
                    header_color = data_color

                    # Reproduce the main table cell & header coloring logic from lines 2567-2586
                    if dtype == "Total":
                        # "Total" rows: shade startColor darker (0.8)
                        data_color = "CDCDCD"  # A0A0A0 at 80% gray
                        header_color = "B7B7B7"  # slightly darker for headers
                    elif dtype == "benchmark":
                        # "benchmark" rows: a blue as in GUI
                        data_color = "B6CDF5"
                        header_color = "A3B8DB"  # a harder blue for headers
                    else:
                        # hierarchy coloring as in the GUI
                        base_rgb = (160,160,160)
                        # In gui: 'depth = code.count("::") if dataType != "Total Fund" else code.count("::") + 1'
                        # But "Total Fund" doesn't exist here; treat as "else"
                        if dtype != "Total Fund":
                            depth = code.count("::")
                        else:
                            depth = code.count("::") + 1
                        maxDepth = max(len(self.sortHierarchy.checkedItems()),1)
                        cRange = 255 - base_rgb[0]
                        ratio = (depth / maxDepth) if maxDepth != 0 else 1
                        def clamp(x): return max(0,min(255,int(x)))
                        def rgb_to_hex(rgb):
                            return "".join(f"{clamp(x):02X}" for x in rgb)
                        # Table: color = int(startColor[i] + cRange * ratio)
                        color_rgb = tuple(
                            clamp(base_rgb[i] + cRange * ratio)
                            for i in range(3)
                        )
                        data_color = rgb_to_hex(color_rgb)
                        # Header: make it "harder" (darker) by multiplying ratio by 1.08 (max 1.0)
                        header_ratio = min(ratio * 1.08, 1.0)
                        header_rgb = tuple(
                            clamp(base_rgb[i] + cRange * header_ratio)
                            for i in range(3)
                        )
                        header_color = rgb_to_hex(header_rgb)

                    # Even/odd row striping: darken data color a bit for odd rows
                    if r % 2 == 1:
                        def hex_to_rgb(h): return tuple(int(h[i:i+2],16) for i in (0,2,4))
                        cur_rgb = hex_to_rgb(data_color)
                        data_color = rgb_to_hex(tuple(
                            int(x*0.93) for x in cur_rgb
                        ))
                        # Make header match "hardness": use the same darkening factor, but even slightly darker
                        header_rgb = hex_to_rgb(header_color)
                        header_color = rgb_to_hex(tuple(
                            int(x*0.91) for x in header_rgb
                        ))

                    data_fill   = PatternFill("solid", data_color, data_color)
                    header_fill = PatternFill("solid", header_color, header_color)

                    # spread header fill across hierarchy cols
                    data_start = num_hier + 1
                    for col in range(level+1, data_start):
                        cell = ws.cell(row=r, column=col, value=row_name if col==level+1 else None)
                        cell.fill = header_fill
                        if col == level+1:
                            cell.alignment = Alignment(indent=level)

                    # data cells with proper formatting
                    for c, colname in enumerate(sorted_cols, start=data_start):
                        val = row_dict.get(colname, None)
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.fill = data_fill
                        if isinstance(val, (int, float)):
                            if colname not in percent_headers:
                                # show with commas, two decimals
                                cell.number_format = "#,##0.00"
                            else:
                                # interpret val as percentage (e.g. 10.5 → 10.5%)
                                cell.value = val / 100.0
                                cell.number_format = "0.00%"

                # 8) autofit column widths
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is not None:
                            text = str(cell.value)
                            max_len = max(max_len, len(text))
                    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

                appliedFilters = {}
                for filter in self.filterOptions:
                    if self.filterDict[filter["key"]].checkedItems() != []:
                        appliedFilters[filter["key"]] = self.filterDict[filter["key"]].checkedItems()
                filterStart = num_hier
                if self.filterDict[filter["key"]].checkedItems() != []: #only write if there are filters applied
                    cell= ws.cell(row=1, column=filterStart, value="Filter:")
                    cell.font = Font(bold=True)
                    cell = ws.cell(row=2, column=filterStart, value="Selections:")
                    cell.font = Font(bold=True)
                    for idx, key in enumerate(appliedFilters, start=filterStart + 1):
                        cell = ws.cell(row=1, column=idx, value=key)
                        cell.alignment = Alignment(wrap_text=True)
                        cell = ws.cell(row=2, column=idx, value=", ".join(appliedFilters[key]))
                        cell.alignment = Alignment(wrap_text=True)

            
                wb.save(path)
            except Exception as e:
                gui_queue.put(lambda error=e, trace = traceback.format_exc(): QMessageBox.critical(self, "Save error", trace))
            else:
                gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
                gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
        executor.submit(processExport)
    def processFunds(self):
        self.cFundsCalculated = True
        self.sleeveFundLinks = {}
        self.cFundToFundLinks = {}
        self.pools = []
        poolList = set()
        funds = load_from_db("funds", db=TRAN_DATABASE_PATH)
        if funds != []:
            consolidatorFunds = {}
            for row in funds: #find sleeve values and consolidated funds
                assetClass = row["assetClass"]
                subAssetClass = row["subAssetClass"]
                sleeve = row["sleeve"]
                if row.get("Fundpipelinestatus") is not None and "Z - Placeholder" in row.get("Fundpipelinestatus"):
                    consolidatorFunds[row["Name"]] = {"cFund" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass, "sleeve" : sleeve}
                    self.cFundToFundLinks[row["Name"]] = []
                if row["sleeve"] not in self.sleeveFundLinks:
                    self.sleeveFundLinks[row["sleeve"]] = [row["Name"]]
                else:
                    self.sleeveFundLinks[row["sleeve"]].append(row["Name"])
                if row["Fundpipelinestatus"] == "I - Internal":
                    self.pools.append({"poolName" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass})
                    poolList.add(row["Name"])
            self.consolidatedFunds = {}
            for row in funds: #assign funds to their consolidators
                if row.get("Parentfund") in consolidatorFunds:
                    self.consolidatedFunds[row["Name"]] = consolidatorFunds.get(row.get("Parentfund"))
                    self.cFundToFundLinks[row.get("Parentfund")].append(row["Name"])
            self.fullLevelOptions["Pool"] = list(poolList)
        else:
            self.consolidatedFunds = {}
    def filterBtnUpdate(self, button, checked):
        if not self.filterCallLock:
            self.buildTableLoadingBox.setVisible(True)
            self.buildTableLoadingBar.setValue(1)
            self.filterCallLock = True
            reloadRequired = False
            for filter in self.filterOptions:
                if filter["key"] not in self.filterBtnExclusions:
                    if not self.filterRadioBtnDict[filter["key"]].isChecked():
                        if self.filterDict[filter["key"]].checkedItems() != []:
                            reloadRequired = True #rebuild the table only if filter selections are being removed
                        self.filterDict[filter["key"]].clearSelection()
                        self.filterDict[filter["key"]].setEnabled(False)
                    else:
                        self.filterDict[filter["key"]].setEnabled(True)
            self.filterCallLock = False
            if reloadRequired or self.currentTableData is None:
                self.buildReturnTable()
            else:
                self.populateReturnsTable(self.currentTableData)
    def resetData(self,*_):
        save_to_db("calculations",None,action="reset", db=TRAN_DATABASE_PATH) #reset calculations so new data will be freshly calculated
        self.poolChangeDates = {"active" : False}
        executor.submit(self.pullData)
    def beginImport(self, *_):
        executor.submit(self.pullData)
    def updateMonthOptions(self):
        start = self.dataTimeStart
        end = datetime.now() - relativedelta(months=1) + relativedelta(hours=8)
        #ends on the previous month. Adds a few hours so index will still be before it and count as a month on the 1st
        index = start
        monthList = []
        while index < end:
            monthList.append(datetime.strftime(index,"%B %Y"))
            index += relativedelta(months=1)
        self.dataEndSelect.addItems(monthList)
        self.dataEndSelect.setCurrentText(monthList[-1])
        self.dataStartSelect.addItems(monthList)
        self.dataStartSelect.setCurrentText(monthList[0])
    def buildReturnTable(self, *_):
        self.buildTableLoadingBox.setVisible(True)
        self.buildTableLoadingBar.setValue(2)
        if not self.cFundsCalculated:
            self.processFunds()
        def buildTable(cancelEvent):
            try:
                print("Building transactions table...")
                self.currentTableData = None #resets so a failed build won't be used
                
                if self.tableBtnGroup.checkedButton().text() == "Complex Table":
                    gui_queue.put(lambda: self.returnOutputType.setCurrentText("Return"))
                    gui_queue.put(lambda: self.dataTypeBox.setVisible(False))
                else:
                    gui_queue.put(lambda: self.dataTypeBox.setVisible(True))
                parameters = ["Total Pool"]
                condStatement = " WHERE [Investor] = ? "
                for filter in self.filterOptions:
                    if filter["key"] != "Investor" and filter["key"] != nameHier["Family Branch"]["local"]:
                        if self.filterDict[filter["key"]].checkedItems() != []:
                            paramTemp = self.filterDict[filter["key"]].checkedItems()
                            for param in paramTemp:
                                parameters.append(param)
                            placeholders = ','.join('?' for _ in paramTemp)
                            condStatement += f" AND [{filter["key"]}] IN ({placeholders})"
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(3))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                data = load_from_db("calculations",condStatement, tuple(parameters), lock=self.lock, db=TRAN_DATABASE_PATH)
                output = {"Total##()##" : {}}
                #output , data = self.calculateUpperLevels(output,data)
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(4))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                complexOutput = copy.deepcopy(output)
                multiPoolFunds = {}
                dataOutputType = self.returnOutputType.currentText()
                for entry in data:
                    if (datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") >  datetime.strptime(self.dataEndSelect.currentText(),"%B %Y") or 
                        datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") <  datetime.strptime(self.dataStartSelect.currentText(),"%B %Y")):
                        #don't build in data outside the selection
                        continue
                    date = datetime.strftime(datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S"), "%B %Y")
                    Dtype = entry["Calculation Type"]
                    level = entry.get("Pool") + "##(" + entry.get("Pool") + ")##"

                    
                    if level not in output.keys():
                        output[level] = {}
                    if entry.get(dataOutputType) not in (None,"None",""):
                        if date not in output[level].keys():
                            #creates value if not exists. If it is not return percent, sums the values
                            output[level][date] = float(entry.get(dataOutputType))
                        elif dataOutputType not in ("Return", "Ownership"):
                            output[level][date] += float(entry.get(dataOutputType))
                        else: #should only reach here if two calculations exist of the same exact row which needs special handling of the return
                            if level not in multiPoolFunds:
                                multiPoolFunds[level] = [entry,]
                            else:
                                multiPoolFunds[level].append(entry)
                    if "dataType" not in output[level].keys():
                        output[level]["dataType"] = Dtype
                if multiPoolFunds and dataOutputType == "Return": #must iterate through data again to correct for returns of multi pool funds
                    multiData = {}
                    for rowKey in multiPoolFunds: #instantiate multiData with the row
                        multiData[rowKey] = {}
                        # date = multiPoolFunds.get(rowKey).get("dateTime")
                        # multiData[rowKey][date] = {"MDdenominator" : 0, "Monthly Gain" : 0}
                    for entry in data:
                        if (datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") >  datetime.strptime(self.dataEndSelect.currentText(),"%B %Y") or 
                            datetime.strptime(entry["dateTime"], "%Y-%m-%d %H:%M:%S") <  datetime.strptime(self.dataStartSelect.currentText(),"%B %Y")):
                            #don't build in data outside the selection
                            continue
                        if entry.get("rowKey") in multiData: #only occurs for the multifunds
                            #sums all gains and MDden for a row for a month
                            dateTime = entry.get("dateTime")
                            if dateTime not in multiData[entry.get("rowKey")]:
                                multiData[entry.get("rowKey")][entry.get("dateTime")] = {"MDdenominator" : float(entry.get("MDdenominator")), "Monthly Gain" : float(entry.get("Monthly Gain"))}
                            else:
                                multiData[entry.get("rowKey")][entry.get("dateTime")]["MDdenominator"] += float(entry.get("MDdenominator"))
                                multiData[entry.get("rowKey")][entry.get("dateTime")]["Monthly Gain"] += float(entry.get("Monthly Gain"))
                    for rowKey in multiData: #set proper return values
                        for date in multiData.get(rowKey):
                            strDate = datetime.strftime(datetime.strptime(date, "%Y-%m-%d %H:%M:%S"), "%B %Y")
                            MDden = multiData.get(rowKey).get(date).get("MDdenominator")
                            returnVal = multiData.get(rowKey).get(date).get("Monthly Gain") / MDden * 100 if MDden != 0 else 0
                            output[rowKey][strDate] = returnVal
                            if self.tableBtnGroup.checkedButton().text() == "Complex Table" and strDate == self.dataEndSelect.currentText():
                                complexOutput[rowKey]["Return"] = returnVal
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(5))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                if self.tableBtnGroup.checkedButton().text() == "Complex Table":
                    output = self.calculateComplexTable(output,complexOutput)
                gui_queue.put(lambda: self.buildTableLoadingBar.setValue(6))
                if cancelEvent.is_set(): #exit if new table build request is made
                    return
                outputKeys = output.keys()
                deleteKeys = []
                for key in outputKeys:
                    if len(output[key].keys()) == 0:
                        deleteKeys.append(key)
                for key in deleteKeys:
                    output.pop(key)
                gui_queue.put(lambda: self.populateReturnsTable(output))
                self.currentTableData = output
            except Exception as e:
                tracebackMsg = traceback.format_exc()
                gui_queue.put(lambda error = e: QMessageBox.warning(self, "Error building returns table", f"Error: {error}. {error.args}. Data entry: \n  \n Traceback:  \n {tracebackMsg}"))
                gui_queue.put(lambda: self.buildTableLoadingBox.setVisible(False))
        if self.buildTableCancel:
            self.buildTableCancel.set()
        if self.buildTableFuture and not self.buildTableFuture.done():
            self.buildTableFuture.cancel()

        cancelEvent = threading.Event()
        self.buildTableCancel = cancelEvent
        self.stack.setCurrentIndex(1)
        future = executor.submit(buildTable, cancelEvent)
        self.buildTableFuture = future
    def buildCode(self, path):
            code = f"##({"::".join(path)})##"
            return code
    def calculateUpperLevels(self, tableStructure,data):
        
        def buildLevel(levelName,levelIdx, struc,data,path : list):
            levelIdx += 1
            entryTemplate = {"dateTime" : None, "Calculation Type" : "Total " + levelName, "Pool" : None, "Fund" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Investor" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None}
            for header in tranAppHeaderOptions:
                if header != "Ownership":
                    entryTemplate[header] = 0

            #check for filtering. If none, use all options
            options = []
            for entry in data: #all available data
                if entry[levelName] not in options: #
                    options.append(entry[levelName])
            options.sort()
            newTotalEntries = []
            if len(sortHierarchy) > levelIdx: #more hierarchy levels to parse
                highTotals = [] #all total values made on the level
                for option in options:
                    tempPath = path.copy()
                    tempPath.append(option)
                    
                    highEntries = {}
                    name = option if levelName != "assetClass" or option != "Cash" else "Cash "
                    code = self.buildCode(tempPath)
                    struc[name + code] = {} #place table space for that level selection
                    levelData = []
                    for entry in data: #separates out only relevant data
                        if entry[levelName] == option:
                            levelData.append(entry)
                    struc, lowTotals, fullEntries = buildLevel(sortHierarchy[levelIdx],levelIdx,struc,levelData,tempPath)
                    newTotalEntries.extend(fullEntries)
                    for total in lowTotals:
                        if total["dateTime"] not in highEntries.keys():
                            highEntries[total["dateTime"]] = copy.deepcopy(entryTemplate)
                            highEntries[total["dateTime"]]["rowKey"] = name + code
                            for label in tranAppDataOptions:
                                highEntries[total["dateTime"]][label] = total[label]
                            if levelName not in ("Investor","Family Branch"):
                                highEntries[total["dateTime"]][levelName] = total[levelName] if total[levelName] != "Cash" or levelName != "assetClass" else "Cash "
                                if levelName == "subAssetClass":
                                    highEntries[total["dateTime"]]["assetClass"] = total["assetClass"] if total["assetClass"] != "Cash" else "Cash "
                        for header in tranAppHeaderOptions:
                            if header != "Ownership":
                                highEntries[total["dateTime"]][header] += float(total[header])
                            elif levelName in ("Pool", "Investor", "Family Branch") and total.get(header) not in (None,"None","",0) and "Pool" in sortHierarchy[:levelIdx]:
                                if highEntries[total["dateTime"]].get(header) is None:
                                    highEntries[total["dateTime"]][header] = float(total[header]) #initialize
                                else:
                                    highEntries[total["dateTime"]][header] += float(total[header]) #aggregate pool ownerships
                    for month in highEntries.keys():
                        highEntries[month]["Return"] = highEntries[month]["Monthly Gain"] / highEntries[month]["MDdenominator"] * 100 if highEntries[month]["MDdenominator"] != 0 else 0
                        highTotals.append(highEntries[month])
                newTotalEntries.extend(highTotals)       
                #high totals: all totals for the exact level
                #newTotalEntries: all totals for every level being tracked
                return struc, highTotals, newTotalEntries
            else: #occurs at level of fund parent
                newEntriesLow = []
                totalDataLow = []
                for option in options:
                    tempPath = path.copy()
                    tempPath.append(option)
                    totalEntriesLow = {}
                    name = option if levelName != "assetClass" or option != "Cash" else "Cash "
                    code = self.buildCode(tempPath)
                    struc[name + code] =  {}
                    levelData = []
                    for entry in data: #separates out only relevant data
                        if entry[levelName] == option:
                            levelData.append(entry)
                    #gui_queue.put(lambda rows = levelData, name = option: self.openTableWindow(rows,f"data for: {name}"))
                    nameList = []
                    investorsAccessed = {}
                    for entry in levelData:
                        fundName = entry["Fund"] if not self.consolidateFundsBtn.isChecked() or entry["Fund"] not in self.consolidatedFunds or entry["Fund"] in self.filterDict["Fund"].checkedItems() else self.consolidatedFunds.get(entry["Fund"]).get("cFund")
                        nameList.append(fundName + code)
                        temp = entry.copy()
                        temp["rowKey"] = fundName + code
                        totalDataLow.append(temp)
                        if entry["dateTime"] not in totalEntriesLow:
                            totalEntriesLow[entry["dateTime"]] = copy.deepcopy(entryTemplate)
                            totalEntriesLow[entry["dateTime"]]["rowKey"] =name + code
                            for label in tranAppDataOptions:
                                totalEntriesLow[entry["dateTime"]][label] = entry[label]
                            if levelName not in ("Investor","Family Branch"):
                                totalEntriesLow[entry["dateTime"]][levelName] = entry[levelName] if entry[levelName] != "Cash" or levelName != "assetClass" else "Cash "
                                if levelName == "subAssetClass":
                                    totalEntriesLow[entry["dateTime"]]["assetClass"] = entry["assetClass"] if entry["assetClass"] != "Cash" else "Cash "
                        for header in tranAppHeaderOptions:
                            if header != "Ownership":
                                totalEntriesLow[entry["dateTime"]][header] += float(entry[header])
                            elif levelName in ("Investor", "Family Branch") and "Pool" in sortHierarchy and entry.get(header) not in (None,"None","") and float(entry.get(header)) != 0:
                                investor = entry.get("Investor")
                                if totalEntriesLow[entry["dateTime"]].get(header) is None:
                                    totalEntriesLow[entry["dateTime"]][header] = float(entry[header]) #assign investor to ownership based on fund
                                    investorsAccessed[entry["dateTime"]] = [investor,]
                                elif investor not in investorsAccessed.get(entry["dateTime"], []): #accounts for family branch level to add the investor level ownerships
                                    totalEntriesLow[entry["dateTime"]][header] += float(entry[header])
                                    investorsAccessed[entry["dateTime"]].append(investor)
                    for name in sorted(nameList):
                        struc[name] = {}
                    for month in totalEntriesLow.keys():
                        totalEntriesLow[month]["Return"] = totalEntriesLow[month]["Monthly Gain"] / totalEntriesLow[month]["MDdenominator"] * 100 if totalEntriesLow[month]["MDdenominator"] != 0 else 0
                        newEntriesLow.append(totalEntriesLow[month])
                totalDataLow.extend(newEntriesLow)
                return struc, newEntriesLow, totalDataLow

        sortHierarchy = self.sortHierarchy.checkedItems()
        levelIdx = 0
        tableStructure, highestEntries, newEntries = buildLevel(sortHierarchy[levelIdx],levelIdx,tableStructure,data, [])
        trueTotalEntries = {}
        for total in highestEntries:
            if total["dateTime"] not in trueTotalEntries.keys():
                trueTotalEntries[total["dateTime"]] = {"dateTime" : None, "Calculation Type" : "Total", "Pool" : None, "Fund" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Investor" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None}
                trueTotalEntries[total["dateTime"]]["rowKey"] = "Total" + self.buildCode([])
                for header in tranAppHeaderOptions:
                    if header != "Ownership":
                        trueTotalEntries[total["dateTime"]][header] = 0
                for label in tranAppDataOptions:
                    trueTotalEntries[total["dateTime"]][label] = total[label]
            for header in tranAppHeaderOptions:
                if header != "Ownership":
                    trueTotalEntries[total["dateTime"]][header] += float(total[header])
        for month in trueTotalEntries.keys():
            trueTotalEntries[month]["Return"] = trueTotalEntries[month]["Monthly Gain"] / trueTotalEntries[month]["MDdenominator"] * 100 if trueTotalEntries[month]["MDdenominator"] != 0 else 0
            newEntries.append(trueTotalEntries[month])
        #data.extend(newEntries)
        return tableStructure,newEntries
                    
    def filterUpdate(self):
        def resetOptions(key,options):
            currentSelections = self.filterDict[key].checkedItems()
            multiBox = self.filterDict[key]
            multiBox.clearItems()
            multiBox.addItems(sorted(options))
            for currentText in currentSelections:
                if currentText in options:
                    multiBox.setCheckedItem(currentText)
        def exitFunc():
            self.filterCallLock = False
            gui_queue.put(lambda: self.buildReturnTable())
        if not self.filterCallLock:
            def processFilter():
                try:
                    #prevents recursion on calls from comboboxes being updated
                    self.filterCallLock = True
                    currentChoices = {}
                    for key in self.filterDict.keys():
                        if key not in self.highOnlyFilters:
                            currentChoices[key] = self.filterDict[key].checkedItems()
                    if all(choices == [] for _, choices in currentChoices.items()):
                        for key in currentChoices.keys():
                            gui_queue.put(lambda: resetOptions(key,self.fullLevelOptions[key]))
                        exitFunc()
                        return
                    for filterSwitch in self.filterOptions:
                        if filterSwitch["key"] not in self.highOnlyFilters:
                            condStatement = ""
                            first = True
                            parameters = []
                            for filter in self.filterOptions:
                                if filter["key"] != filterSwitch["key"] and filter["key"] not in self.highOnlyFilters:
                                    if self.filterDict[filter["key"]].checkedItems() != []:
                                        paramTemp = self.filterDict[filter["key"]].checkedItems()
                                        placeholders = ','.join('?' for _ in paramTemp)
                                        if first:
                                            condStatement = f"WHERE [{filter["dynNameLow"]}] IN ({placeholders})"
                                            first = False
                                        else:
                                            condStatement += f" AND [{filter["dynNameLow"]}] IN ({placeholders})"
                                        for param in paramTemp:
                                            parameters.append(param)
                            lowTran = load_from_db("transactions_low", condStatement,tuple(parameters), lock=self.lock, db=TRAN_DATABASE_PATH)
                            
                            options = {}
                            for filter in self.filterOptions:
                                options[filter["key"]] = []
                            for account in lowTran:
                                for filter in self.filterOptions:
                                    if filter["key"] not in self.highOnlyFilters:
                                        option = account.get(filter["dynNameLow"])
                                        if option and option not in options[filter["key"]]:
                                            options[filter["key"]].append(option)
                            gui_queue.put(lambda key = filterSwitch["key"], opts = options[filterSwitch["key"]]: resetOptions(key,opts))
                except:
                    gui_queue.put(lambda: QMessageBox.warning(self,"Filter Error", "Error occured updating filters"))
                exitFunc()
            executor.submit(processFilter)
            return
    def updateMonths(self):
        start = self.dataTimeStart
        end = datetime.now()
        index = start
        monthList = []
        while index < end:
            monthList.append(index)
            index += relativedelta(months=1)
        dbDates = []
        for monthDT in monthList:
            month = int(monthDT.month)
            year = int(monthDT.year)
            lastDayCurrent = calendar.monthrange(int(year),month)[1]
            lastDayCurrent   = str(lastDayCurrent).zfill(2)
            if month - 1 > 0:
                prevMonth =  month - 1
                prevMyear = year
            else:
                prevMonth = 12
                prevMyear = str(int(year) - 1)
            lastDayPrev = calendar.monthrange(int(prevMyear),prevMonth)[1]
            lastDayPrev   = str(lastDayPrev).zfill(2)
            prevMonth = str(prevMonth).zfill(2)
            month = str(month).zfill(2)
            
            tranStart = f"{year}-{month}-01T00:00:00.000Z"
            bothEnd = f"{year}-{month}-{lastDayCurrent}T00:00:00.000Z"
            accountStart = f"{prevMyear}-{prevMonth}-{lastDayPrev}T00:00:00.000Z"

            
            dateString = monthDT.strftime("%B %Y")

            monthEntry = {"dateTime" : monthDT, "Month" : dateString, "tranStart" : tranStart.removesuffix(".000Z"), "endDay" : bothEnd.removesuffix(".000Z"), "accountStart" : accountStart.removesuffix(".000Z")}
            dbDates.append(monthEntry)
        save_to_db("Months",dbDates, db=TRAN_DATABASE_PATH)

    def pullLevelNames(self):
        allOptions = {}
        fundPoolLink = {}
        for filter in self.filterOptions:
            if filter["key"] not in self.highOnlyFilters:
                allOptions[filter["key"]] = []
        accountsHigh = load_from_db("transactions_high", db=TRAN_DATABASE_PATH)
        if accountsHigh is not None:
            for account in accountsHigh:
                for filter in self.filterOptions:
                    if (filter["key"] in allOptions and "dynNameHigh" in filter.keys() and
                        account.get(filter["dynNameHigh"]) is not None and
                        account.get(filter["dynNameHigh"]) not in allOptions[filter["key"]]):
                        allOptions[filter["key"]].append(account.get(filter["dynNameHigh"]))
        else:
            print("no investor to pool accounts found")
        accountsLow = load_from_db("transactions_low", db=TRAN_DATABASE_PATH)
        if accountsLow is not None:
            for lowAccount in accountsLow:
                for filter in self.filterOptions:
                    if (filter["key"] in allOptions and "dynNameLow" in filter.keys() and
                        lowAccount.get(filter["dynNameLow"]) is not None and
                        lowAccount.get(filter["dynNameLow"]) not in allOptions[filter["key"]]):
                        allOptions[filter["key"]].append(lowAccount.get(filter["dynNameLow"]))
                fundPoolLink[lowAccount["Target name"]] = lowAccount.get("Source name")
        else:
            print("no pool to fund accounts found")
        self.fullLevelOptions = {}
        for filter in self.filterOptions:
            if filter["key"] in allOptions:
                allOptions[filter["key"]].sort()
                self.filterDict[filter["key"]].addItems(allOptions[filter["key"]])
                self.fullLevelOptions[filter["key"]] = allOptions[filter["key"]]
        self.fundPoolLinks = fundPoolLink
        self.filterUpdate()

    def groupingChange(self, *_):
        groupOpts = self.sortHierarchy.checkedItems()
        self.filterCallLock = True
            
        self.filterCallLock = False
        self.buildReturnTable()
    def check_api_key(self, *_):
        key = self.api_input.text().strip()
        if key:
            headers = {
                "Authorization": f"Bearer {key}",
                "Content-Type":  "application/json"
            }
            payload = {
                "advf": [{ "_name": "Fund" }],
                "mode": "compact",
                "page": {"size": 0}
            }
            resp = requests.get(f"{mainURL}/Entity", headers=headers, json=payload)
            if resp.status_code == 200:
                self.api_label.setText('API key valid. Saving to system...')
                subprocess.run(['setx',dynamoAPIenvName,key], check=True)
                os.environ[dynamoAPIenvName] = key
                self.api_key = key
                self.stack.setCurrentIndex(1)
                self.init_data_processing()
            else:
                self.api_label.setText('Invalid API key')
        else:
            self.api_label.setText('API key cannot be empty')

    def show_results(self,*_):
        self.stack.setCurrentIndex(2)

    def pullData(self):
        def checkNewestData(table, rows):
            def buildKey(record):
                value = record[nameHier["Value"]["dynHigh"] if "position" in table else nameHier["CashFlow"]["dynLow"]]
                value = 0 if value is None or value == "None" else value
                key = (
                        record['Source name'] if record['Source name'] is not None else "None",
                        record['Target name'] if record['Target name'] is not None else "None",
                        round(float(value)) if table != "positions_high" else 0,               # normalize to float
                        record['Date'].replace(' ', 'T')      # normalize format if needed
                    )
                return key
            try:
                diffCount = 0
                differences = []
                newRows = []
                previous = load_from_db(table, db=TRAN_DATABASE_PATH) or []

                # Build a set of tuple‐keys for the old data
                oldRecords = set()
                for rec in previous:
                    oldRecords.add(buildKey(rec))

                newRecords = set()
                earliest = None
                for rec in rows:
                    value = rec[nameHier["Value"]["dynHigh"] if "position" in table else nameHier["CashFlow"]["dynLow"]]
                    value = 0 if value is None or value == "None" else value
                    key = buildKey(rec)
                    newRecords.add(key)
                    if key in oldRecords:
                        continue
                    diffCount += 1
                    newRows.append(rec)
                    differences.append(rec)
                    differences.append({"Source name" : key[0],"Target name" : key[1],nameHier["Value"]["dynLow"] : key[2],"Date" : key[3]})
                    # parse the date for comparison
                    dt = datetime.strptime(rec['Date'], "%Y-%m-%dT%H:%M:%S")
                    if earliest is None or dt < earliest:
                        earliest = dt
                    poolTag = "Target name" if "high" in table else "Source name"
                    if dt < self.poolChangeDates.get(rec.get(poolTag),datetime.now()): 
                        self.poolChangeDates[rec.get(poolTag)] = dt # sets each pool value to earliest and instantiates if not existing
                for oldRec in oldRecords:
                    if oldRec not in newRecords: #find if a new record no longer exists in the old. Means old data is removed and must be redone
                        self.foundRetroChange = True
                        self.poolChangeDates["active"] = False
                        print(f"Retroactive changes found in {table}. Resetting whole table.")
                        break
                
                if earliest and not self.foundRetroChange:
                    if earliest < self.earliestChangeDate:
                        self.earliestChangeDate = earliest
                if self.foundRetroChange: #push full api data and reset calc date to redo all data
                    self.earliestChangeDate =  self.dataTimeStart
                    return rows, False
                print(f"Differences in {table} : {diffCount} of {len(rows)}")
                if diffCount > 0 and not demoMode:
                    def openWindow():
                        window = tableWindow(parentSource=self,all_rows=differences,table=table)
                        self.tableWindows[table] = window
                        window.show()
                    gui_queue.put(lambda: openWindow())
                return newRows, True
            except Exception as e:
                print(f"Error searching old data: {e}")
        
        try:
            self.earliestChangeDate = datetime(datetime.now().year,datetime.now().month + 1,datetime.now().day)
            gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(True))
            gui_queue.put(lambda: self.importButton.setEnabled(False))
            self.updateMonths()
            completeLock = threading.Lock()
            self.apiFutures = set()
            self.complete = float(0)
            totalCalls = float(4)
            apiData = {
                "tranCols": "Investment in, Investing Entity, Transaction Type, Effective date, Asset Class (E), Sub-asset class (E), HF Classification, Remaining commitment change, Transaction timing, Amount in system currency, Cash flow change (USD), Parent investor",
                "tranName": "InvestmentTransaction",
                "tranSort": "Effective date:desc",
                "accountCols": "As of Date, Balance Type, Asset Class, Sub-asset class, Value of Investments, Investing entity, Investment in, HF Classification, Parent investor, Value in system currency",
                "accountName": "InvestmentPosition",
                "accountSort": "As of Date:desc",
                "fundCols" : "Fund Name, Asset class category, Parent fund, Fund Pipeline Status",
                "benchCols" : (f"Index, As of date, MTD %, QTD %, YTD %, ITD cumulative %, ITD TWRR %, "
                               f"{', '.join(f'Last {y} yr %' for y in yearOptions)}"), 
            }
            calculationsTest = load_from_db("calculations", db=TRAN_DATABASE_PATH)
            if calculationsTest != []:
                skipCalculations = True
                self.poolChangeDates["active"] = True
                self.foundRetroChange = False
            else:
                skipCalculations = False
            for i in range(2):
                cols_key = 'accountCols' if i == 1 else 'tranCols'
                name_key = 'accountName' if i == 1 else 'tranName'
                sort_key = 'accountSort' if i == 1 else 'tranSort'
                headers = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": apiData[cols_key],
                    "x-sort": apiData[sort_key]
                }
                for j in range(2): #0: fund level, 1: pool to high investor level
                    investmentLevel = "Investing entity" if j == 0 else "Investment in"
                    if i == 0: #transaction
                        if j == 0: #fund level
                            payload = {
                                        "advf": {
                                            "e": [
                                                {
                                                    "_name": "InvestmentTransaction",
                                                    "e": [
                                                        {
                                                            "_name": "InvestorAccount",
                                                            "_not": True
                                                        },
                                                        {
                                                            "_name": "Fund",
                                                            "rule": [
                                                                {
                                                                    "_op": "is",
                                                                    "_prop": "Fund Pipeline Status",
                                                                    "values": [
                                                                        {
                                                                            "id": "d33af081-c4c8-431b-a98b-de9eaf576324",
                                                                            "es": "L_FundPipelineStatus",
                                                                            "name": "I - Internal"
                                                                        }
                                                                    ]
                                                                }
                                                            ]
                                                        }
                                                    ],
                                                    "rule": [
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Cash flow change (USD)"
                                                        },
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Investing entity"
                                                        }
                                                    ]
                                                },
                                                {
                                                    "_name": "InvestmentTransaction",
                                                    "e": [
                                                        {
                                                            "_name": "InvestorAccount",
                                                            "_not": True
                                                        },
                                                        {
                                                            "_name": "Fund",
                                                            "rule": [
                                                                {
                                                                    "_op": "is",
                                                                    "_prop": "Fund Pipeline Status",
                                                                    "values": [
                                                                        {
                                                                            "id": "d33af081-c4c8-431b-a98b-de9eaf576324",
                                                                            "es": "L_FundPipelineStatus",
                                                                            "name": "I - Internal"
                                                                        }
                                                                    ]
                                                                }
                                                            ]
                                                        }
                                                    ],
                                                    "rule": [
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Investing entity"
                                                        },
                                                        {
                                                            "_op": "any_item",
                                                            "_prop": "Transaction type",
                                                            "values": [
                                                                [
                                                                    {
                                                                        "id": "5327639c-8160-4d85-9b23-8c6bf60c5406",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Commitment"
                                                                    },
                                                                    {
                                                                        "id": "37339e7c-1c24-4d13-9d17-86d0efe079b3",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Transfer of commitment"
                                                                    },
                                                                    {
                                                                        "id": "0f8f8671-8579-49d7-b604-05300b6a3990",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Transfer of commitment (out)"
                                                                    },
                                                                    {
                                                                        "id": "5e098d83-70b0-4135-a629-aff19048fb1c",
                                                                        "es": "L_TransactionType",
                                                                        "name": "Secondary - Original commitment (by secondary seller)"
                                                                    }
                                                                ]
                                                            ]
                                                        }
                                                    ]
                                                }
                                            ]
                                        },
                                        "mode": "compact"
                                    }
                        else: #investor level
                            payload = {
                                        "advf": {
                                            "e": [
                                                {
                                                    "_name": "InvestmentTransaction",
                                                    "e": [
                                                        {
                                                            "_name": "InvestorAccount"
                                                        }
                                                    ],
                                                    "rule": [
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Cash flow change (USD)"
                                                        },
                                                        {
                                                            "_op": "not_null",
                                                            "_prop": "Investing entity"
                                                        }
                                                    ]
                                                }
                                            ]
                                        },
                                        "mode": "compact"
                                    }
                    else:
                        continue #removed account positions
                    def bgPullData(payload=payload, headers=headers, i=i, j=j):
                        rows = []
                        idx = 0
                        while rows in ([],None) and idx < 3: #if call fails, tries again
                            idx += 1
                            response = requests.post(f"{mainURL}/Search", headers=headers, data=json.dumps(payload))
                            if response.status_code == 200:
                                try:
                                    data = response.json()
                                except ValueError:
                                    continue
                                if isinstance(data, dict):
                                    rows = data.get('data', data.get('rows', []))
                                elif isinstance(data, list):
                                    rows = data
                                else:
                                    rows = []

                                keys_to_remove = {'_id', '_es'}
                                rows = [
                                    {k: v for k, v in row.items() if k not in keys_to_remove}
                                    for row in rows
                                ]
                            else:
                                print(f"Error in API call. Code: {response.status_code}. {response}")
                                try:
                                    print(f"Error: {response.json()}")
                                    print(f"Headers used:  \n {headers}, \n payload used: \n {payload}")
                                except:
                                    pass
                        if i == 1:
                            if j == 0:
                                pass
                            else:
                                pass
                        else:
                            if j == 0:
                                if skipCalculations: #separate out only new rows to alter db
                                    rows, good = checkNewestData('transactions_low',rows)
                                    if good:
                                        save_to_db('transactions_low', rows, action="add", db=TRAN_DATABASE_PATH)
                                    else:
                                        save_to_db('transactions_low', rows, db=TRAN_DATABASE_PATH)
                                else:
                                    save_to_db('transactions_low', rows, db=TRAN_DATABASE_PATH)
                            else:
                                if skipCalculations: #separate out only new rows to alter db
                                    rows, good = checkNewestData('transactions_high',rows)
                                    if good:
                                        save_to_db('transactions_high', rows, action="add", db=TRAN_DATABASE_PATH)
                                    else:
                                        save_to_db('transactions_high', rows, db=TRAN_DATABASE_PATH)
                                else:
                                    save_to_db('transactions_high', rows, db=TRAN_DATABASE_PATH)
                        with completeLock:
                            self.complete += 1
                        frac = self.complete/totalCalls
                        gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
                    try:
                        submitAPIcall(self,bgPullData)
                    except Exception as e:
                        print(f"Failure to run background thread API call: {e} \n {e.args}")
            fundPayload = {
                            "advf": {
                                "e": [
                                    {
                                        "_name": "Fund"
                                    }
                                ]
                            },
                            "mode": "compact"
                        }
            fundHeaders = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": apiData["fundCols"],
                }
            def bgFundPull():
                response = requests.post(f"{mainURL}/Search", headers=fundHeaders, data=json.dumps(fundPayload))
                if response.status_code == 200:
                    try:
                        data = response.json()
                        if isinstance(data, dict):
                            rows = data.get('data', data.get('rows', []))
                        elif isinstance(data, list):
                            rows = data
                        else:
                            rows = []
                        keys_to_remove = {'_id', '_es'}
                        rows = [{k: v for k, v in row.items() if k not in keys_to_remove} for row in rows]
                        consolidatorFunds = {}
                        for idx, row in enumerate(rows): #find sleeve values and consolidated funds
                            assetCat = row["ExposureAssetClassCategory"]
                            if assetCat is not None and assetCat.count(" > ") == 3:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = assetCat.split(" > ")[2]
                                sleeve = assetCat.split(" > ")[3]
                            elif assetCat is not None and assetCat.count(" > ") == 2:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = assetCat.split(" > ")[2]
                                sleeve = None
                            elif assetCat is not None and assetCat.count(" > ") == 1:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = None
                                sleeve = None
                            else:
                                assetClass = None
                                subAssetClass = None
                                sleeve = None
                            if row.get("Fundpipelinestatus") is not None and "Z - Placeholder" in row.get("Fundpipelinestatus"):
                                consolidatorFunds[row["Name"]] = {"cFund" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass, "sleeve" : sleeve}
                            rows[idx][nameHier["sleeve"]["sleeve"]] =  sleeve
                            rows[idx]["assetClass"] = assetClass
                            rows[idx]["subAssetClass"] = subAssetClass
                        self.consolidatedFunds = {}
                        for row in rows: #assign funds to their consolidators
                            if row.get("Parentfund") in consolidatorFunds:
                                self.consolidatedFunds[row["Name"]] = consolidatorFunds.get(row.get("Parentfund"))
                        if rows != []:
                            save_to_db("funds",rows, db=TRAN_DATABASE_PATH)
                    except Exception as e:
                        print(f"Error proccessing fund API data : {e} {e.args}.  {traceback.format_exc()}")
                    
                else:
                    print(f"Error in API call for fund. Code: {response.status_code}. {response}. {traceback.format_exc()}")
                with completeLock:
                    self.complete += 1
                frac = self.complete/totalCalls
                gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
            submitAPIcall(self,bgFundPull)
            benchmarkPayload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "IndexPerformance"
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
            benchmarkHeaders = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": apiData["benchCols"],
                }
            def bgBenchPull():
                response = requests.post(f"{mainURL}/Search", headers=benchmarkHeaders, data=json.dumps(benchmarkPayload))
                if response.status_code == 200:
                    try:
                        data = response.json()
                        if isinstance(data, dict):
                            rows = data.get('data', data.get('rows', []))
                        elif isinstance(data, list):
                            rows = data
                        else:
                            rows = []
                        keys_to_remove = {'_id', '_es'}
                        rows = [{k: v for k, v in row.items() if k not in keys_to_remove} for row in rows]
                        save_to_db("benchmarks",rows, db=TRAN_DATABASE_PATH)
                    except Exception as e:
                        print(f"Error proccessing benchmark API data : {e} {e.args}.  {traceback.format_exc()}")
                    
                else:
                    print(f"Error in API call for benchmarks. Code: {response.status_code}. {response}. {traceback.format_exc()}")
                with completeLock:
                    self.complete += 1
                frac = self.complete/totalCalls
                gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
            submitAPIcall(self,bgBenchPull)

            wait(self.apiFutures) #wait for all api pulls to complete
            if skipCalculations:
                print("Earliest change: ", self.earliestChangeDate)
                if not self.foundRetroChange:
                    print(f"Changes dates by pools:")
                    for pool in self.poolChangeDates:
                        print(f"        {pool} : {self.poolChangeDates.get(pool)}")
            gui_queue.put(lambda: self.apiLoadingBar.setValue(100))
            
            while not gui_queue.empty(): #wait to assure database has been updated in main thread before continuing
                time.sleep(0.2)
            


            currentTime = datetime.now().strftime("%B %d, %Y @ %I:%M %p")
            changeData = datetime.strftime(self.earliestChangeDate, "%B %d, %Y @ %I:%M %p")
            save_to_db(None,None,query="UPDATE history SET [lastImport] = ?, [changeDate] = ?", inputs=(currentTime,changeData), action="replace", db=TRAN_DATABASE_PATH)
            self.lastImportLabel.setText(f"Last Data Import: {currentTime}")
            gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(False))
            gui_queue.put(lambda: self.calculateReturn())
        except Exception as e:
            QMessageBox.warning(self,"Error Importing Data", f"Error pulling data from dynamo: {e} , {e.args}")
        gui_queue.put(lambda: self.importButton.setEnabled(True))
        gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(False))
    def openTableWindow(self, rows, name = "Table", headers = None):
        window = tableWindow(parentSource=self,all_rows=rows,table=name, headers=headers)
        self.tableWindows[name] = window
        window.show()
    def calculateReturn(self):
        def initalizeCalc():
            try:
                calculationStart = datetime.now()
                gui_queue.put(lambda: self.importButton.setEnabled(False))
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(True))
                self.updateMonths()
                gui_queue.put(lambda: self.pullLevelNames())
                print("Calculating return....")
                fundListDB = load_from_db("funds", db=TRAN_DATABASE_PATH)
                fundList = {}
                for fund in fundListDB:
                    fundList[fund["Name"]] = fund[nameHier["sleeve"]["sleeve"]]
                months = load_from_db("Months", f"ORDER BY [dateTime] ASC", db=TRAN_DATABASE_PATH)
                calculations = []
                monthIdx = 0
                if load_from_db("calculations", db=TRAN_DATABASE_PATH) == []:
                    noCalculations = True
                else:
                    noCalculations = False

                if self.earliestChangeDate > datetime.now() and not noCalculations:
                    #if no new data exists, use old calculations
                    calculations = load_from_db("calculations", db=TRAN_DATABASE_PATH)
                    keys = []
                    for row in calculations:
                        for key in row.keys():
                            if key not in keys:
                                keys.append(key)
                    gui_queue.put( lambda: self.populate(self.calculationTable,calculations,keys = keys))
                    gui_queue.put( lambda: self.buildReturnTable())
                    gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                    gui_queue.put(lambda: self.importButton.setEnabled(True))
                    print("Calculations skipped.")
                    return
                
                # proces pool section----------------------------------------------------------------
                save_to_db("progress",None,action="reset", db=TRAN_DATABASE_PATH)
                self.initializeProgressDB()

                # ------------------- build data cache ----------------------
                tables = [ "transactions_low", "transactions_high", "calculations"]
                table_rows = {t: load_from_db(t, db=TRAN_DATABASE_PATH) for t in tables}
                cache = {}
                for table, rows in table_rows.items():
                    for row in rows:
                        if table in ("transactions_low"):
                            poolKey = row.get("Source name")
                        elif table in ("transactions_high"):
                            poolKey = row.get("Target name")
                        else:
                            poolKey = row.get("Pool")
                        if poolKey is None:
                            continue
                        for m in months:
                            if table == "calculations":
                                if row.get("dateTime") != m["dateTime"]:
                                    continue
                            else:
                                start =  m["tranStart"]
                                date = row.get("Date")
                                if not (start <= date <= m["endDay"]):
                                    continue
                            cache.setdefault(poolKey, {}).setdefault(table, {}).setdefault(m["dateTime"], []).append(row)
                for idx, pool in enumerate(self.pools):
                    self.pools[idx]["cache"] = cache.get(pool.get("poolName"))
                    if self.poolChangeDates.get("active",False): #if the pool changes have been calculated, use it or set to current date if no changes occured
                        self.pools[idx]["earliestChangeDate"] = self.poolChangeDates.get(pool.get("poolName"),datetime.now())
                    else: #if pool changes have not been calculated but calculation requirements were imported, set to earliest global date
                        self.pools[idx]["earliestChangeDate"] =  self.earliestChangeDate 
                    newMonths = []
                    if not noCalculations: #if there are calculations, find all months before the data pull, and then pull those calculations
                        for month in months:
                            #if the calculations for the month have already been complete, pull the old data
                            if self.pools[idx]["earliestChangeDate"] > datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S"):
                                pass
                            else:
                                newMonths.append(month)
                    else:
                        newMonths = months
                    _ = updateStatus(pool.get("poolName"),len(newMonths),threading.Lock(),status="Initialization", db=TRAN_DATABASE_PATH)
                def initializeWorkerPool():
                    self.manager = Manager()
                    self.lock = self.manager.Lock()
                    self.workerStatusQueue = self.manager.Queue()
                    self.workerDBqueue = self.manager.Queue()
                    self.calcFailedFlag = self.manager.Value('b', False)
                    self.cancelCalcBtn.setEnabled(True) #only allows cancelling once the lock for the db exists

                    self.pool = Pool()
                    self.futures = []
                    executor.submit(self.watch_db)

                    commonData = {"noCalculations" : noCalculations,
                                    "months" : months, "fundList" : fundList
                                    }
                    
                    self.calcStartTime = datetime.now()
                    for pool in self.pools:
                        res = self.pool.apply_async(processPoolTransactions, args=(pool, commonData,self.workerStatusQueue, self.workerDBqueue, self.calcFailedFlag))
                        self.futures.append(res)
                    self.pool.close()

                    self.timer.start(int(calculationPingTime * 0.25) * 1000) #check at 0.75 the ping time to prevent queue buildup
                gui_queue.put(lambda: initializeWorkerPool()) #puts on main thread
            except Exception as e:
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                gui_queue.put(lambda: self.importButton.setEnabled(True))
                print(f"Error occured running calculations: {e}")
                print("e.args:", e.args)
                # maybe also:
                print(traceback.format_exc())
        executor.submit(initalizeCalc)
    def initializeProgressDB(self):
        conn = sqlite3.connect(TRAN_DATABASE_PATH)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS progress (
                pool STRING PRIMARY KEY,
                completed INTEGER NOT NULL,
                total INTEGER NOT NULL,
                status STRING NOT NULL
            )
        """)
        conn.commit()
        conn.close()
    def watch_db(self):
        conn = sqlite3.connect(TRAN_DATABASE_PATH)
        c = conn.cursor()
        while True:
            count = 0
            while not self.workerStatusQueue.empty() and count < 300:
                count += 1 #count to allow the loading bar to take the lock and update
                vars = self.workerStatusQueue.get()
                try:
                    failed = updateStatus(vars[0],vars[1],self.lock,status=vars[2],connection=conn, db=TRAN_DATABASE_PATH)
                    if failed:
                        self.calcFailedFlag = failed
                except Exception as e:
                    trace = traceback.format_exc()
                    print(f"Error occured while attempting to run background worker status update: {e}. \n traceback: \n {trace}")
            try:
                with self.lock:
                    c.execute("SELECT * FROM progress")
                    cols = [d[0] for d in c.description]
                    statusLines = [dict(zip(cols, row)) for row in c.fetchall()]
                failed = []
                completed = []
                complete = 0
                total = 0
                for line in statusLines:
                    complete += line.get("completed",0)
                    total += line.get("total",0)
                    if line["status"] == "Failed":
                        failed.append(line)
                    elif line["status"] == "Completed":
                        completed.append(line)
                if len(failed) > 0:
                    print(f"Halting progress watch due to worker '{failed[0].get("pool","Bad Pull")}' failure.")
                    self.queue.append(-86) #will halt the queue
                    break
                elif len(completed) == len(self.pools):
                    print("All workers have declared complete.")
                    self.queue.append(100) #backup in case the numbers below fail
                    break
                if total != 0:
                    percent = int((complete / total) * 100)
                    self.queue.append(percent)
                    if complete >= total:
                        break
            except Exception as e:
                print(f"Error watching database: {e}")
                print(traceback.format_exc())
                pass
            time.sleep(calculationPingTime * 0.01)
        conn.close()
    def updateWorkerDB(self):
        try:
            conn = sqlite3.connect(TRAN_DATABASE_PATH)
            cursor = conn.cursor()
        except:
            print("connection failed")
        dbFailure = False
        maxFails = 4
        print("Initiating background database updates...")
        while True:
            try:
                results = self.workerDBqueue.get_nowait()  # non-blocking, safe for fixed queues
                data = results.get("data")
                failCount = 0
                while True:
                    try:
                        if results.get("type") == "insert":
                            save_to_db(data[0], data[1], action=data[2], connection=conn, lock=self.lock, db=TRAN_DATABASE_PATH)
                            break
                        elif results.get("type") == "update":
                            with self.lock:
                                cursor.executemany(data[0], data[1])
                                conn.commit()
                                break
                        else:
                            print(f"\n\n Database data was not handled correctly: {results} \n\n")
                            break
                    except:
                        failCount += 1
                        print(f"Error updating database. Attempt {failCount} of {maxFails}")
                        if failCount > maxFails:
                            print("Error occured in delayed database updates. Calculation date will be reset")
                            dbFailure = True
                            break
            except queue.Empty:
                break  # all done; queue drained
            except Exception as e:
                print(f"Error occurred updating database from worker threads: {e}, {e.args}")
        print("Background database updates complete")
        if dbFailure: #will force a recalculation on the next opening since the database won't be accurate
            save_to_db(None,None,query="UPDATE history SET [lastCalculation] = ?", inputs=("Database Failure",), action="replace", lock=self.lock, db=TRAN_DATABASE_PATH)
        conn.close()
    def update_from_queue(self):
        if self.queue:
            while self.queue: #cycle through the queue options to get most up to date value. Breaks out if complete or halted
                val = self.queue.pop(0)
                if val in (-86,100):
                    break
            self.calculationLoadingBar.setValue(val)
            timeElapsed = datetime.now() - self.calcStartTime
            secsElapsed = timeElapsed.total_seconds()
            loadingFraction = float(val) / 100 #decimal format percentage
            if loadingFraction > 0:
                est_total_secs = secsElapsed / loadingFraction
                secs_remaining = est_total_secs - secsElapsed
            else:
                secs_remaining = 0
            mins, secs = divmod(int(secs_remaining), 60)
            time_str = f"{mins}m {secs}s" # format as “Xm Ys” or “MM:SS”
            self.calculationLabel.setText(f"Estimated time remaining: {time_str}")
            if val >= 100:
                self.timer.stop()
                executor.submit(self.calcCompletion)
            elif val == -86:
                self.timer.stop()
                if self.cancel:
                    QMessageBox.warning(self,"Calculation Halted", "Calculations are being halted.")
                    self.cancel = False
                else:
                    QMessageBox.warning(self,"Calculation Failure", "A worker thread has failed. Calculations will not be properly completed.")
                self.pool.terminate()
                self.pool.join()
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                gui_queue.put(lambda: self.importButton.setEnabled(True))
                
    def calcCompletion(self):
        try:
            print("Checking worker completion...")
            executor.submit(self.updateWorkerDB)
            self.pool.join()
            print("All workers finished")
            
            calculations = []
            for fut in self.futures:
                try:
                    calculations.extend(fut.get())
                except Exception as e:
                    print(f"Error appending calculations: {e}")
            keys = []
            for row in calculations:
                for key in row.keys():
                    if key not in keys:
                        keys.append(key)
            save_to_db("calculations",calculations, keys=keys, lock=self.lock, db=TRAN_DATABASE_PATH)
            try:
                apiPullTime = load_from_db("history", db=TRAN_DATABASE_PATH)[0]["lastImport"]
                save_to_db(None,None,query="UPDATE history SET [lastCalculation] = ?", inputs=(apiPullTime,), action="replace", lock=self.lock, db=TRAN_DATABASE_PATH)
            except:
                print("failed to update last calculation time")
            gui_queue.put( lambda: self.populate(self.calculationTable,calculations,keys = keys))
            gui_queue.put( lambda: self.buildReturnTable())
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
            gui_queue.put(lambda: self.importButton.setEnabled(True))
            print("Calculations complete.")
            save_to_db("progress",None,action="reset", lock=self.lock, db=TRAN_DATABASE_PATH)
        except:
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
            gui_queue.put(lambda: self.importButton.setEnabled(True))
            print(f"Error occured processing calculation results. Resetting... ")
            print(traceback.format_exc())
    def separateRowCode(self, label):
        header = re.sub(r'##\(.*\)##', '', label, flags=re.DOTALL)
        code = re.findall(r'##\(.*\)##', label, flags=re.DOTALL)[0]
        return header, code
    def headerSortClosed(self):
        self.populateReturnsTable(self.currentTableData)
    def orderColumns(self,keys, exceptions = []):
        mode = self.tableBtnGroup.checkedButton().text()
        if mode == "Monthly Table":
            dates = [datetime.strptime(k, "%B %Y") for k in keys]
            dates = sorted(dates, reverse=True)
            keys = [d.strftime("%B %Y") for d in dates]
        elif mode == "Complex Table":
            newOrder = ["NAV", "Commitment", "Unfunded","MTD","QTD","YTD"] + [f"{y}YR" for y in yearOptions] + ["ITD"]
            ordered = [h for h in newOrder if h in keys]
            ordered += [h for h in keys if h not in newOrder and h not in exceptions]
            keys = ordered
        return keys
    def populateReturnsTable(self, origRows: dict):
        self.buildTableLoadingBar.setValue(7)
        mode = self.tableBtnGroup.checkedButton().text()
        if not origRows:
            # nothing to show
            self.returnsTable.clear()
            self.returnsTable.setRowCount(0)
            self.returnsTable.setColumnCount(0)
            self.buildTableLoadingBox.setVisible(False)
            return

        rows = copy.deepcopy(origRows) #prevents alteration of self.returnsTableData
        for f in self.filterOptions:
            if f["key"] not in self.filterBtnExclusions and not self.filterRadioBtnDict[f["key"]].isChecked():
                to_delete = [k for k,v in rows.items() if v["dataType"] == "Total " + f["key"]]
                for k in to_delete:
                    rows.pop(k)
        
        self.filteredReturnsTableData = copy.deepcopy(rows) #prevents removal of dataType key for data lookup

        # 1) Build a flat list of row-entries:
        #    each entry = (fund_label, unique_code, row_dict)
        row_entries = []
        for fund_label, row_dict in rows.items():
            row_label, code = self.separateRowCode(fund_label)
            row_entries.append((row_label, code, row_dict))

        # 2) Determine columns exactly as before, using cleanedRows for header order
        cleaned = {fund: d.copy() for fund, _, d in row_entries}
        for d in cleaned.values():
            d.pop("dataType", None)

        if not self.headerSort.active or mode == "Monthly Table":
            col_keys = set()
            for d in cleaned.values():
                col_keys |= set(d.keys())
            col_keys = list(col_keys)

            exceptions = ["Return", "Ownership", "MDdenominator", "Monthly Gain"]
            col_keys = self.orderColumns(col_keys, exceptions=exceptions)
            if mode == "Complex Table":
                allKeys = col_keys.copy()
                allKeys.extend(exceptions) #all key options for the header selections
                self.headerSort.set_items(allKeys,[item for item in allKeys if item not in exceptions])
                self.headerSort.setEnabled(True)
            else:
                self.headerSort.setEnabled(False)
        else:
            col_keys = self.headerSort.popup.get_checked_sorted_items()
            self.headerSort.setEnabled(True)

        # 3) Resize & set horizontal headers (we no longer call setVerticalHeaderLabels)
        self.returnsTable.setRowCount(len(row_entries))
        self.returnsTable.setColumnCount(len(col_keys))
        self.returnsTable.setHorizontalHeaderLabels(col_keys)


        # 4) Populate each row
        for r, (fund_label, code, row_dict) in enumerate(row_entries):
            # pull & remove dataType for coloring
            dataType = row_dict.pop("dataType", "")

            startColor = (160, 160, 160)
            if dataType != "Total":
                depth      = code.count("::") if dataType != "Total Pool" else code.count("::") + 1
                # if len(re.findall(r'##\((.*?)\)##', code, flags=re.DOTALL)[0]) > 0:
                #     depth -= 1
                maxDepth   = max(len(self.sortHierarchy.checkedItems()),1)
                cRange     = 255 - startColor[0]
                ratio      = (depth / maxDepth) if maxDepth != 0 else 1
                color = tuple(
                    int(startColor[i] + cRange * ratio)
                    for i in range(3)
                )
                bg = QColor(*color)
            else:
                bg =  QColor(*tuple(
                    int(startColor[i] * 0.8)
                    for i in range(3)
                ))

            # — vertical header: only show the fund, stash the code —
            hdr = QTableWidgetItem(fund_label)
            hdr.setData(Qt.UserRole, code)
            if bg:
                hdr.setBackground(QBrush(bg))
            self.returnsTable.setVerticalHeaderItem(r, hdr)

            # — fill cells —
            for c, col in enumerate(col_keys):
                raw = row_dict.get(col, "")
                if raw not in (None, "", "None"):
                    try:
                        v = round(float(raw), 2)
                        if c in percent_headers or (mode == "Monthly Table" and self.returnOutputType.currentText() in percent_headers):
                            text = f"{v:.2f}%"
                        else:
                            text = f"{v:,.2f}"
                    except:
                        text = str(raw)
                else:
                    text = ""

                item = QTableWidgetItem(text)
                if text:
                    # store raw number for sorting or later retrieval
                    item.setData(Qt.UserRole, v)
                if bg:
                    item.setBackground(QBrush(bg))
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.returnsTable.setItem(r, c, item)
        self.buildTableLoadingBox.setVisible(False)
    def populate(self, table, rows, keys = None):
        if not rows:
            return
        if keys is None:
            headers = list(rows[0].keys())
        else:
            headers = list(keys)

        calcTableModel = DictListModel(rows,headers, self)
        table.setModel(calcTableModel)

class linkBenchmarksWindow(QWidget):
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource=None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle("Link Benchmarks")
        self.resize(800, 500)
        self.setStyleSheet(self.parent.appStyle)
        self.setObjectName("mainPage")
        self._benchmarks = []
        self._links = []
        self.asset_levels = [("Level 1", 1), ("Level 2", 2), ("Level 3", 3)]
        self.selected_asset_level = None
        self.selected_asset = None
        self.selected_benchmark = None

        self.init_ui()

    def init_ui(self):
        mainLayout = QVBoxLayout(self)
        splitter = QSplitter(Qt.Horizontal)

        # --- Left: Benchmark Links Table ---
        leftWidget = QWidget()
        leftLayout = QVBoxLayout(leftWidget)
        self.linksTable = QTableWidget()
        self.linksTable.setColumnCount(4)
        self.linksTable.setHorizontalHeaderLabels(['Benchmark', 'Asset', 'Level', 'Delete'])
        self.linksTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        leftLayout.addWidget(QLabel("Current Benchmark Links:"))
        leftLayout.addWidget(self.linksTable)
        splitter.addWidget(leftWidget)

        # --- Right: Add Benchmark Link ---
        rightWidget = QWidget()
        rightLayout = QVBoxLayout(rightWidget)
        rightLayout.addWidget(QLabel("Add Benchmark Link:"))
        
        # Asset Class Level ComboBox
        self.assetLevelCombo = QComboBox()
        self.assetLevelCombo.addItem("")  # blank/default
        for label, num in self.asset_levels:
            self.assetLevelCombo.addItem(label, num)
        self.assetLevelCombo.currentIndexChanged.connect(self.updateAssetCombo)
        rightLayout.addWidget(QLabel("Asset Class Level:"))
        rightLayout.addWidget(self.assetLevelCombo)

        # Asset ComboBox - will be populated depending on above
        self.assetCombo = QComboBox()
        self.assetCombo.addItem("")  # blank
        rightLayout.addWidget(QLabel("Asset:"))
        rightLayout.addWidget(self.assetCombo)

        # Benchmark ComboBox - will be populated
        self.benchmarkCombo = QComboBox()
        self.benchmarkCombo.addItem("")  # blank
        rightLayout.addWidget(QLabel("Benchmark:"))
        rightLayout.addWidget(self.benchmarkCombo)

        # Confirm Button
        self.confirmBtn = QPushButton("Confirm Link")
        self.confirmBtn.clicked.connect(self.addBenchmarkLink)
        rightLayout.addWidget(self.confirmBtn)
        rightLayout.addStretch()
        splitter.addWidget(rightWidget)

        mainLayout.addWidget(splitter)
        self.setLayout(mainLayout)

        self.refreshBenchmarks()
        self.refreshLinks()

    def refreshLinks(self):
        links = self.parent.db.fetchBenchmarkLinks()
        self._links = links
        self.linksTable.setRowCount(len(links))
        for row, link in enumerate(links):
            self.linksTable.setItem(row, 0, QTableWidgetItem(str(link.get("benchmark", ""))))
            self.linksTable.setItem(row, 1, QTableWidgetItem(str(link.get("asset", ""))))
            self.linksTable.setItem(row, 2, QTableWidgetItem(str(link.get("assetLevel", ""))))
            # Delete button
            btn = QPushButton("Delete")
            btn.clicked.connect(lambda _, r=row: self.deleteLink(r))
            self.linksTable.setCellWidget(row, 3, btn)

    def refreshBenchmarks(self):
        # try to fetch all benchmarks for use in the Benchmark combobox
        try:
            if hasattr(self.parent.db, "fetchBenchmarks"):
                benchmarks = self.parent.db.fetchBenchmarks()
            else:
                benchmarks = []  # fallback - maybe empty or stub
        except Exception as e:
            print(f"Failed to fetch benchmarks: {e}")
            benchmarks = []
        self._benchmarks = benchmarks
        self.updateBenchmarkCombo()

    def updateBenchmarkCombo(self):
        self.benchmarkCombo.clear()
        self.benchmarkCombo.addItem("")
        for b in self._benchmarks:
            # Assume b is a string or has a 'benchmark' key
            if isinstance(b, str):
                self.benchmarkCombo.addItem(b)
            elif isinstance(b, dict) and "benchmark" in b:
                self.benchmarkCombo.addItem(b["benchmark"])
            else:
                self.benchmarkCombo.addItem(str(b))

    def updateAssetCombo(self):
        self.assetCombo.clear()
        self.assetCombo.addItem("")
        asset_level_num = self.assetLevelCombo.currentData()
        asset_key = None
        if asset_level_num == 1:
            asset_key = "assetClass"
        elif asset_level_num == 2:
            asset_key = "subAssetClass"
        elif asset_level_num == 3:
            asset_key = "subAssetSleeve"
        all_opts = getattr(self.parent, "fullLevelOptions", {})
        options = all_opts.get(asset_key, [])
        # Remove duplicates and blank/None values
        assets = sorted({opt for opt in options if opt not in (None,"", "None")})
        self.assetCombo.addItems(assets)

    def deleteLink(self, row):
        if row < 0 or row >= len(self._links):
            return
        link = self._links[row]
        try:
            with self.parent.lock:
                cursor = self.parent.db._conn.cursor()
                cursor.execute(
                    "DELETE FROM benchmarkLinks WHERE benchmark = ? AND asset = ? AND assetLevel = ?",
                    (link["benchmark"], link["asset"], link["assetLevel"])
                )
                self.parent.db._conn.commit()
            self.parent.db.fetchBenchmarkLinks(update=True)
            QMessageBox.information(self, "Success", f"Deleted link: {link['benchmark']} to {link['asset']} at level {link['assetLevel']}.")
        except Exception as e:
            QMessageBox.warning(self, "Delete Error", f"Error deleting link: {e}")
        self.refreshLinks()

    def addBenchmarkLink(self):
        # Get selections
        level_idx = self.assetLevelCombo.currentIndex()
        asset = self.assetCombo.currentText().strip()
        benchmark = self.benchmarkCombo.currentText().strip()
        if level_idx <= 0 or not asset or not benchmark:
            QMessageBox.warning(self, "Incomplete", "Please select asset class level, asset, and benchmark.")
            return
        asset_level = self.assetLevelCombo.currentData()
        try:
            with self.parent.lock:
                cursor = self.parent.db._conn.cursor()
                cursor.execute(
                    "INSERT OR REPLACE INTO benchmarkLinks (benchmark, asset, assetLevel) VALUES (?, ?, ?)",
                    (benchmark, asset, asset_level)
                )
                self.parent.db._conn.commit()
            QMessageBox.information(self, "Success", f"Linked {benchmark} to {asset} at level {asset_level}.")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save link: {e}")
        self.parent.db.fetchBenchmarkLinks(update=True)
        self.refreshLinks()
        self.parent.buildReturnTable()

class exportWindow(QWidget):
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource=None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle("Export Data")
        self.resize(600, 400)
        self.setStyleSheet(self.parent.appStyle)
        self.setObjectName("mainPage")

        # Layouts
        main_layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.filter_boxes = {}
        self.filter_labels = {}

        # --- Filter ComboBoxes for each filterOption ---
        self.filterOptions = getattr(self.parent, "filterOptions", [
            {"key": "Investor", "name": "Investor"},
            {"key": "Pool", "name": "Pool"},
            {"key": "Fund", "name": "Fund"},
            {"key": "assetClass", "name": "Asset Class"},
            {"key": "subAssetClass", "name": "Sub Asset Class"},
            {"key": "Classification", "name": "Classification"},
        ])
        # Use parent's lock if available
        self.lock = getattr(self.parent, "lock", None)

        # Query unique options for each filter from the calculations table
        with self.parent.lock:
            dbPath = getattr(self.parent, "dbPath", DATABASE_PATH)
            conn = sqlite3.connect(dbPath)
            cur = conn.cursor()
            for f in self.filterOptions:
                key = f["key"]
                name = f["name"]
                combo = QComboBox()
                combo.addItem("")  # blank for optional
                try:
                    cur.execute(f"SELECT DISTINCT [{key}] FROM calculations")
                    options = [row[0] for row in cur.fetchall() if row[0] not in (None, "", "None")]
                    options = sorted(set(options))
                    combo.addItems(options)
                except Exception as e:
                    print(f"Error loading filter options for {key}: {e}")
                self.filter_boxes[key] = combo
                self.filter_labels[key] = name
                form_layout.addRow(name + ":", combo)
            conn.close()

        # --- Date selectors ---
        date_layout = QHBoxLayout()
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setSpecialValueText("")  # blank for optional
        self.start_date_edit.setDateRange(QDate(1990, 1, 1), QDate(2100, 12, 31))
        self.start_date_edit.setDate(QDate(2000, 1, 1))
        self.start_date_edit.setMinimumWidth(120)
        self.start_date_edit.setDate(QDate())  # blank

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setSpecialValueText("")  # blank for optional
        self.end_date_edit.setDateRange(QDate(1900, 1, 1), QDate(2100, 12, 31))
        self.end_date_edit.setDate(QDate())  # blank
        self.end_date_edit.setMinimumWidth(120)

        date_layout.addWidget(QLabel("Start Date:"))
        date_layout.addWidget(self.start_date_edit)
        date_layout.addSpacing(20)
        date_layout.addWidget(QLabel("End Date:"))
        date_layout.addWidget(self.end_date_edit)
        form_layout.addRow(date_layout)

        # --- Confirm Button ---
        self.confirm_btn = QPushButton("Export to Excel")
        self.confirm_btn.clicked.connect(self.export_to_excel)
        main_layout.addLayout(form_layout)
        main_layout.addWidget(self.confirm_btn, alignment=Qt.AlignRight)

    def export_to_excel(self):
        # Build WHERE clause from filters
        filters = []
        values = []
        for key, combo in self.filter_boxes.items():
            val = combo.currentText()
            if val:
                filters.append(f"[{key}] = ?")
                values.append(val)
        # Date filters
        start_date = self.start_date_edit.date()
        end_date = self.end_date_edit.date()
        if start_date.isValid():
            filters.append("[dateTime] >= ?")
            # Convert QDate to Python datetime, then format as string
            values.append(start_date.toPyDate().strftime("%Y-%m-%d %H:%M:%S"))
        if end_date.isValid():
            filters.append("[dateTime] <= ?")
            values.append(end_date.toPyDate().strftime("%Y-%m-%d %H:%M:%S"))

        where_clause = ""
        if filters:
            where_clause = "WHERE " + " AND ".join(filters)

        # Query the database
        with self.parent.lock:
            dbPath = getattr(self.parent, "dbPath", DATABASE_PATH)
            conn = sqlite3.connect(dbPath)
            cur = conn.cursor()
            try:
                cur.execute("PRAGMA table_info(calculations)")
                columns = [row[1] for row in cur.fetchall()]
                sql = f"SELECT * FROM calculations {where_clause}"
                print(sql)
                print(tuple(values))
                cur.execute(sql, tuple(values))
                rows = cur.fetchall()
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to query database: {e}")
                conn.close()
                return

        if not rows:
            QMessageBox.information(self, "No Data", "No data found for the selected filters.")
            conn.close()
            return

        # Prompt user for file path
        path, _ = QFileDialog.getSaveFileName(self, "Save as…", "", "Excel Files (*.xlsx)")
        if not path:
            conn.close()
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # Write to Excel
        try:
            df = pd.DataFrame(rows, columns=columns)

            if os.path.exists(path):
                wb = load_workbook(path)
                # Generate a unique sheet name
                base_name = "Export"
                i = 1
                while True:
                    sheet_name = f"{base_name}{i}"
                    if sheet_name not in wb.sheetnames:
                        break
                    i += 1
                # Write DataFrame to the new sheet
                ws = wb.create_sheet(sheet_name)
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                for row_idx, row in enumerate(df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                wb.save(path)
            else:
                df.to_excel(path, index=False)

            QMessageBox.information(self, "Success", f"Data exported to {path}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export to Excel: {e}")
        finally:
            conn.close()

def save_to_db(table, rows, action = "", query = "",inputs = None, keys = None, connection = None, lock = None, db = None):
    try:
        dbPath = db if db else DATABASE_PATH
        if lock is not None:
            lock.acquire()
        if connection is None:
            conn = sqlite3.connect(dbPath)
            cur = conn.cursor()
        else:
            conn = connection
            cur = connection.cursor()
        if action == "reset":
            cur.execute(f"DROP TABLE IF EXISTS {table}")
            conn.commit()
        elif action == "clear":
            cur.execute(f"DELETE FROM {table}")
            conn.commit()
        elif action == "add":
            try:
                for row in rows:
                    cols = list(row.keys())
                    quoted_cols = ','.join(f'"{c}"' for c in cols)
                    placeholders = ','.join('?' for _ in cols)
                    sql = f'INSERT INTO "{table}" ({quoted_cols}) VALUES ({placeholders})'
                    vals = tuple(str(row.get(c, '')) for c in cols)
                    cur.execute(sql,vals)
                    conn.commit()
            except Exception as e:
                print(f"Error inserting row into database: {e}")
                print("e.args:", e.args)
                # maybe also:
                try:
                    print(traceback.format_exc())
                except:
                    pass
        elif action == "calculationUpdate":
            try:
                cur.execute("DELETE FROM calculations WHERE [dateTime] = ?", inputs) #inputs input should be the date for deletion
                for row in rows:
                    cols = list(row.keys())
                    quoted_cols = ','.join(f'"{c}"' for c in cols)
                    placeholders = ','.join('?' for _ in cols)
                    sql = (f"INSERT INTO calculations ({quoted_cols}) VALUES ({placeholders})")
                    vals = tuple(str(row.get(c, '')) for c in cols)
                    cur.execute(sql,vals)
                conn.commit()
            except Exception as e:
                print(f"Error updating calculations in database: {e}")
                print("e.args:", e.args)
                # maybe also:
                try:
                    import traceback
                    print(traceback.format_exc())
                except:
                    pass
        elif action == "replace":
            cur.execute(query,inputs)
            conn.commit()
        elif rows:
            if keys is None:
                cols = list(rows[0].keys())
            else:
                cols = list(keys)
            quoted_cols = ','.join(f'"{c}"' for c in cols)
            col_defs = ','.join(f'"{c}" TEXT' for c in cols)
            if True:
                cur.execute(f'DROP TABLE IF EXISTS "{table}";')
            cur.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({col_defs})')
            cur.execute(f'DELETE FROM "{table}"')
            placeholders = ','.join('?' for _ in cols)
            sql = f'INSERT INTO "{table}" ({quoted_cols}) VALUES ({placeholders})'
            vals = [tuple(str(row.get(c, '')) for c in cols) for row in rows]
            cur.executemany(sql, vals)
            conn.commit()
        else:
            print(f"No rows found for data input to '{table}'")
        if connection is None:
            conn.close()
        else:
            cur.close()
        if lock is not None:
            lock.release()
        return True
    except Exception as e:
        print(f"DB save failed. closing connections {e}, {e.args}") 
        try:
            if lock is not None:
                lock.release()
            cur.close()
        except:
            pass
        return False
def load_from_db(table, condStatement = "",parameters = None, cursor = None, lock = None, db = None):
    try:
        dbPath = db if db else DATABASE_PATH
        if lock is not None:
            lock.acquire()
        # Transactions
        if os.path.exists(dbPath):
            if cursor is None:
                conn = sqlite3.connect(dbPath)
                cur = conn.cursor()
            else:
                cur = cursor
            try:
                if condStatement != "" and parameters is not None:
                    cur.execute(f'SELECT * FROM {table} {condStatement}',parameters)
                elif condStatement != "" and parameters is None:
                    cur.execute(f'SELECT * FROM {table} {condStatement}')
                else:
                    cur.execute(f'SELECT * FROM {table}')
                cols = [d[0] for d in cur.description]
                rows = [dict(zip(cols, row)) for row in cur.fetchall()]
                if cursor is None:
                    conn.close()
                if lock is not None:
                    lock.release()
                return rows
            except Exception as e:
                try:
                    if parameters is not None and table != "calculations":
                        print(f"Error loading from database: {e}, table: {table} condStatment: {condStatement}, parameters: {parameters}")
                    elif table != "calculations":
                        print(f"Error loading from database: {e}, table: {table} condStatment: {condStatement}")
                    else:
                        print(f"Info: {e}, {e.args}")
                    if cursor is None:
                        conn.close()
                except:
                    pass
                if lock is not None:
                    lock.release()
                return []
        else:
            if lock is not None:
                lock.release()
            return []
    except:
        print("DB load failed. closing connections")
        try:
            if lock is not None:
                lock.release()
            if cursor is None:
                cur.close()
        except:
            pass
def updateStatus(pool,totalLoops, lock, status = "Working", connection = None, db = None):
    failure = False
    dbPath = db if db else DATABASE_PATH
    try:
        
        with lock:
            if connection is None:
                conn = sqlite3.connect(dbPath)
            else:
                conn = connection
            c = conn.cursor()

            c.execute("SELECT status FROM progress WHERE status = ?",("Failed",))
            failed = c.fetchall()
            if len(failed) > 0:
                failure = True
            # Update or insert progress for this worker
            if status in ("Working","Initialization") or pool == "DummyFail":
                c.execute("""
                    INSERT INTO progress (pool, completed, total,status)
                    VALUES (?, -1, ?,?)
                    ON CONFLICT(pool) DO UPDATE SET completed = completed + 1, status = ?
                """, (pool, totalLoops,status,status))
            elif status == "Completed":
                c.execute("UPDATE progress SET completed = completed + 1, status = ? WHERE pool = ?", (status,pool))
            else:
                c.execute("UPDATE progress SET status = ? WHERE pool = ?", (status,pool))

            conn.commit()
            if connection is None:
                conn.close()
    except Exception as e:
        print(f"Error updating status: {e}")
    return failure
def processPool(poolData : dict,selfData : dict, statusQueue, dbQueue, failed):
    #Function to take all the information for one pool, calculate all relevant information, and return a list of the calculations
    #Inputs:
    #   poolData: dict with information relevant to this specific pool
    #   selfData: dict with information common to every pool
    #   statusQueue: a multiprocessing Manager queue for all worker threads to send progress bar and status updates. Minimizes database wait time
    #   dbQueue: a multiprocessing manager queue for worker threads to send final database updates to allow the worker to complete and not block the database
    #   failed: a multiprocessing variable. Begins negative. If any worker flags it as true, all workers will see it and halt if they hit the failure checkpoint
    try:
        noCalculations = selfData.get("noCalculations") #boolean of whether or not previous calculations exist to pull from
        months = selfData.get("months") #list of pre-prepared data for each month
        fundList = selfData.get("fundList") #list of funds/investments and some accompanying data (such as asset class level 3)
        calculations = []
        earliestChangeDate = poolData.get("earliestChangeDate") #earliest date for new data from last API pull
        pool = poolData.get("poolName")
        cache = poolData.get("cache") #dataset of all relevant transactions and account balances for the pool
        newMonths = []

        insert_low = [] #lists to store any database changes to do once calculations are complete
        update_low = []
        insert_high = []
        update_high = []


        if not noCalculations: #if there are calculations, find all months before the data pull, and then pull those calculations
            for month in months:
                #if the calculations for the month have already been complete, pull the old data
                if earliestChangeDate > datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S"):
                    calculations.extend(cache.get("calculations", {}).get(month["dateTime"], []))
                else:
                    newMonths.append(month)
        else:
            newMonths = months #check all months if there are no previous calculations
        IRRtrack = {} #dict of each fund's cash flows and dates for IRR calculation
        IRRinvestorTrack = {} #dict of each investor's cash flows and dates for IRR calculation
        for month in newMonths: #loops through every month relevant to the pool
            monthFundIRRtrack = {}
            statusQueue.put((pool,len(newMonths),"Working")) #puts to queue to update loading bar status. Allows computations to continue
            if failed.value: #if other workers failed, halt the process
                print(f"Exiting worker {pool} due to other failure...")
                return []
            totalDays = int(datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S").day  - datetime.strptime(month["tranStart"], "%Y-%m-%dT%H:%M:%S").day) + 1 #total days in month for MD den
            poolFunds = cache.get("positions_low", {}).get(month["dateTime"], []) #account balances for the pool
            #find MD denominator for each investor
            #find total gain per pool
            funds = []
            fundNames = []
            startEntries = {}
            endEntries = {}
            for account in poolFunds: #finds all fund starting and ending balances for the month
                if account["Target name"] not in fundNames:
                    fundNames.append(account["Target name"])
                    funds.append({"fundName" : account["Target name"], "hidden" : False})
                if account["Date"] == month["accountStart"]:
                    if account["Target name"] not in startEntries:
                        startEntries[account["Target name"]] = [account,]
                    else:
                        startEntries[account["Target name"]].append(account)
                elif account["Date"] == month["endDay"]:
                    if account["Target name"] not in endEntries:
                        endEntries[account["Target name"]] = [account,]
                    else:
                        endEntries[account["Target name"]].append(account)

            hiddenFunds = cache.get("transactions_low", {}).get(month["dateTime"], [])
            #funds that do not have account positions but are relevant to the pool (ex: deferred liabilities)
            allPoolTransactions = {}
            for transaction in hiddenFunds:
                if transaction["Target name"] not in fundNames:
                    fundNames.append(transaction["Target name"])
                    funds.append({"fundName" : transaction["Target name"], "hidden" : True})
                if transaction["Target name"] not in allPoolTransactions:
                    allPoolTransactions[transaction["Target name"]] = [transaction,]
                else:
                    allPoolTransactions[transaction["Target name"]].append(transaction)
            poolGainSum = 0
            poolNAV = 0
            poolMDdenominator = 0
            poolWeightedCashFlow = 0
            poolCashFlow = 0
            fundEntryList = []
            for fundDict in funds: #iterate through all funds to find the pool NAV and MD den
                fund = fundDict["fundName"]
                if fund in (None,'None'):
                    continue
                assetClass = None
                subAssetClass = None
                fundClassification = None
                startEntryCache = startEntries.get(fund, [])
                endEntryCache = endEntries.get(fund, [])
                startEntry = copy.deepcopy(startEntryCache)
                endEntry = copy.deepcopy(endEntryCache)
                createFinalValue = False
                noStartValue = False
                if len(startEntry) < 1: #no start value, so NAV = 0
                    startEntry = [{nameHier["Value"]["dynLow"] : 0}]  #nameHier is a dictionary for common references to specific names. 
                    noStartValue = True
                    commitment = 0
                    unfunded = 0
                else: #instantiate starting data
                    assetClass = startEntry[0]["ExposureAssetClass"]
                    subAssetClass = startEntry[0]["ExposureAssetClassSub-assetClass(E)"]
                    fundClassification = startEntry[0]["Target nameExposureHFClassificationLevel2"]
                    commitment = float(startEntry[0].get(nameHier["Commitment"]["local"],0))
                    unfunded = float(startEntry[0].get(nameHier["Unfunded"]["local"],0))
                if len(startEntry) > 1: #combines the values for fund sub classes for calculations
                    split = {}
                    foundDuplicate = False
                    for entry in startEntry: #split the entries by fundclass to check for duplicates
                        fundClass = entry.get(nameHier["FundClass"]["dynLow"])
                        if fundClass not in split:
                            split[fundClass] = [entry,]
                        else:
                            split[fundClass].append(entry)
                            foundDuplicate = True
                    singleEntries = []
                    if foundDuplicate: #if duplicates, loop through to find the best balance type
                        for fundClass in split: #loop by fund
                            if len(split.get(fundClass)) > 1: #check if duplicates
                                foundType = False
                                for balanceType in balanceTypePriority: #loop through balance types by priority
                                    for entry in split.get(fundClass): #loop through the duplicate entries
                                        if entry.get("Balancetype") == balanceType and entry.get(nameHier["Value"]["dynLow"]) not in (None,"None"): #if the balance type is preferred, add the entry and break
                                            singleEntries.append(entry)
                                            foundType = True
                                            break
                                    if foundType: #stop balance type checking if found
                                        break
                                if not foundType: #reaches if nothing was found
                                    for entry in split.get(fundClass): #loop through to find the first with a value
                                        if entry.get(nameHier["Value"]["dynLow"]) not in (None,"None"): #if the balance type is preferred, add the entry and break
                                            singleEntries.append(entry)
                                            foundType = True
                                            break
                                    if not foundType: #final attempt just take the first entry
                                        singleEntries.append(split.get(fundClass)[0])
                            else: #no duplicates for this fund
                                singleEntries.append(split.get(fundClass)[0])
                    else:
                        singleEntries.extend(startEntry)
                    NAV = 0
                    for entry in singleEntries:
                        if entry.get(nameHier["Value"]["dynLow"]) not in (None,0,"None"):
                            NAV += float(entry[nameHier["Value"]["dynLow"]]) #adds values to the first index
                    startEntry[0][nameHier["Value"]["dynLow"]] = str(NAV)
                if len(endEntry) < 1: #no end account balance yet, so create it.  
                    # !!!!!!!!! Should I remove this and just not use the fund if an end balance is not there?
                    createFinalValue = True
                    endEntry = [{nameHier["Value"]["dynLow"] : 0}]
                elif assetClass is None or subAssetClass is None or fundClassification is None: #first of several attempts to find the fund information.
                    assetClass = endEntry[0]["ExposureAssetClass"]
                    subAssetClass = endEntry[0]["ExposureAssetClassSub-assetClass(E)"]
                    fundClassification = endEntry[0]["Target nameExposureHFClassificationLevel2"]
                if len(endEntry) > 1: #combine sub funds for calculations
                    split = {}
                    foundDuplicate = False
                    for entry in endEntry: #split the entries by fundclass to check for duplicates
                        fundClass = entry.get(nameHier["FundClass"]["dynLow"])
                        if fundClass not in split:
                            split[fundClass] = [entry,]
                        else:
                            split[fundClass].append(entry)
                            foundDuplicate = True
                    singleEntries = []
                    if foundDuplicate: #if duplicates, loop through to find the best balance type
                        for fundClass in split: #loop by fund
                            if len(split.get(fundClass)) > 1: #check if duplicates
                                foundType = False
                                for balanceType in balanceTypePriority: #loop through balance types by priority
                                    for entry in split.get(fundClass): #loop through the duplicate entries
                                        if entry.get("Balancetype") == balanceType and entry.get(nameHier["Value"]["dynLow"]) not in (None,"None"): #if the balance type is preferred, add the entry and break
                                            singleEntries.append(entry)
                                            foundType = True
                                            break
                                    if foundType: #stop balance type checking if found
                                        break
                                if not foundType: #reaches if nothing was found
                                    for entry in split.get(fundClass): #loop through to find the first with a value
                                        if entry.get(nameHier["Value"]["dynLow"]) not in (None,"None"): #if the balance type is preferred, add the entry and break
                                            singleEntries.append(entry)
                                            foundType = True
                                            break
                                    if not foundType: #final attempt take first entry
                                        singleEntries.append(split.get(fundClass)[0])
                            else: #no duplicates for this fund
                                singleEntries.append(split.get(fundClass)[0])
                    else:
                        singleEntries.extend(endEntry)
                    NAV = 0
                    for entry in singleEntries:
                        if entry.get(nameHier["Value"]["dynLow"]) not in (None,0,"None"):
                            NAV += float(entry.get(nameHier["Value"]["dynLow"])) #adds values to the first index
                    endEntry[0][nameHier["Value"]["dynLow"]] = str(NAV)
                startEntry = startEntry[0]
                endEntry = endEntry[0]
                fundTransactions = allPoolTransactions.get(fund,[]) 
                cashFlowSum = 0
                weightedCashFlow = 0
                for transaction in fundTransactions: #get fund data, cash flows, and commitment alterations
                    if assetClass is None or assetClass == "None":
                        assetClass = transaction["SysProp_FundTargetNameAssetClass(E)"]
                    if subAssetClass is None or subAssetClass == "None":
                        subAssetClass = transaction["SysProp_FundTargetNameSub-assetClass(E)"]
                    if fundClassification is None or fundClassification == "None":
                        fundClassification = transaction["Target nameExposureHFClassificationLevel2"]
                    if transaction["TransactionType"] not in commitmentChangeTransactionTypes and transaction[nameHier["CashFlow"]["dynLow"]] not in (None, "None"):
                        cashFlowSum -= float(transaction[nameHier["CashFlow"]["dynLow"]])
                        backDate = calculateBackdate(transaction, noStartValue) #Uses dynamo transaction time logic to decide to subtract one day or not
                        weightedCashFlow -= float(transaction[nameHier["CashFlow"]["dynLow"]])  *  (totalDays -int(datetime.strptime(transaction["Date"], "%Y-%m-%dT%H:%M:%S").day) + backDate)/totalDays
                        if transaction.get(nameHier["Unfunded"]["dynLow"]) not in (None,"None"):
                            unfunded += float(transaction[nameHier["Unfunded"]["value"]])
                        if fund not in IRRtrack:
                            IRRtrack[fund] = {"cashFlows" : [], "dates" : []}
                        IRRtrack[fund]["cashFlows"].append(float(transaction[nameHier["CashFlow"]["dynLow"]]))
                        IRRtrack[fund]["dates"].append(datetime.strptime(transaction["Date"], "%Y-%m-%dT%H:%M:%S") - relativedelta(days=backDate))
                        if fund not in monthFundIRRtrack:
                            monthFundIRRtrack[fund] = {"cashFlows" : [], "dates" : []}
                        monthFundIRRtrack[fund]["cashFlows"].append(float(transaction[nameHier["CashFlow"]["dynLow"]]))
                        monthFundIRRtrack[fund]["dates"].append(datetime.strptime(transaction["Date"], "%Y-%m-%dT%H:%M:%S") - relativedelta(days=backDate))
                    elif transaction["TransactionType"] in commitmentChangeTransactionTypes and transaction.get("TransactionType") not in (None,"None"):
                        commitment += float(transaction[nameHier["Commitment"]["dynLow"]])
                        unfunded += float(transaction[nameHier["Commitment"]["dynLow"]])
                try:
                    if startEntry[nameHier["Value"]["dynLow"]] in (None, "None"):
                        startEntry[nameHier["Value"]["dynLow"]] = 0
                    if endEntry[nameHier["Value"]["dynLow"]] in (None, "None"):
                        endEntry[nameHier["Value"]["dynLow"]] = 0
                    if createFinalValue:
                        #implies there is no gain (cash account)
                        endEntry[nameHier["Value"]["dynLow"]] = float(startEntry[nameHier["Value"]["dynLow"]]) + cashFlowSum    
                    fundGain = (float(endEntry[nameHier["Value"]["dynLow"]]) - float(startEntry[nameHier["Value"]["dynLow"]]) - cashFlowSum)
                    fundMDdenominator = float(startEntry[nameHier["Value"]["dynLow"]]) + weightedCashFlow
                    fundNAV = float(endEntry[nameHier["Value"]["dynLow"]])
                    fundReturn = fundGain/fundMDdenominator * 100 if fundMDdenominator != 0 else 0
                    if fund in IRRtrack:
                        IRRitd = calculate_xirr([*IRRtrack[fund]["cashFlows"], fundNAV], [*IRRtrack[fund]["dates"], datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S")])
                    else:
                        IRRitd = None
                    if unfunded < 0:
                        unfunded = 0 #corrects for if original commitment was not logged properly
                    if createFinalValue: #builds an entry to put into the database and cache if it is missing
                        fundEOMentry = {"Date" : month["endDay"], "Source name" : pool, "Target name" : fund , nameHier["Value"]["dynLow"] : endEntry[nameHier["Value"]["dynLow"]],
                                            "Balancetype" : "Calculated_R", "ExposureAssetClass" : assetClass, "ExposureAssetClassSub-assetClass(E)" : subAssetClass,
                                            nameHier["Commitment"]["local"] : commitment, nameHier["Unfunded"]["local"] : unfunded,
                                            nameHier["sleeve"]["local"] : fundList.get(fund,None), nameHier["Classification"]["dynLow"] : fundClassification}
                        insert_low.append(fundEOMentry)
                        # update cache for subsequent months
                        for m in newMonths:
                            if m["accountStart"] <= month["endDay"] <= m["endDay"]:
                                cache.setdefault("positions_low", {}).setdefault(m["dateTime"], []).append(fundEOMentry)
                    else: #update database and cache with the calculated commitment, unfunded, and sleeve (asset lvl 3)
                        query = f"UPDATE positions_low SET [{nameHier['Commitment']["local"]}] = ? , [{nameHier['Unfunded']["local"]}] = ?, [{nameHier["sleeve"]["local"]}] = ? WHERE [Source name] = ? AND [Target name] = ? AND [Date] = ?"
                        inputs = (commitment,unfunded,fundList.get(fund),pool,fund,month["endDay"])
                        update_low.append(inputs)
                        # update cache for all months referencing this date
                        for m in newMonths:
                            if m["accountStart"] <= month["endDay"] <= m["endDay"]:
                                for lst in cache.get("positions_low", {}).get(m["dateTime"], []):
                                    if lst["Target name"] == fund and lst["Date"] == month["endDay"]:
                                        lst[nameHier["Commitment"]["local"]] = commitment
                                        lst[nameHier["Unfunded"]["local"]] = unfunded
                                        lst[nameHier["sleeve"]["local"]] = fundList.get(fund)
                    #sum each fund value into the pool totals
                    poolGainSum += fundGain
                    poolMDdenominator += fundMDdenominator
                    poolNAV += fundNAV
                    poolCashFlow += cashFlowSum
                    poolWeightedCashFlow += weightedCashFlow
                    monthFundEntry = {"dateTime" : month["dateTime"], "Investor" : "Total Fund", "Pool" : pool, "Fund" : fund ,
                                    "assetClass" : assetClass, "subAssetClass" : subAssetClass,
                                    "NAV" : fundNAV, "Monthly Gain" : fundGain, "Return" : fundReturn , 
                                    "MDdenominator" : fundMDdenominator, "Ownership" : "", "Classification" : fundClassification,
                                    "Calculation Type" : "Total Fund",
                                    "IRR ITD" : IRRitd,
                                    nameHier["sleeve"]["local"] : fundList.get(fund),
                                    nameHier["Commitment"]["local"] : commitment,
                                    nameHier["Unfunded"]["local"] : unfunded}
                    if fund not in (None,"None"): #removing blank funds (found duplicate of Monogram in 'HF Direct Investments Pool, LLC - PE (2021)' with most None values)
                        calculations.append(monthFundEntry) #append to calculations for use in report generation and aggregation
                        fundEntryList.append([monthFundEntry, float(startEntry[nameHier["Value"]["dynLow"]])]) #fund data stored on its own for investor calculations


                except Exception as e:
                    print(f"Skipped fund {fund} for {pool} in {month["Month"]} because: {e} {e.args}")
                    #Testing flag. skips fund if the values are zero and cause an error
            if poolNAV == 0 and poolCashFlow == 0:
                #skips the pool if there is no cash flow or value in the pool
                continue
            poolReturn = poolGainSum/poolMDdenominator * 100 if poolMDdenominator != 0 else 0
            monthPoolEntry = {"dateTime" : month["dateTime"], "Investor" : "Total Pool", "Pool" : pool, "Fund" : None ,
                            "assetClass" : poolData.get("assetClass"), "subAssetClass" : poolData.get("subAssetClass") ,
                            "NAV" : poolNAV, "Monthly Gain" : poolGainSum, "Return" : poolReturn , "MDdenominator" : poolMDdenominator,
                                "Ownership" : None, "Calculation Type" : "Total Fund"} 
                            #generic pool data for investors calculations
            investorStartEntries = {}
            investorEndEntries = {}
            investorPositions = cache.get("positions_high", {}).get(month["dateTime"], []) #account balances for investors into the pool for the month
            for pos in investorPositions: #find start and end entries for each investor and sort them
                investor = pos["Source name"]
                if pos["Date"] == month["accountStart"]:
                    if investor not in investorStartEntries:
                        investorStartEntries[investor] = [pos,]
                    else:
                        investorStartEntries[investor].append(pos)
                if pos["Date"] == month["endDay"]:
                    if investor not in investorEndEntries:
                        investorEndEntries[investor] = [pos,]
                    else:
                        investorEndEntries[investor].append(pos)

            allInvestorTransactions = {}
            transactions = cache.get("transactions_high", {}).get(month["dateTime"], []) #all cashflow and commitment based transactions for investors into the pool for the month
            for tran in transactions: #sort by investor
                investor = tran["Source name"]
                if investor not in allInvestorTransactions:
                    allInvestorTransactions[investor] = [tran,]
                else:
                    allInvestorTransactions[investor].append(tran)


            investorMDdenominatorSum = 0
            tempInvestorDicts = {}
            poolOwnershipSum = 0
            for investor in set(investorStartEntries.keys()) | set(investorEndEntries.keys()) | set(allInvestorTransactions.keys()): 
                #iterate through each investor in the pool for the month
                #pool level loop for investors
                investorWeightedCashFlow = 0
                investorCashFlowSum = 0
                tempInvestorDict = {}
                startEntry_cache = investorStartEntries.get(investor)
                if startEntry_cache: #use starting entry
                    startEntry = startEntry_cache[0]
                    noStartValue = False
                else: #if no starting entry, take necessary variables and zero out the value
                    end_cache = investorEndEntries.get(investor)
                    if end_cache:
                        startEntry = copy.deepcopy(end_cache[0])
                        startEntry[nameHier["Value"]["dynHigh"]] = 0
                        noStartValue = True
                    else:
                        continue #ignore the investor completely if there is no starting or ending value
                if startEntry.get(nameHier["Value"]["dynHigh"]) in (None,"None"):
                    startEntry[nameHier["Value"]["dynHigh"]] = 0 #prevent float conversion errors
                investorTransactions = allInvestorTransactions.get(investor,[]) #all investor transactions in the pool for the month
                
                for transaction in investorTransactions: 
                    if transaction.get(nameHier["CashFlow"]["dynHigh"]) not in (None,"None"):
                        investorCashFlowSum -= float(transaction[nameHier["CashFlow"]["dynHigh"]])
                        backDate = calculateBackdate(transaction) #dynamo revert by a day logic
                        investorWeightedCashFlow -= float(transaction[nameHier["CashFlow"]["dynHigh"]])  *  (totalDays -int(datetime.strptime(transaction["Date"], "%Y-%m-%dT%H:%M:%S").day) + backDate)/totalDays
                investorMDdenominator = float(startEntry[nameHier["Value"]["dynHigh"]]) + investorWeightedCashFlow
                tempInvestorDict["MDden"] = investorMDdenominator
                tempInvestorDict["cashFlow"] = investorCashFlowSum
                tempInvestorDict["startVal"] = float(startEntry[nameHier["Value"]["dynHigh"]])
                tempInvestorDict["ExposureAssetClass"] = startEntry["ExposureAssetClass"]
                tempInvestorDict["ExposureAssetClassSub-assetClass(E)"] = startEntry["ExposureAssetClassSub-assetClass(E)"]
                tempInvestorDict[nameHier["Family Branch"]["local"]] = startEntry[nameHier["Family Branch"]["dynHigh"]]
                if tempInvestorDict["startVal"] == 0 and investorCashFlowSum == 0:
                    continue #ignore investors with no value
                EOM = investorEndEntries.get(investor,[])
                if len(EOM) > 0:
                    if EOM[0].get(nameHier["Value"]["dynHigh"]) in (None,"None"):
                        EOM[0][nameHier["Value"]["dynHigh"]] = 0
                if round(tempInvestorDict.get("startVal") + tempInvestorDict.get("cashFlow")) != 0 and len(EOM) > 0 and round(float(EOM[0].get(nameHier["Value"]["dynHigh"],0))) != 0:
                    #only accounts for investor gain (MD den) if they have not exited
                    #exit check: starting value + cashflow is zero OR there is no ending value
                    investorMDdenominatorSum += investorMDdenominator
                tempInvestorDicts[investor] = tempInvestorDict #store investor calculations for secondary iteration for fund level data
            monthPoolEntryInvestorList = [] #stores investor data for third iteration (not needed to be split, but remnant from old logic.)
            ownershipAdjustDict = {}
            for investor in tempInvestorDicts.keys():
                # second investor iteration to find the gain, return,ownership, and NAV values at pool level (i think it is not needed to be split, but remnant from old logic.)
                EOMcheck = investorEndEntries.get(investor,[])
                if len(EOMcheck) > 0:
                    if EOMcheck[0].get(nameHier["Value"]["dynHigh"]) in (None,"None"):
                        EOMcheck[0][nameHier["Value"]["dynHigh"]] = 0 #prevents float conversion errors
                investorMDdenominator = tempInvestorDicts[investor]["MDden"]
                if investorMDdenominatorSum == 0:
                    investorGain = 0 #0 if no true value in the pool. avoids errors
                else:
                    investorGain = poolGainSum * investorMDdenominator / investorMDdenominatorSum
                if investorMDdenominator == 0:
                    investorReturn = 0 #0 if investor has no value in pool. avoids error
                else:
                    investorReturn = investorGain / investorMDdenominator
                if round(tempInvestorDicts[investor]["startVal"] + tempInvestorDicts[investor]["cashFlow"]) == 0 or len(EOMcheck) == 0 or round(float(EOMcheck[0].get(nameHier["Value"]["dynHigh"],0))) == 0: 
                    #zero values if exited investor
                    #exit check: start value and cashflow sums to zero OR no end value OR end value is zero
                    investorEOM = 0
                    investorGain = 0
                    investorMDdenominator = 0
                    investorReturn = 0
                else:
                    investorEOM = tempInvestorDicts[investor]["startVal"] + tempInvestorDicts[investor]["cashFlow"] + investorGain
                monthPoolEntryInvestor = copy.deepcopy(monthPoolEntry) #uses pool data as template
                monthPoolEntryInvestor["Investor"] = investor
                monthPoolEntryInvestor[nameHier["Family Branch"]["local"]] = tempInvestorDicts[investor][nameHier["Family Branch"]["local"]]
                monthPoolEntryInvestor["NAV"] = investorEOM
                monthPoolEntryInvestor["Monthly Gain"] = investorGain
                monthPoolEntryInvestor["Return"] = investorReturn * 100
                monthPoolEntryInvestor["MDdenominator"] = investorMDdenominator
                ownershipPerc = investorEOM/poolNAV * 100 if poolNAV != 0 else 0
                monthPoolEntryInvestor["Ownership"] = ownershipPerc
                poolOwnershipSum += ownershipPerc
                inputs = [investorEOM,"Calculated_R", investor,pool, month["endDay"]]
                ownershipAdjustDict[investor] = [monthPoolEntryInvestor,inputs, EOMcheck]
            for data, inputs, EOMcheck in ownershipAdjustDict.values(): #run through to adjust ownerships if they did not total properly to 100%
                data["Ownership"] = data["Ownership"] * 100 /  poolOwnershipSum if poolOwnershipSum != 0 and ownershipCorrect else data["Ownership"]
                investorEOM = data["Ownership"] / 100 * poolNAV
                data["NAV"] = investorEOM
                monthPoolEntryInvestorList.append(data)
                inputs[0] = investorEOM
                inputs = tuple(inputs)
                if len(EOMcheck) > 0: #only update the database for the investor if they have account balances
                    if round(float(EOMcheck[0].get(nameHier["Value"]["dynHigh"],0))) != round(investorEOM): #don't push an update if the values are the same
                        update_high.append(inputs)
                        for m in newMonths:
                            if m["accountStart"] <= month["endDay"] <= m["endDay"]:
                                for lst in cache.get("positions_high", {}).get(m["dateTime"], []):
                                    if lst["Source name"] == investor and lst["Target name"] == pool and lst["Date"] == month["endDay"]:
                                        lst[nameHier["Value"]["dynHigh"]] = investorEOM
            for investorEntry in monthPoolEntryInvestorList:
                investor = investorEntry["Investor"]
                #final (3rd) investor level iteration to use the pool level results for the investor to calculate the fund level information
                for fundEntry, fundStartValue in fundEntryList:
                    fund = fundEntry["Fund"]
                    investorOwnership = investorEntry["Ownership"] # * 100 /  poolOwnershipSum if poolOwnershipSum != 0 and ownershipCorrect else investorEntry["Ownership"]
                    fundInvestorNAV = investorOwnership / 100 * fundEntry["NAV"]
                    fundInvestorGain = fundEntry["Monthly Gain"] / monthPoolEntry["Monthly Gain"] * investorEntry["Monthly Gain"] if monthPoolEntry["Monthly Gain"] != 0 else 0
                    fundInvestorMDdenominator = investorEntry["MDdenominator"] / monthPoolEntry["MDdenominator"] * fundEntry["MDdenominator"] if monthPoolEntry["MDdenominator"] != 0 else 0
                    fundInvestorReturn = fundInvestorGain / fundInvestorMDdenominator if fundInvestorMDdenominator != 0 else 0
                    fundInvestorOwnership = fundInvestorNAV /  fundEntry["NAV"] if fundEntry["NAV"] != 0 else 0
                    fundInvestorCommitment = fundEntry[nameHier["Commitment"]["local"]] * fundInvestorOwnership
                    fundInvestorUnfunded = fundEntry[nameHier["Unfunded"]["local"]] * fundInvestorOwnership

                    if investorEntry["MDdenominator"] != 0 and investorMDdenominatorSum != 0: 
                        if investor not in IRRinvestorTrack:
                            IRRinvestorTrack[investor] = {}
                        if fund not in IRRinvestorTrack[investor]:
                            IRRinvestorTrack[investor][fund] = {"cashFlows" : [], "dates" : []}
                            if fundStartValue != 0:
                                #if fund start value is not zero, the investment is not new, but the investor is being added
                                #apply ownership to previous fund NAV to get faux starting value for IRR
                                investorStartValue = fundStartValue * investorEntry["Ownership"] / 100
                                IRRinvestorTrack[investor][fund]["cashFlows"].append(investorStartValue * -1)
                                IRRinvestorTrack[investor][fund]["dates"].append(datetime.strptime(month["accountStart"], "%Y-%m-%dT%H:%M:%S"))
                        cashflows = monthFundIRRtrack.get(fundEntry["Fund"], {}).get("cashFlows", [])
                        dates = monthFundIRRtrack.get(fundEntry["Fund"], {}).get("dates", [])
                        for cashflow, date in zip(cashflows, dates):
                            adjustedCashflow = cashflow * investorEntry["MDdenominator"] / investorMDdenominatorSum #ratio the cashflow to their MDdenominator
                            IRRinvestorTrack[investor][fund]["cashFlows"].append(adjustedCashflow)
                            IRRinvestorTrack[investor][fund]["dates"].append(date)
                    if investorEntry["Investor"] in IRRinvestorTrack and fundEntry["Fund"] in IRRinvestorTrack[investorEntry["Investor"]]:
                        fundInvestorIRR = calculate_xirr([*IRRinvestorTrack[investorEntry["Investor"]][fund]["cashFlows"], fundInvestorNAV], [*IRRinvestorTrack[investorEntry["Investor"]][fund]["dates"], datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S")])
                    else:
                        fundInvestorIRR = None
                    monthFundInvestorEntry = {"dateTime" : month["dateTime"], "Investor" : investorEntry["Investor"], "Pool" : pool, "Fund" : fundEntry["Fund"] ,
                                    "assetClass" : fundEntry["assetClass"], "subAssetClass" : fundEntry["subAssetClass"],
                                    "NAV" : fundInvestorNAV, "Monthly Gain" : fundInvestorGain , "Return" :  fundInvestorReturn * 100, 
                                    "MDdenominator" : fundInvestorMDdenominator, "Ownership" : fundInvestorOwnership * 100,
                                    "Classification" : fundEntry["Classification"], nameHier["Family Branch"]["local"] : investorEntry[nameHier["Family Branch"]["local"]],
                                    nameHier["Commitment"]["local"] : fundInvestorCommitment, nameHier["Unfunded"]["local"] : fundInvestorUnfunded, 
                                    "Calculation Type" : "Total Fund",
                                    "IRR ITD" : fundInvestorIRR,
                                    nameHier["sleeve"]["local"] : fundList.get(fund)
                                    }
                    calculations.append(monthFundInvestorEntry) #add fund level data to calculations for use in aggregation and report generation
            #end of months loop
        #commands to add database updates to the queues
        if insert_low:
            dbInputs = {"type" : "insert", "data" : ["positions_low", insert_low, "add"]}
            dbQueue.put(dbInputs)
        if update_low:
            query = f"UPDATE positions_low SET [{nameHier['Commitment']['local']}] = ?, [{nameHier['Unfunded']['local']}] = ?, [{nameHier['sleeve']['local']}] = ? WHERE [Source name] = ? AND [Target name] = ? AND [Date] = ?"
            dbInputs = {"type" : "update", "data" : [query, update_low]}
            dbQueue.put(dbInputs)
        if insert_high:
            dbInputs = {"type" : "insert", "data" : ["positions_high", insert_high, "add"]}
            dbQueue.put(dbInputs)
        if update_high: #inputs = (investorEOM, investor,pool, month["endDay"])
            query = f"UPDATE positions_high SET [{nameHier['Value']['dynHigh']}] = ?, [Balancetype] = ? WHERE [Source name] = ? AND [Target name] = ? AND [Date] = ?"
            dbInputs = {"type" : "update", "data" : [query, update_high]}
            dbQueue.put(dbInputs)
        statusQueue.put((pool,len(newMonths),"Completed")) #push completed status update to the main thread
        return calculations
    except Exception as e: #halt operations for failure or force close/cancel
        statusQueue.put((pool,len(newMonths),"Failed"))
        print(f"Worker for {poolData.get("poolName")} failed.")
        failed = True
        try:
            trace = traceback.format_exc()
            print(trace)
            logging.error(trace)
        except:
            pass
        logging.error(e)
        print("\n")
        return []
def processPoolTransactions(poolData : dict,selfData : dict, statusQueue, dbQueue, failed):
    #Function to take all the information for one pool, calculate all relevant information, and return a list of the calculations
    #Inputs:
    #   poolData: dict with information relevant to this specific pool
    #   selfData: dict with information common to every pool
    #   statusQueue: a multiprocessing Manager queue for all worker threads to send progress bar and status updates. Minimizes database wait time
    #   dbQueue: a multiprocessing manager queue for worker threads to send final database updates to allow the worker to complete and not block the database
    #   failed: a multiprocessing variable. Begins negative. If any worker flags it as true, all workers will see it and halt if they hit the failure checkpoint
    try:
        noCalculations = selfData.get("noCalculations") #boolean of whether or not previous calculations exist to pull from
        months = selfData.get("months") #list of pre-prepared data for each month
        fundList = selfData.get("fundList") #list of funds/investments and some accompanying data (such as asset class level 3)
        calculations = []
        earliestChangeDate = poolData.get("earliestChangeDate") #earliest date for new data from last API pull
        pool = poolData.get("poolName")
        cache = poolData.get("cache") #dataset of all relevant transactions and account balances for the pool
        newMonths = []

        if not noCalculations: #if there are calculations, find all months before the data pull, and then pull those calculations
            for month in months:
                #if the calculations for the month have already been complete, pull the old data
                if earliestChangeDate > datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S"):
                    calculations.extend(cache.get("calculations", {}).get(month["dateTime"], []))
                else:
                    newMonths.append(month)
        else:
            newMonths = months #check all months if there are no previous calculations
        for month in newMonths: #loops through every month relevant to the pool
            statusQueue.put((pool,len(newMonths),"Working")) #puts to queue to update loading bar status. Allows computations to continue
            if failed.value: #if other workers failed, halt the process
                print(f"Exiting worker {pool} due to other failure...")
                return []
            poolCashFlow = 0
            lowTransactions = cache.get("transactions_low", {}).get(month["dateTime"], [])
            for transaction in lowTransactions: #get fund data, cash flows, and commitment alterations
                if transaction["TransactionType"] not in commitmentChangeTransactionTypes and transaction[nameHier["CashFlow"]["dynLow"]] not in (None, "None"):
                    poolCashFlow -= float(transaction[nameHier["CashFlow"]["dynLow"]])
            investorPoolCashFlow = 0
            transactions = cache.get("transactions_high", {}).get(month["dateTime"], []) #all cashflow and commitment based transactions for investors into the pool for the month
            for tran in transactions:
                if tran["TransactionType"] not in commitmentChangeTransactionTypes and tran[nameHier["CashFlow"]["dynHigh"]] not in (None, "None"):
                    investorPoolCashFlow -= float(tran[nameHier["CashFlow"]["dynHigh"]])
            difference = round(poolCashFlow - investorPoolCashFlow,2) * -1
            if difference != 0:
                monthPoolEntry = {"dateTime" : month["dateTime"], "Investor" : "Total Pool", "Pool" : pool, 
                                        "Transaction Sum" : difference,
                                        "Calculation Type" : "Total Pool",
                }
                calculations.append(monthPoolEntry) #append to calculations for use in report generation and aggregation

            #end of months loop
        statusQueue.put((pool,len(newMonths),"Completed")) #push completed status update to the main thread
        return calculations
    except Exception as e: #halt operations for failure or force close/cancel
        statusQueue.put((pool,len(newMonths),"Failed"))
        print(f"Worker for {poolData.get("poolName")} failed.")
        print(traceback.format_exc())
        print("\n")
        return []
            
def calculateBackdate(transaction,noStartValue = False):
    time = transaction.get(nameHier["Transaction Time"]["dynLow"])
    monthDay = datetime.strptime(transaction.get("Date"), "%Y-%m-%dT%H:%M:%S").day
    if noStartValue:
        if time not in (None,"None") and time.lower() == "end of day":
            backDate = 0 #"no start value and end of day"
        else:
            backDate = 1 #"no start value and not end of day"
    elif time in (None,"None"):
        if monthDay == 1:
            backDate = 1 #"First day of month"
        else:
            backDate = 0#"No timing and not first day of month"
    elif time.lower() == "end of day":
        backDate = 0#"End of day"
    else:
        backDate = 1 #"Beginning of day"
    return backDate

def exportTableToExcel(self, rows, headers = None):
    # 1) prompt user
    path, _ = QFileDialog.getSaveFileName(
        self, "Save as…", "", "Excel Files (*.xlsx)"
    )
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"            
    def exportTableToExcel():
        try:
            data = rows  # list of dicts

            if headers is None:
                all_cols = set()
                for row in rows:
                    for key in row:
                        all_cols.add(key)
                all_cols = list(all_cols)
            else:
                all_cols = headers

            # 4) create workbook
            wb = Workbook()
            ws = wb.active

            rowStart = 1
            for idx, colname in enumerate(all_cols, start=1):
                ws.cell(row=rowStart, column=idx, value=colname)


            # 7) populate rows
            for r, row_dict in enumerate(data, start=rowStart + 1):



                # data cells with proper formatting
                for c, colname in enumerate(all_cols, start=1):
                    val = row_dict.get(colname, None)
                    try: #make numerical format if possible
                        val = float(val)
                    except:
                        pass
                    cell = ws.cell(row=r, column=c, value=val)
                    if isinstance(val, (int, float)):
                        if colname not in percent_headers:
                            # show with commas, two decimals
                            cell.number_format = "#,##0.00"
                        else:
                            # interpret val as percentage (e.g. 10.5 → 10.5%)
                            cell.value = val / 100.0
                            cell.number_format = "0.00%"

            # 8) autofit column widths
            for idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    if cell.value is not None:
                        text = str(cell.value)
                        max_len = max(max_len, len(text))
                ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
        except Exception as e:
            gui_queue.put(lambda error = e, trace = traceback.format_exc(): QMessageBox.critical(self, "Processing error", f"{error} \n {trace}"))
        try:
            wb.save(path)
        except Exception as e:
            gui_queue.put(lambda error = e: QMessageBox.critical(self, "Save error", str(error)))
        else:
            gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
            gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
    executor.submit(exportTableToExcel)


class simpleMonthSelector(QWidget):
    currentTextChanged = pyqtSignal()
    def __init__(self, parent = None):
        super().__init__(parent)
        layout = QHBoxLayout(self)
        self.monthSelect = QComboBox()
        self.yearSelect = QComboBox()
        self.changeLock = False
        layout.addWidget(self.monthSelect)
        layout.addWidget(self.yearSelect)
        self.monthSelect.currentTextChanged.connect(self.emitSignal)
        self.yearSelect.currentTextChanged.connect(self.emitSignal)
    def addItems(self, items):
        self.changeLock = True
        months = set()
        years = set()
        for item in items:
            month, year = item.split(" ")
            months.add(month)
            years.add(year)
        self.monthSelect.addItems(sorted(months))
        self.yearSelect.addItems(sorted(years))
        self.changeLock = False
    def setCurrentText(self,text):
        self.changeLock = True
        [month,year] = text.split(" ")
        self.monthSelect.setCurrentText(month)
        self.yearSelect.setCurrentText(year)
        self.changeLock = False
    def currentText(self):
        month = self.monthSelect.currentText()
        year = self.yearSelect.currentText()
        joined = " ".join([month,year])
        return joined
    def emitSignal(self):
        if not self.changeLock:
            self.currentTextChanged.emit()



class displayWindow(QWidget):
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource = None, text = "", title=""):
        super().__init__(parent, flags)
        self.setWindowTitle(title)
        self.parent = parentSource
        if self.parent:
            self.setObjectName("mainPage")
            self.setStyleSheet(self.parent.appStyle)
        #self.resize(1000, 600)

        # Layout and table
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(text))

@attach_logging_to_class
class DictListModel(QAbstractTableModel):
    """
    Simple table model over a list of dicts.
    """
    def __init__(self, rows, headers, parent=None):
        super().__init__(parent)
        self._rows = rows
        self._headers = headers

    def rowCount(self, parent=QModelIndex()):
        return len(self._rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self._headers)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        row = self._rows[index.row()]
        key = self._headers[index.column()]
        # Display the cell text
        if role == Qt.DisplayRole:
            return str(row.get(key, ''))

        # Conditional background coloring
        if role == Qt.BackgroundRole:
            # Example: color 'Value' column green if above threshold
            try:
                Investor = str(row.get('Investor', ''))
                Fund = row.get('Fund', '')
                if Investor == "Total Portfolio":
                    return QBrush(Qt.darkGray)  
                elif Investor == "Total Asset":
                    return QBrush(QColor(181, 135, 235)) 
                elif Investor == "Total subAsset":
                    return QBrush(QColor(213, 193, 236)) 
                elif Investor == "Total Pool":
                    return QBrush(Qt.lightGray) 
                elif Fund is not None and Fund != "None" and Investor == "Total Fund":
                    return QBrush(QColor(181, 235, 135))  
                elif Fund is not None and Fund != "None": #Fund
                    return QBrush(QColor(213, 236, 193)) 
            except (ValueError, TypeError):
                pass

        # Alignment for numbers
        if role == Qt.TextAlignmentRole:
            try:
                float(row.get(key))
                return Qt.AlignVCenter | Qt.AlignRight
            except (ValueError, TypeError):
                return Qt.AlignVCenter | Qt.AlignLeft

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole and orientation == Qt.Horizontal:
            return self._headers[section]
        return None            

@attach_logging_to_class
class underlyingDataWindow(QWidget):
    """
    A window that loads data from four database sources in the parent,
    merges and sorts it by dateTime, and displays it in a QTableWidget
    with a unified set of columns.
    """
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource = None, db = None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle("Underlying Data Viewer")
        self.resize(1000, 600)
        self.db = db
        # Layout and table
        layout = QVBoxLayout(self)
        buttonBox = QWidget()
        buttonLayout = QHBoxLayout()
        exportBtn = QPushButton("Export Data to Excel")
        exportBtn.clicked.connect(self.exportTable)
        buttonLayout.addWidget(exportBtn)
        self.headerOptions = SortButtonWidget()
        self.headerOptions.popup.popup_closed.connect(self.buildTable)
        buttonLayout.addWidget(self.headerOptions)
        buttonBox.setLayout(buttonLayout)
        layout.addWidget(buttonBox)
        self.table = QTableWidget(self)
        layout.addWidget(self.table)
        self.success = False
        self.headerOrder = ["Date", "TradeDate", "_source","Source name","Target name", "CashFlowSys", "ValueInSystemCurrency","Balancetype","TransactionType"]
        self.buildTable()

    def exportTable(self, *_):
        # 1) prompt user
        path, _ = QFileDialog.getSaveFileName(
            self, "Save as…", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        def processExport():
            try:
                data = self.allData  # list of dicts

                all_cols = self.allCols


                if os.path.exists(path):
                    wb = load_workbook(path)
                    # Create a unique sheet name for export
                    base_name = "Export"
                    i = 1
                    while True:
                        sheet_name = f"{base_name}{i}"
                        if sheet_name not in wb.sheetnames:
                            break
                        i += 1
                    ws = wb.create_sheet(sheet_name)
                else:
                    wb = Workbook()
                    ws = wb.active

                rowStart = 1
                for idx, colname in enumerate(all_cols, start=1):
                    ws.cell(row=rowStart, column=idx, value=colname)


                # 7) populate rows
                for r, row_dict in enumerate(data, start=rowStart + 1):        
                    # data cells with proper formatting
                    for c, colname in enumerate(all_cols, start=1):
                        val = row_dict.get(colname, None)
                        try: #make numerical format if possible
                            val = float(val)
                        except:
                            pass
                        cell = ws.cell(row=r, column=c, value=val)
                        if isinstance(val, (int, float)):
                            if colname not in percent_headers:
                                # show with commas, two decimals
                                cell.number_format = "#,##0.00"
                            else:
                                # interpret val as percentage (e.g. 10.5 → 10.5%)
                                cell.value = val / 100.0
                                cell.number_format = "0.00%"

                # 8) autofit column widths
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is not None:
                            text = str(cell.value)
                            max_len = max(max_len, len(text))
                    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
            except Exception as e:
                gui_queue.put(lambda error = e, trace = traceback.format_exc(): QMessageBox.critical(self, "Processing error", f"{error} \n {trace}"))
            try:
                wb.save(path)
            except Exception as e:
                gui_queue.put(lambda error = e: QMessageBox.critical(self, "Save error", str(error)))
            else:
                gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
                gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
        executor.submit(processExport)    
    def buildTable(self):
        if self.parent.tableBtnGroup.checkedButton().text() == "Monthly Table":
            selectedMonth = datetime.strptime(self.parent.selectedCell["month"], "%B %Y")
            tranStart = selectedMonth.replace(day = 1)
            accountStart = tranStart - relativedelta(days= 1)
            allEnd = (tranStart + relativedelta(months=1)) - relativedelta(days=1)
        else:
            endTime = datetime.strptime(self.parent.dataEndSelect.currentText(),"%B %Y")
            allEnd = (endTime.replace(day = 1) + relativedelta(months=1)) - relativedelta(days=1)
            selection = self.parent.selectedCell["month"]
            if selection not in timeOptions or selection == "MTD": #MTD timeframe
                tranStart = endTime.replace(day = 1)
                accountStart = tranStart - relativedelta(days= 1)
            elif selection in ("ITD","IRR ITD"):
                tranStart = self.parent.dataTimeStart
                accountStart = self.parent.dataTimeStart
            else:
                if selection == "QTD":
                    subtract = (int((endTime.month)) % 3 if (int(endTime.month)) % 3 != 0 else 3) - 1
                elif selection == "YTD":
                    subtract = (int((endTime.month)) % 12 if (int(endTime.month)) % 12 != 0 else 12) - 1
                else:
                    subtract = int(selection.removesuffix("YR")) * 12 - 1
                tranStart = (endTime - relativedelta(months=subtract)).replace(day=1)
                accountStart = tranStart - relativedelta(days= 1)

        tranStart = datetime.strftime(tranStart,"%Y-%m-%dT00:00:00")
        accountStart = datetime.strftime(accountStart,"%Y-%m-%dT00:00:00")
        allEnd = datetime.strftime(allEnd,"%Y-%m-%dT00:00:00")
        dataType = self.parent.selectedCell["dataType"]
        if dataType == "Total":
            return
        dataType = dataType.removeprefix("Total ")
        code = self.parent.selectedCell["rowKey"]
        header, code= self.parent.separateRowCode(code)
        if header == "Cash ":
            header = "Cash"
        hier = code.removeprefix("##(").removesuffix(")##").split("::")
        hierSelections = self.parent.sortHierarchy.checkedItems()
        if dataType == "Fund":
            hier.append(header)
            hierSelections.append(dataType)
        if self.db:
            highTables = {"transactions_high" : tranStart}
            lowTables = {"transactions_low": tranStart}
        else:
            highTables = {"positions_high": accountStart,"transactions_high" : tranStart}
            lowTables = {"positions_low": accountStart,"transactions_low": tranStart}
        all_rows = []
        if not self.db and (self.parent.filterDict["Investor"].checkedItems() != [] or self.parent.filterDict["Family Branch"].checkedItems() != []): #investor to pool level entries
            for idx, table in enumerate(highTables.keys()):
                query = "WHERE"
                inputs = []
                for hierIdx, tier in enumerate(hier):
                    if tier == "hiddenLayer":
                        continue #hidden layer should not affect the query
                    for filter in self.parent.filterOptions:
                        dynName = filter.get("dynNameHigh")
                        if hierSelections[hierIdx] == filter["key"] and dynName is not None:
                            if filter["key"] == "assetClass" and idx == 1:
                                dynName = "SysProp_FundTargetNameAssetClass(E)"
                            elif filter["key"] == "subAssetClass" and idx == 1:
                                dynName = "SysProp_FundTargetNameSub-assetClass(E)"
                            query += f" [{dynName}] = ? AND"
                            inputs.append(tier)
                            break #continue to next tier
                if dataType == "Fund":
                    query += " [Target name] = ? AND"
                    inputs.append(self.parent.fundPoolLinks.get(header))
                for filter in self.parent.filterOptions:
                    filterSelections = self.parent.filterDict[filter["key"]].checkedItems()
                    dynName = filter.get("dynNameHigh")
                    if filter["key"] not in ("Classification") and filterSelections != [] and dynName is not None:
                        if filter["key"] == "assetClass" and idx == 1:
                            dynName = "SysProp_FundTargetNameAssetClass(E)"
                        elif filter["key"] == "subAssetClass" and idx == 1:
                            dynName = "SysProp_FundTargetNameSub-assetClass(E)"
                        if filter["key"] == "subAssetSleeve":
                            for sleeve in filterSelections:
                                if self.parent.sleeveFundLinks.get(sleeve) is not None:
                                    placeholders = ','.join('?' for _ in self.parent.sleeveFundLinks.get(tier)) 
                                    query += f" [Target name] in ({placeholders}) AND"
                                    inputs.extend(self.parent.sleeveFundLinks.get(tier))
                                else:
                                    print("Failed to find subAssetSleeve")
                        else:
                            placeholders = ','.join('?' for _ in filterSelections) 
                            query += f" [{dynName}] in ({placeholders}) AND"
                            inputs.extend(filterSelections)
                inputs.extend([highTables[table],allEnd])
                try:
                    rows = load_from_db(table,query.removesuffix("AND") + " AND [Date] BETWEEN ? AND ?", tuple(inputs), db=self.db)
                except Exception as e:
                    print(f"Error in call : {e} ; {e.args}")
                    rows = []
                for row in rows or []:
                    row['_source'] = table
                    all_rows.append(row)
        elif self.db:
            for idx, table in enumerate(highTables.keys()):
                query = "WHERE"
                inputs = []
                for hierIdx, tier in enumerate(hier):
                    if tier == "hiddenLayer":
                        continue #hidden layer should not affect the query
                    for filter in self.parent.filterOptions:
                        dynName = filter.get("dynNameHigh")
                        if hierSelections[hierIdx] == filter["key"] and dynName is not None:
                            if filter["key"] == "assetClass" and idx == 1:
                                dynName = "SysProp_FundTargetNameAssetClass(E)"
                            elif filter["key"] == "subAssetClass" and idx == 1:
                                dynName = "SysProp_FundTargetNameSub-assetClass(E)"
                            query += f" [{dynName}] = ? AND"
                            inputs.append(tier)
                            break #continue to next tier
                for filter in self.parent.filterOptions:
                    filterSelections = self.parent.filterDict[filter["key"]].checkedItems()
                    dynName = filter.get("dynNameHigh")
                    if filter["key"] not in ("Classification") and filterSelections != [] and dynName is not None:
                        if filter["key"] == "assetClass" and idx == 1:
                            dynName = "SysProp_FundTargetNameAssetClass(E)"
                        elif filter["key"] == "subAssetClass" and idx == 1:
                            dynName = "SysProp_FundTargetNameSub-assetClass(E)"
                        if filter["key"] == "subAssetSleeve":
                            for sleeve in filterSelections:
                                if self.parent.sleeveFundLinks.get(sleeve) is not None:
                                    placeholders = ','.join('?' for _ in self.parent.sleeveFundLinks.get(tier)) 
                                    query += f" [Target name] in ({placeholders}) AND"
                                    inputs.extend(self.parent.sleeveFundLinks.get(tier))
                                else:
                                    print("Failed to find subAssetSleeve")
                        else:
                            placeholders = ','.join('?' for _ in filterSelections) 
                            query += f" [{dynName}] in ({placeholders}) AND"
                            inputs.extend(filterSelections)
                inputs.extend([highTables[table],allEnd])
                try:
                    rows = load_from_db(table,query.removesuffix("AND") + " AND [Date] BETWEEN ? AND ?", tuple(inputs))
                except Exception as e:
                    print(f"Error in call : {e} ; {e.args}")
                    rows = []
                for row in rows or []:
                    row['_source'] = table
                    all_rows.append(row)
        for idx, table in enumerate(lowTables.keys()):
            query = "WHERE"
            inputs = []
            for hierIdx, tier in enumerate(hier): #iterate through each tier down to selection
                if tier == "hiddenLayer":
                        continue #hidden layer should not affect the query
                for filter in self.parent.filterOptions: #iterate through filter to find the matching keys
                    dynName = filter.get("dynNameLow")
                    if hierSelections[hierIdx] == filter["key"] and dynName is not None: #matching filter key
                        if filter["key"] == "assetClass" and idx == 1:
                            dynName = "SysProp_FundTargetNameAssetClass(E)"
                        elif filter["key"] == "subAssetClass" and idx == 1:
                            dynName = "SysProp_FundTargetNameSub-assetClass(E)"
                        if filter["key"] == "subAssetSleeve":
                            if self.parent.sleeveFundLinks.get(tier) is not None:
                                placeholders = ','.join('?' for _ in self.parent.sleeveFundLinks.get(tier)) 
                                query += f" [Target name] in ({placeholders}) AND"
                                inputs.extend(self.parent.sleeveFundLinks.get(tier))
                            else:
                                print("Failed to find subAssetSleeve")
                        elif filter["key"] == "Fund" and self.parent.cFundToFundLinks.get(tier) is not None: #account for consolidated funds
                            funds = self.parent.cFundToFundLinks.get(tier)
                            placeholders = ','.join('?' for _ in funds) 
                            inputs.extend(funds)
                            query += f" [Target name] in ({placeholders}) AND"
                        else:
                            query += f" [{dynName}] = ? AND"
                            inputs.append(tier)
                        break #continue to next tier
            for filter in self.parent.filterOptions:
                filterSelections = self.parent.filterDict[filter["key"]].checkedItems()
                dynName = filter.get("dynNameLow")
                if filterSelections != [] and dynName is not None:
                    if filter["key"] == "assetClass" and idx == 1:
                        dynName = "SysProp_FundTargetNameAssetClass(E)"
                    elif filter["key"] == "subAssetClass" and idx == 1:
                        dynName = "SysProp_FundTargetNameSub-assetClass(E)"
                    if filter["key"] == "subAssetSleeve":
                        for sleeve in filterSelections:
                            if self.parent.sleeveFundLinks.get(sleeve) is not None:
                                placeholders = ','.join('?' for _ in self.parent.sleeveFundLinks.get(sleeve)) 
                                query += f" [Target name] in ({placeholders}) AND"
                                inputs.extend(self.parent.sleeveFundLinks.get(sleeve))
                            else:
                                print("Failed to find subAssetSleeve")
                    else:
                        placeholders = ','.join('?' for _ in filterSelections) 
                        query += f" [{dynName}] in ({placeholders}) AND"
                        inputs.extend(filterSelections)
            inputs.extend([lowTables[table],allEnd])
            try:
                rows = load_from_db(table,query.removesuffix("AND") + " AND [Date] BETWEEN ? AND ?", tuple(inputs), db=self.db)
            except Exception as e:
                print(f"Error in call : {e}; {e.args}")
                rows = []
            for row in rows or []:
                row['_source'] = table
                all_rows.append(row) 
        self.allData = all_rows

        if self.db:
            #build organized differences by pool/investor versus transaction type
            diffTableDict = { "Total" : {"Transaction Type" : "Total", "Pool Cashflow" : 0, "Investor Cashflow" : 0}} 
            for transaction in all_rows: #build dict for easy sorting
                if transaction.get("CashFlowSys") not in (None,"None"):
                    tranType = transaction.get("TransactionType")
                    if tranType not in diffTableDict:
                        diffTableDict[tranType] = {"Transaction Type" : tranType, "Pool Cashflow" : 0, "Investor Cashflow" : 0}
                    if transaction.get("_source") == "transactions_low":
                        diffTableDict[tranType]["Pool Cashflow"] += float(transaction.get("CashFlowSys"))
                        diffTableDict["Total"]["Pool Cashflow"] += float(transaction.get("CashFlowSys"))
                    elif transaction.get("_source") == "transactions_high":
                        diffTableDict[tranType]["Investor Cashflow"] += float(transaction.get("CashFlowSys"))
                        diffTableDict["Total"]["Investor Cashflow"] += float(transaction.get("CashFlowSys"))
                    
            diffTable = []
            for tranType in diffTableDict: #calculate differences and put in list of dicts for table
                if tranType != "Total":
                    diffTableDict[tranType]["Difference"] = diffTableDict.get(tranType).get("Pool Cashflow") - diffTableDict.get(tranType).get("Investor Cashflow")
                    diffTable.append(diffTableDict.get(tranType))
            diffTableDict["Total"]["Difference"] = diffTableDict.get("Total").get("Pool Cashflow") - diffTableDict.get("Total").get("Investor Cashflow")
            diffTable.append(diffTableDict.get("Total"))
            diffHeaders = ["Transaction Type", "Pool Cashflow", "Investor Cashflow", "Difference"]
            self.parent.openTableWindow(diffTable, name = f"Transaction types for {hier} in {selectedMonth}", headers = diffHeaders)

        if len(all_rows) == 0:
            print("No rows found")
            return
        # 3) Sort by dateTime column (handles ISO or space-separated)
        def parse_dt(s):
            return datetime.strptime(s, "%Y-%m-%dT00:00:00")

        all_rows.sort(key=lambda r: parse_dt(r.get('Date', '')))

        # 4) Collect the union of all column keys
        if not self.headerOptions.active:
            all_cols = self.headerOrder
            for row in all_rows:
                for key in row.keys():
                    if key not in all_cols:
                        all_cols.append(key)
            self.allCols = all_cols
            self.headerOptions.set_items(all_cols)
        else:
            self.allCols = self.headerOptions.popup.get_checked_sorted_items()

        # 5) Configure the table widget
        self.table.setRowCount(len(all_rows))
        self.table.setColumnCount(len(self.allCols))
        self.table.setHorizontalHeaderLabels(self.allCols)

        # 6) Populate each cell
        for r, row in enumerate(all_rows):
            for c, key in enumerate(self.allCols):
                raw = row.get(key,"")
                try:
                    num = float(raw)
                    text = f"{num:,.2f}"
                    item = QTableWidgetItem(text)
                    item.setData(Qt.UserRole,num)
                except:
                    item = QTableWidgetItem(str(raw))
                self.table.setItem(r, c, item)

        self.success = True

@attach_logging_to_class
class tableWindow(QWidget):
    """
    A window that loads data from four database sources in the parent,
    merges and sorts it by dateTime, and displays it in a QTableWidget
    with a unified set of columns.
    """
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource = None, all_rows = [], table = "", headers = None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle(f"New data in {table}")
        self.resize(1000, 600)

        # Layout and table
        layout = QVBoxLayout(self)
        self.excelBtn = QPushButton("Export to Excel")
        self.excelBtn.clicked.connect(self.export)
        layout.addWidget(self.excelBtn)
        self.table = QTableWidget(self)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)

        self.rows = all_rows

        # 4) Collect the union of all column keys
        if headers is None:
            all_cols = set()
            for row in all_rows:
                all_cols.update(row.keys())
            all_cols = list(all_cols)
        else:
            all_cols = headers
        self.headers = all_cols
        # 5) Configure the table widget
        self.table.setRowCount(len(all_rows))
        self.table.setColumnCount(len(all_cols))
        self.table.setHorizontalHeaderLabels(all_cols)

        # 6) Populate each cell
        for r, row in enumerate(all_rows):
            for c, key in enumerate(all_cols):
                raw = row.get(key,"")
                try:
                    num = float(raw)
                    text = f"{num:,.2f}"
                    item = QTableWidgetItem(text)
                    item.setData(Qt.UserRole,num)
                except:
                    item = QTableWidgetItem(str(raw))
                self.table.setItem(r, c, item)
    def export(self,*_):
        exportTableToExcel(self,self.rows,self.headers)


class ClickableLineEdit(QLineEdit):
    clicked = pyqtSignal()
    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        self.clicked.emit()

class PopupPanel(QWidget):
    closed = pyqtSignal()
    def __init__(self, parent=None):
        super().__init__(parent, flags=Qt.Popup)
        self.parent = parent
        layout = QVBoxLayout(self)
        layout.setContentsMargins(4,4,4,4)
        # ClearFilter button
        self.clear_btn = QPushButton("Clear", self)
        layout.addWidget(self.clear_btn)
        self.all_btn = QPushButton("Select All", self)
        layout.addWidget(self.all_btn)
        self.searchBar = QLineEdit()
        self.searchBar.setPlaceholderText("Search")
        self.searchBar.textChanged.connect(self.parent.updateOptionVisibility)
        layout.addWidget(self.searchBar)
        # scrollable checkbox area
        self.scroll = QScrollArea(self)
        self.scroll.setWidgetResizable(True)
        container = QWidget()
        container.setObjectName("subPanel")
        self.box_layout = QVBoxLayout(container)
        self.box_layout.addStretch()
        container.setLayout(self.box_layout)
        self.scroll.setWidget(container)
        #self.scroll.setFixedHeight(150)
        layout.addWidget(self.scroll)

    def hideEvent(self, event):
        super().hideEvent(event)
        if self.parent.choiceChange: #only emit if there is changes
            self.closed.emit()


class MultiSelectBox(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        # ——— top line edit ———
        self.line_edit = ClickableLineEdit(self)
        self.line_edit.setReadOnly(True)
        self.line_edit.setPlaceholderText("Click to select…")
        self.line_edit.clicked.connect(self._togglePopup)
        self.hierarchy = False
        self.currentItems = []

        # ——— pop-up panel ———
        self.popup = PopupPanel(self)
        self.popup.setObjectName("mainPage")
        # wire up clear button
        self.popup.clear_btn.clicked.connect(self.clearSelection)
        self.popup.all_btn.clicked.connect(self.selectAll)
        # react when popup closes

        # keep track of checkboxes
        self._checkboxes = {}

        # overall layout
        main = QVBoxLayout(self)
        main.addWidget(self.line_edit)
        main.setContentsMargins(0,0,0,0)
        self.setLayout(main)
    def updateOptionVisibility(self):
        for cbKey in self._checkboxes.keys():
            self._checkboxes[cbKey].setVisible(self.popup.searchBar.text().lower() in self._checkboxes[cbKey].text().lower())
    def hierarchyMode(self):
        self.hierarchy = True
    def _togglePopup(self):
        if self.popup.isVisible():
            self.popup.hide()
        else:
            self.choiceChange = False
            self.popup.adjustSize()

            # 1. Get global position under the line edit
            pos = self.line_edit.mapToGlobal(QPoint(0, self.line_edit.height()))
            popup_size = self.popup.sizeHint()

            # 2. Get screen geometry
            screen_geom = QApplication.desktop().availableGeometry(self)

            # 3. Clamp the right edge
            if pos.x() + popup_size.width() > screen_geom.right():
                pos.setX(screen_geom.right() - popup_size.width())

            # 4. Clamp the bottom edge
            if pos.y() + popup_size.height() > screen_geom.bottom():
                pos.setY(screen_geom.bottom() - popup_size.height())

            # 5. Prevent left/top overflow
            pos.setX(max(screen_geom.left(), pos.x()))
            pos.setY(max(screen_geom.top(), pos.y()))

            # 6. Apply
            self.popup.move(pos)
            self.popup.show()
    def addItems(self,items):
        for item in items:
            self.addItem(item)
    def addItem(self, text):
        text = displayLinks.get(text,text) #put checkboxes to show the display version
        if text in self._checkboxes:
            return
        cb = QCheckBox(text, self.popup)
        cb.stateChanged.connect(self._updateLine)
        self.popup.box_layout.insertWidget(
            self.popup.box_layout.count() - 1, cb
        )
        self._checkboxes[text] = cb

    def clearItems(self):
        for cb in self._checkboxes.values():
            self.popup.box_layout.removeWidget(cb)
            cb.deleteLater()
        self._checkboxes.clear()
        self._updateLine()

    def setCheckedItems(self, items):
        for text, cb in self._checkboxes.items():
            text = displayLinks.get(text,text) # check as display version
            if text in items:
                cb.setChecked(True)
        self._updateLine()
    def setCheckedItem(self, item):
        for text, cb in self._checkboxes.items():
            text = displayLinks.get(text,text) # check as display version
            if text == item:
                cb.setChecked(True)
        self._updateLine()

    def checkedItems(self):
        if self.hierarchy:
            items = []
            for item in self.currentItems:
                items.append(displayLinks.get(item,item)) #revert to normal version for output
            return items
        else:
            return [displayLinks.get(t,t) for t, cb in self._checkboxes.items() if cb.isChecked()]

    def clearSelection(self):
        for cb in self._checkboxes.values():
            cb.setChecked(False)
        self.popup.searchBar.setText("")
        self._updateLine()
    def selectAll(self):
        for cb in self._checkboxes.values():
            cb.setChecked(True)
        self._updateLine()

    def _updateLine(self):
        self.choiceChange = True
        temp = self.hierarchy
        self.hierarchy = False
        sel = self.checkedItems()
        sel = [displayLinks.get(item,item) for item in sel.copy()]
        self.hierarchy = temp
        if self.hierarchy:
            for idx, item in enumerate(self.currentItems): #check if item is removed.
                if item not in sel:
                    self.currentItems.pop(idx)
                    break
            for idx, item in enumerate(sel): #check if item is added
                if item not in self.currentItems:
                    self.currentItems.append(item)
                    break
            lines = [f"({i+1}) '{text}'" for i, text in enumerate(self.currentItems)]

            display = "\n".join(lines)
        else:
            # the old single-line, comma-separated format
            display = ", ".join(sel)
        self.line_edit.setText(display)

class SortPopup(QDialog):
    popup_closed = pyqtSignal()

    def __init__(self, items=None, checked_set=None, parent=None):
        super().__init__(parent, Qt.Popup)
        self.setWindowTitle("Sort Items")
        self.setMinimumSize(200, 300)

        self.list_widget = QListWidget(self)
        self.list_widget.setDragDropMode(QListWidget.InternalMove)
        self.list_widget.setDefaultDropAction(Qt.MoveAction)

        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("Select and Sort Headers:"))
        layout.addWidget(self.list_widget)

        self.set_items(items or [], checked_set or set())

        self.list_widget.itemChanged.connect(self.on_item_toggled)

    def set_items(self, items, checked_set):
        self.list_widget.blockSignals(True)
        self.list_widget.clear()
        for item in items:
            list_item = QListWidgetItem(item)
            list_item.setFlags(list_item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsDragEnabled)
            list_item.setCheckState(Qt.Checked if item in checked_set else Qt.Unchecked)
            self.list_widget.addItem(list_item)
        self.list_widget.blockSignals(False)

    def on_item_toggled(self, item: QListWidgetItem):
        # Remove the item from current position
        row = self.list_widget.row(item)
        self.list_widget.takeItem(row)

        if item.checkState() == Qt.Unchecked:
            # Add to end of list
            self.list_widget.addItem(item)
        else:
            # Insert before first unchecked item
            insert_index = self.list_widget.count()
            for i in range(self.list_widget.count()):
                if self.list_widget.item(i).checkState() == Qt.Unchecked:
                    insert_index = i
                    break
            self.list_widget.insertItem(insert_index, item)

        # Reselect the item for visual consistency (optional)
        self.list_widget.setCurrentItem(item)

    def get_checked_sorted_items(self):
        return [
            self.list_widget.item(i).text()
            for i in range(self.list_widget.count())
            if self.list_widget.item(i).checkState() == Qt.Checked
        ]

    def get_all_items(self):
        return [self.list_widget.item(i).text() for i in range(self.list_widget.count())]

    def get_checked_items_set(self):
        return {
            self.list_widget.item(i).text()
            for i in range(self.list_widget.count())
            if self.list_widget.item(i).checkState() == Qt.Checked
        }

    def closeEvent(self, event):
        self.popup_closed.emit()
        super().closeEvent(event)


class SortButtonWidget(QWidget):
    popup_closed = pyqtSignal(list)  # emits checked, sorted items

    def __init__(self, parent=None):
        super().__init__(parent)
        self.items = []
        self.checked_items = set()
        self.active = False

        self.button = QPushButton("Header Options", self)
        self.button.clicked.connect(self.show_popup)

        layout = QVBoxLayout(self)
        layout.addWidget(self.button)
        layout.setContentsMargins(0, 0, 0, 0)

        self.popup = SortPopup(self.items, self.checked_items, self)
        self.popup.popup_closed.connect(self.on_popup_closed)

    def add_item(self, item, checked=True):
        self.items.append(item)
        if checked:
            self.checked_items.add(item)
        self.popup.set_items(self.items,self.checked_items)
        self.active = True

    def set_items(self, items, checked_items=None):
        self.items = list(items)
        self.checked_items = set(checked_items or items)
        self.popup.set_items(self.items,self.checked_items)
        self.active = True

    def show_popup(self):
        if self.popup.isVisible():
            self.popup.hide()
        else:
            self.popup.adjustSize()

            # 1. Get global position under the line edit
            pos = self.button.mapToGlobal(QPoint(0, 0))
            popup_size = self.popup.sizeHint()

            # 2. Get screen geometry
            screen_geom = QApplication.desktop().availableGeometry(self)

            # 3. Clamp the right edge
            if pos.x() + popup_size.width() > screen_geom.right():
                pos.setX(screen_geom.right() - popup_size.width())

            # 4. Clamp the bottom edge
            if pos.y() + popup_size.height() > screen_geom.bottom():
                pos.setY(screen_geom.bottom() - popup_size.height())

            # 5. Prevent left/top overflow
            pos.setX(max(screen_geom.left(), pos.x()))
            pos.setY(max(screen_geom.top(), pos.y()))

            # 6. Apply
            self.popup.move(pos)
            self.popup.show()

    def on_popup_closed(self):
        if self.popup:
            self.items = self.popup.get_all_items()
            self.checked_items = self.popup.get_checked_items_set()
            #self.popup_closed.emit(self.popup.get_checked_sorted_items())
            self.popup.hide()

class SmartStretchTable(QTableWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        # Use interactive resizing by default
        self.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

        # Watch for resize events to re-evaluate
        self.viewport().installEventFilter(self)

        # Optional: Smooth scrolling
        self.setHorizontalScrollMode(self.ScrollPerPixel)

    def eventFilter(self, obj, event):
        if obj == self.viewport():
            QTimer.singleShot(0, self._maybeStretchColumns)
        return super().eventFilter(obj, event)

    def _maybeStretchColumns(self):
        col_count = self.columnCount()
        if col_count == 0:
            return

        total_width = sum(self.columnWidth(i) for i in range(col_count))
        available = self.viewport().width()

        if total_width < available:
            stretch_width = available // col_count
            for i in range(col_count):
                self.setColumnWidth(i, stretch_width)

    def updateData(self, data):
        # Example dynamic population method
        row_count = len(data)
        col_count = len(data[0]) if row_count else 0

        self.setRowCount(row_count)
        self.setColumnCount(col_count)

        for r in range(row_count):
            for c in range(col_count):
                self.setItem(r, c, QTableWidgetItem(str(data[r][c])))

        QTimer.singleShot(0, self._maybeStretchColumns)

if __name__ == '__main__':
    freeze_support()
    key = os.environ.get(dynamoAPIenvName)
    ok = key
    app = QApplication(sys.argv)
    queueTimer = QTimer()
    queueTimer.timeout.connect(poll_queue)
    queueTimer.start(500)

    w = returnsApp(start_index=0 if not ok else 1)
    
    if ok: w.api_key = key
    w.show()
    if ok:
        w.init_data_processing()
    else:
        w.stack.setCurrentIndex(0)
    bgWatch = QTimer()
    bgWatch.timeout.connect(w.watchForUpdateTime)
    hours = 0.8
    bgWatch.start(int(hours * 60 * 60 * 1000))
    sys.exit(app.exec_())
