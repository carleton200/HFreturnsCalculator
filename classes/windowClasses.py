from classes import DatabaseManager
from classes.DatabaseManager import load_from_db
from scripts.basicFunctions import separateRowCode, findSourceName
from scripts.loggingFuncs import attach_logging_to_class
from classes.widgetClasses import SortButtonWidget, MultiSelectBox, simpleMonthSelector
from scripts.exportTableToExcel import exportTableToExcel
from scripts.pyqtFunctions import basicHoldingsReportExport, comboInvestorOpts, controlTable, filt2Query
from scripts.render_report import render_report
from scripts.reportWorkbooks import portfolioSnapshot
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from scripts.instantiate_basics import gui_queue, executor
from scripts.commonValues import sqlPlaceholder, timeOptions, percent_headers, demoMode, nonFundCols
import sqlite3
import logging
import traceback
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout,
    QLabel,  QPushButton, QFormLayout,
    QRadioButton, QButtonGroup, QComboBox, QHBoxLayout,
    QTableWidget, QTableWidgetItem,  QMessageBox,
    QScrollArea, QFileDialog, 
     QHeaderView, QDateEdit, QSplitter
)
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import Qt,  QUrl, QDate
from scripts.instantiate_basics import DATABASE_PATH

class linkBenchmarksWindow(QWidget):
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource=None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle("Link Benchmarks")
        self.resize(800, 500)
        self.setStyleSheet(self.parent.appStyle)
        self.setObjectName("mainPage")
        self._benchmarks = [{},]
        self._links = []
        self.asset_levels = [("Level 1", 1), ("Level 2", 2), ("Level 3", 3)]
        self.selected_asset_level = None
        self.selected_asset = None
        self.selected_benchmark = None
        self.init_ui()

    def init_ui(self):
        try:
            mainLayout = QVBoxLayout(self)
            splitter = QSplitter(Qt.Horizontal)

            # --- Left: Benchmark Links Table ---
            leftWidget = QWidget()
            leftLayout = QVBoxLayout(leftWidget)
            self.linksTable = QTableWidget()
            self.linksTable.setColumnCount(4)
            self.linksTable.setHorizontalHeaderLabels(['Benchmark', 'Asset', 'Level', 'Delete'])
            self.linksTable.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            leftLayout.addWidget(QLabel("Current Benchmark Links:"))
            leftLayout.addWidget(self.linksTable)
            splitter.addWidget(leftWidget)

            # --- Right Side: Vertically Split ---
            rightWidget = QWidget()
            rightVSplitter = QSplitter(Qt.Vertical)

            # --- Top half: Add Benchmark Link ---
            topWidget = QWidget()
            topLayout = QVBoxLayout(topWidget)
            topLayout.addWidget(QLabel("Add Benchmark Link:"))

            # Asset Class Level ComboBox
            self.assetLevelCombo = QComboBox()
            self.assetLevelCombo.addItem("")  # blank/default
            for label, num in self.asset_levels:
                self.assetLevelCombo.addItem(label, num)
            self.assetLevelCombo.currentIndexChanged.connect(self.updateAssetCombo)
            topLayout.addWidget(QLabel("Asset Class Level:"))
            topLayout.addWidget(self.assetLevelCombo)

            # Asset ComboBox - will be populated depending on above
            self.assetCombo = QComboBox()
            self.assetCombo.addItem("")  # blank
            topLayout.addWidget(QLabel("Asset:"))
            topLayout.addWidget(self.assetCombo)

            # Benchmark ComboBox - will be populated
            self.benchmarkCombo = QComboBox()
            self.benchmarkCombo.addItem("")  # blank
            topLayout.addWidget(QLabel("Benchmark:"))
            topLayout.addWidget(self.benchmarkCombo)

            # Confirm Button
            self.confirmBtn = QPushButton("Confirm Link")
            self.confirmBtn.clicked.connect(self.addBenchmarkLink)
            topLayout.addWidget(self.confirmBtn)
            topLayout.addStretch()

            # --- Bottom half: Table Element Benchmark Link ---
            bottomWidget = QWidget()
            bottomLayout = QVBoxLayout(bottomWidget)
            bottomLayout.addWidget(QLabel("Add Table Element Benchmark Link:"))

            # Table Element ComboBox
            self.tableElementCombo = QComboBox()
            self.tableElementCombo.addItem("")  # blank (populate later)
            bottomLayout.addWidget(QLabel("Total Portfolio or Family Branch:"))
            bottomLayout.addWidget(self.tableElementCombo)

            # Table Benchmark ComboBox
            self.tableBenchmarkCombo = QComboBox()
            self.tableBenchmarkCombo.addItem("")  # blank (populate later)
            bottomLayout.addWidget(QLabel("Benchmark:"))
            bottomLayout.addWidget(self.tableBenchmarkCombo)

            # Confirm Button
            self.tableConfirmBtn = QPushButton("Confirm Table Element Link")
            self.tableConfirmBtn.clicked.connect(self.addTableElementBenchmarkLink) # Connect to your logic
            bottomLayout.addWidget(self.tableConfirmBtn)
            bottomLayout.addStretch()
            self.autoAddFamBtn = QPushButton('Auto Connect Family Branches')
            bottomLayout.addWidget(self.autoAddFamBtn)
            self.autoAddFamBtn.clicked.connect(self.autoConnectFams)

            rightVSplitter.addWidget(topWidget)
            rightVSplitter.addWidget(bottomWidget)
            splitter.addWidget(rightVSplitter)

            mainLayout.addWidget(splitter)
            self.setLayout(mainLayout)

            self.refreshBenchmarks()
            self.refreshLinks()
            self.updateTableElementCombo()
        except Exception as e:
            print(f"Error initializing benchmark window: {e} {e.args}")
            logging.error(f"Error initializing benchmark window: {e} {e.args}")
            QMessageBox.critical(self.parent, "Error initializing benchmark window", f"Error initializing benchmark window: {e} {e.args}")

    def refreshLinks(self):
        links = self.parent.db.fetchBenchmarkLinks()
        self._links = links
        self.linksTable.setRowCount(len(links))
        for row, link in enumerate(links):
            self.linksTable.setItem(row, 0, QTableWidgetItem(str(link.get("benchmark", ""))))
            self.linksTable.setItem(row, 1, QTableWidgetItem(str(link.get("asset", ""))))
            self.linksTable.setItem(row, 2, QTableWidgetItem(str(link.get("assetLevel", ""))))
            # Delete button
            btn = QPushButton("Delete")
            btn.clicked.connect(lambda _, r=row: self.deleteLink(r))
            self.linksTable.setCellWidget(row, 3, btn)

    def refreshBenchmarks(self):
        # try to fetch all benchmarks for use in the Benchmark combobox
        try:
            if hasattr(self.parent.db, "fetchBenchmarks"):
                benchmarks = self.parent.db.fetchBenchmarks()
            else:
                benchmarks = {}  # fallback - maybe empty or stub
        except Exception as e:
            print(f"Failed to fetch benchmarks: {e}")
            benchmarks = []
        self._benchmarks = benchmarks
        self.updateBenchmarkCombo()

    def updateBenchmarkCombo(self):
        for combo in (self.benchmarkCombo, self.tableBenchmarkCombo):
            combo.clear()
            combo.addItem("")
            combo.addItems(sorted([b["benchmark"] if isinstance(b,dict) else str(b) for b in self._benchmarks]))
    def updateTableElementCombo(self):
        self.tableElementCombo.clear()
        opts = ["","Total"]
        opts.extend(set(inv['Parentinvestor'] for inv in self.parent.db.fetchInvestors()))
        self.tableElementCombo.addItems(opts)
    def autoConnectFams(self):
        reply = QMessageBox.question(
            self,
            "Auto-connect Family Branch Benchmarks",
            "Would you like to automatically connect all family branches to their related policy and implementation benchmarks? (Only works if the names are formatted exactly the same)",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return

        db = self.parent.db
        family_branches = set(inv['Parentinvestor'] for inv in db.fetchInvestors())
        benchmarks = self._benchmarks

        pending_updates = []  # Collect (famBranch, benchmark) pairs to add
        for famBranch in family_branches:
            famBranchBase = famBranch.split(' Family Branch')[0]
            for suffix in ("Implementation Benchmark", "Policy Benchmark"):
                expected_bench = f"{famBranchBase} {suffix}"
                # Find if this benchmark exists (string match; b can be dict or string)
                match = None
                for b in benchmarks:
                    b_name = b.get("benchmark") if isinstance(b, dict) else str(b)
                    if b_name == expected_bench:
                        match = b_name
                        break
                if match:
                    pending_updates.append((famBranch, match))

        if not pending_updates:
            QMessageBox.information(self, "No matches", "No matching benchmarks were found for any family branch.")
            return

        # Create the update list display for user's confirmation
        updates_text = "\n".join(f"{asset}  →  {benchmark}" for asset, benchmark in pending_updates)
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QScrollArea, QPushButton, QHBoxLayout, QWidget

        # Prepare a scrollable dialog to show all updates
        dialog = QDialog(self)
        dialog.setWindowTitle("Confirm Benchmark Connections")
        dialog.setStyleSheet(self.parent.appStyle)
        vbox = QVBoxLayout(dialog)

        msg_label = QLabel("The following connections will be made:\n")
        vbox.addWidget(msg_label)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        content = QWidget()
        content_layout = QVBoxLayout(content)
        # Set the background of content (the QLabel background) to black
        content.setStyleSheet("background-color: black;")
        label = QLabel(updates_text)
        label.setTextInteractionFlags(Qt.TextSelectableByMouse)
        content_layout.addWidget(label)
        scroll.setWidget(content)
        vbox.addWidget(scroll, stretch=1)

        # OK/Cancel buttons
        btn_hbox = QHBoxLayout()
        ok_btn = QPushButton("Proceed")
        cancel_btn = QPushButton("Cancel")
        btn_hbox.addWidget(ok_btn)
        btn_hbox.addWidget(cancel_btn)
        vbox.addLayout(btn_hbox)

        confirmed = {"val": False}

        def accept():
            confirmed["val"] = True
            dialog.accept()
        def reject():
            dialog.reject()

        ok_btn.clicked.connect(accept)
        cancel_btn.clicked.connect(reject)

        dialog.resize(500, 400)
        dialog.exec_()

        if not confirmed["val"]:
            return

        # Add all benchmark links
        for asset, benchmark in pending_updates:
            self.addTableElementBenchmarkLink(asset=asset, benchmark=benchmark)

        # Mimic the reload method of updateLink
        self.parent.db.fetchBenchmarkLinks(update=True)
        self.refreshLinks()
        self.parent.buildReturnTable()
    def updateAssetCombo(self):
        self.assetCombo.clear()
        self.assetCombo.addItem("")
        asset_level_num = self.assetLevelCombo.currentData()
        asset_key = None
        if asset_level_num == 1:
            asset_key = "assetClass"
        elif asset_level_num == 2:
            asset_key = "subAssetClass"
        elif asset_level_num == 3:
            asset_key = "sleeve"
        options = self.parent.db.fetchFundOptions(asset_key)
        # Remove duplicates and blank/None values
        assets = sorted({opt for opt in options if opt not in (None,"", "None")})
        self.assetCombo.addItems(assets)

    def deleteLink(self, row):
        if row < 0 or row >= len(self._links):
            return
        link = self._links[row]
        try:
            with self.parent.lock:
                cursor = self.parent.db._conn.cursor()
                cursor.execute(
                    "DELETE FROM benchmarkLinks WHERE benchmark = ? AND asset = ? AND assetLevel = ?",
                    (link["benchmark"], link["asset"], link["assetLevel"])
                )
                self.parent.db._conn.commit()
            self.parent.db.fetchBenchmarkLinks(update=True)
            QMessageBox.information(self, "Success", f"Deleted link: {link['benchmark']} to {link['asset']} at level {link['assetLevel']}.")
        except Exception as e:
            QMessageBox.warning(self, "Delete Error", f"Error deleting link: {e}")
        self.parent.db.fetchBenchmarkLinks(update=True)
        self.refreshLinks()
        self.parent.buildReturnTable()
    def addTableElementBenchmarkLink(self, asset = None, benchmark = None):
        if not asset and not benchmark:
            asset = self.tableElementCombo.currentText()
            levelIdx = 0 if asset == "Total" else -1
            benchmark = self.tableBenchmarkCombo.currentText().strip()
            reload = True
        else:
            levelIdx = -1
            reload = False
        if asset == "" or benchmark == "":
            QMessageBox.warning(self, "Incomplete", "Please select asset and benchmark.")
            return
        self.updateLink(levelIdx,asset,benchmark, reload = reload)
    def addBenchmarkLink(self):
        # Get selections
        level_idx = self.assetLevelCombo.currentIndex()
        asset = self.assetCombo.currentText().strip()
        benchmark = self.benchmarkCombo.currentText().strip()
        if level_idx <= 0 or not asset or not benchmark:
            QMessageBox.warning(self, "Incomplete", "Please select asset class level, asset, and benchmark.")
            return
        asset_level = self.assetLevelCombo.currentData()
        self.updateLink(asset_level,asset,benchmark)
        
    def updateLink(self, level,asset,benchmark, reload = True):
        try:
            with self.parent.lock:
                cursor = self.parent.db._conn.cursor()
                cursor.execute(
                    "INSERT OR REPLACE INTO benchmarkLinks (benchmark, asset, assetLevel) VALUES (?, ?, ?)",
                    (benchmark, asset, level)
                )
                self.parent.db._conn.commit()
            if reload:
                QMessageBox.information(self, "Success", f"Linked {benchmark} to {asset} at level {level}.")
        except Exception as e:
            if reload:
                QMessageBox.warning(self, "Error", f"Failed to save link: {e}")
        if reload:
            self.parent.db.fetchBenchmarkLinks(update=True)
            self.refreshLinks()
            self.parent.buildReturnTable()


class exportWindow(QWidget):
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource=None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle("Export Data")
        self.resize(600, 400)
        self.setStyleSheet(self.parent.appStyle)
        self.setObjectName("mainPage")

        # Layouts
        main_layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        self.filter_boxes = {}
        self.filter_labels = {}

        # --- Filter ComboBoxes for each filterOption ---
        self.filterOptions = getattr(self.parent, "filterOptions", [
            {"key": "Source name", "name": "Source name"},
            {"key": "nodePath", "name": "nodePath"},
            {"key": "Target name", "name": "Target name"},
        ])
        # Use parent's lock if available
        self.lock = getattr(self.parent, "lock", None)
        # Query unique options for each filter from the calculations table


        calcKeys = ['Source name','Target name']
        investors = self.parent.db.fetchInvestors()
        famBranches = set(inv['Parentinvestor'] for inv in investors)
        filOptDict = {filKey : set() for filKey in (fil['key'] for fil in self.filterOptions if fil['key'] not in nonFundCols)}
        funds = self.parent.db.fetchFunds()
        dyn2key = self.parent.db.fetchDyn2Key()
        for fund in funds:
            for field, val in ([f,v] for f,v in fund.items() if dyn2key.get(f,"") in filOptDict):
                filOptDict[dyn2key[field]].add(val)
        nodes = [str(n) for n in list(self.parent.db.pullId2Node().keys())]
        with self.parent.lock:
            dbPath = getattr(self.parent, "dbPath", DATABASE_PATH)
            conn = sqlite3.connect(dbPath)
            cur = conn.cursor()
            for f in self.filterOptions:
                key = f["key"]
                name = f["name"]
                combo = MultiSelectBox(dispLib=self.parent.db.userDisplayLib())
                combo.addItem("")  # blank for optional
                try:
                    if key in calcKeys:
                        cur.execute(f"SELECT DISTINCT [{key}] FROM calculations")
                        options = [row[0] for row in cur.fetchall() if row[0] not in (None, "", "None")]
                        options = sorted(set(options))
                        combo.addItems(options)
                    elif key not in nonFundCols:
                        opts = filOptDict[key]
                        combo.addItems(opts)
                    elif key == 'Node':
                        combo.addItems(nodes)
                    elif key == 'Family Branch':
                        combo.addItems(famBranches)
                except Exception as e:
                    print(f"Error loading filter options for {key}: {e}")
                self.filter_boxes[key] = combo
                self.filter_labels[key] = name
                form_layout.addRow(name + ":", combo)
            conn.close()

        # --- Date selectors ---
        date_layout = QHBoxLayout()
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit.setDate(QDate.currentDate())
        self.start_date_edit.setSpecialValueText("")  # blank for optional
        self.start_date_edit.setDateRange(QDate(1990, 1, 1), QDate(2100, 12, 31))
        self.start_date_edit.setDate(QDate(2000, 1, 1))
        self.start_date_edit.setMinimumWidth(120)
        self.start_date_edit.setDate(QDate())  # blank

        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setSpecialValueText("")  # blank for optional
        self.end_date_edit.setDateRange(QDate(1900, 1, 1), QDate(2100, 12, 31))
        self.end_date_edit.setDate(QDate())  # blank
        self.end_date_edit.setMinimumWidth(120)

        date_layout.addWidget(QLabel("Start Date:"))
        date_layout.addWidget(self.start_date_edit)
        date_layout.addSpacing(20)
        date_layout.addWidget(QLabel("End Date:"))
        date_layout.addWidget(self.end_date_edit)
        form_layout.addRow(date_layout)

        # --- Confirm Button ---
        self.confirm_btn = QPushButton("Export to Excel")
        self.confirm_btn.clicked.connect(self.export_to_excel)
        main_layout.addLayout(form_layout)
        main_layout.addWidget(self.confirm_btn, alignment=Qt.AlignRight)

    def export_to_excel(self):
        #Export to excel function for the calculations view page
        # Build WHERE clause from filters
        start_date = self.start_date_edit.date()
        end_date = self.end_date_edit.date()
        where_clause, values = filt2Query(self.parent.db, self.filter_boxes, start_date.toPyDate(), end_date.toPyDate())
        # Query the database
        with self.parent.lock:
            dbPath = getattr(self.parent, "dbPath", DATABASE_PATH)
            conn = sqlite3.connect(dbPath)
            cur = conn.cursor()
            try:
                cur.execute("PRAGMA table_info(calculations)")
                columns = [row[1] for row in cur.fetchall()]
                sql = f"SELECT * FROM calculations {where_clause}"
                print(sql)
                print(tuple(values))
                cur.execute(sql, tuple(values))
                rows = cur.fetchall()
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to query database: {e}")
                conn.close()
                return

        if not rows:
            QMessageBox.information(self, "No Data", "No data found for the selected filters.")
            conn.close()
            return

        # Prompt user for file path
        path, _ = QFileDialog.getSaveFileName(self, "Save as…", "", "Excel Files (*.xlsx)")
        if not path:
            conn.close()
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # Write to Excel
        try:
            df = pd.DataFrame(rows, columns=columns)

            if os.path.exists(path):
                wb = load_workbook(path)
                # Generate a unique sheet name
                base_name = "Export"
                i = 1
                while True:
                    sheet_name = f"{base_name}{i}"
                    if sheet_name not in wb.sheetnames:
                        break
                    i += 1
                # Write DataFrame to the new sheet
                ws = wb.create_sheet(sheet_name)
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                for row_idx, row in enumerate(df.itertuples(index=False), 2):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                wb.save(path)
            else:
                df.to_excel(path, index=False)

            QMessageBox.information(self, "Success", f"Data exported to {path}")
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.abspath(path)))
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export to Excel: {e}")
        finally:
            conn.close()

class displayWindow(QWidget):
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource = None, text = "", title=""):
        super().__init__(parent, flags)
        self.setWindowTitle(title)
        self.parent = parentSource
        if self.parent:
            self.setObjectName("mainPage")
            self.setStyleSheet(self.parent.appStyle)
        #self.resize(1000, 600)

        # Layout and table
        layout = QVBoxLayout(self)
        
        scrollArea = QScrollArea(self) #make a scroll area with a layout inside containing a widget of the text that is styled to the color theme
        textContainer = QWidget()
        textContainer.setObjectName("subPanel")
        textLayout = QVBoxLayout()
        textLayout.addWidget(QLabel(text))
        textContainer.setLayout(textLayout)
        scrollArea.setWidget(textContainer)
        layout.addWidget(scrollArea)

@attach_logging_to_class
class underlyingDataWindow(QWidget):
    """
    A window that loads data from four database sources in the parent,
    merges and sorts it by dateTime, and displays it in a QTableWidget
    with a unified set of columns.
    """
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource = None, db = None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle("Underlying Data Viewer")
        self.resize(1000, 600)
        self.db = db #only input from transactionApp
        # Layout and table
        layout = QVBoxLayout(self)
        buttonBox = QWidget()
        buttonLayout = QHBoxLayout()
        exportBtn = QPushButton("Export Data to Excel")
        exportBtn.clicked.connect(self.exportTable)
        buttonLayout.addWidget(exportBtn)
        self.headerOptions = SortButtonWidget()
        self.headerOptions.popup.popup_closed.connect(self.buildTable)
        buttonLayout.addWidget(self.headerOptions)
        buttonBox.setLayout(buttonLayout)
        layout.addWidget(buttonBox)
        self.table = QTableWidget(self)
        layout.addWidget(self.table)
        self.success = False
        self.headerOrder = ["Date", "TradeDate", "_source","Source name","Target name", "CashFlowSys", "ValueInSystemCurrency","Balancetype","TransactionType"]
        self.buildTable()

    def exportTable(self, *_):
        # 1) prompt user
        path, _ = QFileDialog.getSaveFileName(
            self, "Save as…", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        def processExport():
            try:
                data = self.allData  # list of dicts

                all_cols = self.allCols


                if os.path.exists(path):
                    wb = load_workbook(path)
                    # Create a unique sheet name for export
                    base_name = "Export"
                    i = 1
                    while True:
                        sheet_name = f"{base_name}{i}"
                        if sheet_name not in wb.sheetnames:
                            break
                        i += 1
                    ws = wb.create_sheet(sheet_name)
                else:
                    wb = Workbook()
                    ws = wb.active
                rowStart = 1
                for idx, colname in enumerate(all_cols, start=1):
                    ws.cell(row=rowStart, column=idx, value=colname)
                # 7) populate rows
                for r, row_dict in enumerate(data, start=rowStart + 1):        
                    # data cells with proper formatting
                    for c, colname in enumerate(all_cols, start=1):
                        val = row_dict.get(colname, None)
                        try: #make numerical format if possible
                            val = float(val)
                        except:
                            pass
                        cell = ws.cell(row=r, column=c, value=val)
                        if isinstance(val, (int, float)):
                            if colname not in percent_headers:
                                # show with commas, two decimals
                                cell.number_format = "#,##0.00"
                            else:
                                # interpret val as percentage (e.g. 10.5 → 10.5%)
                                cell.value = val / 100.0
                                cell.number_format = "0.00%"
                # 8) autofit column widths
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is not None:
                            text = str(cell.value)
                            max_len = max(max_len, len(text))
                    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
            except Exception as e:
                gui_queue.put(lambda error = e, trace = traceback.format_exc(): QMessageBox.critical(self, "Processing error", f"{error} \n {trace}"))
            try:
                wb.save(path)
            except Exception as e:
                gui_queue.put(lambda error = e: QMessageBox.critical(self, "Save error", str(error)))
            else:
                gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
                gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
        executor.submit(processExport)    
    def buildTable(self):
        monthTableBool = self.parent.tableBtnGroup.checkedButton().text()  == "Monthly Table"
        if monthTableBool:
            selectedMonth = datetime.strptime(self.parent.selectedCell["month"], "%B %Y")
            tranStart = selectedMonth.replace(day = 1)
            accountStart = tranStart - relativedelta(days= 1)
            allEnd = (tranStart + relativedelta(months=1)) - relativedelta(days=1)
        else:
            endTime = datetime.strptime(self.parent.dataEndSelect.currentText(),"%B %Y")
            allEnd = (endTime.replace(day = 1) + relativedelta(months=1)) - relativedelta(days=1)
            selection = self.parent.selectedCell["month"]
            if selection not in (*timeOptions, 'Distributions TD', 'TVPI','DPI') or selection == "MTD": #MTD timeframe
                tranStart = endTime.replace(day = 1)
                accountStart = tranStart - relativedelta(days= 1)
            elif selection in ("ITD","IRR ITD",'Distributions TD', 'TVPI','DPI'):
                tranStart = self.parent.dataTimeStart
                accountStart = self.parent.dataTimeStart
            else:
                if selection == "QTD":
                    subtract = (int((endTime.month)) % 3 if (int(endTime.month)) % 3 != 0 else 3) - 1
                elif selection == "YTD":
                    subtract = (int((endTime.month)) % 12 if (int(endTime.month)) % 12 != 0 else 12) - 1
                else:
                    subtract = int(selection.removesuffix("YR")) * 12 - 1
                tranStart = (endTime - relativedelta(months=subtract)).replace(day=1)
                accountStart = tranStart - relativedelta(days= 1)
        tranStart = datetime.strftime(tranStart,"%Y-%m-%dT00:00:00")
        accountStart = datetime.strftime(accountStart,"%Y-%m-%dT00:00:00")
        allEnd = datetime.strftime(allEnd,"%Y-%m-%dT00:00:00")
        dataType = self.parent.selectedCell["dataType"]
        if dataType == "Total":
            return
        dataType = dataType.removeprefix("Total ")
        code = self.parent.selectedCell["rowKey"]
        header, code= separateRowCode(code)
        hier = code.removeprefix("##(").removesuffix(")##").split("::")
        hierSelections = self.parent.sortHierarchy.checkedItems()
        if dataType == "Target name":
            hier[-1] = header #sets the final hier (Target) to the actual target name
            hierSelections.append(dataType)
        self.parent.db.buildNodeLib()
        nodeLib = self.parent.db.nodeLib
        nodes = nodeLib.nodes
        if 'Node' in hierSelections: 
            #Check if needs extra node places. occurs from recursive node hierarchy
            nodeIdx = hierSelections.index('Node')
            nodeName = hier[nodeIdx]
            belowNodes = nodeLib.nodePaths.get(nodeName,{}).get('below',set())
            if belowNodes:
                belowNodes = [nodeLib.id2node[nId] for nId in belowNodes]
                # We want to ensure hierSelections matches up with hier; for every extra node at the appropriate positions, insert "Node"
                # We iterate from nodeIdx forwards over hier, checking which items represent nodes that should be declared as such in hierSelections
                i = nodeIdx
                delta = 0  # How many have been inserted (shifts index in hierSelections)
                while i + delta < len(hier):
                    val = hier[i + delta]
                    if val in belowNodes:
                        # Only insert if this isn't already 'Node' at this position (in case already inserted)
                        if i + delta >= len(hierSelections) or hierSelections[i + delta] != 'Node':
                            hierSelections.insert(i + delta, 'Node')
                            delta += 1
                        else:
                            # Already correct, move on
                            pass
                    i += 1
                nodeIdx = hierSelections.index('Node')
                hierSelections = [*hierSelections[:nodeIdx], *(['Node'] * (len(hier) - len(hierSelections))), *hierSelections[nodeIdx:]]
        tables = {"positions": accountStart,"transactions": tranStart} if not self.db else {"transactions": tranStart}
        all_rows = []
        if not self.db: #regular, non transaction mode TODO: make a better boolean here
            fundOptionSets = []
            nodesInPath = []
            filterDict = {}
            inputs = []
            sourceOpts = set()
            invSelections = self.parent.filterDict["Source name"].checkedItems()
            famSelections = self.parent.filterDict["Family Branch"].checkedItems()
            for hierIdx, tier in enumerate(hier): #iterate through each tier down to selection
                tierType = hierSelections[hierIdx]
                if tier == "hiddenLayer":
                    continue #hidden layer should not affect the query
                elif tierType == 'Node':
                    #find all funds under node
                    nodeName = tier
                    if nodeName == '': #occurs from base nodes
                        nodeName = header
                    nodesInPath.append(nodeName)
                    allTargs = set(nodeLib.node2AllTargets[nodeName])
                    fundOptionSets.append(allTargs)
                    nodesInPath.extend([t for t in allTargs if t in nodes]) #adds nodes that are targets of the selected node
                elif tierType == 'Target name':
                    fundOptionSets.append(set(self.parent.cFundToFundLinks.get(tier,[tier,])))
                elif tierType == 'Source name':
                    sourceOpts.add(tier)
                elif tierType == 'Family Branch':
                    sourceOpts.update(self.parent.db.pullInvestorsFromFamilies([tier,]))
                elif tierType not in nonFundCols:
                    filterDict[tierType] = [tier,]
            for key, cb in self.parent.filterDict.items():
                opts = cb.checkedItems()
                if key not in nonFundCols and opts and key not in filterDict.keys():
                    filterDict[key] = opts
            if filterDict:
                fundOptionSets.append(set(self.parent.db.pullFundsFromFilters(filterDict)))
            if fundOptionSets:
                fundOpts = set.intersection(*(set(funds) for funds in fundOptionSets))
                if invSelections != [] or famSelections != []:
                    invs = comboInvestorOpts(self.parent.db,invSelections,famSelections)
                    sourceOpts.update(invs)
                sourceOpts.update(nodesInPath)
                targetOpts = fundOpts.union(set(nodesInPath)) #add nodes so nodes as targets can be selected
                targetPholders = ','.join(sqlPlaceholder for _ in targetOpts)
                condStatement = f"WHERE [Target name] in ({targetPholders})"
                inputs.extend(list(targetOpts))
                if sourceOpts:
                    sourcePholders = ','.join(sqlPlaceholder for _ in sourceOpts)
                    condStatement += f' AND [Source name] in ({sourcePholders})'
                    inputs.extend(list(sourceOpts))
                condStatement += ' AND [Date] BETWEEN ? AND ?'
            else:
                print("WARNING: No relevant funds found for the cell selected.")
                return
            for table, start in tables.items():
                try:
                    if not demoMode:
                        print(f"    Underlying data conditional (1): {condStatement} for: {(*inputs,start,allEnd)}")
                    rows = load_from_db(self.parent.db,table,condStatement, tuple((*inputs,start,allEnd)))
                except Exception as e:
                    print(f"Error in call : {e} ; {e.args}")
                    rows = []
                all_rows.extend(rows)
        elif self.db:
            try:
                condStatement = 'WHERE ([Source name] = ? OR [Target name] = ?)'
                inputs = (header,header, tranStart, allEnd)
                all_rows = load_from_db(self.parent.db,'transactions',condStatement.removesuffix("AND") + " AND [Date] BETWEEN ? AND ?", tuple(inputs))
            except Exception as e:
                print(f"Error in call : {e} ; {e.args}")
                all_rows = []
        if not monthTableBool and selection == 'Distributions TD':
            all_rows = [r for r in all_rows if 'Distribution' in r.get('TransactionType','')]
        self.allData = all_rows

        if self.db:
            #build organized differences by pool/investor versus transaction type
            diffTableDict = { "Total" : {"Transaction Type" : "Total", "Below Cashflow" : 0, "Above Cashflow" : 0}} 
            for transaction in all_rows: #build dict for easy sorting
                if transaction.get("CashFlowSys") not in (None,"None"):
                    tranType = transaction.get("TransactionType")
                    if tranType not in diffTableDict:
                        diffTableDict[tranType] = {"Transaction Type" : tranType, "Below Cashflow" : 0, "Above Cashflow" : 0}
                    if transaction.get("Source name") == header:
                        diffTableDict[tranType]["Below Cashflow"] += float(transaction.get("CashFlowSys"))
                        diffTableDict["Total"]["Below Cashflow"] += float(transaction.get("CashFlowSys"))
                    elif transaction.get("Target name") == header:
                        diffTableDict[tranType]["Above Cashflow"] += float(transaction.get("CashFlowSys"))
                        diffTableDict["Total"]["Above Cashflow"] += float(transaction.get("CashFlowSys"))
                    
            diffTable = []
            for tranType in diffTableDict: #calculate differences and put in list of dicts for table
                if tranType != "Total":
                    diffTableDict[tranType]["Difference"] = diffTableDict.get(tranType).get("Below Cashflow") - diffTableDict.get(tranType).get("Above Cashflow")
                    diffTable.append(diffTableDict.get(tranType))
            diffTableDict["Total"]["Difference"] = diffTableDict.get("Total").get("Below Cashflow") - diffTableDict.get("Total").get("Above Cashflow")
            diffTable.append(diffTableDict.get("Total"))
            diffHeaders = ["Transaction Type", "Below Cashflow", "Above Cashflow", "Difference"]
            self.parent.openTableWindow(diffTable, name = f"Transaction types for {hier} in {selectedMonth}", headers = diffHeaders)

        if len(all_rows) == 0:
            print("No rows found")
            return
        # 3) Sort by dateTime column (handles ISO or space-separated)
        def parse_dt(s):
            return datetime.strptime(s, "%Y-%m-%dT00:00:00")

        all_rows.sort(key=lambda r: parse_dt(r.get('Date', '')))

        # 4) Collect the union of all column keys
        if not self.headerOptions.active:
            all_cols = self.headerOrder
            for row in all_rows:
                for key in row.keys():
                    if key not in all_cols:
                        all_cols.append(key)
            self.allCols = all_cols
            self.headerOptions.set_items(all_cols)
        else:
            self.allCols = self.headerOptions.popup.get_checked_sorted_items()

        # 5) Configure the table widget
        self.table.setRowCount(len(all_rows))
        self.table.setColumnCount(len(self.allCols))
        self.table.setHorizontalHeaderLabels(self.allCols)

        # 6) Populate each cell
        for r, row in enumerate(all_rows):
            for c, key in enumerate(self.allCols):
                raw = row.get(key,"")
                try:
                    num = float(raw)
                    text = f"{num:,.2f}"
                    item = QTableWidgetItem(text)
                    item.setData(Qt.UserRole,num)
                except:
                    item = QTableWidgetItem(str(raw))
                self.table.setItem(r, c, item)

        self.success = True

@attach_logging_to_class
class tableWindow(QWidget):
    """
    A window that loads basic rows and displays as a table
    """
    def __init__(self, parent=None, flags=Qt.WindowFlags(), parentSource = None, all_rows = [], table = "", headers = None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.setWindowTitle(f"New data in {table}")
        self.resize(1000, 600)

        # Layout and table
        layout = QVBoxLayout(self)
        self.excelBtn = QPushButton("Export to Excel")
        self.excelBtn.clicked.connect(self.export)
        layout.addWidget(self.excelBtn)
        self.table = QTableWidget(self)
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)

        self.rows = all_rows

        # 4) Collect the union of all column keys
        if headers is None:
            all_cols = set()
            for row in all_rows:
                all_cols.update(row.keys())
            all_cols = list(all_cols)
        else:
            all_cols = headers
        self.headers = all_cols
        # 5) Configure the table widget
        self.table.setRowCount(len(all_rows))
        self.table.setColumnCount(len(all_cols))
        self.table.setHorizontalHeaderLabels(all_cols)

        # 6) Populate each cell
        for r, row in enumerate(all_rows):
            for c, key in enumerate(all_cols):
                raw = row.get(key,"")
                try:
                    num = float(raw)
                    text = f"{num:,.2f}"
                    item = QTableWidgetItem(text)
                    item.setData(Qt.UserRole,num)
                except:
                    item = QTableWidgetItem(str(raw))
                self.table.setItem(r, c, item)
    def export(self,*_):
        exportTableToExcel(self,self.rows,self.headers)

class reportExportWindow(QWidget):
    def __init__(self, db: DatabaseManager, exportType, parent = None, flags = Qt.WindowFlags(), parentSource = None):
        super().__init__(parent, flags)
        self.parent = parentSource
        self.exportType = exportType
        self.setWindowTitle('Report Export Options')

        self.fam2inv = db.pullInvestorsFromFamilies
        self.db = db

        self.setStyleSheet(self.parent.appStyle)
        self.setObjectName('mainPage')
        layout = QVBoxLayout()
        self.setLayout(layout)

        #Date choice
        dateBox = QWidget()
        dateBox.setObjectName("borderFrame")
        dateLayout = QVBoxLayout()
        self.dateSelect = simpleMonthSelector()
        self.populateMonths()
        dateLayout.addWidget(QLabel('As of Date:'))
        dateLayout.addWidget(self.dateSelect)
        dateBox.setLayout(dateLayout)
        layout.addWidget(dateBox)

        #Source choice
        sourceBox = QWidget()
        sourceBox.setObjectName('borderFrame')
        sourceBoxLayout = QVBoxLayout()
        sourceBox.setLayout(sourceBoxLayout)
        layout.addWidget(sourceBox)
        self.rGroup = QButtonGroup()
        self.rGroup.buttonClicked.connect(self.swapChoice)
        allBtn = QRadioButton('Full Portfolio')
        famBtn = QRadioButton("Select by Family Branch")
        invBtn = QRadioButton("Select by Investor")
        for rb in (allBtn,famBtn,invBtn):
            self.rGroup.addButton(rb)
        sourceBoxLayout.addWidget(allBtn)
        splitter = QSplitter(Qt.Horizontal)
        sourceBoxLayout.addWidget(splitter)
        famSideL = QWidget()
        invSideR = QWidget()
        splitter.addWidget(famSideL)
        splitter.addWidget(invSideR)
        famSideLlay = QVBoxLayout()
        invSideRlay = QVBoxLayout()
        famSideL.setLayout(famSideLlay)
        invSideR.setLayout(invSideRlay)

        famSideLlay.addWidget(famBtn)
        invSideRlay.addWidget(invBtn)
        famSelect = MultiSelectBox()
        invSelect = MultiSelectBox()
        famSideLlay.addWidget(famSelect)
        invSideRlay.addWidget(invSelect)

        #Classification choice
        classBox = QWidget()
        classBox.setObjectName('borderFrame')
        classBoxLayout = QHBoxLayout()
        classBox.setLayout(classBoxLayout)
        layout.addWidget(classBox)
        classBoxLayout.addWidget(QLabel('Classification:'))
        self.classChoice = QComboBox()
        fund2trait = db.fetchFund2Trait()
        classOpts = set(f.get('Classification') for f in fund2trait.values() if f.get('Classification') is not None)
        classOpts = ['',*sorted(list(classOpts))]
        self.classChoice.addItems(classOpts)
        self.classChoice.setCurrentText('HFC')
        classBoxLayout.addWidget(self.classChoice)

        self.confirmBtn = QPushButton('Generate Report')
        layout.addWidget(self.confirmBtn)
        self.confirmBtn.clicked.connect(self.beginExport)

        investors = db.fetchInvestors()
        for cb, invKey in ([invSelect, 'Name'],[famSelect,'Parentinvestor']):
            cb.addItems(sorted(set((inv[invKey] for inv in investors))))
        
        self.invSelect = invSelect
        self.famSelect = famSelect
        allBtn.click()
    def populateMonths(self):
        start = self.parent.dataTimeStart
        end = datetime.now() - relativedelta(months=1) + relativedelta(hours=8)
        #ends on the previous month. Adds a few hours so index will still be before it and count as a month on the 1st
        index = start
        monthList = []
        while index < end:
            monthList.append(datetime.strftime(index,"%B %Y"))
            index += relativedelta(months=1)
        self.dateSelect.addItems(monthList)
        self.dateSelect.setCurrentText(monthList[-1])
    def swapChoice(self,button):
        try:
            btnText = button.text()
            if btnText == 'Full Portfolio':
                self.invSelect.setVisible(False)
                self.famSelect.setVisible(False)
            elif btnText == 'Select by Family Branch':
                self.invSelect.setVisible(False)
                self.famSelect.setVisible(True)
            elif btnText == 'Select by Investor':
                self.famSelect.setVisible(False)
                self.invSelect.setVisible(True)
            else:
                raise ValueError('Error: Button Selection could not be connected to options')
        except Exception as e:
            QMessageBox.warning(self,'Error', f"{e}")
    def beginExport(self,*_):
        try:
            self.enabled = False
            self.confirmBtn.setEnabled(False)
            source = self.rGroup.checkedButton().text()
            print(f"Exporting for : {source}")
            classification = self.classChoice.currentText()
            if source == 'Full Portfolio':
                self.invSelect.selectAll()
                investors = self.invSelect.checkedItems()
                sourceName = findSourceName([],[])
            elif source == 'Select by Family Branch':
                fams = self.famSelect.checkedItems()
                investors = self.fam2inv(fams)
                sourceName = findSourceName(fams,[])
            elif source == 'Select by Investor':
                investors = self.invSelect.checkedItems()
                sourceName = findSourceName([],investors)
            else:
                raise ValueError('Error: Button Selection could not be connected to options')
            if not investors:
                raise ValueError('No investors found')
            endDate = self.dateSelect.currentText()
            endDate = datetime.strptime(endDate,'%B %Y')
            if 'Full Report' in self.exportType:
                placeholders = ','.join('?' for _ in investors)
                date = self.dateSelect.currentText()
                dateDt = datetime.strptime(date,'%B %Y')
                date = datetime.strftime(dateDt,'%Y-%m-%d 00:00:00')
                inputs = (*investors,)
                condStatement = f' WHERE [Source name] in ({placeholders})'
                calcs = self.parent.db.loadCalcs(condStatement,inputs)
                trans = self.parent.db.loadFromDB('transactions',condStatement,inputs)
                snapshotWorkbook = portfolioSnapshot(calcs, trans, self.parent,investors,dateDt)
                render_report(self.parent,snapshotWorkbook)
            else:
                filters = {'Source name' : investors}
                if classification != '':
                    filters['Classification'] = [classification,]
                controlTable(self.parent,reset = True, filterChoices=filters, endDate = endDate)
                basicHoldingsReportExport(self.parent, sourceName = sourceName, classification = classification)
            self.confirmBtn.setEnabled(True)
            self.enabled = True
        except Exception as e:
            print("Error occured in report export initialization")
            QMessageBox.warning(self,'Error in report export', f"An error occured initializing the report export \n {e.args}")
            print(traceback.format_exc())
            self.confirmBtn.setEnabled(True)
