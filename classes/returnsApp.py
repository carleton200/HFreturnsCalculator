from math import e
from scripts.loggingFuncs import attach_logging_to_class
from classes.DatabaseManager import DatabaseManager, load_from_db, save_to_db
from scripts.instantiate_basics import ASSETS_DIR, DATABASE_PATH, gui_queue, executor, APIexecutor, HELP_PATH
from classes.widgetClasses import CheckboxIntInputWidget, simpleMonthSelector, MultiSelectBox, SortButtonWidget
from scripts.commonValues import (currentVersion, headerSortExclusions, nameHier, headerOptions, nonAggregatingCols, nonDefaultHeaders, ownershipCorrect, masterFilterOptions, importInterval, 
                    currentVersion, demoMode, fullRecalculations, calculationPingTime, dashInactiveMinutes, nonFundCols, mainTableNames,
                    nodePathSplitter,assetClass1Order, assetClass2Order,headerOptions, dataOptions, assetLevelLinks, textCols,
                    yearOptions, percent_headers, mainURL, dynamoAPIenvName)
from scripts.processInvestments import processInvestments
from scripts.basicFunctions import (calc_DPI_TVPI, findSign, updateStatus, annualizeITD, submitAPIcall, get_connected_node_groups, 
                                 descendingNavSort, accountBalanceKey, filt2Query, separateRowCode, findSourceName)
from classes.windowClasses import reportExportWindow, underlyingDataWindow, linkBenchmarksWindow, tableWindow, exportWindow, displayWindow
from classes.tableWidgets import DictListModel, SmartStretchTable
from TreeScripts.dash_launcher import _run_dash_app_process
from classes.transactionApp import transactionApp
from scripts.pyqtFunctions import basicHoldingsReportExport
from scripts.processClump import processClump
from classes.nodeLibrary import nodeLibrary
from openpyxl.utils import get_column_letter
import statistics


import os
import re
import json
import time
import copy
import requests
import calendar
import traceback
import threading
import subprocess
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from concurrent.futures import wait
from collections import defaultdict
from multiprocessing import Pool, Manager
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from openpyxl.styles import PatternFill, Alignment, Font


from PyQt5.QtWidgets import (
                                QApplication, QDialog, QInputDialog, QWidget, QStackedWidget, QVBoxLayout,
                                QLabel, QLineEdit, QPushButton,
                                QRadioButton, QButtonGroup, QComboBox, QHBoxLayout,
                                QTableWidget, QTableWidgetItem, QProgressBar, QTableView, QCheckBox, QMessageBox,
                                QFileDialog, QGridLayout,
                                QFrame
                            )
from PyQt5.QtGui import QBrush, QColor, QDesktopServices
from PyQt5.QtCore import Qt, QTimer, QUrl

@attach_logging_to_class
class returnsApp(QWidget):
    def __init__(self, start_index=0):
        super().__init__()
        self.setWindowTitle('CRSPR')
        self.setGeometry(100, 100, 1000, 600)

        os.makedirs(ASSETS_DIR, exist_ok=True)
        self.db = DatabaseManager(DATABASE_PATH)
        self.start_index = start_index
        self.api_key = None
        self.filterCallLock = False
        self.cancel = False
        self.lock = threading.Lock()
        self.db._lock = self.lock #attach the lock to the database manager
        self.tableWindows = {}
        self.dataTimeStart = datetime(2000,1,1)
        self.earliestChangeDate = datetime.now() + relativedelta(months=1)
        self.nodeChangeDates = {"active" : False}
        self.currentTableData = None
        self.currentTableFlags = None
        self.fullLevelOptions = {}
        self.buildTableCancel = None
        self.buildTableFuture = None
        self.cFundsCalculated = False
        self.previousGrouping = set()

        self.timer = QTimer()
        self.timer.timeout.connect(self.update_from_queue)
        self.queue = []
        
        # Create multiprocessing Manager for Dash app lifecycle tracking
        from multiprocessing import Manager
        self.dash_manager = Manager()
        self.dash_active_flag = self.dash_manager.dict({'active': True})

        # main stack
        self.main_layout = QVBoxLayout()
        self.appStyle = ("""
                        QWidget#borderFrame {
                            border: 2px solid #3E85E9;
                            border-radius: 4px;
                            padding: 4px;
                        }
                        QWidget#titleBox {
                            border: 4px solid #0665EA;
                            border-radius: 5px;
                            padding: 4px;
                        }
                        QWidget#mainPage, QMessageBox, QDialog {
                            background-color: #383838
                        }
                        QLineEdit[status="disabled"]{
                            background-color: #383838;
                            color : #383838;
                        }
                        QPushButton {
                            background-color: #3E85E9;
                            border: 2px solid transparent;
                            border-radius: 12px;
                            padding: 4px
                        }
                        QPushButton:hover {
                                background-color: #1771EE;
                        }
                        QPushButton#exportBtn {
                            background-color: #51AE2B;
                        }
                        QPushButton#exportBtn:hover {
                            background-color: #429321;
                        }
                        QPushButton#cancelBtn {
                            background-color: #D63131;
                        }
                        QPushButton#helpBtn {
                            background-color: #FFDE59;
                            color: black;
                        }
                        QPushButton#cntrlBtn {
                            background-color: #03fc98;
                            color: black;
                        }
                        QLabel, QRadioButton, QCheckBox, QProgressBar {
                            color: white
                        }
                        QTableWidget, QWidget#subPanel, QHeaderView::corner, QTableCornerButton::section {
                        background-color : #514F4F
                        }
                        QHeaderView::section {
                            background-color: #A8A2A2;
                        }
                        QListWidget {
                            background-color : #514F4F;
                            color: white
                        }
                        QLineEdit{
                            border: 2px solid transparent;
                            border-radius: 12px;
                            background-color: #514F4F;
                            color : white;
                        }
                        QComboBox {
                            background-color: #514F4F;
                            color : white;
                        }
                    """)
        self.setStyleSheet(self.appStyle)
        self.setObjectName("mainPage")
        self.checkVersion()
        self.stack = QStackedWidget()
        self.init_global_widgets()

        self.init_api_key_page() #0
        self.init_returns_page() #1
        self.init_calculation_page() #2

        self.stack.setCurrentIndex(start_index)
        self.main_layout.addWidget(self.stack)
        self.setLayout(self.main_layout)
    def closeEvent(self, event):
        """Called when the widget is being closed."""
        try:
            # Signal Dash apps to shut down
            if hasattr(self, 'dash_active_flag'):
                self.dash_active_flag['active'] = False
                print("Signaled Dash apps to shut down")
            
            # Close any pools
            if hasattr(self, 'pool'):
                self.pool.close()
                self.pool.join()
            # Close database connection
            if hasattr(self, 'db') and hasattr(self.db, '_conn'):
                print("Closing database connection...")
                self.db._conn.close()
                print("Database connection closed.")
            
            # Shutdown the manager (this will also signal subprocesses)
            if hasattr(self, 'dash_manager'):
                self.dash_manager.shutdown()
            
        except Exception as e:
            print(f"Error during cleanup: {e}")
        
        # Accept the close event to allow the window to close
        event.accept()
    def init_global_widgets(self):
        headerBox = QWidget()
        headerLayout = QHBoxLayout()
        self.lastImportLabel = QLabel("Last Data Import: ")
        headerLayout.addWidget(self.lastImportLabel)
        headerLayout.addStretch()
        headerLayout.addWidget(QLabel("Carleton's Really Speedy Performance Reporting"))
        headerLayout.addStretch()
        headerLayout.addWidget(QLabel(f"Version: {currentVersion}"))
        self.helpBtn = QPushButton("Help")
        self.helpBtn.clicked.connect(self.helpClicked)
        self.helpBtn.setObjectName("helpBtn")
        headerLayout.addWidget(self.helpBtn)
        headerBox.setLayout(headerLayout)
        self.main_layout.addWidget(headerBox)
        self.apiLoadingBarBox = QWidget()
        t2 = QVBoxLayout()
        t2.addWidget(QLabel("Pulling transaction and account data from server..."))
        self.apiLoadingBar = QProgressBar()
        self.apiLoadingBar.setRange(0,100)
        t2.addWidget(self.apiLoadingBar)
        self.apiLoadingBarBox.setLayout(t2)
        self.apiLoadingBarBox.setVisible(False)
        self.main_layout.addWidget(self.apiLoadingBarBox)
        loadLay = QGridLayout()
        self.calculationLoadingBar = QProgressBar()
        self.calculationLoadingBar.setRange(0,100)
        self.calculationLabel = QLabel()
        self.cancelCalcBtn = QPushButton("Cancel Calculations")
        self.cancelCalcBtn.setObjectName("cancelBtn")
        self.cancelCalcBtn.setEnabled(False)
        self.cancelCalcBtn.clicked.connect(self.cancelCalc)
        loadLay.addWidget(self.calculationLabel,0,0,1,5)
        loadLay.addWidget(self.calculationLoadingBar, 1,0, 1,5)
        loadLay.addWidget(self.cancelCalcBtn, 2, 2)
        self.calculationLoadingBox = QWidget()
        self.calculationLoadingBox.setLayout(loadLay)
        self.calculationLoadingBox.setVisible(False)
        self.main_layout.addWidget(self.calculationLoadingBox)

    def init_api_key_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        self.api_label = QLabel('Enter Dynamo API Key:')
        self.api_input = QLineEdit()
        btn = QPushButton('Submit')
        btn.clicked.connect(self.check_api_key)
        layout.addWidget(self.api_label)
        layout.addWidget(self.api_input)
        layout.addWidget(btn)
        page.setLayout(layout)
        self.stack.addWidget(page)

    def init_calculation_page(self):
        page = QWidget()
        layout = QVBoxLayout()
        self.info_label = QLabel('Results')
        layout.addWidget(self.info_label)
        btn_to_form = QPushButton('Go to Results')
        exportBtn = QPushButton('Export Data')
        exportBtn.setObjectName("exportBtn")
        exportBtn.clicked.connect(self.exportCalculations)
        layout.addWidget(exportBtn)
        btn_to_form.clicked.connect(lambda: self.stack.setCurrentIndex(1))
        layout.addWidget(btn_to_form)
        

        hl = QHBoxLayout()
        self.calculationTable = QTableView(); self.calculationTable.setSortingEnabled(True)
        hl.addWidget(self.calculationTable)
        layout.addLayout(hl)

        

        page.setLayout(layout)
        self.stack.addWidget(page)

    def init_returns_page(self):
        
        page = QWidget()
        layout = QVBoxLayout()
        controlsBox = QWidget()
        controlsLayout = QHBoxLayout()
        controlsLayout.addStretch(1)
        minControlsBtn = QPushButton('Toggle Table Controls')
        minControlsBtn.clicked.connect(self.toggleMinControls)
        minControlsBtn.setObjectName('cntrlBtn')
        controlsLayout.addWidget(minControlsBtn)
        self.importButton = QPushButton('Reimport Data')
        self.importButton.clicked.connect(self.beginImport)
        self.clearButton = QPushButton('Full Recalculation')
        self.clearButton.clicked.connect(self.resetData)
        # Tie clearButton visibility to importButton
        original_setEnabled = self.importButton.setEnabled
        def setEnabled_wrapper(visible):
            original_setEnabled(visible)
            self.clearButton.setEnabled(visible)
        self.importButton.setEnabled = setEnabled_wrapper
        if not demoMode:
            controlsLayout.addWidget(self.clearButton, stretch=0)
        controlsLayout.addWidget(self.importButton, stretch=0)
        btn_to_results = QPushButton('See Calculation Database')
        btn_to_results.clicked.connect(self.show_results)
        controlsLayout.addWidget(btn_to_results, stretch=0)
        tranAppBtn = QPushButton('Transaction App')
        tranAppBtn.clicked.connect(self.openTranApp)
        controlsLayout.addWidget(tranAppBtn, stretch=0)
        dashAppBtn = QPushButton('Tree Hierarchy Viewer')
        dashAppBtn.clicked.connect(self.openDashApp)
        controlsLayout.addWidget(dashAppBtn, stretch=0)
        exportReportBtn = QPushButton("PDF Export")
        exportReportBtn.clicked.connect(self.exportReport)
        exportReportBtn.setObjectName('exportBtn')
        controlsLayout.addWidget(exportReportBtn)
        exportBtn = QPushButton("Excel Export")
        exportBtn.clicked.connect(self.exportPage)
        exportBtn.setObjectName("exportBtn")
        controlsLayout.addWidget(exportBtn, stretch=0)
        controlsLayout.addStretch(1)
        controlsBox.setLayout(controlsLayout)
        layout.addWidget(controlsBox)

        self.tableControlsBox = QWidget()
        tableControlsLayout = QVBoxLayout()
        self.tableControlsBox.setLayout(tableControlsLayout)

        optionsBox = QWidget()
        optionsBox.setObjectName("borderFrame")
        optionsGrid = QGridLayout()
        optionsTitle = QLabel("Options")
        optionsTitle.setObjectName("titleBox")
        optionsGrid.addWidget(optionsTitle,0,0,5,1)
        self.tableBtnGroup = QButtonGroup()
        self.complexTableBtn = QRadioButton("Complex Table")
        self.monthlyTableBtn = QRadioButton("Monthly Table")
        buttonBox = QWidget()
        buttonLayout = QVBoxLayout()
        for idx, rb in enumerate((self.complexTableBtn,self.monthlyTableBtn)):
            self.tableBtnGroup.addButton(rb)
            buttonLayout.addWidget(rb)
        self.returnOutputType = QComboBox()
        self.returnOutputType.addItems([opt for opt in headerOptions if opt not in textCols])
        self.returnOutputType.currentTextChanged.connect(self.buildReturnTable)
        self.dataTypeBox = QWidget()
        dataTypeLayout = QHBoxLayout()
        dataTypeLayout.addWidget(QLabel("Data type:"))
        dataTypeLayout.addWidget(self.returnOutputType)
        self.dataTypeBox.setLayout(dataTypeLayout)
        buttonLayout.addWidget(self.dataTypeBox)
        buttonBox.setLayout(buttonLayout)
        optionsGrid.addWidget(buttonBox, 0,1,2,1)
        self.tableBtnGroup.buttonClicked.connect(self.buildReturnTable)
        self.complexTableBtn.setChecked(True)
        
        self.dataStartSelect = simpleMonthSelector()
        self.dataEndSelect = simpleMonthSelector()
        for idx, [text, CB] in enumerate((["Start: ", self.dataStartSelect], ["End: ", self.dataEndSelect])):
            optionsGrid.addWidget(QLabel(text),idx,2)
            optionsGrid.addWidget(CB,idx,3)
        optionsGrid.addWidget(QLabel("Benchmarks:"),0,4)
        self.benchmarkSelection = MultiSelectBox()
        self.benchmarkSelection.popup.closed.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.benchmarkSelection,1,4)
        optionsGrid.addWidget(QLabel("Group by: "),0,5)
        self.sortHierarchy = MultiSelectBox(dispLib=self.db.userDisplayLib())
        self.sortHierarchy.hierarchyMode()
        self.sortHierarchy.setCheckedItems(["assetClass","subAssetClass"])
        self.sortHierarchy.popup.closed.connect(self.groupingChange)
        optionsGrid.addWidget(self.sortHierarchy,1,5)
        self.consolidateFundsBtn = QCheckBox("Consolidate Funds")
        self.consolidateFundsBtn.setChecked(True)
        self.consolidateFundsBtn.clicked.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.consolidateFundsBtn,0,6)
        self.exitedFundsInput = CheckboxIntInputWidget('Hide Funds With 0 NAV for ', 1, ' Months')
        self.exitedFundsInput.setChecked(False)
        self.exitedFundsInput.valChange.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.exitedFundsInput,1,6)
        self.headerSort = SortButtonWidget()
        self.headerSort.exclusions = headerSortExclusions
        self.headerSort.popup.popup_closed.connect(self.headerSortClosed)
        optionsGrid.addWidget(self.headerSort,0,7)
        self.sortStyle = QPushButton("Sort Style: NAV")
        self.sortStyle.clicked.connect(self.sortStyleClicked)
        optionsGrid.addWidget(self.sortStyle,1,7)
        # Add a horizontal line across the optionsGrid after the top row of options
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        optionsGrid.addWidget(line, 2, 1, 1, optionsGrid.columnCount() - 1)
        self.assetClass3Visibility = MultiSelectBox(dispLib=self.db.userDisplayLib())
        self.assetClass3Visibility.popup.closed.connect(self.assetClass3VisibilityChanged)
        optionsGrid.addWidget(QLabel("Hidden Asset Level 3s:"),3,1)
        optionsGrid.addWidget(self.assetClass3Visibility,4,1)
        self.showBenchmarkLinksBtn = QCheckBox("Show Benchmark Links")
        self.showBenchmarkLinksBtn.setChecked(True)
        self.showBenchmarkLinksBtn.clicked.connect(self.buildReturnTable)
        optionsGrid.addWidget(self.showBenchmarkLinksBtn,3,2)
        self.linkBenchmarksBtn = QPushButton("Link Benchmarks")
        self.linkBenchmarksBtn.clicked.connect(self.linkBenchmarks)
        optionsGrid.addWidget(self.linkBenchmarksBtn,4,2)

        #assetClassOrderSorts
        self.AC1sort = SortButtonWidget(btnName='Asset Level 1 Sorting')
        self.AC1sort.set_items(assetClass1Order,[])
        self.AC1sort.popup.popup_closed.connect(lambda: self.ACsortChange(lvl=1))
        self.AC2sort = SortButtonWidget(btnName='Asset Level 2 Sorting')
        self.AC2sort.set_items(assetClass2Order,[])
        self.AC2sort.popup.popup_closed.connect(lambda: self.ACsortChange(lvl=2))
        optionsGrid.addWidget(self.AC1sort,3,5)
        optionsGrid.addWidget(self.AC2sort,4,5)

        optionsBox.setLayout(optionsGrid)
        tableControlsLayout.addWidget(optionsBox)

        mainFilterBox = QWidget()
        mainFilterBox.setObjectName("borderFrame")
        mainFilterLayout = QGridLayout()
        filterTitle = QLabel("Filters")
        filterTitle.setObjectName("titleBox")
        mainFilterLayout.addWidget(filterTitle,0,0,3,1)
        resetFiltersBtn = QPushButton("Reset Filters")
        def filterReset(*_):
            self.instantiateFilters()
            self.buildReturnTable()
        resetFiltersBtn.clicked.connect(filterReset)
        
        mainFilterLayout.addWidget(resetFiltersBtn,3,0)

        self.filterOptions = masterFilterOptions
        self.filterBtnExclusions = ["Source name","Classification", nameHier["subClassification"]["local"], nameHier["Family Branch"]["local"]]
        self.filterDict = {}
        self.filterRadioBtnDict = {}
        self.filterBtnGroup = QButtonGroup()
        self.filterBtnGroup.setExclusive(False)
        for col, filter in enumerate(self.filterOptions, start=1):
            row = int((col - col % 5) / 5) * 2
            col = int(col - row * 5 / 2 + 1)
            if filter["key"] not in self.filterBtnExclusions:
                #investor level is not filterable. It is total portfolio or shows the investors data
                self.filterRadioBtnDict[filter["key"]] = QCheckBox(f"{filter["name"]}:")
                self.filterRadioBtnDict[filter["key"]].setChecked(True)
                self.filterBtnGroup.addButton(self.filterRadioBtnDict[filter["key"]])
                mainFilterLayout.addWidget(self.filterRadioBtnDict[filter["key"]],row, col)
            else:
                mainFilterLayout.addWidget(QLabel(f"{filter["name"]}:"), row, col)
            if filter["key"] != "Target name":
                self.sortHierarchy.addItem(filter["key"])
            self.filterDict[filter["key"]] = MultiSelectBox(dispLib=self.db.userDisplayLib())
            self.filterDict[filter["key"]].popup.closed.connect(lambda: self.filterUpdate())
            mainFilterLayout.addWidget(self.filterDict[filter["key"]],row + 1,col)
        self.sortHierarchy.setCheckedItems(["assetClass","subAssetClass"])
        self.filterBtnGroup.buttonToggled.connect(self.filterBtnUpdate)
        mainFilterBox.setLayout(mainFilterLayout)
        tableControlsLayout.addWidget(mainFilterBox)
        layout.addWidget(self.tableControlsBox)
        t1 = QVBoxLayout() #build table loading bar
        self.buildTableLoadingBox = QWidget()
        self.tableLoadingLabel = QLabel("Building returns table...")
        t1.addWidget(self.tableLoadingLabel)
        self.buildTableLoadingBar = QProgressBar()
        self.buildTableLoadingBar.setRange(0,8)
        t1.addWidget(self.buildTableLoadingBar)
        self.buildTableLoadingBox.setLayout(t1)
        self.buildTableLoadingBox.setVisible(False)
        layout.addWidget(self.buildTableLoadingBox)
        self.returnsTable = SmartStretchTable() #table
        self.returnsTable.setSelectionMode(QTableWidget.ContiguousSelection)  # Required
        self.returnsTable.setSelectionBehavior(QTableWidget.SelectItems)
        layout.addWidget(self.returnsTable)
        unDataBox = QWidget()
        unDataLayout = QHBoxLayout()
        unDataLayout.addStretch(1)
        self.viewUnderlyingDataBtn = QPushButton("View Underlying Data")
        self.viewUnderlyingDataBtn.clicked.connect(self.viewUnderlyingData)
        unDataLayout.addWidget(self.viewUnderlyingDataBtn,stretch=0)
        self.openDashWithSelectionBtn = QPushButton("View in Tree Hierarchy")
        self.openDashWithSelectionBtn.clicked.connect(self.openDashAppWithSelection)
        unDataLayout.addWidget(self.openDashWithSelectionBtn, stretch=0)
        unDataLayout.addStretch(1)
        unDataBox.setLayout(unDataLayout)
        layout.addWidget(unDataBox)
        if ownershipCorrect:
            layout.addWidget(QLabel("Notice: Yellow highlights on NAV or monthly table indicate the values or sub-group values are affected by investor ownership corrections"))
        


        page.setLayout(layout)
        self.stack.addWidget(page)

        self.instantiateFilters()
        self.updateMonthOptions()
        if self.start_index != 0:
            self.filterUpdate()
        self.dataEndSelect.currentTextChanged.connect(self.buildReturnTable)
        self.dataStartSelect.currentTextChanged.connect(self.buildReturnTable)
    def init_data_processing(self):
        self.calcSubmitted = False
        self.lastImportDB = load_from_db(self.db,"history")
        if len(self.lastImportDB) != 1:
            self.lastImportDB = None
        if self.lastImportDB is None:
            print("No previous import found")
            #pull data is there is no data pulled yet
            self.importButton.setEnabled(False)
            executor.submit(self.pullData)
        else:
            lastImportString = self.lastImportDB[0]["lastImport"]
            lastImport = datetime.strptime(lastImportString, "%B %d, %Y @ %I:%M %p")  
            self.lastImportLabel.setText(f"Last Data Import: {lastImportString}")
            now = datetime.now()
            if lastImport.month != now.month or now > (lastImport + importInterval):
                print(f"Reimporting due to time elapsing. \n     Last import: {lastImport}\n    Current time: {now}")
                #pull data if in a new month or 1 days have elapsed
                self.importButton.setEnabled(False)
                executor.submit(self.pullData)
            else:
                calculations = load_from_db(self.db,"calculations")
                if calculations != []:
                    self.populate(self.calculationTable,calculations)
                    self.buildReturnTable()
                else:
                    executor.submit(self.pullData)
    def watchForUpdateTime(self):
        try:
            print("Checking if update required.")
            lastImportString = self.lastImportDB[0]["lastImport"]
            lastImport = datetime.strptime(lastImportString, "%B %d, %Y @ %I:%M %p")  
            now = datetime.now()
            if lastImport.month != now.month or now > (lastImport + importInterval):
                print(f"Reimporting due to the time elapsing. \n     Last import: {lastImport}\n    Current time: {now}")
                #pull data if in a new month or 1 days have elapsed
                executor.submit(self.pullData)
        except:
            print("Background watch failed")

    def helpClicked(self,*_):
        try:
            with open(HELP_PATH, 'r', encoding='utf-8') as f:
                text = f.read()
            helpMessage = displayWindow(parentSource=self, text=text, title="Help Page")
            helpMessage.show()
            self.helpPage = helpMessage
        except:
            QMessageBox.warning(self,"Error","Error opening help page.")
    def toggleMinControls(self,*_):
        self.tableControlsBox.setVisible(not self.tableControlsBox.isVisible())
    def exportCalculations(self,*_):
        window = exportWindow(parentSource=self)
        window.show()
        self.exportWindow = window
    def openTranApp(self,*_):
        tranApp = transactionApp(self.db, apiKey=self.api_key)
        tranApp.stack.setCurrentIndex(1)
        tranApp.init_data_processing()
        tranApp.show()
        self.tranApp = tranApp
    def openDashApp(self, from_selection=False, *_):
        """
        Launch Dash Tree Hierarchy Viewer app.
        
        Args:
            from_selection: If True, use current table selection for node/date pre-selection.
                          If False, launch with no pre-selection (full view).
        """
        target_node = None
        target_date = None
        
        # If launching from selection, extract node and date from current table selection
        if from_selection:
            row = self.returnsTable.currentRow()
            col = self.returnsTable.currentColumn()
            
            if row < 0 or col < 0:
                QMessageBox.warning(
                    self,
                    "No Selection",
                    "Please select a cell in the returns table first."
                )
                return
            
            # Get row data
            key = list(self.filteredReturnsTableData.keys())[row]
            row_data = self.filteredReturnsTableData[key]
            data_type = row_data.get("dataType", "")
            
            # Check if this is a Node or Target name
            if "Node" in data_type or data_type == "Total Target name":
                vh_item = self.returnsTable.verticalHeaderItem(row)
                entity_name = vh_item.text() if vh_item else None
                
                if entity_name:
                    # For nodes, extract the node path
                    if "Node" in data_type:
                        # Node paths use nodePathSplitter = " > "
                        target_node = entity_name
                    else:
                        # Direct investment/fund name
                        target_node = entity_name
            
            # Get date from selected column
            hh_item = self.returnsTable.horizontalHeaderItem(col)
            month_str = hh_item.text() if hh_item else None
            
            if month_str:
                try:
                    # Convert "Month Year" to datetime, get end of month
                    dt = datetime.strptime(month_str, "%B %Y")
                    # Get last day of month
                    next_month = dt + relativedelta(months=1)
                    target_date = (next_month - relativedelta(days=1)).strftime("%Y-%m-%d")
                except:
                    pass
            
            # If no date from column, use dataEndSelect
            if not target_date:
                try:
                    end_date_str = self.dataEndSelect.currentText()
                    dt = datetime.strptime(end_date_str, "%B %Y")
                    next_month = dt + relativedelta(months=1)
                    target_date = (next_month - relativedelta(days=1)).strftime("%Y-%m-%d")
                except:
                    pass
        
        # Create non-modal loading message and store as instance variable
        self.dash_loading_msg = QMessageBox(self)
        self.dash_loading_msg.setWindowTitle("Loading Data")
        self.dash_loading_msg.setText("Loading position data from database...\nThis may take a moment.")
        self.dash_loading_msg.setStandardButtons(QMessageBox.NoButton)
        self.dash_loading_msg.setModal(False)  # Make it non-modal so it doesn't block
        self.dash_loading_msg.show()
        QApplication.processEvents()
        
        # Load data in background thread, then launch in separate process
        def load_and_launch():
            try:
                # Load data using DatabaseManager
                data_df = self.db.load_dash_data()
                
                if data_df is None or data_df.empty:
                    # Close loading message and show warning
                    gui_queue.put(self._close_dash_loading_msg)
                    gui_queue.put(lambda: QMessageBox.warning(
                        self,
                        "No Data",
                        "No position data available to display in Tree Hierarchy Viewer."
                    ))
                    return
                
                # Pickle the data for passing to subprocess
                import pickle
                data_pickle = pickle.dumps(data_df)
                
                # Close loading message
                gui_queue.put(self._close_dash_loading_msg)
                
                # Launch in separate process with pre-loaded data
                from multiprocessing import Process
                inactivity_timeout = dashInactiveMinutes  # Minutes of inactivity before auto-shutdown
                # Pass the shared active flag dict to the subprocess
                p = Process(target=_run_dash_app_process, args=(data_pickle, target_node, target_date, self.dash_active_flag, inactivity_timeout))
                p.daemon = False  # Allow it to continue after main app closes
                p.start()
                
                # Store reference to prevent garbage collection
                if not hasattr(self, 'dash_processes'):
                    self.dash_processes = []
                self.dash_processes.append(p)
                
                # Build success message
                if from_selection and (target_node or target_date):
                    msg = "Tree Hierarchy Viewer is starting...\n\n"
                    if target_node:
                        msg += f"Pre-selected node: {target_node}\n"
                    if target_date:
                        msg += f"Pre-selected date: {target_date}\n"
                    msg += "\nIt will open in your browser shortly."
                else:
                    msg = "Tree Hierarchy Viewer is starting...\nIt will open in your browser shortly."
                
                gui_queue.put(lambda: QMessageBox.information(self, "Dash App Launched", msg))
                
            except Exception as e:
                # Ensure loading message is closed on exception
                gui_queue.put(self._close_dash_loading_msg)
                gui_queue.put(lambda: QMessageBox.critical(
                    self,
                    "Error",
                    f"Failed to launch Tree Hierarchy Viewer:\n{str(e)}"
                ))
                print(f"Error launching Dash app: {e}")
                import traceback
                traceback.print_exc()
        
        # Run in background thread
        import threading
        thread = threading.Thread(target=load_and_launch, daemon=True)
        thread.start()
    
    def openDashAppWithSelection(self, *_):
        """Launch Dash app with current table selection (wrapper for backward compatibility)"""
        self.openDashApp(from_selection=True)
    
    def _close_dash_loading_msg(self):
        """Helper method to safely close the Dash loading message"""
        if hasattr(self, 'dash_loading_msg') and self.dash_loading_msg is not None:
            try:
                
                if self.dash_loading_msg.isVisible():
                    self.dash_loading_msg.close()
                self.dash_loading_msg.hide()
                self.dash_loading_msg.destroy()
                self.dash_loading_msg = None
            except Exception as e:
                print(f"Error closing Dash loading message: {e}")
                # Ensure it's set to None even if close fails
                self.dash_loading_msg = None
    
    def cancelCalc(self, *_):
        _ = updateStatus(self,"DummyFail",99, status="Failed")
        self.cancel = True
    def viewUnderlyingData(self,*_):
        row = self.returnsTable.currentRow()
        col = self.returnsTable.currentColumn()
        key = list(self.filteredReturnsTableData.keys())[row]
        vh_item = self.returnsTable.verticalHeaderItem(row)
        row = vh_item.text() if vh_item else f"Row {row}"

        # Get the horizontal (column) header text
        hh_item = self.returnsTable.horizontalHeaderItem(col)
        col = hh_item.text() if hh_item else f"Column {col}"
        self.selectedCell = {"entity": row, "month" : col, "rowKey" : key, "dataType" : self.filteredReturnsTableData[key]["dataType"] }
        try:
            window = underlyingDataWindow(parentSource=self)
            self.udWindow = window
            if window.success:
                window.show()
        except Exception as e:
            print(f"Error in data viewing window: {e} {traceback.format_exc()}")
    def exportReport(self,*_):
        msgBox = QMessageBox(self)
        msgBox.setWindowTitle("Choose Export Type")
        msgBox.setText("Select export type for PDF export:")
        combo = QComboBox(msgBox)
        opt1 = 'Current Selection - Only Holdings'
        opt2 = 'Current Selection - Full Report'
        opt3 = 'Select Investor or Family Branch - Only Holdings'
        opt4 = 'Select Investor or Family Branch - Full Report'
        if not demoMode:
            reportOpts = [opt1,opt2,opt3,opt4]
        else:
            reportOpts = [opt1,opt3]
        combo.addItems(reportOpts)
        msgBox.layout().addWidget(combo, 1, 1)
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        
        result = msgBox.exec_()
        if result == QMessageBox.Ok:
            export_type = combo.currentText()
        else:
            return  # User cancelled
        if not self.complexTableBtn.isChecked():
            self.complexTableBtn.click() #must be in complex table mode. For now.
        if export_type in (opt3,opt4):
            window = reportExportWindow(self.db, export_type, parentSource = self)
            window.show()
            self.reportExportWindow = window
        elif export_type == opt1:
            classification = ', '.join(self.filterDict['Classification'].checkedItems())
            famChoices = self.filterDict['Family Branch'].checkedItems()
            invChoices = self.filterDict['Source name'].checkedItems()
            sourceName = findSourceName(famChoices,invChoices)
            basicHoldingsReportExport(self, classification=classification, sourceName=sourceName)
        return
    def exportPage(self,*_):
        # Use a built-in PyQt combobox messagebox for export type selection
        msgBox = QMessageBox(self)
        msgBox.setWindowTitle("Choose Export Type")
        msgBox.setText("Select export type for Excel export:")
        combo = QComboBox(msgBox)
        ACopt = 'Lookup Friendly'
        combo.addItems(['Current Table', ACopt])
        msgBox.layout().addWidget(combo, 1, 1)
        msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        
        result = msgBox.exec_()
        if result == QMessageBox.Ok:
            export_type = combo.currentText()
        else:
            return  # User cancelled
        self.exportCurrentTable(ACcols = export_type == ACopt)
    def exportCurrentTable(self,*_, ACcols = False):
        #excel export for the returns app excel export
        # helper to darken a 6-digit hex color by a given factor
        def darken_color(hex_color, factor=0.01):
            h = hex_color.strip("#")
            r = int(h[0:2], 16)
            g = int(h[2:4], 16)
            b = int(h[4:6], 16)
            dr = max(0, int(r * factor))
            dg = max(0, int(g * factor))
            db = max(0, int(b * factor))
            return f"{dr:02X}{dg:02X}{db:02X}"
        # 1) prompt user
        path, _ = QFileDialog.getSaveFileName(
            self, "Save as…", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        def processExport():
            try:
                data = self.filteredReturnsTableData

                # 2) determine hierarchy levels present
                all_types = {row.get("dataType") for row in data.values()}
                if self.sortHierarchy.checkedItems() != []:
                    full_hierarchy = ["Total"] + ["Total " + level for level in self.sortHierarchy.checkedItems()] + ["Total Target name"]
                else:
                    full_hierarchy = ["Total", "Total assetClass", "Total Target name"]
                hierarchy_levels = [lvl for lvl in full_hierarchy if lvl in all_types]

                headerSort : SortButtonWidget = self.headerSort
                currentHeaders = list(self.filteredHeaders)
                mode = self.tableBtnGroup.checkedButton().text()
                if not headerSort.active or mode == "Monthly Table" or any(opt not in headerSort.options() for opt in currentHeaders):
                    col_keys = currentHeaders
                    exceptions = nonDefaultHeaders
                    col_keys = self.orderColumns(col_keys, exceptions=exceptions)
                    if mode == "Complex Table":
                        allKeys = col_keys.copy()
                        allKeys.extend(exceptions) #all key options for the header selections
                        headerSort.set_items(allKeys,[item for item in allKeys if item not in exceptions])
                        headerSort.setEnabled(True)
                    else:
                        headerSort.setEnabled(False)
                else:
                    col_keys = headerSort.popup.get_checked_sorted_items()
                    headerSort.setEnabled(True)      
                sorted_cols = col_keys
                if ACcols:
                    insertCols = []
                    if 'Total Family Branch' in hierarchy_levels:
                        insertCols = ['Family Branch']
                        famCol = 2
                    else:
                        famCol = None
                    if 'Total Source name' in hierarchy_levels:
                        sourceCol = 3 if famCol else 2
                        insertCols.append('Investor')
                    else:
                        sourceCol = None
                    sorted_cols = ['Level','Name',*insertCols,'AC1','AC2','AC3','Investment',*sorted_cols]
                    fund2trait = self.db.fund2trait
                # 4) create workbook or add sheet if already exists
                if os.path.exists(path):
                    wb = load_workbook(path)
                    # Create a unique sheet name for export
                    base_name = "Export"
                    i = 1
                    while True:
                        sheet_name = f"{base_name}{i}"
                        if sheet_name not in wb.sheetnames:
                            break
                        i += 1
                    ws = wb.create_sheet(sheet_name)
                else:
                    wb = Workbook()
                    ws = wb.active

                rowStart = 3
                # 5) header row
                startCol = 1 if ACcols else 2
                freezeCol = 3 if ACcols else 2
                for idx, colname in enumerate(sorted_cols, start=startCol):
                    ws.cell(row=rowStart, column=idx, value=colname)

                split_cell = f"{get_column_letter(freezeCol)}4"
                ws.freeze_panes = split_cell

                # 7) populate rows
                sortHier = self.sortHierarchy.checkedItems()
                maxDepth   = max(len(sortHier),1) + 1
                nodeDicts = self.db.fetchNodes()
                nodes = set(nD['name'] for nD in nodeDicts)
                for r, (row_name, row_dict) in enumerate(data.items(), start=rowStart + 1):
                    row_name, code = separateRowCode(row_name)
                    dtype = row_dict.get("dataType")
                    if dtype != "benchmark": #keeps benchmark as the previous hierarchy level
                        level = hierarchy_levels.index(dtype) if dtype in hierarchy_levels else 0
                        # fills
                        data_color = "FFFFFF"
                        if dtype != "Total Target name":
                            depth      = code.count("::") if dtype != "Total" else code.count("::") - 1
                            data_color = darken_color(data_color,depth/maxDepth/3 + 2/3)
                        else:
                            depth = maxDepth
                        header_color = data_color
                        data_fill   = PatternFill("solid", data_color, data_color)
                        header_fill = PatternFill("solid", header_color, header_color)

                    if not ACcols:
                        cell = ws.cell(row=r, column=1, value=row_name.strip())
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                        if dtype == "benchmark":
                            cell.font = Font(color="0000FF")
                        cell.alignment = Alignment(indent=level)
                    else:
                        itemDepth = depth + 2 if dtype != "Total Target name" else depth + 1#TODO: solve to find real depth w node issues
                        if dtype != "Total Target name" and 'Node' in sortHier and sortHier.index('Node') < depth:
                            itemHier = code.removeprefix("##(").removesuffix(")##").split("::")
                            prevNodeCnt = 0
                            for nodeTier in (tier for tier in itemHier if tier in nodes):
                                prevNodeCnt += 1 #find num of nodes passed to adjust depth
                            if prevNodeCnt > 0:
                                itemDepth -= prevNodeCnt - 1 #reduce depth by all previous nodes more than 1
                        if dtype != "benchmark":
                            row_dict['Level'] = f'L{itemDepth}' #move to 1 indexed. Push past total
                        else:
                            row_dict['Level'] = f'B{itemDepth}'
                        row_dict['Name'] = row_name
                        itemHier = code.removeprefix("##(").removesuffix(")##").split("::")
                        if dtype == 'Total':
                            row_dict['Level'] = 'L1'
                        elif dtype == "Total Target name":
                            row_dict['Investment'] = row_name
                            target = list(self.cFundToFundLinks.get(row_name,[row_name,]))[0] #convert any consolidated funds to one of their sub-funds
                            fundTraits = fund2trait.get(target,{})
                            AC1 = fundTraits.get('assetClass','Not Found')
                            AC2 = fundTraits.get('subAssetClass','Not Found')
                            AC3 = fundTraits.get('subAssetSleeve','Not Found')
                            for txt,var in (['AC1',AC1],['AC2',AC2],['AC3',AC3]):
                                row_dict[txt] = var
                        else:
                            for rowTxt, AClvl in (['AC1','assetClass'],['AC2','subAssetClass'],['AC3','sleeve']):
                                if AClvl in sortHier and sortHier.index(AClvl) <= len(itemHier) - 1: #check viability
                                    try:
                                        row_dict[rowTxt] = itemHier[sortHier.index(AClvl)]
                                    except:
                                        print('Failed assigning AC levels')
                                        raise
                        for stripKey in (key for key in row_dict if key in ('AC1','AC2','AC3','Name')):
                            row_dict[stripKey] = row_dict[stripKey].strip() #remove spaces from these cols. (ex: 'Cash  ' --> 'Cash')
                        sourceSearch = []
                        if famCol:
                            sourceSearch.append(['Family Branch','Family Branch'])
                        if sourceCol:
                            sourceSearch.append(['Investor','Source name'])
                        if sourceSearch:
                            for rowTxt, lvl in sourceSearch:
                                if lvl in sortHier and sortHier.index(lvl) <= len(itemHier) - 1: #check viability
                                    try:
                                        row_dict[rowTxt] = itemHier[sortHier.index(lvl)]
                                    except:
                                        print('Failed assigning investors and family branches')
                                        raise

                                    


                    # data cells with proper formatting
                    for c, colname in enumerate(sorted_cols, start=startCol):
                        val = row_dict.get(colname, None)
                        cell = ws.cell(row=r, column=c, value=val)
                        cell.fill = data_fill
                        if dtype == "benchmark":
                            cell.font = Font(color="0000FF")
                        if isinstance(val, (int, float)):
                            if colname not in percent_headers:
                                # show with commas, two decimals
                                cell.number_format = "#,##0.00"
                            else:
                                # interpret val as percentage (e.g. 10.5 → 10.5%)
                                cell.value = val / 100.0
                                cell.number_format = "0.00%"

                # 8) autofit column widths
                for idx, col_cells in enumerate(ws.columns, start=1):
                    max_len = 0
                    for cell in col_cells:
                        if cell.value is not None:
                            text = str(cell.value)
                            max_len = max(max_len, len(text))
                    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2

                filterSelections = {}
                for filter in self.filterOptions:
                    selections = self.filterDict.get(filter.get("key")).checkedItems()
                    if selections != []:
                        filterSelections[filter.get("key")] = ", ".join(selections)

                # Determine date range to display in column A
                date_start = ""
                date_end = ""
                # Try to retrieve selected date range values if possible
                try:
                    if hasattr(self, "dataStartSelect") and hasattr(self.dataStartSelect, "currentText"):
                        date_start = self.dataStartSelect.currentText()
                    if hasattr(self, "dataEndSelect") and hasattr(self.dataEndSelect, "currentText"):
                        date_end = self.dataEndSelect.currentText()
                except Exception:
                    pass
                date_range_str = ""
                if date_start or date_end:
                    if date_start and date_end:
                        date_range_str = f"{date_start} to {date_end}"
                    elif date_start:
                        date_range_str = f"Start: {date_start}"
                    elif date_end:
                        date_range_str = f"End: {date_end}"

                if filterSelections or date_range_str:
                    # Date range in column A
                    ws.cell(row=1, column=1, value="Date Range:")
                    ws.cell(row=2, column=1, value=date_range_str)
                    ws.cell(row=1, column=1).font = Font(bold=True)
                    ws.cell(row=2, column=1).font = Font(bold=True)
                    # Filters start at column 2
                    ws.cell(row=1, column=2, value="Filters:")
                    ws.cell(row=2, column=2, value="Selections:")
                    ws.cell(row=1, column=2).font = Font(bold=True)
                    ws.cell(row=2, column=2).font = Font(bold=True)
                    for idx, filter in enumerate(filterSelections, start=3):
                        ws.cell(row=1, column=idx, value=filter)
                        cell = ws.cell(row=2, column=idx, value=filterSelections.get(filter))
                        cell.alignment = Alignment(wrap_text=True)

            
                wb.save(path)
            except Exception as e:
                gui_queue.put(lambda error=e, trace = traceback.format_exc(): QMessageBox.critical(self, "Save error", trace))
            else:
                gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
                gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
        executor.submit(processExport)
    def processFunds(self):
        print('Funds are processing')
        self.cFundsCalculated = True
        self.sleeveFundLinks = {}
        self.cFundToFundLinks = {}
        sleeves = set()
        funds = load_from_db(self.db,"funds")
        if funds != []:
            self.consolidatedFunds = {}
            self.cFundToFundLinks = defaultdict(set)
            for f in funds:
                name = f['Name']
                parent = f.get('Parentfund')
                if parent not in (None,'None',''):
                    self.consolidatedFunds[name] = parent
                    self.cFundToFundLinks[parent].add(name)
            return
            consolidatorFunds = {}
            for row in funds: #find sleeve values and consolidated funds
                assetClass = row["assetClass"]
                subAssetClass = row["subAssetClass"]
                sleeve = row["sleeve"]
                sleeves.add(sleeve)
                if row.get("Fundpipelinestatus") is not None and "Z - Placeholder" in row.get("Fundpipelinestatus"):
                    consolidatorFunds[row["Name"]] = {"cFund" : row["Name"], "assetClass" : assetClass, "subAssetClass" : subAssetClass, "sleeve" : sleeve}
                    self.cFundToFundLinks[row["Name"]] = []
                if row["sleeve"] not in self.sleeveFundLinks:
                    self.sleeveFundLinks[row["sleeve"]] = [row["Name"]]
                else:
                    self.sleeveFundLinks[row["sleeve"]].append(row["Name"])
            self.consolidatedFunds = {}
            for row in funds: #assign funds to their consolidators
                if row.get("Parentfund") in consolidatorFunds:
                    self.consolidatedFunds[row["Name"]] = consolidatorFunds.get(row.get("Parentfund"))
                    self.cFundToFundLinks[row.get("Parentfund")].append(row["Name"])
            self.fullLevelOptions["subAssetSleeve"] = list(sleeves)
        else:
            self.consolidatedFunds = {}
    def filterBtnUpdate(self, button, checked):
        if not self.filterCallLock:
            self.buildTableLoadingBox.setVisible(True)
            self.buildTableLoadingBar.setValue(1)
            self.filterCallLock = True
            reloadRequired = False
            for filter in self.filterOptions:
                if filter["key"] not in self.filterBtnExclusions:
                    if not self.filterRadioBtnDict[filter["key"]].isChecked():
                        if self.filterDict[filter["key"]].checkedItems() != []:
                            reloadRequired = True #rebuild the table only if filter selections are being removed
                        self.filterDict[filter["key"]].clearSelection()
                        self.filterDict[filter["key"]].setEnabled(False)
                        self.filterDict[filter["key"]].line_edit.setProperty("status","disabled")
                        self.filterDict[filter["key"]].line_edit.style().unpolish(self.filterDict[filter["key"]].line_edit)
                        self.filterDict[filter["key"]].line_edit.style().polish(self.filterDict[filter["key"]].line_edit)
                    else:
                        self.filterDict[filter["key"]].setEnabled(True)
                        self.filterDict[filter["key"]].line_edit.setProperty("status","enabled")
                        self.filterDict[filter["key"]].line_edit.style().unpolish(self.filterDict[filter["key"]].line_edit)
                        self.filterDict[filter["key"]].line_edit.style().polish(self.filterDict[filter["key"]].line_edit)
            self.filterCallLock = False
            if reloadRequired or self.currentTableData is None:
                self.buildReturnTable()
            else:
                self.populateReturnsTable(self.currentTableData, self.currentTableFlags)
    def resetData(self,*_):
        if not self.testAPIconnection():
            QMessageBox.warning(self,"API Failure", "API connection has failed. Server is down or API key is bad. \n Previous calculations are left in place for viewing.")
            return
        for table in ("calculations","positions","transactions"):
            save_to_db(self.db,table,None,action="clear") #reset all tables so everything will be fresh data
        self.nodeChangeDates = {"active" : False}
        executor.submit(self.pullData)
    def beginImport(self, *_):
        self.importButton.setEnabled(False)
        print("Initiating import...")
        executor.submit(self.pullData)
    def updateMonthOptions(self):
        start = self.dataTimeStart
        end = datetime.now() - relativedelta(months=1) + relativedelta(hours=8)
        #ends on the previous month. Adds a few hours so index will still be before it and count as a month on the 1st
        index = start
        monthList = []
        while index < end:
            monthList.append(datetime.strftime(index,"%B %Y"))
            index += relativedelta(months=1)
        self.dataEndSelect.addItems(monthList)
        self.dataEndSelect.setCurrentText(monthList[-1])
        self.dataStartSelect.addItems(monthList)
        self.dataStartSelect.setCurrentText(monthList[0])
    def linkBenchmarks(self,*_):
        #open the link benchmarks window as its own window that is non blocking
        self.linkBenchmarksWindow = linkBenchmarksWindow(parentSource=self)
        self.linkBenchmarksWindow.show()
    def buildReturnTable(self, *_):
        self.buildTableLoadingBox.setVisible(True)
        self.buildTableLoadingBar.setValue(2)
        if not self.cFundsCalculated:
            self.processFunds()
        
        if self.buildTableCancel:
            self.buildTableCancel.set()
        if self.buildTableFuture and not self.buildTableFuture.done():
            self.buildTableFuture.cancel()

        cancelEvent = threading.Event()
        self.buildTableCancel = cancelEvent
        self.stack.setCurrentIndex(1)
        future = executor.submit(self.buildTable, cancelEvent)
        self.buildTableFuture = future
    def buildTable(self, cancelEvent):
        try:
            print("Building return table...")
            self.currentTableData = None #resets so a failed build won't be used
            complexMode = self.tableBtnGroup.checkedButton().text() == "Complex Table"
            gui_queue.put(lambda: self.dataTypeBox.setVisible(not complexMode))
            startDate = datetime.strptime(self.dataStartSelect.currentText(), "%B %Y")
            endDate = datetime.strptime(self.dataEndSelect.currentText(), "%B %Y")
            condStatement, parameters = filt2Query(self.db, self.filterDict,startDate,endDate)
            gui_queue.put(lambda: self.buildTableLoadingBar.setValue(3))
            if cancelEvent.is_set(): #exit if new table build request is made
                return
            data = load_from_db(self.db,"calculations",condStatement, tuple(parameters))
            for idx in range(len(data)):
                data[idx]['ownershipAdjust'] = data[idx]['ownershipAdjust'] == 'True'
            output = {"Total##()##" : {}}
            flagOutput = {"Total##()##" : {}}
            if self.benchmarkSelection.checkedItems() != [] or self.showBenchmarkLinksBtn.isChecked():
                output = self.applyBenchmarks(output)
            output , data = self.calculateUpperLevels(output,data)
            for benchmark in self.pendingBenchmarks: #remove the benchmarks used only in benchmark links
                if benchmark not in self.benchmarkChoices and benchmark + self.buildCode([]) in output.keys():
                    output.pop(benchmark + self.buildCode([]))
            gui_queue.put(lambda: self.buildTableLoadingBar.setValue(4))
            if cancelEvent.is_set(): #exit if new table build request is made
                return
            complexOutput = copy.deepcopy(output)
            multiData = {}
            # Cache frequently used values and avoid repeated lookups/parsing in the loop
            headerOptions_local = headerOptions
            complexMode_local = complexMode
            end_month_str = self.dataEndSelect.currentText()
            cFunds_checked = self.consolidateFundsBtn.isChecked()
            consolidatedFunds_local = self.consolidatedFunds
            investor_checked = self.filterDict["Source name"].checkedItems() != []
            fam_checked = self.filterDict["Family Branch"].checkedItems() != []
            # Map "%" to NAV only for non-complex mode
            dataOutputType = self.returnOutputType.currentText() if not complexMode_local else "Return"
            if not complexMode_local and dataOutputType == "%":
                dataOutputType = "NAV"
            # Cache month string conversions by timestamp
            month_cache = {}
            get_month = month_cache.get
            set_month = month_cache.__setitem__
            # Local refs for speed
            out_ref = output
            c_out_ref = complexOutput
            flag_ref = flagOutput
            dtS = self.dataTimeStart
            hideMonths , hideMonthsNum = self.exitedFundsInput.getStatus()
            exitedCheck = {key : dtS for key in out_ref} #dict of all rowkeys and their earliest date with a NAV != 0
            fund2LastDate = self.db.fetchFund2Date(dateType = 'last')
            fund2Inception = self.db.fetchFund2Date(dateType = 'inception')
            for entry in data:
                e_get = entry.get
                # month string from dateTime (with small cache)
                dt_str = e_get("dateTime","")
                if dt_str is None:
                    print(f"Warning: data with no date in table build: {entry}")
                    continue #can't use data with no date. Shouldn't have gotten this fay anyways
                month_str = get_month(dt_str)
                if month_str is None:
                    month_str = datetime.strftime(datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S"), "%B %Y")
                    set_month(dt_str, month_str)
                Dtype = entry["Calculation Type"]
                targ = e_get('Target name')
                if complexMode_local and Dtype == 'Total Target name': 
                    if targ in fund2LastDate:
                        entry['Last Actual Date'] = fund2LastDate[targ]
                    if targ in fund2Inception:
                        entry['Inception'] = fund2Inception[targ]
                level = entry["rowKey"]
                if hideMonths and e_get('NAV',0) != 0: #set the date to the most recent with a NAV
                    exitedCheck[level] = max(exitedCheck.get(level,dtS), datetime.strptime(dt_str,'%Y-%m-%d %H:%M:%S'))
                if dataOutputType == "IRR ITD" and ((cFunds_checked and e_get("Target name") in consolidatedFunds_local) or e_get("Calculation Type") != "Total Target name"):
                    # skip IRR for consolidated funds or non-total
                    continue
                # ensure output[level]
                lvl_out = out_ref.get(level)
                if lvl_out is None:
                    lvl_out = {}
                    out_ref[level] = lvl_out
                val = e_get(dataOutputType)
                if val not in (None, "None", ""):
                    # flags
                    level_flags = flag_ref.setdefault(level, {})
                    level_flags.setdefault(month_str, False)
                    level_flags[month_str] = (e_get("ownershipAdjust", False)) or level_flags[month_str]
                    # aggregate
                    if month_str not in lvl_out:
                        lvl_out[month_str] = float(val)  if dataOutputType not in textCols else val
                    elif dataOutputType not in ("Return", "Ownership"):
                        lvl_out[month_str] += float(val)
                    else:
                        # same row needs special handling later
                        multiData.setdefault(level, {})
                if "dataType" not in lvl_out:
                    lvl_out["dataType"] = Dtype
                # complex table accumulation only at end month
                if complexMode_local and month_str == end_month_str:
                    lvl_c_out = c_out_ref.get(level)
                    if lvl_c_out is None:
                        lvl_c_out = {}
                        c_out_ref[level] = lvl_c_out
                    if "dataType" not in lvl_c_out:
                        lvl_c_out["dataType"] = Dtype
                    level_flags = flag_ref.setdefault(level, {})
                    nav_flag = level_flags.setdefault("NAV", False)
                    level_flags["NAV"] = (e_get("ownershipAdjust", False)) or nav_flag
                    consolidated = (cFunds_checked and e_get("Target name") in consolidatedFunds_local)
                    if headerOptions_local and headerOptions_local[0] not in lvl_c_out:
                        for option in headerOptions_local:
                            if (option in ('IRR ITD',*textCols) and (consolidated or e_get("Calculation Type") != "Total Target name")):
                                continue #skip IRR for consolidated or aggregate levels  OR ownership for consolidated funds #TODO: build lowestAggregates to handle consolidated ownership
                            ov = e_get(option)
                            lvl_c_out[option] = float(ov if ov not in (None, "None", "") else 0) if option not in textCols else ov
                    else:
                        print('multiple per rowkey')
                        for option in headerOptions_local:
                            if option not in ("Ownership", "IRR ITD"):
                                ov = e_get(option)
                                lvl_c_out[option] += float(ov if ov not in (None, "None", "") else 0)
                    if e_get("Ownership") not in (None, "None") and (investor_checked or fam_checked):
                        own = float(entry["Ownership"]) if entry.get("Ownership") not in (None, "None") else 0.0
                        if "Ownership" not in lvl_c_out:
                            lvl_c_out["Ownership"] = own
                        else:
                            lvl_c_out["Ownership"] += own
                        # else:
                        #     complexOutput[level]["Ownership"] += float(entry["Ownership"])
            for tableStruc in (output,complexOutput): 
                #remove bad table entries with no dataType (means data was somehow irrelevant. (ex: fund starts after the selected range))
                keys = tableStruc.keys()
                pops = [key for key in keys if "dataType" not in tableStruc[key]]
                for pop in pops:
                    tableStruc.pop(pop)
            if multiData and dataOutputType == "Return": #must iterate through data again to correct for returns of multi pool funds
                for entry in (entry for entry in data if entry.get("rowKey") in multiData):
                    #only occurs for the multifunds
                    #sums all gains and MDden for a row for a month
                    dateTime = entry.get("dateTime")
                    if dateTime not in multiData[entry.get("rowKey")]:
                        multiData[entry.get("rowKey")][entry.get("dateTime")] = {"MDdenominator" : float(entry.get("MDdenominator")), "Monthly Gain" : float(entry.get("Monthly Gain"))}
                    else:
                        multiData[entry.get("rowKey")][entry.get("dateTime")]["MDdenominator"] += float(entry.get("MDdenominator"))
                        multiData[entry.get("rowKey")][entry.get("dateTime")]["Monthly Gain"] += float(entry.get("Monthly Gain"))
                for rowKey in multiData: #set proper return values
                    for date in multiData.get(rowKey):
                        strDate = datetime.strftime(datetime.strptime(date, "%Y-%m-%d %H:%M:%S"), "%B %Y")
                        MDden = multiData.get(rowKey).get(date).get("MDdenominator")
                        returnVal = multiData.get(rowKey).get(date).get("Monthly Gain") / MDden * 100 if MDden != 0 else 0
                        output[rowKey][strDate] = returnVal
                        if complexMode and strDate == self.dataEndSelect.currentText():
                            complexOutput[rowKey]["Return"] = returnVal
            if hideMonths:
                monthThresh = datetime.strptime(self.dataEndSelect.currentText(), '%B %Y') - relativedelta(months=hideMonthsNum)
                deleteKeys = {key for key in set(output.keys()) | set(complexOutput.keys()) if exitedCheck.get(key,dtS) <= monthThresh and output[key]['dataType'] != 'benchmark'}
                for table in (output,complexOutput): #delete the keys from both tables of funds that are empty
                    for dKey in (key for key in deleteKeys if key in table):
                        table.pop(dKey)
            if complexMode:
                for rowKey in (key for key in complexOutput if complexOutput[key].get("NAV",0.0) != 0):
                    complexOutput[rowKey]["%"] = complexOutput[rowKey].get("NAV",0.0) / complexOutput["Total##()##"].get("NAV",0.0) * 100 if complexOutput["Total##()##"]["NAV"] != 0 else 0
            elif self.returnOutputType.currentText() == "%":
                for rowKey in reversed(output): #iterate through backwards so total is affected last
                    for date in [header for header in output[rowKey].keys() if header != "dataType"]:
                        output[rowKey][date] = float(output[rowKey][date]) / float(output["Total##()##"][date]) * 100 if  float(output["Total##()##"][date]) != 0 else 0                
            gui_queue.put(lambda: self.buildTableLoadingBar.setValue(5))
            if cancelEvent.is_set(): #exit if new table build request is made
                return
            if  complexMode:
                output = self.calculateComplexTable(output,complexOutput)
            gui_queue.put(lambda: self.buildTableLoadingBar.setValue(6))
            if cancelEvent.is_set(): #exit if new table build request is made
                return
            for key in (key for key in output.keys() if len(output[key].keys()) == 0):
                output.pop(key) #remove empty entries
            gui_queue.put(lambda: self.populateReturnsTable(output,flagStruc=flagOutput))
            self.currentTableData = output
            self.currentTableFlags = flagOutput
        except Exception as e:
            tracebackMsg = traceback.format_exc()
            gui_queue.put(lambda error = e: QMessageBox.warning(self, "Error building returns table", f"Error: {error}. {error.args}. Data entry: \n  \n Traceback:  \n {tracebackMsg}"))
            gui_queue.put(lambda: self.buildTableLoadingBox.setVisible(False))
    def calculateComplexTable(self,monthOutput,complexOutput):
        # Precompute end-of-period and month sequences
        endTime = datetime.strptime(self.dataEndSelect.currentText(), "%B %Y")
        end_month_str = datetime.strftime(endTime, "%B %Y")
        MTDtime = [end_month_str]
        # avoid repeated int() calls
        end_month = endTime.month
        q_len = (end_month % 3) if (end_month % 3) != 0 else 3
        y_len = (end_month % 12) if (end_month % 12) != 0 else 12
        QTDtimes = [datetime.strftime(endTime - relativedelta(months=i), "%B %Y") for i in range(q_len)]
        YTDtimes = [datetime.strftime(endTime - relativedelta(months=i), "%B %Y") for i in range(y_len)]
        timeSections = {"MTD": MTDtime, "QTD": QTDtimes, "YTD": YTDtimes}
        # Year windows
        YR_times = {yr: [datetime.strftime(endTime - relativedelta(months=i), "%B %Y") for i in range(12 * yr)] for yr in yearOptions}

        # Local refs for speed
        mo = monthOutput
        co = complexOutput

        # Cache for parsing month strings
        month_dt_cache = {}
        def parse_month_str(s):
            dt = month_dt_cache.get(s)
            if dt is None:
                dt = datetime.strptime(s, "%B %Y")
                month_dt_cache[s] = dt
            return dt

        for level, lvl_month in mo.items():
            lvl_co = co.get(level)
            if not lvl_co or lvl_co.get('dataType') == "benchmark":
                # Skip filtered rows and benchmarks (imported separately)
                continue
            #TVPI and DPI
            lvl_co = calc_DPI_TVPI(lvl_co)

            # Aggregate MTD/QTD/YTD compounded performance
            for timeFrame, monthOpts in timeSections.items():
                cPerf = 1.0
                # multiply only months present
                for monthO in monthOpts:
                    val = lvl_month.get(monthO)
                    if val is not None:
                        cPerf *= (1.0 + float(val) / 100.0)
                    if cPerf < 0:
                        cPerf = -1.999
                        break
                lvl_co[timeFrame] = (cPerf - 1.0) * 100.0 if cPerf > 0 else 'N/A'

            # Yearly windows compounded and annualized
            lvl_month_keys = lvl_month.keys()
            nav_zero = (lvl_co.get("NAV", 1) == 0)
            for yearKey, months in YR_times.items():
                # Ensure all months are present
                if not all(m in lvl_month_keys for m in months):
                    continue
                # Skip if NAV==0 and any month return is 0
                if nav_zero and any(float(lvl_month[m]) == 0 for m in months):
                    continue
                cYR = 1.0
                for m in months:
                    cYR *= (1.0 + float(lvl_month[m]) / 100.0)
                    if cYR < 0:
                        cYR = -1.999
                        break
                lvl_co[f"{yearKey}YR"] = (((cYR ** (1 / int(yearKey))) - 1) * 100.0) if cYR > 0 else 'N/A'

            # ITD
            try:
                if lvl_month.get("dataType", "") != "benchmark":
                    if end_month_str in lvl_month:
                        # months up to endTime
                        # build list of (dt, str) once, ignoring 'dataType'
                        month_pairs = []
                        for m in lvl_month_keys:
                            if m == "dataType":
                                continue
                            dt = parse_month_str(m)
                            if dt <= endTime:
                                month_pairs.append((dt, m))
                        if len(month_pairs) >= 2:
                            month_pairs.sort(key=lambda x: x[0])
                            cITD = 1.0
                            monthCount = 0
                            for _, m in month_pairs:
                                monthCount += 1
                                cITD *= (1.0 + float(lvl_month[m]) / 100.0)
                                if cITD < 0:
                                    cITD = -1.999
                                    monthCount = 14
                                    break
                            lvl_co["ITD"] = annualizeITD(cITD, monthCount)
                        else:
                            # ITD is just the previous month if no more months are found
                            lvl_co["ITD"] = lvl_month[MTDtime]
                else:
                    lvl_co["ITD"] = lvl_month["ITD"]
            except Exception:
                pass

        return co
    def applyBenchmarks(self, output):
        if self.showBenchmarkLinksBtn.isChecked(): #activate the benchmark links so they are all used if relevant
            benchmarkLinks = self.db.fetchBenchmarkLinks()
            self.pendingBenchmarks = set(link.get("benchmark") for link in benchmarkLinks)
        self.benchmarkChoices = self.benchmarkSelection.checkedItems()
        allBenchmarkChoices = set(set(self.benchmarkChoices) | set(self.pendingBenchmarks))
        code = self.buildCode([])
        placeholders = ','.join('?' for _ in allBenchmarkChoices)
        if allBenchmarkChoices:
            benchmarks = load_from_db(self.db,"benchmarks",f"WHERE [Index] IN ({placeholders})",tuple(allBenchmarkChoices))
        else:
            benchmarks = []
        for bench in benchmarks:
            name = bench["Index"] + code
            if (datetime.strptime(bench["Asofdate"], "%Y-%m-%dT%H:%M:%S") < datetime.strptime(self.dataStartSelect.currentText(), "%B %Y") or
                datetime.strptime(bench["Asofdate"], "%Y-%m-%dT%H:%M:%S") > datetime.strptime(self.dataEndSelect.currentText(), "%B %Y") + relativedelta(months=1) ) :
                continue #skip if outside selected range
            date = datetime.strftime(datetime.strptime(bench["Asofdate"], "%Y-%m-%dT%H:%M:%S"), "%B %Y")
            if output.get(name) is None:
                output[name] =  {}
            if self.tableBtnGroup.checkedButton().text() != "Complex Table" and self.returnOutputType.currentText() == "Return": #show monthly return benchmarks
                output[name][date] = float(bench.get("MTDnet",0) if bench.get("MTDnet",0) !=  "None" else 0) * 100
                if output[name].get("dataType") is None:
                    output[name]["dataType"] = "benchmark"
            elif self.tableBtnGroup.checkedButton().text() == "Complex Table" and date == self.dataEndSelect.currentText():
                #populate the complex fields
                if output[name].get("dataType") is None:
                    output[name]["dataType"] = "benchmark"
                for time in ("MTD","QTD","YTD"):
                    if bench.get(f"{time}net", "None") != "None":
                        output[name][time] = float(bench.get(f"{time}net")) * 100
                if bench.get("ITDTWR","None") != "None":
                    output[name]["ITD"] = float(bench.get("ITDTWR")) * 100
                for year in yearOptions:
                    if bench.get(f"Last{year}yrnet","None") not in ("None",None):
                        output[name][f"{year}YR"] = float(bench.get(f"Last{year}yrnet")) * 100
        return output
    def buildCode(self, path):
            code = f"##({"::".join(path)})##"
            return code
    def calculateUpperLevels(self, tableStructure,data):
        # Hot-path caches and precomputations for performance on large inputs
        headerOptions_local = headerOptions
        AC1order = self.db.fetchACorder(1)
        AC2order = self.db.fetchACorder(2)
        fund2traitGet = self.db.fetchFund2Trait().get
        id2Node = self.db.pullId2Node()
        filteredTargets = self.filterDict["Target name"].checkedItems()
        dataOptions_local = dataOptions
        sortHierarchy = self.sortHierarchy.checkedItems()
        buildCode = self.buildCode
        showBenchLinks = self.showBenchmarkLinksBtn.isChecked()
        consolidateFunds = self.consolidateFundsBtn.isChecked()
        consolidatedFunds_map = self.consolidatedFunds
        # Precompute end-of-period datetime once for NAV sorting comparisons
        try:
            end_period_dt = datetime.strptime(self.dataEndSelect.currentText(), "%B %Y")
        except Exception:
            end_period_dt = None
        def applyLinkedBenchmarks(struc,code, levelName, option):
            for entry in benchmarkLinks:
                if levelName == assetLevelLinks[entry.get("assetLevel")].get("Link") and option == entry.get("asset"):
                    benchmark = entry.get("benchmark")
                    if benchmark + self.buildCode([]) in struc.keys(): #pull the benchmark data in if it exists
                        temp = struc[benchmark + self.buildCode([])]
                        struc[benchmark + code] = temp.copy()
                    else:
                        struc[benchmark + code] = {} #place table space for that level selection. Will not populate if previous failed
            return struc
        def buildLevel(hier, levelName,levelIdx, struc,data,path : list, insertedOption = None, noHeader: bool = False):
            levelIdx += 1
            entryTemplate = {"dateTime" : None, "Calculation Type" : "Total " + levelName, "Node" : None, "Target name" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Source name" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None, "IRR ITD" : None}
            for header in headerOptions_local:
                if header not in ("Ownership", "IRR ITD"):
                    entryTemplate[header] = 0

            # Group data once by the current level to avoid repeated scans
            upperEntries = []
            allEntries = []
            grouped_by_level = defaultdict(list)
            if levelName not in nonFundCols:
                for _e in data:
                    grouped_by_level[fund2traitGet(_e['Target name'],{}).get(levelName,"")].append(_e)
            elif levelName == 'Source name':
                for _e in data:
                    grouped_by_level[_e[levelName]].append(_e)
            elif levelName == nameHier["Family Branch"]["local"]:
                inv2fam = self.db.investor2family
                for _e in data:
                    grouped_by_level[inv2fam.get(_e['Source name'])].append(_e)
            elif levelName == 'Node':
                nodePathOptDict = {}
                grouped_by_level = defaultdict(list)
                for _e in data: #split into highest node levels
                    nodePath = _e['nodePath']
                    baseNode = nodePath.split(nodePathSplitter)[0].strip() if nodePath not in (None,'None') else nodePath
                    grouped_by_level[baseNode].append(_e)
                    nodePathOptDict.setdefault(baseNode,set()).update(nodePath.split(nodePathSplitter))
            # Derive options from grouped keys
            options = [key for key in list(grouped_by_level.keys()) if key]
            if levelName == "subAssetClass":
                # Sort options so those in assetClass2Order come first (in the given order),
                # then anything else not in assetClass2Order appears after (sorted alphabetically)
                options = [v for v in AC2order if v in options] + sorted([v for v in options if v not in AC2order])
            elif levelName == "assetClass":
                options = [v for v in AC1order if v in options] + sorted([v for v in options if v not in AC1order])
            elif levelName == 'Node':
                def nodeSortKey(n):
                    return id2Node[int(n)] if n not in (None, 'None') else n
                options.sort(key=lambda n: nodeSortKey(n))
            else:
                options.sort()
            if len(hier) > levelIdx: #more hierarchy levels to parse
                for option in options:
                    if levelName == 'Node' and len(nodePathOptDict[option]) > 1: #Break apart to lower node levels and remove from processing in this loop
                        struc, upperEntriesExtend, allEntriesExtend = nodeRecursion(hier, levelName,levelIdx, struc,grouped_by_level[option],path, insertedOption,option)
                        upperEntries.extend(upperEntriesExtend)
                        allEntries.extend(allEntriesExtend)
                        continue

                    highEntries = {}
                    if levelName == 'Node':
                        name = id2Node.get(int(option),"Node Name Not Found") if option not in (None, 'None', ' -1 ') else "No Node"
                    else:
                        name = option 

                    tempPath = path.copy()
                    tempPath.append(name)

                    code = buildCode(tempPath)
                    if not noHeader:
                        struc[name + code] = {} #place table space for that level selection
                        if showBenchLinks:
                            struc = applyLinkedBenchmarks(struc,code, levelName, option)
                                
                                
                    #separates out only relevant data
                    levelData = grouped_by_level.get(option, [])
                    if len(sortHierarchy) > levelIdx + 1 and levelName == "subAssetClass" and sortHierarchy[levelIdx] == "subAssetSleeve" and option in self.db.fetchOptions("asset3Visibility").keys():
                        #will skip the subAssetSleeve for hidden ones and send the entire section of data to the next level
                        tempPath.append("hiddenLayer")
                        struc, lowTotals, fullEntries = buildLevel(hier, hier[levelIdx + 1],levelIdx + 1,struc,levelData,tempPath)
                    else:
                        struc, lowTotals, fullEntries = buildLevel(hier, hier[levelIdx],levelIdx,struc,levelData,tempPath, insertedOption = option)
                    allEntries.extend(fullEntries)
                    for total in lowTotals:
                        dt =  total['dateTime']
                        if dt not in highEntries.keys():
                            highEntries[dt] = copy.deepcopy(entryTemplate)
                            highEdT = highEntries[dt]
                            highEdT["rowKey"] = name + code
                            highEdT["dateTime"] = dt
                            for label in dataOptions_local: #instantiates basic string values
                                highEdT[label] = total[label]
                            if levelName not in ("Source name","Family Branch"):
                                highEdT[levelName] = total[levelName]
                                if levelName == "subAssetClass":
                                    highEdT["assetClass"] = total["assetClass"]
                        else:
                            highEdT = highEntries[dt]
                        if not highEdT.get("ownershipAdjust", False) and total.get('ownershipAdjust',False):
                            highEdT["ownershipAdjust"] = True
                        for header in headerOptions_local:
                            if header not in ("Ownership", "IRR ITD"):
                                highEdT[header] += float(total[header])
                            elif header == "Ownership" and levelName in nonFundCols and total.get(header) not in (None,"None","",0) and "Node" in sortHierarchy[:levelIdx]:
                                #allows aggregation of ownership if above level is parent investor or overall pool
                                if highEdT.get(header) is None:
                                    highEdT[header] = float(total[header]) #initialize
                                else:
                                    highEdT[header] += float(total[header]) #aggregate pool ownerships
                    for month in highEntries.keys():
                        highMonth = highEntries[month]
                        mdDen = highMonth['MDdenominator']
                        gain = highMonth["Monthly Gain"]
                        highMonth["Return"] = abs(gain / mdDen * 100) * findSign(gain) if mdDen != 0 else 0
                        highMonth = calc_DPI_TVPI(highMonth)
                    upperEntries.extend([hEntry for _,hEntry in highEntries.items()])
                if not noHeader:
                    allEntries.extend(upperEntries)       
                #high totals: all totals for the exact level
                #newTotalEntries: all totals for every level being tracked
                return struc, upperEntries, allEntries
            else: #occurs at level of fund parent
                if levelName == "subAssetSleeve" and sortHierarchy[levelIdx - 2] == "subAssetClass" and insertedOption in self.db.fetchOptions("asset3Visibility").keys():
                    options = ["hiddenLayer"]
                NAVsort = "NAV" in self.sortStyle.text()
                for option in options:
                    if levelName == 'Node' and len(nodePathOptDict[option]) > 1: #Break apart to lower node levels and remove from processing in this loop
                        struc, upperEntriesExtend, allEntriesExtend = nodeRecursion(hier, levelName,levelIdx, struc,grouped_by_level[option],path, insertedOption,option)
                        upperEntries.extend(upperEntriesExtend)
                        allEntries.extend(allEntriesExtend)
                        continue
                    totalEntriesLow = {}
                    if levelName == 'Node':
                        name = id2Node.get(int(option),"Node Name Not Found") if option not in (None, 'None') else "No Node"
                    else:
                        name = option 
                    code = buildCode([*path,name])
                    if option != "hiddenLayer":
                        if not noHeader:
                            struc[name + code] = {} #place table space for that level selection
                            if showBenchLinks:
                                struc = applyLinkedBenchmarks(struc,code, levelName, option)
                        levelData = grouped_by_level.get(option, []) #separates out only relevant data
                    else:
                        levelData = data #dont filter the data for hidden layer
                    nameList = {} #used for sorting by descending NAV
                    investorsAccessed = {}
                    lowestAggregates = defaultdict(dict)
                    for entry in levelData:
                        dt = entry['dateTime']
                        target_raw = entry["Target name"]
                        if not consolidateFunds or target_raw not in consolidatedFunds_map or target_raw in filteredTargets:
                            targetName = target_raw
                        else:
                            targetName = consolidatedFunds_map.get(target_raw)
                        targetTraitGet = fund2traitGet(target_raw,{}).get
                        name_key = targetName + code
                        nameList[name_key] = nameList.get(name_key, 0.0)
                        if NAVsort and end_period_dt is not None and datetime.strptime(dt, "%Y-%m-%d %H:%M:%S") == end_period_dt:
                            #store list of names and NAVs to sort by descending NAV
                            nav_val = entry.get("NAV", 0.0)
                            nameList[name_key] += float(nav_val) if nav_val not in (None,"None","" ,0) else 0.0
                        temp = entry.copy()
                        temp["rowKey"] = name_key
                        temp["Calculation Type"] = "Total Target name"
                        if name_key not in lowestAggregates[dt]: #note: aggregating like this will have datapoints with random investor attached if not sourced by investor. Irrelevant when made
                            lowAgDt = lowestAggregates[dt]
                            lowAgDt[name_key] = {k : v for k,v in temp.items() if v not in (None,'None','')}
                            for header in (h for h in headerOptions_local if h in lowAgDt[name_key]):
                                lowAgDt[name_key][header] = float(temp[header]) #convert any nums to float
                        else:
                            lowAgDict = lowestAggregates[dt][name_key]
                            lowAgDict['ownershipAdjust'] = lowAgDict.get('ownershipAdjust',False) or temp.get('ownershipAdjust',False)
                            for h in (h for h in headerOptions_local if temp.get(h) not in (None,'None','',0)):
                                if h not in lowAgDict:
                                    lowAgDict[h] = float(temp[h])
                                elif h not in ('IRR ITD','Ownership'):
                                    lowAgDict[h] += float(temp[h])
                                elif h == 'Ownership':
                                    if lowAgDict[h] == 0: #if no ownership, assume empty fund, replace. Replace is for empty funds in a consolidated fund
                                        lowAgDict['Target name'] = temp['Target name']
                                        lowAgDict[h] = float(temp[h])
                                    elif lowAgDict['Target name'] == temp['Target name']: #only aggregate within the same target. All within consolidated should be equal
                                        lowAgDict[h] += float(temp[h])
                                elif temp.get(h) != 0: #save options to make the median later
                                    if isinstance(lowAgDict[h],list):
                                        lowAgDict[h].append(float(temp[h]))
                                    else:
                                        lowAgDict[h] = [lowAgDict[h],float(temp[h])]

                        if dt not in totalEntriesLow:
                            totalEntriesLow[dt] = copy.deepcopy(entryTemplate)
                            totalLowDt = totalEntriesLow[dt]
                            totalLowDt["rowKey"] = name + code
                            totalLowDt["dateTime"] = dt
                            for label in dataOptions_local:
                                totalLowDt[label] = targetTraitGet(label,"")
                            if levelName not in ("Source name","Family Branch"):
                                lvl_val = targetTraitGet(levelName,"")
                                totalLowDt[levelName] = lvl_val
                                if levelName == "subAssetClass":
                                    totalLowDt["assetClass"] = targetTraitGet("assetClass","")
                        else:
                            totalLowDt = totalEntriesLow[dt]
                        if not totalLowDt.get("ownershipAdjust", False) and entry.get('ownershipAdjust',False):
                            totalLowDt["ownershipAdjust"] = True
                        for header in headerOptions_local:
                            v = entry.get(header)
                            if header not in ("Ownership", "IRR ITD") and v not in (None,"None",""):
                                totalLowDt[header] += float(v)
                            elif header == "Ownership" and levelName in ("Source name", "Family Branch", 'Node') and "Node" in sortHierarchy and v not in (None,"None","") and float(v) != 0:
                                investor = entry.get("Source name")
                                if totalLowDt.get(header) is None:
                                    totalLowDt[header] = float(v) #assign investor to ownership based on fund
                                    investorsAccessed[dt] = {investor}
                                else:
                                    accessed = investorsAccessed.get(dt, set())
                                    if investor not in accessed: #accounts for family branch level to add the investor level ownerships
                                        totalLowDt[header] += float(v)
                                        accessed.add(investor)
                                        investorsAccessed[dt] = accessed
                    for dt in lowestAggregates:
                        for _, entry in lowestAggregates[dt].items():
                            gain = entry['Monthly Gain']
                            MDden = entry['MDdenominator']
                            entry['Return'] = abs(gain / MDden * 100) * findSign(gain) if MDden != 0 else 0
                            if 'IRR ITD' in entry and isinstance(entry['IRR ITD'],list):
                                entry['IRR ITD'] = statistics.median(entry['IRR ITD'])
                            entry = calc_DPI_TVPI(entry)
                        allEntries.extend([v for _,v in lowestAggregates[dt].items()])
                        
                    if not NAVsort:
                        for name in sorted(nameList.keys()): #sort by alphabetical order
                            struc[name] = {}
                    else:
                        for name in descendingNavSort(nameList): #sort by descending NAV
                            struc[name] = {}
                    for month in totalEntriesLow.keys():
                        e = totalEntriesLow[month]
                        gain = e["Monthly Gain"]
                        MDden = e["MDdenominator"]
                        e["Return"] = abs(gain / MDden * 100) * findSign(gain) if MDden != 0 else 0
                        e = calc_DPI_TVPI(e)
                        
                    upperEntries.extend(totalEntriesLow.values())
                if not noHeader:
                    allEntries.extend(upperEntries)
                return struc, upperEntries, allEntries
        def nodeRecursion(hier, levelName,levelIdx, struc,data,path, insertedOption,baseNodeId):
            baseNode = id2Node[int(baseNodeId)]
            rowKey = baseNode + buildCode(path)
            struc[rowKey] = {}
            nodeSumEntryDict: dict[dict] = {} #datetime : params: vals
            baseNodeData = []
            lowNodeData = []
            for _e in data:
                temp = _e.copy()
                nPath = temp['nodePath']
                if nPath.strip() == baseNodeId:
                    baseNodeData.append(temp)
                else: #send back through the recusion with one lower level of nodePath. Continues until it reaches the lowest node
                    lowerPath = nodePathSplitter.join(part.strip() for part in nPath.split(nodePathSplitter)[1:])
                    temp['nodePath'] = lowerPath
                    lowNodeData.append(temp)
            #send both types back through buildLevel. Isolated node data will build as normal. data to be split more will appear back here and split again
            #TODO: base data needs a false heading to sum at. Lower nodes will sum at their node headings
            newHier = [*hier[:levelIdx], 'Node', *hier[levelIdx:]]
            tempPath = path.copy()
            tempPath.append(baseNode)
            struc, baseUpper, baseAllData = buildLevel(newHier,newHier[levelIdx],levelIdx,struc,baseNodeData,tempPath,insertedOption, noHeader = True)
            struc, recursUpper, recursAllData = buildLevel(newHier,newHier[levelIdx],levelIdx,struc,lowNodeData,tempPath,insertedOption)
            entryTemplate = {"dateTime" : None, "Calculation Type" : "Total " + levelName, "Node" : baseNode, "Target name" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Source name" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None, "IRR ITD" : None, 'rowKey' : rowKey}
            for header in (header for header in headerOptions_local if header not in ("Ownership", "IRR ITD")):
                entryTemplate[header] = 0.0
            for _e in (*baseUpper,*recursUpper):
                _eGet = _e.get
                dt = _e['dateTime']
                if dt not in nodeSumEntryDict.keys():
                    nodeSumEntryDict[dt] = copy.deepcopy(entryTemplate)
                    nodeDt = nodeSumEntryDict[dt]
                    nodeDt["dateTime"] = dt
                    for label in dataOptions_local: #instantiates basic string values
                        nodeDt[label] = _e[label]
                    if levelName not in ("Source name","Family Branch"):
                        val =  _e[levelName]
                        nodeDt[levelName] = val
                        if levelName == "subAssetClass":
                            aCval = _e['assetClass']
                            nodeDt["assetClass"] = aCval
                else:
                    nodeDt = nodeSumEntryDict[dt]
                if not nodeDt.get("ownershipAdjust", False) and _e.get('ownershipAdjust',False):
                    nodeDt["ownershipAdjust"] = True
                for header in headerOptions_local:
                    if header not in ("Ownership", "IRR ITD"):
                        nodeDt[header] += float(_e[header])
                    elif header == "Ownership" and levelName in nonFundCols and _eGet(header) not in (None,"None","",0) and "Node" in sortHierarchy[:levelIdx]:
                        #allows aggregation of ownership if above level is parent investor or overall pool
                        if nodeDt.get(header) is None:
                            nodeDt[header] = float(_e[header]) #initialize
                        else:
                            nodeDt[header] += float(_e[header]) #aggregate pool ownerships
            for month in nodeSumEntryDict:
                nodeMonth = nodeSumEntryDict[month]
                mdDen = nodeMonth['MDdenominator']
                gain = nodeMonth['Monthly Gain']
                nodeMonth['Return'] =  abs(gain / mdDen * 100) * findSign(gain) if mdDen != 0 else 0.0
            nodeUpperEntries = nodeSumEntryDict.values()
            return struc, nodeUpperEntries, [*baseAllData, *recursAllData]

        if self.showBenchmarkLinksBtn.isChecked():
            benchmarkLinks = self.db.fetchBenchmarkLinks()
            tableStructure = applyLinkedBenchmarks(tableStructure,self.buildCode(["Total",]), "Total", "Total") #apply benchmark links to total
        levelIdx = 0
        buildHier = sortHierarchy
        tableStructure, highestEntries, newEntries = buildLevel(buildHier, buildHier[0],levelIdx,tableStructure,data, [])
        trueTotalEntries = {}
        for total in highestEntries:
            if total["dateTime"] not in trueTotalEntries.keys():
                trueTotalEntries[total["dateTime"]] = {"dateTime" : None, "Calculation Type" : "Total", "Node" : None, "Fund" : None ,
                                            "assetClass" : None, "subAssetClass" : None, "Source name" : None,
                                            "Return" : None , nameHier["sleeve"]["local"] : None,
                                            "Ownership" : None, "IRR ITD" : None}
                trueTotalEntries[total["dateTime"]]["rowKey"] = "Total" + self.buildCode([])
                for header in headerOptions_local:
                    if header != "Ownership":
                        trueTotalEntries[total["dateTime"]][header] = 0
                for label in (*dataOptions_local,'dateTime'):
                    trueTotalEntries[total["dateTime"]][label] = total[label]
            if not trueTotalEntries[total["dateTime"]].get("ownershipAdjust", False) and total.get('ownershipAdjust',False):
                trueTotalEntries[total["dateTime"]]["ownershipAdjust"] = True
            for header in headerOptions_local:
                if header not in ("Ownership", "IRR ITD"):
                    trueTotalEntries[total["dateTime"]][header] += float(total[header])
        for month in trueTotalEntries.keys():
            gain = trueTotalEntries[month]["Monthly Gain"]
            MDden = trueTotalEntries[month]["MDdenominator"]
            trueTotalEntries[month]["Return"] = abs(gain / MDden * 100) * findSign(gain) if MDden != 0 else 0
            newEntries.append(trueTotalEntries[month])
        return tableStructure,newEntries
                    
    def filterUpdate(self):
        self.buildReturnTable()
        return
        from functools import partial

        self.buildTableLoadingBar.setValue(0)
        self.tableLoadingLabel.setText("Waiting on database connection...")
        self.buildTableLoadingBox.setVisible(True)

        if self.filterCallLock:
            return

        def resetOptions(key, new_options):
            multiBox = self.filterDict[key]
            old_options = set(multiBox._checkboxes.keys())
            new_options_set = set(new_options)

            if old_options != new_options_set:
                currentSelections = multiBox.checkedItems()
                multiBox.clearItems()
                multiBox.addItems(sorted(new_options))
                for text in currentSelections:
                    if text in new_options_set:
                        multiBox.setCheckedItem(text)

        def exitFunc():
            self.filterCallLock = False
            gui_queue.put(lambda: self.tableLoadingLabel.setText("Building returns table..."))
            gui_queue.put(lambda: self.buildReturnTable())

        def processFilter():
            try:
                self.filterCallLock = True

                # Get fund2trait mapping
                fund2trait = self.db.fetchFund2Trait()

                # Get current selections for all filters (except highOnlyFilters)
                currentChoices = {
                    key: self.filterDict[key].checkedItems()
                    for key in self.filterDict
                    if key not in self.highOnlyFilters
                }

                # If no filters are selected, reset all to full options
                if all(not choices for choices in currentChoices.values()):
                    gui_queue.put(self.instantiateFilters)
                    exitFunc()
                    return

                # Process each filter to determine available options
                for targetFilter in self.filterOptions:
                    targetKey = targetFilter["key"]
                    if targetKey in self.highOnlyFilters:
                        continue

                    # Get selected values for all OTHER filters (excluding the target filter)
                    other_selections = {
                        key: selections
                        for key, selections in currentChoices.items()
                        if key != targetKey and selections
                    }

                    # Filter funds based on other filter selections
                    filtered_funds = []
                    for fund_name, traits in fund2trait.items():
                        # Check if this fund matches all other filter selections
                        matches = True
                        for filter_key, selected_values in other_selections.items():
                            # Only check filters that exist in fund2trait (skip Node and other non-fund columns)
                            if filter_key in traits:
                                fund_value = traits.get(filter_key, "")
                                if fund_value not in selected_values:
                                    matches = False
                                    break
                        if matches:
                            filtered_funds.append((fund_name, traits))

                    # Collect available options for the target filter from filtered funds
                    available_options = set()
                    
                    if targetKey == "Target name":
                        # Special case: "Target name" filter uses fund names directly
                        available_options = {fund_name for fund_name, _ in filtered_funds}
                    else:
                        # For other filters, get values from traits
                        for fund_name, traits in filtered_funds:
                            trait_value = traits.get(targetKey)
                            if trait_value is not None and trait_value not in (None, "", "None"):
                                available_options.add(trait_value)

                    # Update the target filter with available options
                    gui_queue.put(partial(resetOptions, targetKey, sorted(available_options)))

            except Exception as e:
                gui_queue.put(lambda: QMessageBox.warning(self, "Filter Error", f"Error occurred updating filters:\n{e}"))

            exitFunc()

        executor.submit(processFilter)
    def ACsortChange(self,lvl : int):
        AC = 'assetClass' if lvl == 1 else 'subAssetClass'
        #Update database
        newOrder = self.AC1sort.popup.get_checked_sorted_items() if lvl == 1 else self.AC2sort.popup.get_checked_sorted_items()
        if newOrder: #only if choices are made
            newOpts = [{'id' : item, 'value' : idx} for idx, item in enumerate(newOrder)]
            self.db.saveNewOptions(f'{AC}_sort',newOpts)

            #update the table
            sortHier = self.sortHierarchy.checkedItems()
            if AC in sortHier:
                self.buildReturnTable() #if a relevant sort item was changed, rebuild the table

        

    def assetClass3VisibilityChanged(self):
        hiddenItems = self.assetClass3Visibility.checkedItems()
        self.db.saveAsset3Visibility(hiddenItems)
        self.buildReturnTable()
    def updateMonths(self):
        start = self.dataTimeStart
        end = datetime.now()
        index = start
        monthList = []
        while index < end:
            monthList.append(index)
            index += relativedelta(months=1)
        dbDates = []
        for monthDT in monthList:
            month = int(monthDT.month)
            year = int(monthDT.year)
            lastDayCurrent = calendar.monthrange(int(year),month)[1]
            lastDayCurrent   = str(lastDayCurrent).zfill(2)
            if month - 1 > 0:
                prevMonth =  month - 1
                prevMyear = year
            else:
                prevMonth = 12
                prevMyear = str(int(year) - 1)
            lastDayPrev = calendar.monthrange(int(prevMyear),prevMonth)[1]
            lastDayPrev   = str(lastDayPrev).zfill(2)
            prevMonth = str(prevMonth).zfill(2)
            month = str(month).zfill(2)
            
            tranStart = f"{year}-{month}-01T00:00:00.000Z"
            bothEnd = f"{year}-{month}-{lastDayCurrent}T00:00:00.000Z"
            accountStart = f"{prevMyear}-{prevMonth}-{lastDayPrev}T00:00:00.000Z"

            
            dateString = monthDT.strftime("%B %Y")

            monthEntry = {"dateTime" : monthDT, "Month" : dateString, "tranStart" : tranStart.removesuffix(".000Z"), "endDay" : bothEnd.removesuffix(".000Z"), "accountStart" : accountStart.removesuffix(".000Z")}
            dbDates.append(monthEntry)
        save_to_db(self.db,"Months",dbDates)
    def instantiateFilters(self,*_, keepChoices: bool = False):
        if keepChoices:
            CBs = list(self.filterDict.values())
            CBs.extend([self.assetClass3Visibility,self.benchmarkSelection])
            choices = [cb.checkedItems() for cb in CBs]
        investors = self.db.fetchInvestors()
        for filKey, invKey in (['Source name', 'Name'],['Family Branch','Parentinvestor']):
            self.filterDict[filKey].clearItems()
            self.filterDict[filKey].addItems(sorted(set((inv[invKey] for inv in investors))))
        filOptDict = {filKey : set() for filKey in (fil['key'] for fil in self.filterOptions if fil['key'] not in nonFundCols)}
        funds = self.db.fetchFunds()
        dyn2key = self.db.fetchDyn2Key()
        for fund in funds:
            for field, val in ([f,v] for f,v in fund.items() if dyn2key.get(f,"") in filOptDict):
                filOptDict[dyn2key[field]].add(val)
        for filKey in filOptDict.keys():
            self.filterDict[filKey].clearItems()
            self.filterDict[filKey].addItems(sorted(filOptDict[filKey]))
        self.filterDict['Node'].clearItems()
        self.filterDict['Node'].addItems([str(n) for n in list(self.db.pullId2Node().keys())])
        self.assetClass3Visibility.addItems(sorted(filOptDict['subAssetClass']))
        self.assetClass3Visibility.setCheckedItems([key for key, val in self.db.fetchOptions('asset3Visibility').items() if val == 'hide'])
        self.benchmarkSelection.addItems(sorted(self.db.fetchBenchmarks()))
        
        AC1order = self.db.fetchACorder(1)
        AC1opts = [*AC1order,*[opt for opt in filOptDict['assetClass'] if opt not in AC1order]]
        AC2order = self.db.fetchACorder(2)
        AC2opts = [*AC2order,*[opt for opt in filOptDict['subAssetClass'] if opt not in AC2order]]

        self.AC1sort.set_items(AC1opts,AC1order)
        self.AC2sort.set_items(AC2opts,AC2order)
        if not keepChoices: #defaults
            self.filterDict["Classification"].setCheckedItem("HFC")
        else: #set back to previous values
            for idx, cb in enumerate(CBs):
                cb.setCheckedItems(choices[idx])


    def groupingChange(self):
        groupOpts = self.sortHierarchy.checkedItems()
        if groupOpts == []:
            self.sortHierarchy.setCheckedItems(["assetClass","subAssetClass"])
        self.filterCallLock = True
            
            
        self.filterCallLock = False
        self.buildReturnTable()
    def testAPIconnection(self, key=None):
        apiKey = self.api_key if key is None else key
        headers = {
            "Authorization": f"Bearer {apiKey}",
            "Content-Type":  "application/json"
        }
        payload = {
            "advf": [{ "_name": "Fund" }],
            "mode": "compact",
            "page": {"size": 0}
        }
        resp = requests.get(f"{mainURL}/Entity", headers=headers, json=payload)
        if resp.status_code == 200:
            return True
        else:
            return False
    def check_api_key(self, *_):
        key = self.api_input.text().strip()
        if key:
            if self.testAPIconnection(key=key):
                self.api_label.setText('API key valid. Saving to system...')
                subprocess.run(['setx',dynamoAPIenvName,key], check=True)
                os.environ[dynamoAPIenvName] = key
                self.api_key = key
                self.stack.setCurrentIndex(1)
                self.init_data_processing()
            else:
                self.api_label.setText('Invalid API key or Dynamo is not responding')
        else:
            self.api_label.setText('API key cannot be empty')

    def show_results(self,*_):
        self.stack.setCurrentIndex(2)

    def pullData(self):
        if not self.testAPIconnection():
            gui_queue.put(lambda: QMessageBox.warning(self,"API Failure", "API connection has failed. Server is down or API key is bad. \n Previous calculations are left in place for viewing."))
            return
        def checkNewestData(table, rows, nodes : list[str], sources, targets):
            #iterate through the freshly imported rows, check if they match with the previous data. 
            #inputs: table name, rows of newly imported data
            #outputs: newImportedRows, oldDatabaseRows, self.earliestChangeDate is updated if a new earliest change date is found
            earliest = None
            if fullRecalculations: #use all old data, and track the earliest data of entry
                for rec in rows:
                    rowNodes = [node for node in nodes if node in (rec.get("Target name"), rec.get("Source name"))] #nodes that the entry connects to
                    if not rowNodes and rec.get('Target name') in targets and rec.get('Source name') in sources:
                        rowNodes = ['noNodeData',]
                    elif not rowNodes:
                        print(f"Warning: no nodes or direct investment found attached to a datapoint: {rec}")
                    dt = datetime.strptime(rec['Date'], "%Y-%m-%dT%H:%M:%S")
                    if earliest is None or dt < earliest: #sets overall values to earliest
                        earliest = dt.replace(day=1)
                    with earlyChangeDateLock:
                        for node in rowNodes:
                            if dt < self.nodeChangeDates.get(node,datetime.now()): 
                                self.nodeChangeDates[node] =  dt.replace(day=1) # sets each pool value to earliest and instantiates if not existing
                self.earliestChangeDate = min([dt for key, dt in self.nodeChangeDates.items() if key != 'active'])
                return {'old' : [], 'new' : rows}
            def buildKey(record): #TODO: check for transactions of the same value in the same source to target. Could ignore new ones
                value = record[nameHier["Value"]["dynHigh"] if table == "positions" else nameHier["CashFlow"]["dynLow"]]
                value = 0 if value is None or value == "None" else value
                key = (
                        record['Source name'] if record['Source name'] is not None else "None",
                        record['Target name'] if record['Target name'] is not None else "None",
                        round(float(value)) if table != "positions" else 0,               # normalize to float
                        record['Date'].replace(' ', 'T')      # normalize format if needed
                    )
                return key
            try:
                diffCount = 0
                differences = []
                previous = load_from_db(self.db,table) or []

                # Build a set of tuple‐keys for the old data
                oldRecords = set()
                for rec in previous:
                    oldRecords.add(buildKey(rec))

                newRecords = set()
                earliest = None
                for rec in rows:
                    rowNodes = [node for node in nodes if node in (rec.get("Target name"), rec.get("Source name"))] #nodes that the entry connects to
                    if not rowNodes and rec.get('Target name') in targets and rec.get('Source name') in sources:
                        rowNodes = ['noNodeData',]
                    elif not rowNodes:
                        print(f"Warning: no nodes or direct investment found attached to a datapoint: {rec}")
                    value = rec[nameHier["Value"]["dynHigh"] if "positions" in table else nameHier["CashFlow"]["dynLow"]]
                    value = 0 if value is None or value == "None" else value
                    key = buildKey(rec)
                    newRecords.add(key)
                    if table == "positions": #updates new data to have required fields
                        rec[nameHier["Unfunded"]["local"]] = 0
                        rec[nameHier["Commitment"]["local"]] = 0
                    if key in oldRecords:
                        continue
                    diffCount += 1
                    differences.append(rec)
                    differences.append({"Source name" : key[0],"Target name" : key[1],nameHier["Value"]["dynLow"] : key[2],"Date" : key[3]})
                    # parse the date for comparison
                    dt = datetime.strptime(rec['Date'], "%Y-%m-%dT%H:%M:%S")
                    if earliest is None or dt < earliest: #sets overall values to earliest
                        earliest = dt.replace(day=1)
                    with earlyChangeDateLock:
                        for node in rowNodes:
                            if dt < self.nodeChangeDates.get(node,datetime.now()): 
                                self.nodeChangeDates[node] =  dt.replace(day=1) # sets each pool value to earliest and instantiates if not existing
                for oldRec in oldRecords:
                    #find if a new record no longer exists in the old. Means old data is altered and must be redone from that timeframe
                    if oldRec not in newRecords: 
                        dt = datetime.strptime(oldRec[3], "%Y-%m-%dT%H:%M:%S")
                        if earliest is None or dt < earliest:
                            earliest = dt.replace(day=1)

                with earlyChangeDateLock:
                    if earliest and earliest < self.earliestChangeDate:
                        self.earliestChangeDate = earliest
                print(f"Differences in {table} : {diffCount} of {len(rows)}")
                if diffCount > 0 and not demoMode:
                    def openWindow():
                        window = tableWindow(parentSource=self,all_rows=differences,table=table)
                        self.tableWindows[table] = window
                        window.show()
                    gui_queue.put(lambda: openWindow())
                return {"old": previous, "new": rows}
            except Exception as e:
                print(traceback.format_exc())
                print(f"Error searching old data: {e}")
        try:
            self.earliestChangeDate = datetime.now() + relativedelta(months=1)
            earlyChangeDateLock = threading.Lock()
            gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(True))
            gui_queue.put(lambda: self.importButton.setEnabled(False))
            self.updateMonths()
            completeLock = threading.Lock()
            self.apiFutures = set()
            self.complete = float(0)
            totalCalls = float(6)
            importedTables = {}
            apiData = {
                "tranCols": "Investment in, Investing Entity, Transaction Type, Effective date, Remaining commitment change, Transaction timing, Cash flow change (USD), ValueInSystemCurrency, HF Cash Flow Type",
                "tranName": "InvestmentTransaction",
                "tranSort": "Effective date:desc",
                "accountCols": "As of Date, Balance Type, Investing entity, Investment in, Value in system currency, Fund class, Sub-account",
                "accountName": "InvestmentPosition",
                "accountSort": "As of Date:desc",
                "fundCols" : "Parent fund, Fund Name, Fund Pipeline Status, Asset class category, HF Classification, HF sub-classification",
                "secCols" : "Parent Security, Security Name, Asset class category, HF Classification, HF Sub-Classification",
                "investorCols" : "Parent investor, Account name",
                "InvestorName" : "",
                "benchCols" : (f"Index, As of date, MTD %, QTD %, YTD %, ITD cumulative %, ITD TWRR %, "
                               f"{', '.join(f'Last {y} yr %' for y in yearOptions)}"), 
            }
            calculationsTest = load_from_db(self.db,"calculations")
            if calculationsTest != []:
                skipCalculations = True
                self.nodeChangeDates = {"active" : True}
                self.foundRetroChange = False
            else:
                skipCalculations = False
            accountTranTableFutures = []
            #key for the table naming convention {i : {j : table name}}
            def apiHeader(cols, sort = None):
                header = {
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json",
                    "x-columns": cols,
                }
                if sort:
                    header["x-sort"] : sort
                return header
            positionsPayload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "InvestmentPosition",
                                                "rule": [
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investing entity"
                                                    },
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investment in"
                                                    },
                                                    {
                                                        "_op": "is_null",
                                                        "_prop": "Holding"
                                                    }
                                                ]
                                            },
                                            {
                                                "_name": "InvestmentPosition",
                                                "e": [
                                                    {
                                                        "_name": "Security",
                                                        "rule": [
                                                            {
                                                                "_op": "is_null",
                                                                "_prop": "HoldingsInsightID"
                                                            },
                                                            {
                                                                "_op": "is",
                                                                "_prop": "Security status",
                                                                "values": [
                                                                    {
                                                                        "id": "71d185fd-d891-4021-be50-179ad11ec21f",
                                                                        "es": "L_SecurityStatus",
                                                                        "name": "Portfolio"
                                                                    }
                                                                ]
                                                            }
                                                        ]
                                                    }
                                                ]
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
            transactionsPayload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "InvestmentTransaction",
                                                "rule": [
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investing entity"
                                                    },
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investment in"
                                                    },
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Cash flow change (USD)"
                                                    },
                                                    {
                                                        "_op": "is_null",
                                                        "_prop": "Cash flow model"
                                                    },
                                                    {
                                                        "_op": "is_null",
                                                        "_prop": "HoldingsInsightID"
                                                    }
                                                ]
                                            },
                                            {
                                                "_name": "InvestmentTransaction",
                                                "rule": [
                                                    {
                                                        "_op": "any_item",
                                                        "_prop": "Transaction type",
                                                        "values": [
                                                            [
                                                                {
                                                                    "id": "5327639c-8160-4d85-9b23-8c6bf60c5406",
                                                                    "es": "L_TransactionType",
                                                                    "name": "Commitment"
                                                                },
                                                                {
                                                                    "id": "37339e7c-1c24-4d13-9d17-86d0efe079b3",
                                                                    "es": "L_TransactionType",
                                                                    "name": "Transfer of commitment"
                                                                },
                                                                {
                                                                    "id": "0f8f8671-8579-49d7-b604-05300b6a3990",
                                                                    "es": "L_TransactionType",
                                                                    "name": "Transfer of commitment (out)"
                                                                },
                                                                {
                                                                    "id": "5e098d83-70b0-4135-a629-aff19048fb1c",
                                                                    "es": "L_TransactionType",
                                                                    "name": "Secondary - Original commitment (by secondary seller)"
                                                                }
                                                            ]
                                                        ]
                                                    },
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investment in"
                                                    },
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Investing entity"
                                                    },
                                                    {
                                                        "_op": "is_null",
                                                        "_prop": "Cash flow model"
                                                    },
                                                    {
                                                        "_op": "is_null",
                                                        "_prop": "HoldingsInsightID"
                                                    },
                                                    {
                                                        "_op": "not_null",
                                                        "_prop": "Amount in system currency"
                                                    }
                                                ]
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
            def bgPullData(tableName, payload, headers):
                    rows = []
                    idx = 0
                    while rows in ([],None) and idx < 3: #if call fails, tries again
                        idx += 1
                        response = requests.post(f"{mainURL}/Search", headers=headers, data=json.dumps(payload))
                        if response.status_code == 200:
                            try:
                                data = response.json()
                            except ValueError:
                                continue
                            if isinstance(data, dict):
                                rows = data.get('data', data.get('rows', []))
                            elif isinstance(data, list):
                                rows = data
                            else:
                                rows = []

                            keys_to_remove = {'_id', '_es'}
                            rows = [
                                {k: v for k, v in row.items() if k not in keys_to_remove}
                                for row in rows
                            ]
                        else:
                            print(f"Error in API call. Code: {response.status_code}. {response}")
                            try:
                                print(f"Error: {response.json()}")
                            except:
                                pass
                    if len(rows) == 0: #prevents bad calculations from missing data. Appears if partial re-calculation but new data is corrupted
                        raise RuntimeError("API import did not function properly. Try again.")
                    targets, sources, nodes = nodeLibrary.findNodes(None,rows)
                    tables = checkNewestData(tableName,rows, nodes, sources, targets)
                    with completeLock:
                        self.complete += 1
                    frac = self.complete/totalCalls
                    gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
                    return tableName,tables
            try:
                accountTranTableFutures.append(APIexecutor.submit(bgPullData,'transactions',transactionsPayload,apiHeader(apiData["tranCols"], apiData["tranSort"])))
                accountTranTableFutures.append(APIexecutor.submit(bgPullData,'positions',positionsPayload,apiHeader(apiData["accountCols"], apiData["accountSort"])))
            except RuntimeError:
                raise
            except Exception as e:
                print(f"Failure to run background thread API call: {e} \n {e.args}")
            fundPayload = {
                            "advf": {
                                "e": [
                                    {
                                        "_name": "Fund",
                                        "rule": [
                                            {
                                                "_op": "is",
                                                "_prop": "Fund Pipeline Status",
                                                "values": [
                                                    {
                                                        "id": "a6d25ed8-4027-4642-8fb7-710b01a213f7",
                                                        "es": "L_FundPipelineStatus",
                                                        "name": "P - Portfolio"
                                                    }
                                                ]
                                            }
                                        ]
                                    }
                                ]
                            },
                            "mode": "compact"
                        }
            secPayload = {
                            "advf": {
                                "e": [
                                    {
                                        "_name": "Security",
                                        "rule": [
                                            {
                                                "_op": "is_null",
                                                "_prop": "HoldingsInsightID"
                                            },
                                            {
                                                "_op": "is",
                                                "_prop": "Security status",
                                                "values": [
                                                    {
                                                        "id": "71d185fd-d891-4021-be50-179ad11ec21f",
                                                        "es": "L_SecurityStatus",
                                                        "name": "Portfolio"
                                                    }
                                                ]
                                            }
                                        ]
                                    }
                                ]
                            },
                            "mode": "compact"
                        }
            def bgFundSecPull(sec: bool = False):
                try:
                    if not sec:
                        response = requests.post(f"{mainURL}/Search", headers=apiHeader(apiData["fundCols"]), data=json.dumps(fundPayload))
                    else:
                        response = requests.post(f"{mainURL}/Search", headers=apiHeader(apiData["secCols"]), data=json.dumps(secPayload))
                except Exception as e:
                    print(f"Fund api call failed: {e.args}")
                if response.status_code == 200:
                    try:
                        tableName = "funds" if not sec else "securities"
                        data = response.json()
                        if isinstance(data, dict):
                            rows = data.get('data', data.get('rows', []))
                        elif isinstance(data, list):
                            rows = data
                        else:
                            rows = []
                        keys_to_remove = {'_id', '_es'}
                        rows = [{k: v for k, v in row.items() if k not in keys_to_remove} for row in rows]
                        for idx, row in enumerate(rows): #find sleeve values and consolidated funds
                            assetCat = row["ExposureAssetClassCategory"]
                            if assetCat is not None and assetCat.count(" > ") == 3:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = assetCat.split(" > ")[2]
                                sleeve = assetCat.split(" > ")[3]
                            elif assetCat is not None and assetCat.count(" > ") == 2:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = assetCat.split(" > ")[2]
                                sleeve = None
                            elif assetCat is not None and assetCat.count(" > ") == 1:
                                assetClass = assetCat.split(" > ")[1]
                                subAssetClass = None
                                sleeve = None
                            else:
                                assetClass = None
                                subAssetClass = None
                                sleeve = None
                            rows[idx][nameHier["sleeve"]["sleeve"]] =  sleeve
                            rows[idx]["assetClass"] = assetClass
                            rows[idx]["subAssetClass"] = subAssetClass
                        if rows != []:
                            save_to_db(self.db,tableName,rows)
                        else:
                            print(f"Warning: No {tableName} found from API pull")
                    except Exception as e:
                        print(f"Error proccessing {tableName} API data : {e} {e.args}.  {traceback.format_exc()}")
                    
                else:
                    print(f"Error in API call for {tableName}. Code: {response.status_code}. {response}. {traceback.format_exc()}")
                with completeLock:
                    self.complete += 1
                frac = self.complete/totalCalls
                gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
            submitAPIcall(self,bgFundSecPull)
            submitAPIcall(self,bgFundSecPull,True)
            totalCalls -= 1
            benchmarkPayload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "IndexPerformance"
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
            investorPayload = {
                                    "advf": {
                                        "e": [
                                            {
                                                "_name": "InvestorAccount"
                                            }
                                        ]
                                    },
                                    "mode": "compact"
                                }
            def basicAPIpull(name,payload,cols,sort = None):
                response = requests.post(f"{mainURL}/Search", headers=apiHeader(cols, sort = sort), data=json.dumps(payload))
                if response.status_code == 200:
                    try:
                        data = response.json()
                        if isinstance(data, dict):
                            rows = data.get('data', data.get('rows', []))
                        elif isinstance(data, list):
                            rows = data
                        else:
                            rows = []
                        keys_to_remove = {'_id', '_es'}
                        rows = [{k: v for k, v in row.items() if k not in keys_to_remove} for row in rows]
                        save_to_db(self.db,name,rows)
                    except Exception as e:
                        print(f"Error proccessing {name} API data : {e} {e.args}.  {traceback.format_exc()}")
                    
                else:
                    print(f"Error in API call for {name}. Code: {response.status_code}. {response}. {traceback.format_exc()}")
                with completeLock:
                    self.complete += 1
                frac = self.complete/totalCalls
                gui_queue.put(lambda val = frac: self.apiLoadingBar.setValue(int(val * 100)))
            submitAPIcall(self,basicAPIpull,'benchmarks',benchmarkPayload,apiData['benchCols'])
            submitAPIcall(self,basicAPIpull,'investors',investorPayload,apiData['investorCols'])
            wait(accountTranTableFutures)
            for future in accountTranTableFutures:
                #must be careful. There are a maximum of 5 threads but there are 6 calls, and 2 are waited for after
                table, tableData = future.result()
                if not skipCalculations or fullRecalculations:
                    importedTables[table] = tableData["new"] #all calculations are from scratch anyways, so use the new data
                else:
                    mergedTable = []
                    poolTag = "Target name" if "high" in table else "Source name"
                    if not self.nodeChangeDates.get("active",False): #if inactive, use generic starting date
                        changeDate = self.earliestChangeDate
                    for rec in tableData["new"]:
                        pool = rec[poolTag]
                        if self.nodeChangeDates.get("active",False): #if active, specifiy date by pool
                            changeDate = self.nodeChangeDates.get(pool,datetime.now())
                        if changeDate < datetime.strptime(rec["Date"], "%Y-%m-%dT%H:%M:%S"): #new data past the editing date
                            mergedTable.append(rec)
                    for rec in tableData["old"]:
                        pool = rec[poolTag]
                        if self.nodeChangeDates.get("active",False): #if active, specifiy date by pool
                            changeDate = self.nodeChangeDates.get(pool,datetime.now())
                        if changeDate >= datetime.strptime(rec["Date"], "%Y-%m-%dT%H:%M:%S"): #old data before the editing date to be kept
                            mergedTable.append(rec)
                    importedTables[table] = mergedTable
            wait(self.apiFutures)
            if skipCalculations:
                print("Earliest change: ", self.earliestChangeDate)
                if self.nodeChangeDates.get("active", False):
                    print(f"Change dates by node:")
                    for node in self.nodeChangeDates:
                        print(f"        {node} : {self.nodeChangeDates.get(node)}")
            gui_queue.put(lambda: self.apiLoadingBar.setValue(100))
            executor.submit(self.db.postAPIupdate) #make sure the new api data is in the caches

            self.apiCallTime = datetime.now().strftime("%B %d, %Y @ %I:%M %p")
            self.processFunds()
            gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(False))
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(True)) #secondary early change to make it appear faster if running slow
            gui_queue.put(lambda: self.calculateReturn(importedTables))
        except RuntimeError as e:
            gui_queue.put(lambda error = e: QMessageBox.warning(self,"Error Importing Data", f"Error pulling data from dynamo: {error} , {error.args}"))
        except Exception as e:
            print(traceback.format_exc())
            trace = traceback.format_exc() if traceback.format_exc() and not demoMode else ""
            gui_queue.put(lambda error = e: QMessageBox.warning(self,"Error Importing Data", f"Error pulling data from dynamo: {error} , {error.args} \n \n {trace}"))
        gui_queue.put(lambda: self.importButton.setEnabled(True))
        gui_queue.put(lambda: self.apiLoadingBarBox.setVisible(False))
    def openTableWindow(self, rows, name = "Table"):
        window = tableWindow(parentSource=self,all_rows=rows,table=name)
        self.tableWindows[name] = window
        window.show()
    def calculateReturn(self, dynImportData : dict):
        def initalizeCalc():
            try:
                gui_queue.put(lambda: self.importButton.setEnabled(False))
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(True))
                self.updateMonths()
                print("Calculating return....")
                fundListDB = load_from_db(self.db,"funds")
                fundList = {}
                for fund in fundListDB:
                    fundList[fund["Name"]] = fund[nameHier["sleeve"]["sleeve"]]
                months = load_from_db(self.db,"Months", f"ORDER BY [dateTime] ASC")
                calculations = []
                if load_from_db(self.db,"calculations") == []:
                    noCalculations = True
                else:
                    noCalculations = False

                if self.earliestChangeDate > datetime.now() and not noCalculations:
                    #if no new data exists, use old calculations
                    calculations = load_from_db(self.db,"calculations")
                    keys = list({key for row in calculations for key in row.keys()})
                    gui_queue.put( lambda: self.populate(self.calculationTable,calculations,keys = keys))
                    gui_queue.put( lambda: self.buildReturnTable())
                    gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                    gui_queue.put(lambda: self.importButton.setEnabled(True))
                    save_to_db(self.db,None,None,query="UPDATE history SET [lastImport] = ?", inputs=(self.apiCallTime,), action="replace")
                    self.lastImportLabel.setText(f"Last Data Import: {self.apiCallTime}")
                    self.lastImportDB[0]['lastImport'] = self.apiCallTime
                    print("Calculations skipped.")
                    return
                
                # proces pool section----------------------------------------------------------------
                self.workerProgress = {}

                nodeLib = nodeLibrary([*dynImportData['transactions'],*dynImportData['positions']])
                # INSERT_YOUR_CODE
                nodeClumps = get_connected_node_groups(nodeLib.nodePaths)
                clumpIdxs = {node : idx for idx, clump in enumerate(nodeClumps) for node in clump}

                # ------------------- build data cache ----------------------
                tables = mainTableNames
                table_rows = {t: dynImportData[t] for t in tables}
                cache = {}
                for table, rows in table_rows.items(): 
                    #split the data by nodes for calculations
                    for row in rows:
                        if row['Target name'] in nodeLib.targets and row['Source name'] in nodeLib.sources:
                            #investments directly from investor to fund
                            for m in months: #find the month the account balance or transaction belongs in
                                start = m["accountStart"] if table == "positions" else m["tranStart"]
                                date = row.get("Date")
                                if not (start <= date <= m["endDay"]):
                                    continue
                                cache.setdefault(-1, {}).setdefault('noNodeData', {}).setdefault(table, {}).setdefault(m["dateTime"], []).append(row)
                        else:
                            for direction in ["Target name" , "Source name", 'node']:
                                potNode = row.get(direction)
                                if potNode not in nodeLib.nodes:
                                    continue
                                else:
                                    if table == 'positions': #if the node is the source, it is below. Otherwise, above
                                        tableName = 'positions_below' if 'Source' in direction else 'positions_above'
                                    elif table == 'transactions':
                                        tableName = 'transactions_below' if 'Source' in direction else 'transactions_above'
                                    for m in months: #find the month the account balance or transaction belongs in
                                        start = m["accountStart"] if table == "positions" else m["tranStart"]
                                        date = row.get("Date")
                                        if not (start <= date <= m["endDay"]):
                                            continue
                                        cache.setdefault(clumpIdxs[potNode], {}).setdefault(potNode, {}).setdefault(tableName, {}).setdefault(m["dateTime"], []).append(row)
                self.cachedDynTables = {table : [] for table in mainTableNames}
                self.cachedLinkedCalculations = []
                if fullRecalculations:
                    self.nodeChangeDates['active'] = False #no more using cached data. Full calculations every time
                    self.earliestChangeDate = self.dataTimeStart
                if self.nodeChangeDates.get("active",False): #iterate through nodes that have custom calculation dates
                    runClumps = {idx : [] for idx in range(nodeClumps)}
                    for idx, cNodes in enumerate(nodeClumps):
                        if any(node in self.nodeChangeDates for node in cNodes) or (idx == 0 and not any(node in self.nodeChangeDates for node in nodeLib.nodes)):
                            runClumps[clumpIdxs[node]] = [{'name' : node} for node in cNodes]
                        else:
                            for node in cNodes:
                                self.cachedLinkedCalculations.extend([calcRow for _, rows in  cache[idx].get(node,{}).get("calculations", {}).items() for calcRow in rows])
                                for table in mainTableNames: #add the dynTable data to maintain the pool data and add it again after calculations
                                    if "positions" == table: #remove the duplicate account balances (EOM = next BOM)
                                        uniqueBalances = {accountBalanceKey(dynRow): dynRow for month in  cache[idx].get(node,{}).get(table, {}) for dynRow in cache[idx].get(node,{}).get(table, {}).get(month)}
                                        self.cachedDynTables[table].extend([entry for _,entry in uniqueBalances.items()])
                                    else:
                                        self.cachedDynTables[table].extend([dynRow for month in  cache[idx].get(node,{}).get(table, {}) for dynRow in cache[idx].get(node,{}).get(table, {}).get(month)])
                else:
                    runClumps = [[{'name' : node} for node in cNodes] for cNodes in nodeClumps]
                nodeCount = 0
                for clumpList in runClumps:
                    for idx, nodeDict in enumerate[dict](clumpList):
                        clumpList[idx]["cache"] = cache[clumpIdxs[nodeDict['name']]].get(nodeDict['name'])
                        if self.nodeChangeDates.get("active",False): #if the pool changes have been calculated, use it or set to current date if no changes occured
                            clumpList[idx]["earliestChangeDate"] = self.nodeChangeDates.get(nodeDict.get("name"),datetime.now())
                        else: #if pool changes have not been calculated but calculation requirements were imported, set to earliest global date
                            clumpList[idx]["earliestChangeDate"] =  self.earliestChangeDate 
                    #must calculate the entire clump from the same time frame as one node change may likely affect the others
                    clumpEarliestChangeDate = min((nodeDict['earliestChangeDate'] for nodeDict in clumpList))
                    for idx, nodeDict in enumerate[dict](clumpList):
                        clumpList[idx]["earliestChangeDate"] =  clumpEarliestChangeDate
                    newMonths = []
                    if not noCalculations: #if there are calculations, find all months before the data pull, and then pull those calculations
                        for month in months:
                            #if the calculations for the month have already been complete, pull the old data
                            if clumpEarliestChangeDate > datetime.strptime(month["endDay"], "%Y-%m-%dT%H:%M:%S"):
                                pass
                            else:
                                newMonths.append(month)
                    else:
                        newMonths = months
                    for idx, nodeDict in enumerate[dict](clumpList):
                        nodeCount += 1
                        _ = updateStatus(self, nodeDict['name'],len(newMonths), status="Initialization")
                nodeCount += 1
                _ = updateStatus(self, 'noNodeData',len(months), status="Initialization")
                def initializeWorkerPool():
                    self.manager = Manager()
                    self.lock = self.manager.Lock()
                    self.workerStatusQueue = self.manager.Queue()
                    self.workerDBqueue = self.manager.Queue()
                    self.calcFailedFlag = self.manager.Value('b', False)
                    self.cancelCalcBtn.setEnabled(True) #only allows cancelling once the lock for the db exists

                    self.pool = Pool()
                    self.futures = []
                    self.noNodeFuture = None
                    executor.submit(self.watch_db,nodeCount)

                    commonData = {"noCalculations" : noCalculations,
                                    "months" : months, "fundList" : fundList
                                    }
                    
                    self.calcStartTime = datetime.now()
                    print("Building worker pool...")
                    noNodeDataDict = {'name' : 'noNodeData', 'cache' : cache[-1].get('noNodeData',{})}
                    res = self.pool.apply_async(processInvestments, args=(noNodeDataDict, commonData,self.workerStatusQueue, self.workerDBqueue, self.calcFailedFlag))
                    self.futures.append(res)
                    for clumpData in runClumps:
                        res = self.pool.apply_async(processClump, args=(clumpData,nodeLib, commonData,self.workerStatusQueue, self.workerDBqueue, self.calcFailedFlag))
                        self.futures.append(res)
                    print("Workers all built. Processing...")
                    self.pool.close()

                    self.timer.start(int(calculationPingTime * 0.25) * 1000) #check at 0.75 the ping time to prevent queue buildup
                self.cachedNodePaths = nodeLib.nodePaths
                gui_queue.put(lambda: initializeWorkerPool()) #puts on main thread
            except Exception as e:
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                gui_queue.put(lambda: self.importButton.setEnabled(True))
                print(f"Error occured running calculations: {e}")
                print("e.args:", e.args)
                # maybe also:
                print(traceback.format_exc())
        executor.submit(initalizeCalc)
    def watch_db(self, totalNodes : int):
        while True:
            count = 0
            while not self.workerStatusQueue.empty() and count < 300:
                count += 1 #count to allow the loading bar to take the lock and update
                vars = self.workerStatusQueue.get()
                try:
                    failed = updateStatus(self, vars[0],vars[1],status=vars[2])
                    if failed:
                        self.calcFailedFlag = failed
                except Exception as e:
                    trace = traceback.format_exc()
                    print(f"Error occured while attempting to run background worker status update: {e}. \n traceback: \n {trace}")
            try:
                if self.workerProgress == {}:
                    QMessageBox.warning(self,"Calculation Issue", "Progress tracking has been deleted early. Calculations are being halted. This may result from multi clicking 'Reimport Data' before it can process.")
                    self.workerProgress = {"DummyFail" : {'pool' : 'dummyFail', 'completed' : 0, 'total' : 99, 'status' : "Failure"}}
                statusLines = [entry for _, entry in self.workerProgress.items()]
                failed = []
                completed = []
                complete = 0
                total = 0
                for line in statusLines:
                    complete += line.get("completed",0)
                    total += line.get("total",0)
                    if line["status"] == "Failed":
                        failed.append(line)
                    elif line["status"] == "Completed":
                        completed.append(line)
                if len(failed) > 0:
                    print(f"Halting progress watch due to worker '{failed[0].get("pool","Bad Pull")}' failure.")
                    print(self.workerProgress)
                    self.queue.append(-86) #will halt the queue
                    break
                elif len(completed) == totalNodes:
                    print("All workers have declared complete.")
                    self.queue.append(100) #backup in case the numbers below fail
                    break
                if total != 0:
                    percent = int((complete / total) * 100)
                    self.queue.append(percent)
                    if complete >= total:
                        break
            except Exception as e:
                print(f"Error watching database: {e}")
                print(traceback.format_exc())
                pass
            time.sleep(calculationPingTime * 0.01)
    def update_from_queue(self):
        if self.queue:
            while self.queue: #cycle through the queue options to get most up to date value. Breaks out if complete or halted
                val = self.queue.pop(0)
                if val in (-86,100):
                    break
            self.calculationLoadingBar.setValue(val)
            timeElapsed = datetime.now() - self.calcStartTime
            secsElapsed = timeElapsed.total_seconds()
            loadingFraction = float(val) / 100 #decimal format percentage
            if loadingFraction > 0:
                est_total_secs = secsElapsed / loadingFraction
                secs_remaining = est_total_secs - secsElapsed
            else:
                secs_remaining = 0
            mins, secs = divmod(int(secs_remaining), 60)
            time_str = f"{mins}m {secs}s" # format as “Xm Ys” or “MM:SS”
            self.calculationLabel.setText(f"Estimated time remaining: {time_str}")
            if val >= 100:
                self.timer.stop()
                executor.submit(self.calcCompletion)
            elif val == -86:
                self.timer.stop()
                if self.cancel:
                    QMessageBox.warning(self,"Calculation Halted", "Calculations are being halted.")
                    self.cancel = False
                else:
                    QMessageBox.warning(self,"Calculation Failure", "A worker thread has failed. Calculations will not be properly completed.")
                self.pool.terminate()
                self.pool.join()
                gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
                gui_queue.put(lambda: self.importButton.setEnabled(True))
                
    def calcCompletion(self):
        try:
            print("Checking worker completion...")
            self.pool.join()
            print("All workers finished")
            
            nodeCalculations = []
            allDynTables = {table: [] for table in mainTableNames}
            for fut in self.futures:
                try:
                    nCalcs, dynTables = fut.get()
                    nodeCalculations.extend(nCalcs)
                    for table in dynTables:
                        allDynTables[table].extend(dynTables[table])
                except Exception as e:
                    print(traceback.format_exc())
                    print(f"Error appending calculations: {e}")
            nodeCalculations.extend(self.cachedLinkedCalculations)
            for table in dynTables: #add dynamo table data in for pools that were not calculated again
                allDynTables[table].extend(self.cachedDynTables[table])
            keys = list({key for row in nodeCalculations for key in row.keys()})
            print("Updating database...")
            save_to_db(self.db,"calculations",nodeCalculations, keys=keys)
            save_to_db(self.db, "nodes", [node for _, node in self.cachedNodePaths.items()])
            executor.submit(self.db.postCalcUpdate) #make sure the cached node data is up to date
            for table in mainTableNames:
                save_to_db(self.db,table, allDynTables[table])
            print("Database updated.")
            try:
                save_to_db(self.db,None,None,query="UPDATE history SET [lastImport] = ?", inputs=(self.apiCallTime,), action="replace")
                gui_queue.put(lambda: self.lastImportLabel.setText(f"Last Data Import: {self.apiCallTime}"))
                self.lastImportDB[0]['lastImport'] = self.apiCallTime
            except Exception as e:
                gui_queue.put( lambda: QMessageBox.warning(self,"Warning",f"Failed to update internal data for last import time. Data will likely reimport soon: {e} {e.args}"))
                print(f"failed to update last import time {e} {e.args}")
            gui_queue.put(lambda: self.instantiateFilters(keepChoices=True))
            gui_queue.put( lambda: self.populate(self.calculationTable,nodeCalculations,keys = keys))
            gui_queue.put( lambda: self.buildReturnTable())
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
            gui_queue.put(lambda: self.importButton.setEnabled(True))
            print("Calculations complete.")
            self.workerProgress = {}
        except:
            gui_queue.put(lambda: self.calculationLoadingBox.setVisible(False))
            gui_queue.put(lambda: self.importButton.setEnabled(True))
            print(f"Error occured processing calculation results. Resetting... ")
            print(traceback.format_exc())
    def checkVersion(self):
        self.currentVersionAccess = False
        self.globalVersion = None
        try:
            row = load_from_db(self.db,"history")[0]
            self.globalVersion = row["currentVersion"]
            if row["currentVersion"] == currentVersion:
                self.currentVersionAccess = True
            else:
                QMessageBox.warning(self,"Outdated Version", f"Your current version ({currentVersion}) is outdated. The current version is {row["currentVersion"]}. \n" + 
                                    " Please request the newer version. \n Your version will run, but cannot access the shared data. This will be signifigantly slower and may have bugs/errors.")
        except:
            QMessageBox.warning(self,"Error checking version", f"An error occured checking that your version is up to date. This session has been granted limited access. \n "+
                                "Your session will run, but cannot access the shared data. This will be signifigantly slower and may have bugs/errors. \n \n Try restarting the app." +
                                " If this error persists, contact an admin.")
    
    def sortStyleClicked(self, *args):
        if "NAV" in self.sortStyle.text():
            self.sortStyle.setText("Sort Style: Alphabetical")
        else:
            self.sortStyle.setText("Sort Style: NAV")
        self.buildReturnTable()
    def headerSortClosed(self):
        self.populateReturnsTable(self.currentTableData, self.currentTableFlags)
    def orderColumns(self,keys, exceptions = []):
        mode = self.tableBtnGroup.checkedButton().text()
        if mode == "Monthly Table":
            dates = [datetime.strptime(k, "%B %Y") for k in keys]
            dates = sorted(dates, reverse=True)
            keys = [d.strftime("%B %Y") for d in dates]
        elif mode == "Complex Table":
            newOrder = ["%", "NAV", "Commitment", "Unfunded","MTD","QTD","YTD"] + [f"{y}YR" for y in yearOptions] + ["ITD"]
            ordered = [h for h in newOrder if h in keys]
            ordered += [h for h in keys if h not in newOrder and h not in exceptions]
            keys = ordered
        return keys
    def populateReturnsTable(self, origRows: dict, flagStruc : dict = {}):
        try:
            self.buildTableLoadingBar.setValue(7)
            mode = self.tableBtnGroup.checkedButton().text()
            if not origRows:
                # nothing to show
                self.returnsTable.clear()
                self.returnsTable.setRowCount(0)
                self.returnsTable.setColumnCount(0)
                self.buildTableLoadingBox.setVisible(False)
                return

            rows = copy.deepcopy(origRows) #prevents alteration of self.returnsTableData
            for f in self.filterOptions: #remove dataTypes the user has chosen not to see
                if f["key"] not in self.filterBtnExclusions and not self.filterRadioBtnDict[f["key"]].isChecked():
                    for k,v in rows.items():
                        if "dataType" not in v:
                            print(f"Bad row. Key: {k} \n       row: {v}")
                    to_delete = [k for k,v in rows.items() if v["dataType"] == "Total " + f["key"]]
                    for k in to_delete:
                        rows.pop(k)
            to_delete = []
            for row in rows.keys():
                row_label, _ = separateRowCode(row)
                if row_label == "hiddenLayer":
                    to_delete.append(row)
            for row in to_delete:
                rows.pop(row)
            
            self.filteredReturnsTableData = copy.deepcopy(rows) #prevents removal of dataType key for data lookup

            # 1) Build a flat list of row-entries:
            #    each entry = (fund_label, unique_code, row_dict)
            row_entries = []
            for fund_label, row_dict in rows.items():
                row_label, code = separateRowCode(fund_label)
                row_entries.append((row_label, code, row_dict, fund_label))

            # 2) Determine columns exactly as before, using cleanedRows for header order
            cleaned = {row_label: d.copy() for row_label, _, d, _ in row_entries}
            for d in cleaned.values():
                d.pop("dataType", None)
            headerSort : SortButtonWidget = self.headerSort
            currentHeaders = set(key for cRow in cleaned.values() for key in cRow.keys())
            if not headerSort.active or mode == "Monthly Table" or any(opt not in headerSort.options() and opt not in headerSortExclusions for opt in currentHeaders):
                col_keys = set()
                for d in cleaned.values():
                    col_keys |= set(d.keys())
                col_keys = list(col_keys)

                exceptions = nonDefaultHeaders
                col_keys = self.orderColumns(col_keys, exceptions=exceptions)
                if mode == "Complex Table":
                    totalRowKeys = list(cleaned[list(cleaned.keys())[0]].keys())
                    chosenKeys = [key for key in col_keys if key in (*totalRowKeys,*nonAggregatingCols) and key not in exceptions] #headers in the total, but not the exceptions
                    allKeys = chosenKeys.copy() #start the sortable options w the chosen ones
                    for keySet in (col_keys,exceptions): #extend allKeys by the ones not chosen for later selection option
                        allKeys.extend([key for key in keySet if key not in allKeys])
                    headerSort.set_items(allKeys,chosenKeys)
                    headerSort.setEnabled(True)
                    col_keys = chosenKeys
                else:
                    totalRowKeys = list(cleaned[list(cleaned.keys())[0]].keys())
                    if totalRowKeys: #only if the values are aggregated to the total
                        col_keys = [key for key in col_keys if key in totalRowKeys and key not in exceptions] #prevents benchmarks alone extending the tables
                    headerSort.setEnabled(False)
            else:
                col_keys = headerSort.popup.get_checked_sorted_items()
                headerSort.setEnabled(True)
            self.filteredHeaders = col_keys
            # 3) Resize & set horizontal headers (we no longer call setVerticalHeaderLabels)
            self.returnsTable.setRowCount(len(row_entries))
            self.returnsTable.setColumnCount(len(col_keys))
            self.returnsTable.setHorizontalHeaderLabels(col_keys)

            bg = None
            # 4) Populate each row
            colorDepths = [code.count("::") + 1 * (row_dict['dataType'] == 'Total Target name') for _, code, row_dict,_ in row_entries]
            maxDepth = max(colorDepths)
            fundsPresent = any(row_dict['dataType'] == 'Total Target name' for _, _, row_dict,_ in row_entries)
            trackIdx = 0
            trackDepth = 0
            for i in range(len(colorDepths)):
                d = colorDepths[i]
                if trackDepth < d: #further depth
                    trackIdx = i
                    trackDepth = d
                if fundsPresent and (trackDepth > d or i == len(colorDepths) - 1) and d != maxDepth: #back up  from depth or the end, but did not go full depth
                    if i == len(colorDepths) - 1:
                        i += 1
                    colorDepths[trackIdx:i] = [maxDepth] * (i-trackIdx) #set the depth for low section all the way down
                trackDepth = d
            if not fundsPresent:
                maxDepth += 1 #if funds are off, don't allow upper sorts to be white
            colorDepths = [c/maxDepth for c in colorDepths] if maxDepth != 0 else colorDepths
            self.tableColorDepths = colorDepths
            for r, (fund_label, code, row_dict, rowKey) in enumerate(row_entries):
                # pull & remove dataType for coloring
                dataType = row_dict.pop("dataType", "")
                if dataType != "benchmark": #benchmark will use previous rounds color
                    startColor = (160, 160, 160)
                    if dataType == "Total":
                        color = tuple(
                            int(startColor[i] * 0.8)
                            for i in range(3)
                        )
                        bg =  QColor(*color)
                    else:
                        cRange     = 255 - startColor[0]
                        color = tuple(
                            int(startColor[i] + cRange * colorDepths[r])
                            for i in range(3)
                        )
                        bg = QColor(*color)
                
                    

                # — vertical header: only show the fund, stash the code —
                hdr = QTableWidgetItem(fund_label)
                hdr.setData(Qt.UserRole, code)
                if dataType not in  ("Total Target name","benchmark"):
                    font = hdr.font()
                    font.setBold(True)
                    hdr.setFont(font)
                if bg:
                    hdr.setBackground(QBrush(bg))
                    if dataType == "benchmark":
                        hdr.setBackground(QBrush(QColor("0000FF")))
                self.returnsTable.setVerticalHeaderItem(r, hdr)

                # — fill cells —
                for c, col in enumerate(col_keys):
                    raw = row_dict.get(col, "")
                    if raw not in (None, "", "None"):
                        try:
                            v = round(float(raw), 2)
                            if c in percent_headers or (mode == "Monthly Table" and self.returnOutputType.currentText() in percent_headers):
                                text = f"{v:.2f}%"
                            else:
                                text = f"{v:,.2f}"
                        except:
                            text = str(raw)
                    else:
                        text = ""

                    item = QTableWidgetItem(text)
                    if text:
                        # store raw number for sorting or later retrieval
                        item.setData(Qt.UserRole, v)
                    if bg:
                        if flagStruc.get(rowKey,{}).get(col,False):
                            colorL = (color[0], color[1], int(color[2] * 0.8))
                            item.setBackground(QBrush(QColor(*colorL))) #yellow tints the cell for ownership adjustment
                        else:
                            item.setBackground(QBrush(bg))
                    if dataType == "benchmark":
                        item.setForeground(QColor(0,0,255))
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.returnsTable.setItem(r, c, item)
            self.buildTableLoadingBox.setVisible(False)
        except Exception as e:
            QMessageBox.warning(self,'Build Table Failed',f'Error occured attempting to format the table. Please try again. \n {e.args} {traceback.format_exc()}')
    def populate(self, table, rows, keys = None):
        if not rows:
            return
        if keys is None:
            headers = list(rows[0].keys())
        else:
            headers = list(keys)

        self.calcTableModel = DictListModel(rows,headers, self)
        table.setModel(self.calcTableModel)

