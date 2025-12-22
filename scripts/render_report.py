import sys
from datetime import datetime
from pathlib import Path

import markdown
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import pandas as pd
from jinja2 import Environment, FileSystemLoader, select_autoescape

try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except OSError:
    print("Warning: WeasyPrint (GTK) not available. PDF generation will be skipped.")
    WEASYPRINT_AVAILABLE = False
except ImportError:
    print("Warning: WeasyPrint not installed. PDF generation will be skipped.")
    WEASYPRINT_AVAILABLE = False


def get_base_dir():
    # Script mode
    if hasattr(sys.modules["__main__"], "__file__"):
        return Path(sys.modules["__main__"].__file__).resolve().parent
    # Notebook / Spyder / IPython mode
    return Path(".").resolve()


def format_number(value):
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return f"{value:,.2f}"  


def get_holdings_row_class(row_type):
    """Map type to CSS class for holdings table rows."""
    type_lower = str(row_type).lower().strip()
    
    if type_lower == 'total':
        return 'holdings-total'
    elif type_lower == 'category' or type_lower == 'subclass':
        return 'holdings-subclass'
    elif type_lower == 'pool':
        return 'holdings-pool'
    elif type_lower == 'benchmark':
        return 'holdings-benchmark'
    else:
        return 'holdings-investment'  # fallback


def create_page_groups(portfolio_holdings_rows, rows_per_page=32, first_page_rows=32):
    """
    Split rows into page groups with a fixed number of rows per page.
    first_page_rows: Number of rows for the very first page.
    rows_per_page: Number of rows for all subsequent pages.
    """
    pages = []
    current_page = []
    
    # Use the first page limit initially
    current_limit = first_page_rows
    
    for i, row in enumerate(portfolio_holdings_rows):
        # Check if adding this row would exceed the current limit
        if len(current_page) >= current_limit:
            pages.append(current_page)
            current_page = [row]
            # After the first page is filled, switch to the standard limit for all subsequent pages
            current_limit = rows_per_page
        else:
            current_page.append(row)
    
    # Add the last page if it has any rows
    if current_page:
        pages.append(current_page)
    
    return pages


def create_benchmark_chart(benchmarks_df, column, title, output_path, y_lim=None, colors=None):
    """
    Create a vertical bar chart for benchmark returns.
    
    benchmarks_df: DataFrame with Benchmark column and return columns
    column: Column name to plot (e.g., 'MTD', '1Y')
    title: Chart title
    output_path: Path to save the chart image
    y_lim: Tuple (ymin, ymax) for shared y-axis limits
    colors: Optional list of colors for each bar
    """
    # Filter out rows with NaN values for this column
    data = benchmarks_df[['Benchmark', column]].copy()
    data = data.dropna(subset=[column])
    
    if len(data) == 0:
        return None
    
    # Default colors matching the image description
    # MSCI ACWI: dark blue, Barclays: light gray, 60/40: dark red, MSCI EM: yellow, CPPI: teal, PE: gray
    if colors is None:
        color_map = {
            'MSCI ACWI': '#1f4e78',
            'Barclays US Aggregate': '#d0d0d0',
            '60% MSCI ACWI / 40% BB Aggregate': '#8b0000',
            'MSCI EM': '#ffd700',
            'Commercial Property Prices Index': '#008080',
            'PE Benchmark': '#808080'
        }
        colors = [color_map.get(bench, '#808080') for bench in data['Benchmark']]
    
    fig, ax = plt.subplots(figsize=(6, 5))
    
    # Convert values to percentages (multiply by 100)
    data_percent = data[column] * 100
    
    # Create vertical bar chart
    x_pos = range(len(data))
    bars = ax.bar(x_pos, data_percent, color=colors)
    
    # Format axes
    ax.set_title(title, fontsize=10, fontweight='bold', pad=20)
    ax.set_xticks(x_pos)
    ax.set_xticklabels(data['Benchmark'], rotation=45, ha='right', fontsize=8)
    
    # Remove y-axis display
    ax.set_yticks([])
    ax.spines['left'].set_visible(False)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    # Calculate max absolute value for positioning labels (in percentage)
    max_abs_val = data_percent.abs().max()
    if y_lim:
        # y_lim is already in percentage format (multiplied by 100)
        max_abs_val = max(abs(y_lim[0]), abs(y_lim[1]))
    
    # Add value labels on bars (with 1 decimal place)
    for i, (bar, val) in enumerate(zip(bars, data_percent)):
        if val >= 0:
            ax.text(i, val + max_abs_val * 0.02, f'{val:.1f}%', 
                   ha='center', fontsize=8, va='bottom')
        else:
            ax.text(i, val - max_abs_val * 0.02, f'({abs(val):.1f}%)', 
                   ha='center', fontsize=8, va='top', color='red')
    
    # Add zero line
    ax.axhline(y=0, color='black', linewidth=0.5)
    
    # Set y-axis limits (use shared limits if provided, already in percentage)
    if y_lim:
        ax.set_ylim(y_lim)
    else:
        # Calculate dynamic limits based on actual data
        min_val = data_percent.min()
        max_val = data_percent.max()
        
        # Add dynamic padding based on actual min/max values
        padded_min = min_val - abs(min_val) * 0.05
        y_min = min(padded_min, -3.0)
        
        padded_max = max_val + abs(max_val) * 0.05
        y_max = max(padded_max, 3.0)
        
        ax.set_ylim(y_min, y_max)
    
    # Adjust layout
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    return output_path


def render_report(workbook = None):
    print(f"Report generation started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    base_dir = get_base_dir()
    print("Base dir:", base_dir)

    # 1) Load sheets
    if not workbook:
        from openpyxl import load_workbook
        excel_path = base_dir / "HighLevelPortfolioSnapshot.xlsx"
        dfs = load_workbook(excel_path)
    else:
        dfs = workbook

    # 2) Assets and Flows
    af = dfs["assets_and_flows"].copy()
    af = af.iloc[:, :4]
    af.columns = ["label", "month", "year_1", "inception"]
    assets_and_flows_rows = af.to_dict(orient="records")

    # 3) Portfolio Returns — read directly from Excel with the real sheet name
    #    First, find the sheet whose normalized name matches "portfolio_returns"
    xl = pd.ExcelFile(excel_path)

    target_norm = "portfolio_returns"
    sheet_name_match = None
    for sn in xl.sheet_names:
        norm = sn.strip().lower().replace(" ", "_")
        if norm == target_norm:
            sheet_name_match = sn
            break

    if sheet_name_match is None:
        raise ValueError(
            f"Could not find a sheet matching '{target_norm}'. "
            f"Available sheets: {xl.sheet_names}"
        )


    # Read that sheet. Usually header row is the first row with
    # ['', 'Month', 'QTD', 'YTD', '1 Year', '3 Year', 'Inception']
    pr = xl.parse(sheet_name_match, header=0)
    # Keep first 7 columns: label + 6 metrics
    pr = pr.iloc[:, :7]
    pr.columns = [
        "label",
        "month",
        "qtd",
        "ytd",
        "one_year",
        "three_year",
        "inception",
    ]
    portfolio_returns_rows = pr.to_dict(orient="records")

    # 3b) Overall Family Breakdown (right panel)
    ofb = dfs["overall_family_breakdown"].copy()

    # Assume first 5 cols: Asset, LM $MM, Δ, CM $MM, %
    ofb = ofb.iloc[:, :5]
    ofb.columns = ["asset", "lm_mm", "delta_mm", "cm_mm", "pct"]
    overall_family_breakdown_rows = ofb.to_dict(orient="records")

    # 3c) HF Foundations
    # Find key that matches 'foundations' case-insensitively
    foundations_key = next((k for k in dfs.keys() if k.lower() == "foundations"), None)
    if foundations_key:
        hff = dfs[foundations_key].copy()
        # Expected columns: Asset Class, $MM, % Allocation
        # Take first 3 columns
        hff = hff.iloc[:, :3]
        hff.columns = ["asset_class", "mm", "pct"]
        hf_foundations_rows = hff.to_dict(orient="records")
    else:
        print("Warning: 'foundations' sheet not found in loaded keys:", list(dfs.keys()))
        hf_foundations_rows = []

    # 3d) Sports (under Overall Family Breakdown)
    # Look for "overall_family_breakdown2"
    sports_key = next((k for k in dfs.keys() if "breakdown2" in k.lower()), None)
    if sports_key:
        sprT = dfs[sports_key].copy()
        # Expected columns: Sports, Share %, Team Value, Debt, Equity, Family Share
        # Take first 6 columns
        sprT = sprT.iloc[:, :6]
        sprT.columns = ["sports", "share_pct", "team_value", "debt", "equity", "family_share"]
        sports_rows = sprT.to_dict(orient="records")
    else:
        print("Warning: 'overall_family_breakdown2' sheet not found in loaded keys:", list(dfs.keys()))
        sports_rows = []

    # 3e) Returns vs Benchmark
    rvb_key = next((k for k in dfs.keys() if "returns_vs_benchmark" in k.lower()), None)
    if rvb_key:
        rvb = dfs[rvb_key].copy()
        # Expecting ~20 columns
        # 0: Asset Class, 1: $MM
        # 2-4: Alloc vs Target (%, Tgt %, Delta $)
        # 5-7: Month (RTN, BM, Delta)
        # 8-10: YTD
        # 11-13: 1 Year
        # 14-16: 3 Year
        # 17-19: Inception
        rvb = rvb.iloc[:, :20]
        rvb.columns = [
            "asset_class", "mm",
            "alloc_pct", "tgt_pct", "alloc_delta",
            "m_rtn", "m_bm", "m_delta",
            "ytd_rtn", "ytd_bm", "ytd_delta",
            "y1_rtn", "y1_bm", "y1_delta",
            "y3_rtn", "y3_bm", "y3_delta",
            "inc_rtn", "inc_bm", "inc_delta"
        ]
        returns_vs_benchmark_rows = rvb.to_dict(orient="records")
    else:
        print("Warning: 'returns_vs_benchmark' sheet not found in loaded keys:", list(dfs.keys()))
        returns_vs_benchmark_rows = []

    # 3f) Portfolio Holdings (Pages 9-18)
    holdings_path = base_dir / "HF Capital Portfolio Holdings.xlsx"
    format_path = base_dir / "Portfolio Holding Format.xlsx"
    
    # First, load the format mapping from the separate format file
    type_mapping = {}
    if format_path.exists():
        try:
            dfs_format = load_workbook(format_path)
            # Get the first sheet (or look for a specific sheet name)
            format_key = list(dfs_format.keys())[0]
            format_df = dfs_format[format_key].copy()
            # Expected columns: Entry, Type (or similar)
            # Take first 2 columns
            format_df = format_df.iloc[:, :2]
            format_df.columns = ["entry", "type"]
            # Create mapping: entry -> type
            for _, row in format_df.iterrows():
                entry = str(row.get("entry", "")).strip()
                entry_type = str(row.get("type", "")).strip().title()
                if entry and entry_type and entry.lower() != "entry":
                    type_mapping[entry] = entry_type
            print(f"Loaded {len(type_mapping)} entries from format mapping file")
        except Exception as e:
            print(f"Warning: Error loading format file '{format_path}': {e}")
    else:
        print(f"Warning: Format file '{format_path}' not found.")
    
    if holdings_path.exists():
        try:
            dfs_holdings = load_workbook(holdings_path)
            
            # Load the actual holdings data from "Sheet"
            holdings_key = "Sheet"
            if holdings_key not in dfs_holdings:
                # Try to find it case-insensitively or use first sheet
                holdings_key = next((k for k in dfs_holdings.keys() if k.lower() == "sheet"), None)
                if holdings_key is None:
                    holdings_key = list(dfs_holdings.keys())[0]
                    print(f"Warning: 'Sheet' not found, using first sheet: '{holdings_key}'")
            
            holdings_df = dfs_holdings[holdings_key].copy()

            
            # Expected columns: Asset Class/Investment, %, NAV, MTD, YTD, 1YR, 3YR, ITD
            holdings_df = holdings_df[['Asset_Class_Investment', 'Pct', 'NAV', 'MTD', 'YTD', '1YR', '3YR', 'ITD']]
            holdings_df.columns = [
                "investment", "pct_alloc", "nav", "mtd", "ytd",
                "one_year", "three_year", "itd"
            ]
            
            # Convert to dict and add type information from mapping
            portfolio_holdings_rows = holdings_df.to_dict(orient="records")
            
            for row in portfolio_holdings_rows:
                investment = str(row.get("investment", "")).strip()
                # Look up type in mapping
                row["type"] = type_mapping.get(investment, "Investment")  # Default to "Investment" if not found
                row["type"] = str(row["type"]).strip().title()  # Normalize
                # Add CSS class based on type
                row["row_class"] = get_holdings_row_class(row["type"])
                # Determine if border should appear below (for Total and Category rows)
                row["show_border"] = row["type"].lower() in ['total', 'category', 'subclass']
            
            # Split rows into page groups
            portfolio_holdings_pages = create_page_groups(portfolio_holdings_rows)
            
            print(f"Loaded {len(portfolio_holdings_rows)} portfolio holdings rows")
            print(f"Split into {len(portfolio_holdings_pages)} pages")
        except Exception as e:
            print(f"Warning: Error loading portfolio holdings: {e}")
            import traceback
            traceback.print_exc()
            portfolio_holdings_pages = []
    else:
        print(f"Warning: '{holdings_path}' not found.")
        portfolio_holdings_pages = []

    # 3g) Benchmarks (World Markets page)
    benchmarks_path = base_dir / "benchmarks.xlsx"
    benchmarks_rows = []
    benchmark_chart_mtd_path = None
    benchmark_chart_1y_path = None
    
    if benchmarks_path.exists():
        try:
            benchmarks_df = pd.read_excel(benchmarks_path, sheet_name=0)
            # Expected columns: Benchmark, Info, MTD, QTD, YTD, 1Y, 3Y, 5Y, 10Y
            benchmarks_df = benchmarks_df[['Benchmark', 'Info', 'MTD', 'QTD', 'YTD', '1Y', '3Y', '5Y', '10Y']]
            benchmarks_rows = benchmarks_df.to_dict(orient="records")
            
            # Create output directory for charts
            output_dir = base_dir / "output"
            output_dir.mkdir(exist_ok=True)
            
            # Calculate shared y-axis limits across both MTD and 1Y columns
            # Multiply by 100 to convert from decimal to percentage
            mtd_data = benchmarks_df['MTD'].dropna() * 100
            y1_data = benchmarks_df['1Y'].dropna() * 100
            
            if len(mtd_data) > 0 or len(y1_data) > 0:
                all_values = pd.concat([mtd_data, y1_data])
                min_val = all_values.min()
                max_val = all_values.max()
                
                # Add dynamic padding based on actual min/max values
                # Use smaller padding (5%) and only where needed
                padded_min = min_val - abs(min_val) * 0.05
                y_min = min(padded_min, -3.0)
                
                padded_max = max_val + abs(max_val) * 0.05
                y_max = max(padded_max, 3.0)
                
                shared_y_lim = (y_min, y_max)
            else:
                shared_y_lim = None
            
            # Create bar charts with shared y-axis
            benchmark_chart_mtd_path = create_benchmark_chart(
                benchmarks_df, 'MTD', 'Benchmarks, Month-To-Date Returns',
                output_dir / "benchmark_chart_mtd.png",
                y_lim=shared_y_lim
            )
            benchmark_chart_1y_path = create_benchmark_chart(
                benchmarks_df, '1Y', 'Benchmarks, 1 Year Returns',
                output_dir / "benchmark_chart_1y.png",
                y_lim=shared_y_lim
            )
            
            print(f"Loaded {len(benchmarks_rows)} benchmark rows")
            print(f"Created benchmark charts")
        except Exception as e:
            print(f"Warning: Error loading benchmarks: {e}")
            import traceback
            traceback.print_exc()
            benchmarks_rows = []
    else:
        print(f"Warning: '{benchmarks_path}' not found.")
        benchmarks_rows = []

    # 3h) JPM LOC Data
    jpm_loc_path = base_dir / "JMP LOC.xlsx"
    jpm_usage_rows = []
    jpm_collateral_rows = []
    jpm_letters_of_credit = None

    if jpm_loc_path.exists():
        try:
            print(f"Loading JPM LOC data from {jpm_loc_path}...")
            # Load 'Current Usage' sheet
            usage_df = pd.read_excel(jpm_loc_path, sheet_name=0) # Assuming first sheet
            # It seems the sheet might have headers on row 0. Based on screenshot:
            # Type, Size, Cost, Amount Drawn
            # 'Letters of Credit' is likely a separate row below 'Total' or a separate variable.
            # Let's clean the DF.
            
            # Look for 'Type' in columns to identify header row if needed, but assuming standard format
            usage_df.columns = [str(c).strip() for c in usage_df.columns]
            
            # Filter main table rows
            # Rows usually: Secured, Unsecured, Total
            # And maybe 'Letters of Credit' is in there too
            
            # We want to extract 'Letters of Credit' specifically
            loc_row = usage_df[usage_df.iloc[:, 0].astype(str).str.contains("Letters of Credit", case=False, na=False)]
            if not loc_row.empty:
                # Assuming the value is in the 'Size' column (2nd column, index 1)
                jpm_letters_of_credit = loc_row.iloc[0, 1]
            
            # Usage table rows (exclude Letters of Credit if it was in the same table)
            usage_table = usage_df[~usage_df.iloc[:, 0].astype(str).str.contains("Letters of Credit", case=False, na=False)].copy()
            # Rename columns to standard keys
            # Expecting: Type, Size, Cost, Amount Drawn
            # Map by position to be safe if names vary slightly
            if len(usage_table.columns) >= 4:
                usage_table = usage_table.iloc[:, :4]
                usage_table.columns = ["type", "size", "cost", "amount_drawn"]
                jpm_usage_rows = usage_table.to_dict(orient="records")
            else:
                print("Warning: JPM Usage table has fewer than 4 columns")

            # Load 'Current Collateral' sheet
            collateral_df = pd.read_excel(jpm_loc_path, sheet_name=1) # Assuming second sheet
            # Expected columns: Collateral Account, Market Value, Lending Value, %
            if len(collateral_df.columns) >= 4:
                collateral_df = collateral_df.iloc[:, :4]
                collateral_df.columns = ["account", "market_value", "lending_value", "pct"]
                jpm_collateral_rows = collateral_df.to_dict(orient="records")
            else:
                print("Warning: JPM Collateral table has fewer than 4 columns")
                
            print(f"Loaded {len(jpm_usage_rows)} JPM usage rows")
            print(f"Loaded {len(jpm_collateral_rows)} JPM collateral rows")
            
        except Exception as e:
            print(f"Warning: Error loading JPM LOC data: {e}")
            # Don't print full traceback for permission error if expected
            if "Permission denied" not in str(e):
                import traceback
                traceback.print_exc()
    else:
        print(f"Warning: '{jpm_loc_path}' not found.")

    # 4) Load Narrative
    narrative_path = base_dir / "narrative.md"
    if narrative_path.exists():
        narrative_text = narrative_path.read_text(encoding="utf-8")
        narrative_html = markdown.markdown(narrative_text)
    else:
        print(f"Warning: '{narrative_path}' not found.")
        narrative_html = "<p>No narrative found.</p>"

    # 4b) Load JPM Commentary
    jpm_commentary_path = base_dir / "jpm_commentary.md"
    if jpm_commentary_path.exists():
        jpm_text = jpm_commentary_path.read_text(encoding="utf-8")
        jpm_commentary_html = markdown.markdown(jpm_text)
    else:
        print(f"Warning: '{jpm_commentary_path}' not found.")
        jpm_commentary_html = "<p>No JPM commentary found.</p>"

    # 5) Templates
    templates_dir = base_dir / "templates"

    env = Environment(
        loader=FileSystemLoader(str(templates_dir)),
        autoescape=select_autoescape(["html", "xml"]),
    )
    env.filters["format_number"] = format_number
    env.filters["is_nan"] = lambda x: pd.isna(x) if isinstance(x, (float, int)) or x is None else False

    template = env.get_template("snapshot.html")

    # 6) Render (single call – include all panels)
    # Convert chart paths to relative paths for template
    chart_mtd_rel = "output/benchmark_chart_mtd.png" if benchmark_chart_mtd_path else None
    chart_1y_rel = "output/benchmark_chart_1y.png" if benchmark_chart_1y_path else None
    
    html_str = template.render(
        as_of="September 30, 2025",
        assets_and_flows=assets_and_flows_rows,
        portfolio_returns=portfolio_returns_rows,
        overall_family_breakdown=overall_family_breakdown_rows,
        hf_foundations=hf_foundations_rows,
        sports=sports_rows,
        returns_vs_benchmark=returns_vs_benchmark_rows,
        benchmarks=benchmarks_rows,
        benchmark_chart_mtd=chart_mtd_rel,
        benchmark_chart_1y=chart_1y_rel,
        portfolio_holdings_pages=portfolio_holdings_pages,
        narrative_html=narrative_html,
        jpm_usage=jpm_usage_rows,
        jpm_collateral=jpm_collateral_rows,
        jpm_letters_of_credit=jpm_letters_of_credit,
        jpm_commentary_html=jpm_commentary_html,
    )

    # 7) Output
    output_dir = base_dir / "output"
    output_dir.mkdir(exist_ok=True)
    out_path = output_dir / "snapshot.html"
    out_path.write_text(html_str, encoding="utf-8")
    print(f"Wrote {out_path}")

    # 7) PDF Output
    if WEASYPRINT_AVAILABLE:
        pdf_path = output_dir / "snapshot.pdf"
        try:
            HTML(string=html_str, base_url=str(base_dir)).write_pdf(pdf_path)
            print(f"Wrote {pdf_path}")
        except Exception as e:
            print(f"Error generating PDF: {e}")
    else:
        print("Skipped PDF generation (WeasyPrint not available)")

    print(f"Report generation completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    render_report()
