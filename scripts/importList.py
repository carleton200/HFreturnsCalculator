import sys
import os
import json
import subprocess
import traceback
import sqlite3
from pandas.core.apply import com
import requests
import calendar
import warnings
import pandas as pd
import time
import copy
import re
import pyxirr
import pyodbc
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, wait
import queue
import threading
import logging, functools
from dateutil.relativedelta import relativedelta
from multiprocessing import Pool, freeze_support, Manager
from PyQt5.QtWidgets import (
    QApplication, QWidget, QStackedWidget, QVBoxLayout,
    QLabel, QLineEdit, QPushButton, QFormLayout,
    QRadioButton, QButtonGroup, QComboBox, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QProgressBar, QTableView, QCheckBox, QMessageBox,
    QScrollArea, QFileDialog, QListWidgetItem, QListWidget, QDialog, QSizePolicy, QGridLayout,
    QFrame, QTextEdit, QHeaderView, QDateEdit, QSplitter
)
from PyQt5.QtGui import QBrush, QColor, QDesktopServices
from PyQt5.QtCore import Qt, QTimer, QAbstractTableModel, QModelIndex, pyqtSignal, QPoint, QUrl, QDate