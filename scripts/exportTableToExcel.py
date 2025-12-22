import traceback
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import QUrl
from scripts.instantiate_basics import gui_queue, executor
from scripts.commonValues import percent_headers
def exportTableToExcel(self, rows, headers = None):
    #Excel export for the generic table window
    # 1) prompt user
    path, _ = QFileDialog.getSaveFileName(
        self, "Save as…", "", "Excel Files (*.xlsx)"
    )
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"            
    def exportTableToExcel():
        try:
            data = rows  # list of dicts

            if headers is None:
                all_cols = set()
                for row in rows:
                    for key in row:
                        all_cols.add(key)
                all_cols = list(all_cols)
            else:
                all_cols = headers

            # 4) create workbook
            wb = Workbook()
            ws = wb.active

            rowStart = 1
            for idx, colname in enumerate(all_cols, start=1):
                ws.cell(row=rowStart, column=idx, value=colname)


            # 7) populate rows
            for r, row_dict in enumerate(data, start=rowStart + 1):



                # data cells with proper formatting
                for c, colname in enumerate(all_cols, start=1):
                    val = row_dict.get(colname, None)
                    try: #make numerical format if possible
                        val = float(val)
                    except:
                        pass
                    cell = ws.cell(row=r, column=c, value=val)
                    if isinstance(val, (int, float)):
                        if colname not in percent_headers:
                            # show with commas, two decimals
                            cell.number_format = "#,##0.00"
                        else:
                            # interpret val as percentage (e.g. 10.5 → 10.5%)
                            cell.value = val / 100.0
                            cell.number_format = "0.00%"

            # 8) autofit column widths
            for idx, col_cells in enumerate(ws.columns, start=1):
                max_len = 0
                for cell in col_cells:
                    if cell.value is not None:
                        text = str(cell.value)
                        max_len = max(max_len, len(text))
                ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
        except Exception as e:
            gui_queue.put(lambda error = e, trace = traceback.format_exc(): QMessageBox.critical(self, "Processing error", f"{error} \n {trace}"))
        try:
            wb.save(path)
        except Exception as e:
            gui_queue.put(lambda error = e: QMessageBox.critical(self, "Save error", str(error)))
        else:
            gui_queue.put(lambda: QMessageBox.information(self, "Saved", f"Excel saved to:\n{path}"))
            gui_queue.put(lambda: QDesktopServices.openUrl(QUrl.fromLocalFile(path)))
    executor.submit(exportTableToExcel)
